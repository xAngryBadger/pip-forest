"""Excel export — timeline, cascata, occupancy, color styling, profiles, batch consolidation."""

import json
import math
import os
from collections import defaultdict

import pandas as pd

from .config import salvar_config, OUTPUT_DIR, PERFIS_DIR
from .constants import CT317_HARDCODE_HH_BASE, _FASE_CORES_HEX
from .text_utils import normalizar_chave, atividades_por_filtro, _slug_ficheiro_seguro
from .tarifas import resolver_chave_tarifa
from .scheduler import (
    eh_limpeza_quimica_pos_plantio,
    _match_filtros_fase,
    _fases_ordem_config,
)
from .turmas import menu_vincular_atividades_turma
from .datas import (
    _converter_dia_simulado_para_data,
    _DIAS_SEMANA_CURTO,
    _DIAS_SEMANA_COMPLETO,
    _formatar_data_dia,
)
from .ui import (
    G, Y, C, DM, BL, RS,
    sub, aviso, ok, prompt, selecionar,
    linha, pedir_int,
)


def _fase_nome_pt(fase_id):
    m = {
        "rocada": "Rocada",
        "formiga": "Formiga",
        "coroamento": "Coroamento",
        "coveamento": "Coveamento",
        "adubacao_quimica": "Adubacao quim.",
        "plantio": "Plantio",
        "irrigacao": "Irrigacao",
        "limpeza_quimica": "Limpeza quim.",
        "demais": "Demais",
        "reforco": "Reforco",
        "pool": "Pelotao unif.",
    }
    return m.get(fase_id, str(fase_id).capitalize())


def _classificar_fase_nome(atv, seq_cfg, modo, atvs_plantio, atvs_irrig):
    """Retorna (fase_id, fase_valor) para rotulagem no Excel."""
    from collections import OrderedDict as _OD

    if eh_limpeza_quimica_pos_plantio(atv, seq_cfg):
        return "limpeza_quimica", 8.0
    if atv in atvs_plantio or _match_filtros_fase(
        atv, seq_cfg.get("filtros_plantio") or ["plantio"], None
    ):
        return "plantio", 6.0
    if atv in atvs_irrig or _match_filtros_fase(
        atv, seq_cfg.get("filtros_irrigacao") or ["irrig"], None
    ):
        return "irrigacao", 7.0
    fases = _fases_ordem_config(seq_cfg, modo)
    for i, fase in enumerate(fases):
        if _match_filtros_fase(atv, fase.get("filtros") or [], fase.get("exclusoes")):
            return fase.get("id", f"fase_{i}"), float(i)
    return "demais", 5.5


def _gerar_aba_timeline(cronograma, seq_cfg, modo_seq, atividades_reais, fazenda, dia_ref=None, mes_ref=None, ano_ref=None):
    """Retorna DataFrame para aba TIMELINE_CASCATA com colunas de visualização.
    
    Se dia_ref, mes_ref, ano_ref forem fornecidos, adiciona colunas de data real.
    """
    atvs_plantio = set(
        atividades_por_filtro(
            atividades_reais, seq_cfg.get("filtros_plantio") or ["plantio"]
        )
    )
    atvs_irrig = set(
        atividades_por_filtro(
            atividades_reais, seq_cfg.get("filtros_irrigacao") or ["irrig"]
        )
    )
    rows = []
    for c in cronograma:
        atv = c.get("Atividade", "")
        fase_id, fase_val = _classificar_fase_nome(
            atv, seq_cfg, modo_seq, atvs_plantio, atvs_irrig
        )
        modo_exec = c.get("Modo", "Normal")
        if modo_exec == "Reforco":
            fase_id_display = "reforco"
        elif modo_exec == "PoolPosBloqueio":
            fase_id_display = "pool"
        else:
            fase_id_display = fase_id
        
        dia_simulado = c.get("Dia", 1)
        
        # Calcular data real se parametros fornecidos
        data_real = None
        dia_semana = ""
        if dia_ref and mes_ref and ano_ref:
            data_tuple = _converter_dia_simulado_para_data(
                dia_simulado, dia_ref, mes_ref, ano_ref
            )
            if data_tuple:
                data_real = data_tuple[0]  # "20/04/2025"
                dia_semana = data_tuple[1]  # "Seg"
        
        row = {
            "Dia": dia_simulado,
            "Semana": int(math.ceil(float(dia_simulado) / 5.0)),
            "Fazenda": fazenda,
            "Talhao": c.get("Talhao", ""),
            "Atividade": atv,
            "Fase": _fase_nome_pt(fase_id),
            "Fase_ID": fase_id,
            "Fase_Ordem": fase_val,
            "Turma": c.get("Turma", ""),
            "Operarios": c.get("Operarios", 0),
            "HH": c.get("HH", 0),
            "Custo_MO": c.get("Custo_MO", 0),
            "Modo": modo_exec,
            "Cor_Hex": _FASE_CORES_HEX.get(fase_id_display, "BDC3C7"),
        }
        
        # Adicionar colunas de data real se calculadas
        if data_real:
            row["Data"] = data_real
            row["Dia_Semana"] = dia_semana
            
        rows.append(row)
    
    df = pd.DataFrame(rows) if rows else pd.DataFrame()
    
    # Reordenar colunas: Data, Dia_Semana primeiro se existirem
    if not df.empty and "Data" in df.columns:
        cols = ["Data", "Dia_Semana"] + [c for c in df.columns if c not in ["Data", "Dia_Semana"]]
        df = df[cols]
    
    return df


def _gerar_aba_cascata_explicada(cronograma, jornada, dia_ref=None, mes_ref=None, ano_ref=None):
    """
    Trilha explicativa da cascata por dia/turma/atividade.
    Mostra capacidade, consumo, saldo e pendencia (carry-over) de forma didatica.
    Se dia_ref, mes_ref, ano_ref forem fornecidos, adiciona colunas de data real.
    """
    if not cronograma:
        return pd.DataFrame()

    rows_src = []
    for i, c in enumerate(cronograma):
        try:
            dia = int(c.get("Dia", 0) or 0)
            hh = float(c.get("HH", 0) or 0.0)
            ops = float(c.get("Operarios", 0) or 0.0)
        except (TypeError, ValueError):
            continue
        if dia <= 0 or hh <= 0.0:
            continue
        rows_src.append(
            {
                "_ord": i,
                "Dia": dia,
                "Semana": int(math.ceil(float(dia) / 5.0)),
                "Fazenda": c.get("Fazenda", ""),
                "Talhao": c.get("Talhao", ""),
                "Atividade": c.get("Atividade", ""),
                "Turma": c.get("Turma", ""),
                "Operarios": ops,
                "HH": hh,
            }
        )
    if not rows_src:
        return pd.DataFrame()

    # demanda total por atividade/talhao/turma para calcular pendencia durante o consumo
    demanda_total = defaultdict(float)
    for r in rows_src:
        k = (str(r["Talhao"]), str(r["Atividade"]), str(r["Turma"]))
        demanda_total[k] += float(r["HH"])

    df_rows = (
        pd.DataFrame(rows_src)
        .sort_values(["Dia", "Turma", "_ord"])
        .reset_index(drop=True)
    )
    out = []
    consumido_atividade = defaultdict(float)

    for (dia, turma), grp in df_rows.groupby(["Dia", "Turma"], sort=True):
        ops_dia = max(float(x) for x in grp["Operarios"].tolist()) if len(grp) else 0.0
        cap_dia = max(0.0, float(ops_dia) * float(jornada))
        usado_dia = 0.0
        
        # Calcular data real se parametros fornecidos
        data_real = None
        dia_semana = ""
        if dia_ref and mes_ref and ano_ref:
            data_tuple = _converter_dia_simulado_para_data(
                dia, dia_ref, mes_ref, ano_ref
            )
            if data_tuple:
                data_real = data_tuple[0]
                dia_semana = data_tuple[1]
        
        for _, r in grp.iterrows():
            hh_inicio = max(0.0, cap_dia - usado_dia)
            hh_cons = float(r["HH"])
            usado_dia += hh_cons
            hh_saldo = max(0.0, cap_dia - usado_dia)
            k = (str(r["Talhao"]), str(r["Atividade"]), str(turma))
            consumido_atividade[k] += hh_cons
            pend = max(0.0, float(demanda_total[k]) - float(consumido_atividade[k]))
            op = float(r["Operarios"] or 0.0)
            hh_equiv_op = (hh_cons / op) if op > 0.01 else 0.0
            
            row = {
                "Tipo_Linha": "ATIVIDADE",
                "Dia": int(dia),
                "Semana": int(r["Semana"]),
                "Fazenda": r["Fazenda"],
                "Talhao": r["Talhao"],
                "Atividade": r["Atividade"],
                "Turma": turma,
                "Operarios": int(round(op)) if op > 0 else 0,
                "Jornada_h_dia": round(float(jornada), 2),
                "Capacidade_Dia_HH": round(cap_dia, 2),
                "HH_Disponivel_Inicio_Dia": round(hh_inicio, 2),
                "HH_Atividade_Demandado": round(float(demanda_total[k]), 2),
                "HH_Atividade_Consumido": round(hh_cons, 2),
                "HH_Consumido_Por_Operador_Equiv": round(hh_equiv_op, 3),
                "HH_Saldo_Apos_Atividade": round(hh_saldo, 2),
                "HH_Pendente_Atividade": round(pend, 2),
                "Fechou_Dia": "S" if hh_saldo <= 0.01 else "N",
                "Calculo_Dia": f"{cap_dia:.2f} - {usado_dia - hh_cons:.2f} - {hh_cons:.2f} = {hh_saldo:.2f}",
            }
            
            # Adicionar data real se calculada
            if data_real:
                row["Data"] = data_real
                row["Dia_Semana"] = dia_semana
                
            out.append(row)
            
        # Resumo do dia
        resumo_row = {
            "Tipo_Linha": "RESUMO_DIA",
            "Dia": int(dia),
            "Semana": int(math.ceil(float(dia) / 5.0)),
            "Fazenda": "",
            "Talhao": "",
            "Atividade": "__RESUMO_DIA__",
            "Turma": turma,
            "Operarios": int(round(ops_dia)) if ops_dia > 0 else 0,
            "Jornada_h_dia": round(float(jornada), 2),
            "Capacidade_Dia_HH": round(cap_dia, 2),
            "HH_Disponivel_Inicio_Dia": round(cap_dia, 2),
            "HH_Atividade_Demandado": "",
            "HH_Atividade_Consumido": round(usado_dia, 2),
            "HH_Consumido_Por_Operador_Equiv": round((usado_dia / ops_dia), 3)
            if ops_dia > 0.01
            else 0.0,
            "HH_Saldo_Apos_Atividade": round(max(0.0, cap_dia - usado_dia), 2),
            "HH_Pendente_Atividade": "",
            "Fechou_Dia": "S" if max(0.0, cap_dia - usado_dia) <= 0.01 else "N",
            "Calculo_Dia": f"{cap_dia:.2f} - {usado_dia:.2f} = {max(0.0, cap_dia - usado_dia):.2f}",
        }
        
        # Adicionar data real no resumo
        if data_real:
            resumo_row["Data"] = data_real
            resumo_row["Dia_Semana"] = dia_semana
            
        out.append(resumo_row)
        
    df = pd.DataFrame(out)
    
    # Reordenar colunas: Data, Dia_Semana primeiro se existirem
    if not df.empty and "Data" in df.columns:
        cols = ["Data", "Dia_Semana"] + [c for c in df.columns if c not in ["Data", "Dia_Semana"]]
        df = df[cols]
        
    return df


def _gerar_aba_ocupacao_turmas(cronograma, turmas, jornada, dias_simulado, dia_ref=None, mes_ref=None, ano_ref=None):
    """Retorna DataFrame pivot: dia x turma com HH, Cap, Uso%, Status.
    Se dia_ref, mes_ref, ano_ref forem fornecidos, adiciona colunas de data real.
    """
    if not cronograma or dias_simulado < 1:
        return pd.DataFrame()
    turma_nomes = sorted(set(t["nome"] for t in turmas))
    turma_ops = {t["nome"]: t["operarios"] for t in turmas}
    hh_dia_turma = defaultdict(lambda: defaultdict(float))
    for c in cronograma:
        hh_dia_turma[c["Dia"]][c.get("Turma", "")] += float(c.get("HH", 0))
    rows = []
    for dia in range(1, dias_simulado + 1):
        # Calcular data real se parametros fornecidos
        data_real = None
        dia_semana = ""
        if dia_ref and mes_ref and ano_ref:
            data_tuple = _converter_dia_simulado_para_data(
                dia, dia_ref, mes_ref, ano_ref
            )
            if data_tuple:
                data_real = data_tuple[0]
                dia_semana = data_tuple[1]
        
        row = {"Dia": dia, "Semana": int(math.ceil(dia / 5.0))}
        
        # Adicionar data real se calculada
        if data_real:
            row["Data"] = data_real
            row["Dia_Semana"] = dia_semana
        hh_total_dia = 0.0
        cap_total_dia = 0.0
        for tn in turma_nomes:
            hh = hh_dia_turma[dia].get(tn, 0.0)
            cap = turma_ops.get(tn, 0) * jornada
            pct = (hh / cap * 100) if cap > 0.01 else 0.0
            row[f"{tn}_HH"] = round(hh, 2)
            row[f"{tn}_Cap"] = round(cap, 2)
            row[f"{tn}_Uso%"] = round(pct, 1)
            if pct >= 90:
                row[f"{tn}_Status"] = "ALTO"
            elif pct >= 50:
                row[f"{tn}_Status"] = "MEDIO"
            elif pct > 0.01:
                row[f"{tn}_Status"] = "BAIXO"
            else:
                row[f"{tn}_Status"] = "OCIOSO"
            hh_total_dia += hh
            cap_total_dia += cap
        row["Total_HH"] = round(hh_total_dia, 2)
        row["Total_Cap"] = round(cap_total_dia, 2)
        row["Total_Uso%"] = round(
            (hh_total_dia / cap_total_dia * 100) if cap_total_dia > 0.01 else 0.0, 1
        )
        rows.append(row)
    return pd.DataFrame(rows)


def _df_crono_operacional(df_crono, dia_ref=None, mes_ref=None, ano_ref=None):
    """Remove colunas monetarias do cronograma para export operacional.
    Se dia_ref, mes_ref, ano_ref forem fornecidos, adiciona colunas de data real.
    """
    drop = [c for c in ("Custo_MO",) if c in df_crono.columns]
    df = df_crono.drop(columns=drop, errors="ignore")
    
    # Adicionar colunas de data real se parametros fornecidos
    if dia_ref and mes_ref and ano_ref and "Dia" in df.columns:
        datas_reais = []
        dias_semana = []
        for _, row in df.iterrows():
            dia_simulado = row.get("Dia", 1)
            data_tuple = _converter_dia_simulado_para_data(
                dia_simulado, dia_ref, mes_ref, ano_ref
            )
            if data_tuple:
                datas_reais.append(data_tuple[0])
                dias_semana.append(data_tuple[1])
            else:
                datas_reais.append(f"Dia_{dia_simulado}")
                dias_semana.append("")
        
        # Inserir colunas no inicio
        df.insert(0, "Dia_Semana", dias_semana)
        df.insert(0, "Data", datas_reais)
    
    return df


def _escrever_cronograma_e_cascata(
    writer, df_crono_op, df_timeline, sheet_name="CRONOGRAMA_E_CASCATA"
):
    """
    Cronograma e timeline na mesma folha (linha em branco entre blocos).
    Retorna a linha 1-based do cabecalho da timeline (para colorir Fase), ou None.
    """
    df_crono_op.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
    if df_timeline is None or getattr(df_timeline, "empty", True):
        return None
    start = len(df_crono_op) + 2
    df_timeline.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start)
    return start + 1


def _aplicar_cores_timeline_excel(wb, sheet_name="TIMELINE_CASCATA", header_row=1):
    """Colorir coluna Cor_Hex como fill real na coluna da Fase (header_row = linha do cabecalho da timeline, 1-based)."""
    try:
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    header = [cell.value for cell in ws[header_row]]
    if "Cor_Hex" not in header or "Fase" not in header:
        return
    idx_cor = header.index("Cor_Hex") + 1
    idx_fase = header.index("Fase") + 1
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        hex_val = str(row[idx_cor - 1].value or "BDC3C7")
        if len(hex_val) == 6:
            fill = PatternFill(
                start_color=hex_val, end_color=hex_val, fill_type="solid"
            )
            row[idx_fase - 1].fill = fill
            row[idx_fase - 1].font = Font(color="FFFFFF", bold=True)


def _aplicar_cores_ocupacao_excel(wb, sheet_name="OCUPACAO_TURMAS_DIA"):
    """Colorir Status (ALTO/MEDIO/BAIXO/OCIOSO) na aba de ocupação."""
    try:
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    header = [cell.value for cell in ws[1]]
    status_cols = [i for i, h in enumerate(header) if h and str(h).endswith("_Status")]
    fills = {
        "ALTO": PatternFill(
            start_color="E74C3C", end_color="E74C3C", fill_type="solid"
        ),
        "MEDIO": PatternFill(
            start_color="F39C12", end_color="F39C12", fill_type="solid"
        ),
        "BAIXO": PatternFill(
            start_color="3498DB", end_color="3498DB", fill_type="solid"
        ),
        "OCIOSO": PatternFill(
            start_color="95A5A6", end_color="95A5A6", fill_type="solid"
        ),
    }
    font_w = Font(color="FFFFFF", bold=True)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for ci in status_cols:
            val = str(row[ci].value or "")
            if val in fills:
                row[ci].fill = fills[val]
                row[ci].font = font_w






def _salvar_perfil_equipe(turmas, executores, jornada, nome_perfil):
    os.makedirs(PERFIS_DIR, exist_ok=True)
    dados = {
        "nome": nome_perfil,
        "executores": executores,
        "jornada": jornada,
        "turmas": [
            {
                "nome": t["nome"],
                "operarios": t["operarios"],
                "atividades": list(t.get("atividades") or []),
            }
            for t in turmas
        ],
    }
    caminho = os.path.join(PERFIS_DIR, f"{_slug_ficheiro_seguro(nome_perfil)}.json")
    with open(caminho, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)
    return caminho


def _listar_perfis_equipe():
    if not os.path.isdir(PERFIS_DIR):
        return []
    out = []
    for fn in sorted(os.listdir(PERFIS_DIR)):
        if fn.endswith(".json"):
            try:
                with open(os.path.join(PERFIS_DIR, fn), "r", encoding="utf-8") as f:
                    d = json.load(f)
                out.append(d)
            except Exception:
                pass
    return out


def _carregar_perfil_equipe_menu():
    perfis = _listar_perfis_equipe()
    if not perfis:
        aviso("Nenhum perfil de equipe salvo ainda.")
        return None
    nomes = [p.get("nome", "?") for p in perfis]
    sel = selecionar("PERFIL DE EQUIPE", nomes)
    if not sel:
        return None
    for p in perfis:
        if p.get("nome") == sel:
            return p
    return None


# ──────────────────────────────────────────────
#  EQUIPE PADRAO + BATCH ORCHESTRATOR
# ──────────────────────────────────────────────


def _checkpoint_editar_template(turmas, atividades_reais_global):
    """Checkpoint before each farm: allow quick edits to team template."""
    sub()
    print(G + BL + "  CHECKPOINT — Equipe Padrao" + RS)
    for t in turmas:
        print(
            G
            + f"  - {t['nome']}: "
            + C
            + f"{t['operarios']} ops, {len(t.get('atividades', []))} atividades"
            + RS
        )
    print(DM + "  [0] Continuar sem alterar" + RS)
    print(DM + "  [1] Editar operarios de uma turma" + RS)
    print(DM + "  [2] Adicionar nova turma" + RS)
    print(DM + "  [3] Redistribuir atividades (S/N) de uma turma" + RS)
    sub()
    op = prompt("Opcao", "0").strip()
    if op == "1":
        nomes = [t["nome"] for t in turmas]
        nm = selecionar("TURMA PARA EDITAR", nomes)
        if nm:
            for t in turmas:
                if t["nome"] == nm:
                    t["operarios"] = pedir_int(
                        f"  Novos operarios para '{nm}'", t["operarios"]
                    )
    elif op == "2":
        nome = prompt("Nome da nova turma", f"Turma {len(turmas) + 1}")
        qtd = pedir_int("Quantos operarios", 1)
        nova = {"nome": nome, "operarios": qtd, "atividades": []}
        if atividades_reais_global:
            menu_vincular_atividades_turma(nova, atividades_reais_global)
        turmas.append(nova)
        ok(f"Turma '{nome}' adicionada.")
    elif op == "3":
        nomes = [t["nome"] for t in turmas]
        nm = selecionar("TURMA PARA REDISTRIBUIR", nomes)
        if nm and atividades_reais_global:
            for t in turmas:
                if t["nome"] == nm:
                    menu_vincular_atividades_turma(t, atividades_reais_global)
    return turmas


def _recomendar_equipes_padrao(
    total_hh, dias_meta, cap_ep_dia, jornada, prazo_absoluto
):
    """Compute how many standard-team sets are needed and return recommendation dict."""
    if cap_ep_dia <= 0.01 or dias_meta <= 0:
        return None
    hh_capacidade_ep = cap_ep_dia * dias_meta
    if hh_capacidade_ep >= total_hh:
        return {
            "status": "suficiente",
            "ep_necessarias": 1,
            "ep_extras": 0,
            "trabalhadores_extras": 0,
        }
    ep_necessarias = math.ceil(total_hh / hh_capacidade_ep)
    ep_extras = ep_necessarias - 1
    trab_necessarios = math.ceil(total_hh / (dias_meta * jornada))
    trab_extras = max(0, trab_necessarios - int(cap_ep_dia / jornada))
    return {
        "status": "insuficiente",
        "ep_necessarias": ep_necessarias,
        "ep_extras": ep_extras,
        "trabalhadores_extras": trab_extras,
        "trab_total_necessario": trab_necessarios,
    }


def _imprimir_recomendacao_ep(rec, fazenda, prazo_absoluto):
    """Print equipe padrao recommendation for one farm."""
    if not rec:
        return
    if rec["status"] == "suficiente":
        print(G + f"  Equipe padrao SUFICIENTE para '{fazenda}'." + RS)
    else:
        print(Y + f"  Equipe padrao INSUFICIENTE para '{fazenda}'." + RS)
        if prazo_absoluto:
            print(
                C
                + f"  [SUGESTAO] +{rec['ep_extras']} equipe(s) padrao (total {rec['ep_necessarias']}) cumpririam a meta."
                + RS
            )
            print(
                C
                + f"  [ALTERNATIVA] +{rec['trabalhadores_extras']} trabalhador(es) extra(s) (total {rec['trab_total_necessario']})."
                + RS
            )


def _exportar_excel_consolidado_lote(
    resultados, empresa_filtro=None, nome_arquivo_micro="", extras=None
):
    """Exporta workbook consolidado do lote com cascata inter-fazendas e timeline unificada."""
    if not resultados:
        return
    extras = extras or {}
    try:
        pasta = OUTPUT_DIR
        os.makedirs(pasta, exist_ok=True)
        emp_slug = (
            _slug_ficheiro_seguro(empresa_filtro)
            if empresa_filtro
            else "Todas_empresas"
        )
        nome_xlsx = f"Consolidado_SmartScheduler_{emp_slug}.xlsx"
        caminho = os.path.join(pasta, nome_xlsx)
        meta_rows = [
            {"Campo": "Empresa_filtro_EQUIPE", "Valor": empresa_filtro or "(todas)"},
            {
                "Campo": "Microplanejamento",
                "Valor": os.path.basename(nome_arquivo_micro)
                if nome_arquivo_micro
                else "",
            },
        ]
        for k, v in extras.items():
            meta_rows.append({"Campo": str(k), "Valor": str(v)})

        dias_acum_total = max(
            (int(x.get("dia_fim_acumulado", 0)) for x in resultados), default=0
        )
        dias_meta_val = int(extras.get("Dias_meta", 0) or 0)
        resumo_rows = [
            {"Metrica": "Fazendas processadas", "Valor": len(resultados)},
            {
                "Metrica": "HH total (soma)",
                "Valor": round(sum(float(x.get("total_hh", 0)) for x in resultados), 1),
            },
            {
                "Metrica": "Dias acumulados lote continuo",
                "Valor": dias_acum_total,
            },
            {
                "Metrica": "Dias meta",
                "Valor": dias_meta_val,
            },
            {
                "Metrica": "Saldo meta (dias)",
                "Valor": max(0, dias_meta_val - dias_acum_total),
            },
            {
                "Metrica": "Status meta global",
                "Valor": "DENTRO" if dias_acum_total <= dias_meta_val else "EXCEDIDO",
            },
        ]
        d_mec_vals = [
            int(x.get("dias_mecanizado") or 0)
            for x in resultados
            if x.get("dias_mecanizado")
        ]
        if d_mec_vals:
            resumo_rows.append(
                {"Metrica": "Dias cenario mecanizado (max)", "Valor": max(d_mec_vals)}
            )

        rows_faz = []
        for x in resultados:
            rec = x.get("rec_ep") or {}
            row_faz = {
                "Fazenda": x.get("fazenda"),
                "Dias_simulado": x.get("dias_simulado"),
                "Dia_inicio_acum": x.get("dia_inicio_acumulado"),
                "Dia_fim_acum": x.get("dia_fim_acumulado"),
                "Meta_consumida_%": x.get("pct_meta_consumida"),
                "Saldo_meta_dias": x.get("saldo_meta_apos"),
                "Status_meta": x.get("status_meta_continuo"),
                "Total_HH": x.get("total_hh"),
            }
            rows_faz.append(row_faz)

        curva_rows = []
        for x in resultados:
            curva_rows.append(
                {
                    "Fazenda": x.get("fazenda"),
                    "Dia_fim_acumulado": x.get("dia_fim_acumulado", 0),
                    "Meta_dias": dias_meta_val,
                    "Consumido_%": x.get("pct_meta_consumida", 0),
                    "HH_acumulado": round(
                        sum(
                            float(r.get("total_hh", 0))
                            for r in resultados[: resultados.index(x) + 1]
                        ),
                        1,
                    ),
                }
            )

        crono_all_rows = []
        for x in resultados:
            offset = int(x.get("dia_inicio_acumulado", 1)) - 1
            for c in x.get("cronograma") or []:
                row = dict(c)
                row["Dia_Lote"] = int(c.get("Dia", 0)) + offset
                row["Semana_Lote"] = int(math.ceil(row["Dia_Lote"] / 5.0))
                crono_all_rows.append(row)

        with pd.ExcelWriter(caminho, engine="openpyxl") as w:
            pd.DataFrame(meta_rows).to_excel(w, sheet_name="METADADOS", index=False)
            pd.DataFrame(resumo_rows).to_excel(w, sheet_name="RESUMO", index=False)
            pd.DataFrame(rows_faz).to_excel(
                w, sheet_name="CASCATA_FAZENDAS", index=False
            )
            pd.DataFrame(curva_rows).to_excel(
                w, sheet_name="CURVA_CONSUMO_META", index=False
            )
            if crono_all_rows:
                pd.DataFrame(crono_all_rows).to_excel(
                    w, sheet_name="CRONOGRAMA_LOTE", index=False
                )
            try:
                wb = w.book
                from openpyxl.styles import Font, PatternFill

                if "CASCATA_FAZENDAS" in wb.sheetnames:
                    ws = wb["CASCATA_FAZENDAS"]
                    header = [cell.value for cell in ws[1]]
                    if "Status_meta" in header:
                        idx_st = header.index("Status_meta") + 1
                        fills_st = {
                            "OK": PatternFill(
                                start_color="27AE60",
                                end_color="27AE60",
                                fill_type="solid",
                            ),
                            "RISCO": PatternFill(
                                start_color="F39C12",
                                end_color="F39C12",
                                fill_type="solid",
                            ),
                            "EXCEDIDO": PatternFill(
                                start_color="E74C3C",
                                end_color="E74C3C",
                                fill_type="solid",
                            ),
                        }
                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                            val = str(row[idx_st - 1].value or "")
                            if val in fills_st:
                                row[idx_st - 1].fill = fills_st[val]
                                row[idx_st - 1].font = Font(color="FFFFFF", bold=True)
            except Exception:
                pass
        ok(f"Consolidado Excel exportado: {nome_xlsx}")
    except Exception as ex:
        aviso(f"Nao foi possivel exportar consolidado Excel: {ex}")


