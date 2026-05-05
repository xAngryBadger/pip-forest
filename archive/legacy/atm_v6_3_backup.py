"""
SRF — Sistema de Restauracao Florestal v6.3
Smart Scheduler com Comparativo Manual/Mecanizado
Uso : python atm_v6_3.py
ATM_DEMO=1 python atm_v6_3.py
python atm_v6_3.py --demo
Modo DEMO: se existir USEESTAPLANILHAULIANOPOLIS.xlsx, gera/atualiza ulianopolisswg.xlsx;
tarifas CT 313 como no fluxo normal; [1] usa a fazenda com mais linhas (micro municipio Ulianopolis).
"""

import atexit
import calendar
import datetime
import copy
import io
import json
import math
import os
import re
import subprocess
import sys
import unicodedata
from collections import OrderedDict, defaultdict
from contextlib import redirect_stderr, redirect_stdout
from statistics import median

import pandas as pd

try:
    from srf_monitor_state import (
        append_relatorio as _monitor_append_relatorio,
    )
    from srf_monitor_state import (
        build_rendimentos_from_demandas as _monitor_build_rendimentos,
    )
    from srf_monitor_state import (
        default_state_path as _monitor_default_state_path,
    )
    from srf_monitor_state import (
        merge_emit as _monitor_merge_emit,
    )
except Exception:
    _monitor_append_relatorio = None
    _monitor_build_rendimentos = None
    _monitor_default_state_path = None
    _monitor_merge_emit = None

try:
    from rich.console import Console
    from rich.table import Table
except ImportError:
    print("Instale: pip install rich pandas openpyxl")
    sys.exit(1)

# ──────────────────────────────────────────────
#  CORES & UI (estilo ATM v3)
# ──────────────────────────────────────────────
try:
    import colorama

    colorama.init()
    G = "\033[92m"
    Y = "\033[93m"
    R = "\033[91m"
    C = "\033[96m"
    DM = "\033[2m"
    BL = "\033[1m"
    RS = "\033[0m"
except ImportError:
    G = Y = R = C = DM = BL = RS = ""

W = 66
console = Console()

ASCII_ART = r"""
        ,@@@@@@@,
    ,,,.   ,@@@@@@/@@,  .oo8888o.
    ,&%%&%&&%,@@@@@/@@@@@@,8888\88/8o
,%&\%&&%&&%,@@@\@@@/@@@88\88888/88'
%&&%&%&/%&&%@@\@@/ /@@@88888\88888'
%&&%/ %&%%&&@@\ V /@@' `88\8 `/88'
`&%\ ` /%&'    |.|        \ '|8'
    |o|        | |         | |
    |.|        | |         | |
    \\/ ._\//_/__/  ,\_//__\\/.  \_//__/_
"""

VERSION = "6.3"
APP_NAME = "SRF - Sistema de Restauracao Florestal"
DIR = os.path.dirname(os.path.abspath(__file__))
CFGP = os.path.join(DIR, "config.json")
DOSSIER_DIRNAME = "dossiês"
ROOT_DIR = os.path.dirname(os.path.dirname(DIR))
DATA_DIR = os.path.join(ROOT_DIR, "data")
INPUT_DIR = os.path.join(DATA_DIR, "planilhas")
OUTPUT_DIR = os.path.join(DATA_DIR, DOSSIER_DIRNAME)
PROFILES_DIR = os.path.join(DATA_DIR, "perfis_equipe")

PRECO_FINAL_JSON_DEFAULT = "preco_final.json"
PRECO_FINAL_JSON_DOWNLOADS = os.path.join(
    os.path.expanduser("~"), "Downloads", PRECO_FINAL_JSON_DEFAULT
)
_PRECO_FINAL_JSON_CACHE = {"path": "", "mtime": None, "mapa": {}}

# ATM 6.1: foco operacional (atividades + HH). Valores em R$ ficam desativados temporariamente.
MODO_SOMENTE_HH = True
CT_REAL_FILENAME = "ct317real.xlsx"

# Modo DEMO (Ulianópolis): ATM_DEMO=1 ou --demo
# Fonte de verdade para reconstruir o demo: USEESTAPLANILHAULIANOPOLIS.xlsx (municipio Ulianopolis)
DEMO_MICRO_FILENAME = "ulianopolisswg.xlsx"
DEMO_MICRO_SOURCE_FILENAME = "USEESTAPLANILHAULIANOPOLIS.xlsx"


def _is_demo_mode():
    v = os.environ.get("ATM_DEMO", "").strip().lower()
    if v in ("1", "true", "yes", "sim", "on"):
        return True
    return "--demo" in sys.argv


def _is_legacy_mode():
    v = os.environ.get("ATM_LEGACY", "").strip().lower()
    if v in ("1", "true", "yes", "sim", "on"):
        return True
    return "--legacy" in sys.argv


def _is_beta_mode():
    # Fluxo beta promovido a padrão; legado fica opt-in por flag/env.
    if _is_legacy_mode():
        return False
    v = os.environ.get("ATM_BETA", "").strip().lower()
    if v in ("0", "false", "no", "off"):
        return False
    return True


def _is_demo_micro_path(path_or_name):
    if not path_or_name:
        return False
    return os.path.basename(str(path_or_name)).lower() == DEMO_MICRO_FILENAME.lower()


# ──────────────────────────────────────────────
#  MAPEAMENTO FIXO DE COLUNAS CONHECIDAS
#  (fallback semantico so se nenhuma bater)
# ──────────────────────────────────────────────
KNOWN_COLUMNS = {
    "fazenda": ["NOME FAZENDA", "CÓDIGO FAZENDA"],
    "chave": ["CHAVE POLÍGONO", "CHAVE POLIGONO"],
    "area": [
        "ÁREA TRABALHADA ESTIMADA (HECTARE)",
        "ÁREA POLÍGONO (HECTARE)",
        "AREA POLIGONO (HECTARE)",
        "AREA TRABALHADA ESTIMADA (HECTARE)",
    ],
    "atividade": ["ATIVIDADES", "ATIVIDADE"],
}


def linha(c="="):
    print(G + c * W + RS)


def sub(c="-"):
    print(DM + c * W + RS)


def cabecalho(sub_titulo=""):
    os.system("cls" if os.name == "nt" else "clear")
    print(G + ASCII_ART + RS)
    linha()
    print(G + BL + f"  [ SRF ]  {APP_NAME}  v{VERSION}".center(W) + RS)
    if sub_titulo:
        print(DM + G + sub_titulo.center(W) + RS)
    print(DM + G + datetime.datetime.now().strftime("  %d/%m/%Y  %H:%M").center(W) + RS)
    linha()

def subcabecalho(sub_titulo=""):
    """Versão incremental que não limpa a tela, mantém conteúdo anterior."""
    print("\n" + "─" * W)
    print(G + BL + f"  [ SRF ]  {APP_NAME}  v{VERSION}".center(W) + RS)
    if sub_titulo:
        print(DM + G + sub_titulo.center(W) + RS)
    print(DM + G + datetime.datetime.now().strftime("  %d/%m/%Y  %H:%M").center(W) + RS)
    print("─" * W + "\n")


def aviso(m):
    print(Y + f"\n  !  {m}" + RS)


def erro(m):
    print(R + f"\n  X  {m}" + RS)


def ok(m):
    print(G + f"\n  +  {m}" + RS)


def prompt(msg, default=None):
    suf = f" [{default}]" if default is not None else ""
    try:
        v = input(G + "  >> " + C + msg + suf + G + ": " + RS).strip()
    except (EOFError, KeyboardInterrupt):
        print()
        sys.exit(0)
    return v if v else (str(default) if default is not None else "")


def pedir_float(msg, default, allow_zero=False):
    while True:
        v = prompt(msg, default)
        try:
            f = float(str(v).replace(",", "."))
            if f > 0 or (allow_zero and f >= 0):
                return f
        except ValueError:
            pass
        aviso("Valor invalido.")


def _parse_jornada(valor):
    s = str(valor).strip().replace(",", ".")
    try:
        return float(s)
    except ValueError:
        pass
    for sep in (":", "h", "e", "E", "H"):
        if sep in s:
            partes = s.split(sep, 1)
            try:
                h = float(partes[0].strip())
                m = float(partes[1].strip())
                if h >= 0 and 0 <= m < 60:
                    return h + m / 60.0
            except (ValueError, IndexError):
                pass
    return None


def pedir_jornada(msg, default):
    while True:
        v = prompt(msg, default)
        resultado = _parse_jornada(v)
        if resultado is not None and resultado > 0:
            return resultado
        aviso("Valor invalido. Use decimal (6.5) ou horario (6:30 = 6h30).")


def pedir_int(msg, default, allow_zero=False):
    while True:
        v = prompt(msg, default)
        try:
            i = int(v)
            if i > 0 or (allow_zero and i >= 0):
                return i
        except ValueError:
            pass
        aviso("Valor invalido.")


def selecionar(titulo, itens, zero_label="Voltar"):
    print(G + f"\n  -- {titulo} " + "--" * max(0, (W - len(titulo) - 6) // 2) + RS)
    for i, it in enumerate(itens, 1):
        print(G + f"  [{i:2}] " + C + str(it) + RS)
    print(G + f"  [ 0] " + DM + zero_label + RS)
    while True:
        v = prompt("Escolha").strip()
        if v == "0":
            return None
        if v.isdigit() and 1 <= int(v) <= len(itens):
            return itens[int(v) - 1]
        aviso("Opcao invalida.")


def selecionar_paginado(titulo, itens, page_size=5, zero_label="Voltar"):
    total = len(itens)
    page = 0
    max_page = math.ceil(total / page_size) - 1
    while True:
        start = page * page_size
        end = min(start + page_size, total)
        print(
            G
            + f"\n  -- {titulo} (pag {page + 1}/{max_page + 1}) "
            + "--" * max(0, (W - len(titulo) - 16) // 2)
            + RS
        )
        for i in range(start, end):
            print(G + f"  [{i + 1:2}] " + C + str(itens[i]) + RS)
        nav = []
        if page > 0:
            nav.append("[-] Anterior")
        if page < max_page:
            nav.append("[+] Proxima")
        nav.append("[0] " + zero_label)
        print(DM + "  " + "   ".join(nav) + RS)
        v = prompt("Escolha").strip()
        if v == "0":
            return -1
        if v == "+" and page < max_page:
            page += 1
            continue
        if v == "-" and page > 0:
            page -= 1
            continue
        if v.isdigit() and 1 <= int(v) <= total:
            return int(v) - 1
        aviso("Opcao invalida.")


def confirmar(msg, default=True):
    s = "S/n" if default else "s/N"
    v = prompt(f"{msg} [{s}]").strip().lower()
    if not v:
        return default
    return v in ("s", "sim", "y", "yes")


def remover_acentos(texto):
    if not isinstance(texto, str):
        return ""
    return (
        "".join(
            c
            for c in unicodedata.normalize("NFD", texto)
            if unicodedata.category(c) != "Mn"
        )
        .lower()
        .strip()
    )

_RE_PUNCT = re.compile(r"[^a-z0-9 ]+")
_RE_SPACES = re.compile(r"\s+")


def normalizar_chave(texto):
    """remover_acentos + strip punctuation + collapse whitespace. Canonical lookup key."""
    s = remover_acentos(texto)
    s = _RE_PUNCT.sub(" ", s)
    return _RE_SPACES.sub(" ", s).strip()


def _normalizar_chave_atividade_semantica(texto):
    """
    Normaliza texto de atividade preservando semantica de tokens de fase.
    Regra de negocio (supervisor): PL=Plantio, CD=Conducao.
    Mantem retrocompatibilidade com chaves legadas ao ser usado em conjunto com
    _candidatos_chave_atividade().
    """
    base = normalizar_chave(texto)
    if not base:
        return base
    toks = base.split()
    out = []
    for i, t in enumerate(toks):
        prev = toks[i - 1] if i > 0 else ""
        if t == "pl" and prev in ("impl", "implant", "implantacao"):
            out.append("plantio")
        elif t == "cd" and prev in ("impl", "implant", "implantacao"):
            out.append("conducao")
        else:
            out.append(t)
    return " ".join(out).strip()


def _candidatos_chave_atividade(texto):
    """
    Gera variantes de chave para lookup robusto:
    1) legado (PL/CD literal), 2) semantico (Plantio/Conducao).
    """
    legado = normalizar_chave(texto)
    semantico = _normalizar_chave_atividade_semantica(texto)
    if semantico and semantico != legado:
        return [legado, semantico]
    return [legado]


def _formatar_periodo_meta(mes_ref, ano_ref, prazo_meses):
    """Retorna (inicio, fim) do periodo meta em texto (MM/AAAA)."""
    try:
        mes_ref = int(mes_ref)
        ano_ref = int(ano_ref)
        prazo_meses = int(round(float(prazo_meses)))
    except Exception:
        return None
    inicio = f"{mes_ref:02d}/{ano_ref}"
    if prazo_meses <= 0:
        return (inicio, inicio)
    mes_fim = mes_ref + (prazo_meses - 1)
    ano_fim = ano_ref + (mes_fim - 1) // 12
    mes_fim = ((mes_fim - 1) % 12) + 1
    fim = f"{mes_fim:02d}/{ano_fim}"
    return (inicio, fim)


def _formatar_data_dia(dia, mes, ano):
    """Formata data DD/MM/AAAA; assume valores inteiros validos."""
    return f"{int(dia):02d}/{int(mes):02d}/{int(ano)}"


# Mapeamento de dias da semana para abreviacoes brasileiras
_DIAS_SEMANA_CURTO = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
_DIAS_SEMANA_COMPLETO = [
    "Segunda-feira",
    "Terça-feira",
    "Quarta-feira",
    "Quinta-feira",
    "Sexta-feira",
    "Sábado",
    "Domingo",
]


def _converter_dia_simulado_para_data(
    dia_simulado: int, dia_ref: int, mes_ref: int, ano_ref: int
):
    """
    Converte dia simulado (1, 2, 3...) para data real.
    Considera todos os dias do calendario (incluindo fins de semana).

    Retorna: (data_str, dia_semana_curto, dia_semana_completo, data_obj)
    Ex: (1, 20, 4, 2025) -> ("20/04/2025", "Seg", "Segunda-feira", date_obj)
    """
    try:
        from datetime import date, timedelta

        dia_simulado = int(dia_simulado)
        dia_ref = int(dia_ref)
        mes_ref = int(mes_ref)
        ano_ref = int(ano_ref)

        # Data de inicio
        data_inicio = date(ano_ref, mes_ref, dia_ref)

        # Adiciona (dia_simulado - 1) dias (dia 1 = data_inicio)
        data_real = data_inicio + timedelta(days=dia_simulado - 1)

        # Formata data como DD/MM/AAAA
        data_str = f"{data_real.day:02d}/{data_real.month:02d}/{data_real.year}"

        # Obtem dia da semana (0=Segunda, 6=Domingo)
        dia_semana_idx = data_real.weekday()
        dia_semana_curto = _DIAS_SEMANA_CURTO[dia_semana_idx]
        dia_semana_completo = _DIAS_SEMANA_COMPLETO[dia_semana_idx]

        return (data_str, dia_semana_curto, dia_semana_completo, data_real)
    except Exception:
        return (f"Dia_{dia_simulado}", "-", "-", None)


def _calcular_data_fim_por_meses(dia_inicio, mes_ref, ano_ref, prazo_meses):
    """
    Calcula data final (dia/mes/ano) a partir de um dia inicial e prazo em meses.
    Ajusta o dia para o maximo do mes final.
    """
    try:
        dia_inicio = int(dia_inicio)
        mes_ref = int(mes_ref)
        ano_ref = int(ano_ref)
        prazo_meses = int(round(float(prazo_meses)))
    except Exception:
        return None
    if prazo_meses <= 0:
        return (dia_inicio, mes_ref, ano_ref)
    mes_fim = mes_ref + (prazo_meses - 1)
    ano_fim = ano_ref + (mes_fim - 1) // 12
    mes_fim = ((mes_fim - 1) % 12) + 1
    ultimo_dia = calendar.monthrange(ano_fim, mes_fim)[1]
    dia_fim = min(max(1, dia_inicio), int(ultimo_dia))
    return (dia_fim, mes_fim, ano_fim)


# ──────────────────────────────────────────────
# CONTEXT DASHBOARD (PERSISTENTE)
# ──────────────────────────────────────────────
class ContextoSessao:
    """Armazena todas as escolhas importantes durante a sessão para exibição no dashboard."""

    def __init__(self):
        self.fazenda_selecionada = None
        self.equipe_selecionada = None
        self.talhoes_selecionados = []
        self.total_talhoes_fazenda = 0
        self.area_total_fazenda = 0.0
        self.data_inicio = None
        self.data_termino = None
        self.atividades_distribuidas = 0
        self.total_atividades = 0
        self.modo_atual = None
        self.orcamento_estrito = True
        self.tarifas_carregadas = 0
        self.timestamp_atualizacao = None
        self._console = Console()

    def atualizar_fazenda(self, nome_fazenda, df_fazenda=None):
        """Atualiza informações da fazenda selecionada."""
        self.fazenda_selecionada = nome_fazenda
        if df_fazenda is not None:
            self.total_talhoes_fazenda = (
                df_fazenda["chave"].nunique() if "chave" in df_fazenda.columns else 0
            )
            if "area" in df_fazenda.columns:
                self.area_total_fazenda = df_fazenda["area"].sum()
            elif "area_ha" in df_fazenda.columns:
                self.area_total_fazenda = df_fazenda["area_ha"].sum()
            else:
                self.area_total_fazenda = 0.0
        else:
            self.total_talhoes_fazenda = 0
            self.area_total_fazenda = 0.0
        self.timestamp_atualizacao = datetime.datetime.now()

    def atualizar_equipe(self, nome_equipe):
        """Atualiza equipe selecionada."""
        self.equipe_selecionada = nome_equipe
        self.timestamp_atualizacao = datetime.datetime.now()

    def definir_escopo_talhoes(self, talhoes_selecionados, todos_talhoes):
        """Define os talhões selecionados e atualiza metadados."""
        self.talhoes_selecionados = (
            list(set(talhoes_selecionados))
            if isinstance(talhoes_selecionados, list)
            else []
        )
        self.total_talhoes_fazenda = (
            len(set(todos_talhoes)) if isinstance(todos_talhoes, list) else 0
        )
        self.timestamp_atualizacao = datetime.datetime.now()

    def atualizar_atividades(self, distribuidas, total):
        """Atualiza contagem de atividades distribuídas vs total."""
        self.atividades_distribuidas = distribuidas
        self.total_atividades = total
        self.timestamp_atualizacao = datetime.datetime.now()

    def definir_datas(self, inicio, termino):
        """Define datas de início e término da operação."""
        self.data_inicio = inicio
        self.data_termino = termino
        self.timestamp_atualizacao = datetime.datetime.now()

    def atualizar_modo(self, modo):
        """Atualiza o modo atual (single, lote, multi_equipe)."""
        self.modo_atual = modo
        self.timestamp_atualizacao = datetime.datetime.now()

    def atualizar_configuracoes(self, cfg):
        """Atualiza configurações importantes do sistema."""
        self.orcamento_estrito = (
            cfg.get("orcamento_estrito", True) if isinstance(cfg, dict) else True
        )
        self.tarifas_carregadas = (
            len(cfg.get("tarifas", {})) if isinstance(cfg, dict) else 0
        )
        self.timestamp_atualizacao = datetime.datetime.now()

    def limpar_contexto(self):
        """Limpa o contexto para iniciar nova sessão."""
        self.__init__()

    def _criar_tabela_dashboard(self):
        """Cria a tabela Rich formatada com informações do contexto."""
        table = Table(
            title="[bold cyan]Dashboard de Contexto[/bold cyan]",
            show_header=True,
            header_style="bold cyan",
            border_style="bright_blue",
            padding=(0, 1),
        )

        # Colunas principais
        table.add_column("Fazenda", style="bold green", width=30, justify="left")
        table.add_column("Equipe", style="bold yellow", width=20, justify="left")
        table.add_column("Talhoes", style="bold blue", width=15, justify="center")
        table.add_column(
            "Atividades", style="bold magenta", width=15, justify="center"
        )
        table.add_column("Datas", style="bold red", width=25, justify="left")

        # Preparar dados de cada coluna
        fazenda_info = "[dim]Não selecionada[/dim]"
        if self.fazenda_selecionada:
            fazenda_info = f"[bold green]{self.fazenda_selecionada}[/bold green]"
            meta_parts = []
            if self.total_talhoes_fazenda > 0:
                qtd_talhoes = (
                    len(self.talhoes_selecionados)
                    if self.talhoes_selecionados
                    else self.total_talhoes_fazenda
                )
                meta_parts.append(f"{qtd_talhoes}/{self.total_talhoes_fazenda} talhões")
            if self.area_total_fazenda > 0:
                meta_parts.append(f"{self.area_total_fazenda:,.1f}ha")
            if meta_parts:
                fazenda_info += f"\n[dim]{' | '.join(meta_parts)}[/dim]"

        equipe_info = self.equipe_selecionada or "[dim]Não selecionada[/dim]"

        talhoes_info = "[dim]0/0[/dim]"
        if self.total_talhoes_fazenda > 0:
            qtd = (
                len(self.talhoes_selecionados)
                if self.talhoes_selecionados
                else self.total_talhoes_fazenda
            )
            talhoes_info = f"[bold]{qtd}/{self.total_talhoes_fazenda}[/bold]"

        atividades_info = "[dim]0/0[/dim]"
        if self.total_atividades > 0:
            atividades_info = (
                f"[bold]{self.atividades_distribuidas}/{self.total_atividades}[/bold]"
            )

        datas_info = "[dim]Não definidas[/dim]"
        if self.data_inicio or self.data_termino:
            datas_parts = []
            if self.data_inicio:
                datas_parts.append(f"Início: {self.data_inicio}")
            if self.data_termino:
                datas_parts.append(f"Término: {self.data_termino}")
            datas_info = "\n".join(datas_parts)

        # Adicionar linha principal
        table.add_row(
            fazenda_info, equipe_info, talhoes_info, atividades_info, datas_info
        )

        # Segunda linha com informações adicionais (se houver)
        info_adicional = []
        if self.modo_atual:
            info_adicional.append(
                f"[cyan]Modo:[/cyan] [white]{self.modo_atual}[/white]"
            )
        if self.tarifas_carregadas > 0:
            info_adicional.append(
                f"[cyan]Tarifas:[/cyan] [white]{self.tarifas_carregadas}[/white]"
            )
        if not self.orcamento_estrito:
            info_adicional.append("[cyan]Orçamento:[/cyan] [white]Flexível[/white]")

        if info_adicional:
            table.add_row(
                "\n" + "\n".join(info_adicional), "", "", "", "", end_section=True
            )

        return table

    def exibir(self, console_inst=None, mostrar_sempre=True):
        """Exibe o dashboard formatado no console."""
        # Não mostrar se nada foi selecionado e mostrar_sempre=False
        if not mostrar_sempre and not self.fazenda_selecionada:
            return

        console_inst = console_inst or self._console
        table = self._criar_tabela_dashboard()
        console_inst.print("\n", table, "\n")


# Instância global do contexto
contexto_sessao = ContextoSessao()
_MONITOR_STATE_PATH = (
    _monitor_default_state_path(os.getpid())
    if callable(_monitor_default_state_path)
    else None
)


def _emitir_monitor_state(partial):
    if _MONITOR_STATE_PATH and callable(_monitor_merge_emit):
        try:
            _monitor_merge_emit(_MONITOR_STATE_PATH, partial)
        except Exception:
            pass


def _emitir_monitor_relatorio(titulo, texto):
    if _MONITOR_STATE_PATH and callable(_monitor_append_relatorio):
        try:
            _monitor_append_relatorio(_MONITOR_STATE_PATH, titulo, texto)
        except Exception:
            pass


def _emitir_monitor_atual():
    """Emite o estado atual do contexto para os monitores."""
    if not (_MONITOR_STATE_PATH and callable(_monitor_merge_emit)):
        return
    
    try:
        # Construir estado parcial do contexto atual - FORMATO ESPERADO PELOS MONITORES
        estado = {}
        
        # Operação: fazenda, modo, equipe (para feed "meta")
        if contexto_sessao.fazenda_selecionada:
            estado["operacao"] = {
                "fazenda_atual": str(contexto_sessao.fazenda_selecionada),
                "modo": str(contexto_sessao.modo_atual or ""),
                "equipe_atual": str(contexto_sessao.equipe_selecionada or ""),
                "status_geral": "em_execucao",
            }
            
            # Adicionar mensagem curta com contexto
            msg_parts = []
            if contexto_sessao.fazenda_selecionada:
                msg_parts.append(str(contexto_sessao.fazenda_selecionada))
            if contexto_sessao.equipe_selecionada:
                msg_parts.append(f"Eq:{contexto_sessao.equipe_selecionada}")
            if msg_parts:
                estado["operacao"]["mensagem_curta"] = " | ".join(msg_parts)
        
        # Lote: talhões, atividades (para feed "meta" - lote section)
        if contexto_sessao.talhoes_selecionados is not None:
            estado["lote"] = {
                "talhoes_selecionados": len(contexto_sessao.talhoes_selecionados),
                "talhoes_total": contexto_sessao.total_talhoes_fazenda,
                "area_ha": contexto_sessao.area_total_fazenda,
                # Adicionar campos que o monitor espera
                "dias_meta": 0,  # placeholder - será atualizado depois se necessário
                "dias_consumidos": 0,
                "saldo_dias": 0,
                "status_meta_continuo": "OK",
                "prazo_absoluto": True,
            }
        
        # Atividades distribuídas (vamos usar para rendimentos_sessao simplificado)
        # Nota: O monitor de rendimentos espera uma lista de {atividade, hh_ha, origem, chave_tarifa}
        # Vamos simplificar e enviar o que temos por enquanto
        if contexto_sessao.total_atividades > 0 and contexto_sessao.atividades_distribuidas > 0:
            # Criar uma entrada simplificada para mostrar que há atividades vinculadas
            estado["rendimentos_sessao"] = [{
                "atividade": f"{contexto_sessao.atividades_distribuidas}/{contexto_sessao.total_atividades} atividades",
                "hh_ha": 0.0,  # vamos calcular depois se necessário
                "origem": "sessao",
                "chave_tarifa": "progresso"
            }]
        
        # Timestamp
        if contexto_sessao.timestamp_atualizacao:
            estado["timestamp"] = contexto_sessao.timestamp_atualizacao.timestamp()
            estado["timestamp_iso"] = contexto_sessao.timestamp_atualizacao.strftime("%Y-%m-%dT%H:%M:%S")
        
        _monitor_merge_emit(_MONITOR_STATE_PATH, estado)
    except Exception:
        pass  # Silent fail to not disrupt main flow


def _emitir_monitor_rendimentos(atividade_nome: str, vincular: bool, hh_ha: float = 0.0, origem: str = "", chave_tarifa: str = ""):
    """Emite atualização de rendimento quando uma atividade é vinculada/desvinculada."""
    if not (_MONITOR_STATE_PATH and callable(_monitor_merge_emit)):
        return
    
    try:
        # Para o feed de rendimentos, vamos manter uma lista acumulativa
        # Primeiro, tentar obter o estado atual para manter o que já temos
        estado_atual = {}
        rendimentos_existentes = []
        
        # Obter estado atual existente (se houver)
        if _MONITOR_STATE_PATH and os.path.exists(_MONITOR_STATE_PATH):
            try:
                import json
                with open(_MONITOR_STATE_PATH, 'r', encoding='utf-8') as f:
                    dados_existentes = json.load(f)
                    rendimentos_existentes = dados_existentes.get('rendimentos_sessao', []).copy()
            except:
                rendimentos_existentes = []
        
        # Processar a atividade atual
        if atividade_nome and vincular and hh_ha > 0:
            # Verificar se esta atividade já existe na lista (para atualizar ou adicionar)
            atividade_encontrada = False
            novas_rendimentos = []
            for rend in rendimentos_existentes:
                if rend.get('atividade') == str(atividade_nome):
                    # Atualizar existente
                    novas_rendimentos.append({
                        'atividade': str(atividade_nome),
                        'hh_ha': float(hh_ha),
                        'origem': str(origem) if origem else rend.get('origem', 'sessao'),
                        'chave_tarifa': str(chave_tarifa) if chave_tarifa else rend.get('chave_tarifa', 'vinculada')
                    })
                    atividade_encontrada = True
                else:
                    # Manter outras atividades como estavam
                    novas_rendimentos.append(rend)
            
            # Se não encontrou, adicionar nova
            if not atividade_encontrada:
                novas_rendimentos.append({
                    'atividade': str(atividade_nome),
                    'hh_ha': float(hh_ha),
                    'origem': str(origem) if origem else 'sessao',
                    'chave_tarifa': str(chave_tarifa) if chave_tarifa else 'vinculada'
                })
            
            estado = {"rendimentos_sessao": novas_rendimentos}
        elif not vincular:
            # Se desvincular, remover da lista
            novas_rendimentos = [
                rend for rend in rendimentos_existentes 
                if rend.get('atividade') != str(atividade_nome)
            ]
            estado = {"rendimentos_sessao": novas_rendimentos}
        else:
            # Caso nenhum dos acima, manter estado atual
            estado = {"rendimentos_sessao": rendimentos_existentes}
        
        # Também atualizar o estado geral para manter sincronização
        estado_geral = {}
        if contexto_sessao.fazenda_selecionada:
            estado_geral["operacao"] = {
                "fazenda_atual": str(contexto_sessao.fazenda_selecionada),
                "modo": str(contexto_sessao.modo_atual or ""),
                "equipe_atual": str(contexto_sessao.equipe_selecionada or ""),
                "status_geral": "em_execucao",
            }
            
            msg_parts = []
            if contexto_sessao.fazenda_selecionada:
                msg_parts.append(str(contexto_sessao.fazenda_selecionada))
            if contexto_sessao.equipe_selecionada:
                msg_parts.append(f"Eq:{contexto_sessao.equipe_selecionada}")
            if msg_parts:
                estado_geral["operacao"]["mensagem_curta"] = " | ".join(msg_parts)
        
        if contexto_sessao.talhoes_selecionados is not None:
            estado_geral["lote"] = {
                "talhoes_selecionados": len(contexto_sessao.talhoes_selecionados),
                "talhoes_total": contexto_sessao.total_talhoes_fazenda,
                "area_ha": contexto_sessao.area_total_fazenda,
                "dias_meta": 0,
                "dias_consumidos": 0,
                "saldo_dias": 0,
                "status_meta_continuo": "OK",
                "prazo_absoluto": True,
            }
        
        if contexto_sessao.timestamp_atualizacao:
            estado_geral["timestamp"] = contexto_sessao.timestamp_atualizacao.timestamp()
            estado_geral["timestamp_iso"] = contexto_sessao.timestamp_atualizacao.strftime("%Y-%m-%dT%H:%M:%S")
        
        # Mesclar estado geral com rendimentos
        estado_geral.update(estado)
        _monitor_merge_emit(_MONITOR_STATE_PATH, estado_geral)
    except Exception:
        pass  # Silent fail


def _abrir_monitor_janela(feed="meta", pid=None):
    """
    Abre uma janela separada com o monitor SRF.
    Usa subprocess para iniciar um terminal novo.
    """
    try:
        target_pid = int(pid or os.getpid())
        script_monitor = os.path.join(DIR, "srf_monitor.py")

        if not os.path.isfile(script_monitor):
            aviso(f"Script do monitor nao encontrado: {script_monitor}")
            return False

        # Comando base
        cmd = [
            sys.executable,
            script_monitor,
            "--feed", str(feed),
            "--pid", str(target_pid),
        ]

        if os.name == "nt":
            # Windows: usar start para nova janela
            subprocess.Popen(
                ["start", "cmd", "/k"] + cmd,
                shell=True,
                creationflags=subprocess.CREATE_NEW_CONSOLE,
            )
        else:
            # Linux/Mac: tentar terminais comuns (Kitty/Foot primeiro — típico em Hyprland/Sway)
            terminais = [
                ["kitty", "-e"] + cmd,
                ["foot"] + cmd,
                ["wezterm", "start", "--"] + cmd,
                ["alacritty", "-e"] + cmd,
                ["gnome-terminal", "--"] + cmd,
                ["konsole", "--hold", "-e"] + cmd,
                ["xfce4-terminal", "-e"] + cmd,
                ["xterm", "-hold", "-e"] + cmd,
            ]
            for term_cmd in terminais:
                try:
                    subprocess.Popen(
                        term_cmd,
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                    )
                    break
                except FileNotFoundError:
                    continue
            else:
                aviso("Nenhum terminal compativel encontrado para abrir o monitor.")
                return False

        ok(f"Monitor aberto (PID {target_pid}, feed={feed})")
        return True
    except Exception as e:
        aviso(f"Erro ao abrir monitor: {e}")
        return False


def dashboard_header(console_inst=None, mostrar_sempre=True):
    """Wrapper para exibir o dashboard (mantém compatibilidade com chamadas existentes)."""
    global contexto_sessao
    contexto_sessao.exibir(console_inst, mostrar_sempre)


# ──────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────
def _default_sequencia_dict():
    return {
        "modo": "implantacao",
        "offset_limpeza_quimica_dias": 30,
        "filtros_plantio": ["plantio"],
        "filtros_irrigacao": ["irrig"],
        "limpeza_quimica_filtros": ["limpeza", "quim"],
        "limpeza_quimica_exclusoes": ["impl", "impl."],
        "implantacao_outras_fase": 5.5,
        "implantacao_fases": [
            {"id": "rocada", "filtros": ["rocada", "roçada"], "exclusoes": []},
            {
                "id": "formiga",
                "filtros": ["formiga", "combate a formiga", "combate a formigas"],
                "exclusoes": [],
            },
            {"id": "coroamento", "filtros": ["coroamento", "coroa"], "exclusoes": []},
            {"id": "coveamento", "filtros": ["coveamento", "coveam"], "exclusoes": []},
            {
                "id": "adubacao_quimica",
                "filtros": ["adubacao quim", "adubação quím", "melhora quim"],
                "exclusoes": [],
            },
        ],
        # Ordem exata solicitada para manutencao SWG (conforme lista de atividades do Excel).
        "swg_fases": [
            {
                "id": "swg_rocada_manual",
                "filtros": ["rocada manual", "roçada manual"],
                "exclusoes": [],
            },
            {
                "id": "swg_limpeza_area",
                "filtros": ["limpeza de area", "limpeza de área"],
                "exclusoes": [],
            },
            {
                "id": "swg_capina_coroa",
                "filtros": ["capina manual coroa", "capina manual"],
                "exclusoes": [],
            },
            {
                "id": "swg_combate_formigas",
                "filtros": ["combate a formigas", "combate a formiga", "formigas"],
                "exclusoes": [],
            },
            {
                "id": "swg_coveamento",
                "filtros": [
                    "coveam area nao subsol",
                    "coveam área não subsol",
                    "coveamento",
                ],
                "exclusoes": [],
            },
            {
                "id": "swg_adubacao_base",
                "filtros": [
                    "adubacao quim man de base",
                    "adubação quim man de base",
                    "adubacao",
                ],
                "exclusoes": [],
            },
            {
                "id": "swg_plantio_manual",
                "filtros": ["plantio manual", "plantio"],
                "exclusoes": [],
            },
            {
                "id": "swg_irrigacao_inicial",
                "filtros": ["irrigacao inicial", "irrigação inicial", "irrigacao"],
                "exclusoes": [],
            },
        ],
    "personalizado_ordem": [],
}


# ──────────────────────────────────────────────
# CONFIGURAÇÃO DE TERRITÓRIOS (V6 NOVO)
# ──────────────────────────────────────────────
def _territorio_config():
    """
    Configuração de territórios/cidades para distribuição automática de equipes.
    Baseado nos requisitos:
    - SWG: 3 equipes com 3 operários úteis cada (1 coordenação não trabalha)
    - Inovesa: 4 equipes com 4 operários úteis cada (1 coordenação não trabalha)
    - Territórios: Açailândia (3), Dom Eliseu (2), Ulianópolis (1), Cidelêndia (1)
    """
    return {
        "cidades": {
            "acailandia": {"nome": "Açailândia", "n_equipes_swg": 3, "n_equipes_inovesa": 4},
            "dom_eliseu": {"nome": "Dom Eliseu", "n_equipes_swg": 2, "n_equipes_inovesa": 3},
            "ulianopolis": {"nome": "Ulianópolis", "n_equipes_swg": 1, "n_equipes_inovesa": 2},
            "cidelandia": {"nome": "Cidelêndia", "n_equipes_swg": 1, "n_equipes_inovesa": 2},
        },
        "empresas": {
            "swg": {
                "nome": "SWG",
                "coordenadores_por_equipe": 1,
                "operarios_por_equipe": 3,  # 3 úteis + 1 coordenação = 4 total
                "equipes_por_cidade": {
                    "acailandia": 3,
                    "dom_eliseu": 2,
                    "ulianopolis": 1,
                    "cidelandia": 1,
                },
            },
            "inovesa": {
                "nome": "Inovesa",
                "coordenadores_por_equipe": 1,
                "operarios_por_equipe": 4,  # 4 úteis + 1 coordenação = 5 total
                "equipes_por_cidade": {
                    "acailandia": 4,
                    "dom_eliseu": 3,
                    "ulianopolis": 2,
                    "cidelandia": 2,
                },
            },
        },
    }


def _detectar_cidade_por_fazenda(nome_fazenda: str) -> str:
    """
    Detecta a cidade/território baseado no nome da fazenda.
    Retorna o código da cidade ou None se não detectar.
    """
    nome_norm = normalizar_chave(str(nome_fazenda))
    cidade_keywords = {
        "acailandia": ["acailandia", "acailand", "ailandia"],
        "dom_eliseu": ["dom eliseu", "eliseu", "dom_eliseu"],
        "ulianopolis": ["ulianopolis", "ulianopol", "ulianópolis"],
        "cidelandia": ["cidelandia", "cideland", "cidelndia", "buritirana"],
    }
    for cidade, keywords in cidade_keywords.items():
        for kw in keywords:
            if kw in nome_norm:
                return cidade
    return None


def _distribuir_fazendas_por_territorio(fazendas: list) -> dict:
    """
    Distribui fazendas por território/cidade automaticamente.
    Retorna: {"acailandia": [fazenda1, fazenda2], "dom_eliseu": [...], ...}
    """
    distribuicao = {"acailandia": [], "dom_eliseu": [], "ulianopolis": [], "cidelandia": []}
    nao_identificadas = []

    for faz in fazendas:
        cidade = _detectar_cidade_por_fazenda(faz)
        if cidade:
            distribuicao[cidade].append(faz)
        else:
            nao_identificadas.append(faz)

    return distribuicao, nao_identificadas


def _calcular_equipes_territorio(cidade: str, empresa: str = "auto") -> dict:
    """
    Calcula configuração de equipes para um território.
    empresa: "swg", "inovesa", ou "auto" (detecta da fazenda)
    """
    cfg_territorio = _territorio_config()
    if cidade not in cfg_territorio["cidades"]:
        return None

    info_cidade = cfg_territorio["cidades"][cidade]

    # Se empresa é auto, assume baseado na cidade (logica customizavel)
    if empresa == "auto":
        empresa = "swg"  # Default SWG (pode ser modificado)

    info_empresa = cfg_territorio["empresas"][empresa]
    n_equipes = info_empresa["equipes_por_cidade"].get(cidade, 1)
    operarios_por_eq = info_empresa["operarios_por_equipe"]
    coordenadores = info_empresa["coordenadores_por_equipe"]
    total_por_eq = operarios_por_eq + coordenadores

    return {
        "cidade": cidade,
        "nome_cidade": info_cidade["nome"],
        "empresa": empresa,
        "nome_empresa": info_empresa["nome"],
        "n_equipes": n_equipes,
        "operarios_por_equipe": operarios_por_eq,
        "coordenadores_por_equipe": coordenadores,
        "total_por_equipe": total_por_eq,
        "total_operarios": n_equipes * operarios_por_eq,
        "total_coordenadores": n_equipes * coordenadores,
        "total_geral": n_equipes * total_por_eq,
    }


def _sugerir_config_territorio(fazendas: list, modo_simplificado: bool = True) -> dict:
    """
    Sugere configuração completa de equipes baseada na distribuição de fazendas.
    modo_simplificado: True = mostra apenas resumo, False = detalhado
    """
    distribuicao, nao_id = _distribuir_fazendas_por_territorio(fazendas)
    sugestoes = []

    for cidade, fazs in distribuicao.items():
        if not fazs:
            continue
        config = _calcular_equipes_territorio(cidade)
        if config:
            config["fazendas"] = fazs
            config["n_fazendas"] = len(fazs)
            sugestoes.append(config)

    return {
        "distribuicao": distribuicao,
        "nao_identificadas": nao_id,
        "sugestoes": sugestoes,
        "total_equipes": sum(s["n_equipes"] for s in sugestoes),
        "total_operarios": sum(s["total_operarios"] for s in sugestoes),
    }


def _merge_sequencia_defaults(seq):
    """Preenche chaves ausentes em cfg['sequencia'] (muta seq)."""
    d0 = _default_sequencia_dict()
    for k, v in d0.items():
        if k not in seq:
            seq[k] = v
        elif (
            k in ("implantacao_fases", "swg_fases", "personalizado_ordem")
            and not seq[k]
        ):
            seq[k] = v


def carregar_config():
    if not os.path.exists(CFGP):
        with open(CFGP, "w", encoding="utf-8") as f:
            json.dump({"de_para": {}, "tarifas": {}, "atividades": {}}, f)
    with open(CFGP, "r", encoding="utf-8") as f:
        cfg = json.load(f)
    for k in ("de_para", "tarifas", "atividades"):
        if k not in cfg:
            cfg[k] = {}
    if "orcamento_estrito" not in cfg:
        cfg["orcamento_estrito"] = True
    if "filtros_bloqueio_global" not in cfg:
        cfg["filtros_bloqueio_global"] = ["plantio", "irrig"]
    if "fazendas_ct" not in cfg:
        cfg["fazendas_ct"] = []
    if "sequencia" not in cfg or not isinstance(cfg.get("sequencia"), dict):
        cfg["sequencia"] = {}
    _merge_sequencia_defaults(cfg["sequencia"])
    if "preco_final_json_path" not in cfg:
        cfg["preco_final_json_path"] = ""
    if "comparativo" not in cfg or not isinstance(cfg.get("comparativo"), dict):
        cfg["comparativo"] = {}
    if "execucao_compacta" not in cfg["comparativo"]:
        # Quando True, o segundo cenário (mecanizado) roda silencioso e exibe só o comparativo final.
        cfg["comparativo"]["execucao_compacta"] = True
    mapa_json = _carregar_mapa_preco_final_json(cfg)
    if mapa_json:
        _aplicar_mapa_preco_final_em_tarifas(cfg.get("tarifas", {}), mapa_json)
    return cfg


def salvar_config(cfg):
    with open(CFGP, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


def parse_intervalos_escolha(texto, max_n):
    """
    Converte '1,3,5-8' em indices 0-based unicos e ordenados (numeracao 1..max_n).
    Espacos ignorados; intervalos inclusive.
    """
    out = set()
    if not texto or not str(texto).strip() or max_n < 1:
        return []
    for part in str(texto).replace(" ", "").split(","):
        if not part:
            continue
        if "-" in part:
            a, b = part.split("-", 1)
            try:
                lo, hi = int(a), int(b)
                if lo > hi:
                    lo, hi = hi, lo
                for k in range(lo, hi + 1):
                    if 1 <= k <= max_n:
                        out.add(k - 1)
            except ValueError:
                continue
        else:
            try:
                k = int(part)
                if 1 <= k <= max_n:
                    out.add(k - 1)
            except ValueError:
                continue
    return sorted(out)


def mediana_rendimento_hh(tarifas):
    """Mediana dos rendimento_hh > 0 em config.tarifas, ou None."""
    vals = []
    for v in (tarifas or {}).values():
        if not isinstance(v, dict):
            continue
        try:
            x = float(v.get("rendimento_hh", 0))
            if x > 0:
                vals.append(x)
        except (TypeError, ValueError):
            pass
    if not vals:
        return None
    vals.sort()
    n = len(vals)
    mid = n // 2
    if n % 2:
        return vals[mid]
    return (vals[mid - 1] + vals[mid]) / 2.0


def resolver_rendimento_hh(
    cfg, tarifas, t_nome, strict=False, session_hh=None, atv_micro=None
):
    """
    HH/ha para a chave t_nome em tarifas.
    session_hh: dict opcional {nome_micro ou chave_tarifa: hh_ha} valido so na execucao atual (nao grava config).
    Modo strict: sem mediana/8 silenciosos — retorna None se invalido.
    Excecao: atividades mecanizadas (tipo Mecanizada ou HM>0) retornam 0.0 em strict.
    """
    if session_hh:
        try:
            if atv_micro and atv_micro in session_hh:
                return float(session_hh[atv_micro])
            if t_nome and t_nome in session_hh:
                return float(session_hh[t_nome])
        except (TypeError, ValueError):
            pass
    if t_nome in (tarifas or {}):
        row = tarifas[t_nome]
        r = row.get("rendimento_hh")
        try:
            rf = float(r)
            if rf >= 0:
                if strict and rf <= 0:
                    tipo = str(row.get("tipo", "")).lower()
                    hm = float(row.get("rendimento_hm", 0) or 0)
                    if "mecaniz" in tipo or hm > 0:
                        return 0.0
                    return None
                return rf
        except (TypeError, ValueError):
            if strict:
                return None
    if strict:
        return None
    ex = cfg.get("rendimento_hh_fallback")
    if ex is not None:
        try:
            e = float(ex)
            if e > 0:
                return e
        except (TypeError, ValueError):
            pass
    med = mediana_rendimento_hh(tarifas)
    if med is not None and med > 0:
        return med
    return 8.0


def resolver_rendimento_hm(cfg, tarifas, t_nome, strict=False):
    """
    HM/ha para a chave t_nome em tarifas.
    Em strict: ausencia/invalido retorna 0.0 (HM nao bloqueia fluxo humano).
    """
    if t_nome in (tarifas or {}):
        row = tarifas[t_nome]
        r = row.get("rendimento_hm", 0)
        try:
            rf = float(r)
            if rf >= 0:
                return rf
        except (TypeError, ValueError):
            pass
    if strict:
        return 0.0
    ex = cfg.get("rendimento_hm_fallback")
    if ex is not None:
        try:
            e = float(ex)
            if e > 0:
                return e
        except (TypeError, ValueError):
            pass
    return 0.0


# ──────────────────────────────────────────────
#  CT RAW -> STG NORMALIZER
# ──────────────────────────────────────────────
STG_FILENAME = "CT_317_NORMALIZADA.xlsx"

# Padrao de prototipo: mapeamento fixo EXAME -> CT_317 (normalizado sem acentos).
# Fonte: leitura real das abas MICROPLANEJAMENTO_ABRIL_JUNHO e STG_TARIFAS.
#
# Declividade (ROÇADA MANUAL CLASSE I..V na CT): classe I = terreno mais plano (menor HH/ha e
# menor preco/ha); classe V = declive maximo (maior HH e maior preco — maior esforco operacional).
# O micro nao traz a classe — o padrao SRF usa sempre CLASSE I / limpeza plana onde aplicavel.
# Ver aviso no scheduler.
DEFAULT_DEPARA_EXAME_CT317 = {
    # Dicionario gerado a partir dos JSONs oficiais:
    #   - ~/Downloads/planejamento_abr_jun.json (micro)
    #   - ~/Downloads/preco_final.json (CT)
    # Chave: atividade do micro normalizada; valor: operacao canonica do preco_final.
    "adubacao quim man de base impl pl app rl": "ADUBAÇÃO QUÍM MAN DE BASE Impl. PL - APP/ RL",
    "capina manual coroa impl cd app rl i": "CAPINA MANUAL COROA Impl. CD APP/ RL I",
    "capina manual coroa impl pl app rl i": "CAPINA MANUAL COROA Impl. PL - APP/ RL I",
    "capina quim man total manut app rl": "CAPINA QUÍM MAN TOTAL Manut. APP/RL",
    "combate a formigas impl cd app rl": "COMBATE À FORMIGAS Impl. CD APP/RL",
    "combate a formigas impl pl app rl": "COMBATE À FORMIGAS Impl. PL APP/ RL",
    "combate a formigas manut app rl": "COMBATE À FORMIGAS Manut. APP/RL",
    "conducao de regeneracao": "CONDUÇÃO DE REGENERAÇÃO",
    "controle de invasoras app rl": "CONTROLE DE INVASORAS APP/RL I",
    "coveam area nao subsol impl pl app rl": "COVEAM ÁREA NÃO SUBSOL Impl. PL APP/ RL",
    "coveamento motocoveador pl app rl": "COVEAMENTO - MOTOCOVEADOR PL APP/RL",
    "eliminacao de exoticas impl cd app rl": "ELIMINAÇÃO DE EXÓTICAS Impl. CD - APP/RL",
    "irrigacao inicial man impl pl app rl": "IRRIGAÇÃO INICIAL MAN Impl. PL - APP/ RL",
    "limpeza de area qu man app rl": "LIMPEZA DE ÁREA QUIM. MAN APP/RL",
    "limpeza de area quim impl cd app rl": "LIMPEZA DE AREA QUIM. Impl. CD APP/RL",
    "nucleacao em faixas app rl": "NUCLEAÇÃO EM FAIXAS APP/RL",
    "plantio manual app rl": "PLANTIO MANUAL APP/RL",
    "preparo de solo mec c grade app rl": "PREPARO DE SOLO MEC C/ GRADE APP/RL",
    "preparo de solo mec s adub app rl": "PREPARO DE SOLO MEC S/ ADUB APP/RL",
    "rocada manual impl cd app rl i": "ROÇADA MANUAL Impl. CD APP/RL I",
    "rocada manual impl pl app rl i": "ROÇADA MANUAL Impl. PL APP/RL I",
}

# Fallback minimo de HH/HM para garantir operacao da CT317 real mesmo quando
# o layout da planilha variar e algum item nao for lido pelo parser.
CT317_HARDCODE_HH_BASE = {
    "ADUBAÇÃO QUÍM MAN 3 MESES APP/RL": {"rendimento_hh": 5.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "ADUBAÇÃO QUÍM MAN DE BASE Impl. PL - APP/ RL": {"rendimento_hh": 10.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "ADUBAÇÃO QUÍM MAN DE BASE MAdap. APP/RL": {"rendimento_hh": 11.6, "rendimento_hm": 0.0, "tipo": "Manual"},
    "ADUBAÇÃO QUÍM MAN DE BASE Manut. APP/RL": {"rendimento_hh": 10.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "APLIC QUIM (LIQ) C/ DRONE TOT APP/RL": {"rendimento_hh": 0.0, "rendimento_hm": 0.5, "tipo": "Mecanizada"},
    "APLIC QUIM (LIQ) C/ DRONE (EXTRA)": {"rendimento_hh": 0.0, "rendimento_hm": 0.35, "tipo": "Mecanizada"},
    "CAPINA MANUAL COROA Impl. CD APP/ RL I": {"rendimento_hh": 6.2, "rendimento_hm": 0.0, "tipo": "Manual"},
    "CAPINA MANUAL COROA Impl. CD APP/ RL II": {"rendimento_hh": 8.5, "rendimento_hm": 0.0, "tipo": "Manual"},
    "CAPINA MANUAL COROA Impl. CD APP/ RL III": {"rendimento_hh": 11.9, "rendimento_hm": 0.0, "tipo": "Manual"},
    "CAPINA MANUAL COROA Impl. CD APP/ RL IV": {"rendimento_hh": 17.1, "rendimento_hm": 0.0, "tipo": "Manual"},
    "CAPINA MANUAL COROA Impl. CD APP/ RL V": {"rendimento_hh": 34.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "CAPINA MANUAL COROA Impl. PL - APP/ RL I": {"rendimento_hh": 6.2, "rendimento_hm": 0.0, "tipo": "Manual"},
    "CAPINA MANUAL COROA Impl. PL - APP/ RL II": {"rendimento_hh": 8.5, "rendimento_hm": 0.0, "tipo": "Manual"},
    "CAPINA MANUAL COROA Impl. PL - APP/ RL III": {"rendimento_hh": 11.9, "rendimento_hm": 0.0, "tipo": "Manual"},
    "CAPINA MANUAL COROA Impl. PL - APP/ RL IV": {"rendimento_hh": 17.1, "rendimento_hm": 0.0, "tipo": "Manual"},
    "CAPINA MANUAL COROA Impl. PL - APP/ RL V": {"rendimento_hh": 34.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "CAPINA MANUAL COROAMENTO MAdap. APP/RL I": {"rendimento_hh": 6.2, "rendimento_hm": 0.0, "tipo": "Manual"},
    "CAPINA MANUAL COROAMENTO MAdap. APP/RL II": {"rendimento_hh": 8.5, "rendimento_hm": 0.0, "tipo": "Manual"},
    "CAPINA MANUAL COROAMENTO MAdap. APP/RL III": {"rendimento_hh": 11.9, "rendimento_hm": 0.0, "tipo": "Manual"},
    "CAPINA MANUAL COROAMENTO MAdap. APP/RL IV": {"rendimento_hh": 17.1, "rendimento_hm": 0.0, "tipo": "Manual"},
    "CAPINA MANUAL COROAMENTO MAdap. APP/RL V": {"rendimento_hh": 34.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "CAPINA MANUAL COROAMENTO Manut. APP/RL I": {"rendimento_hh": 6.2, "rendimento_hm": 0.0, "tipo": "Manual"},
    "CAPINA MANUAL COROAMENTO Manut. APP/RL II": {"rendimento_hh": 8.5, "rendimento_hm": 0.0, "tipo": "Manual"},
    "CAPINA MANUAL COROAMENTO Manut. APP/RL III": {"rendimento_hh": 11.9, "rendimento_hm": 0.0, "tipo": "Manual"},
    "CAPINA MANUAL COROAMENTO Manut. APP/RL IV": {"rendimento_hh": 17.1, "rendimento_hm": 0.0, "tipo": "Manual"},
    "CAPINA MANUAL COROAMENTO Manut. APP/RL V": {"rendimento_hh": 34.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "CAPINA QUIM MAN COROA Manut. APP/RL": {"rendimento_hh": 5.4, "rendimento_hm": 0.0, "tipo": "Manual"},
    "CAPINA QUIM MAN TOTAL MAdap. APP/RL": {"rendimento_hh": 9.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "CAPINA QUIM MAN TOTAL Manut. APP/RL": {"rendimento_hh": 9.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "CAPINA QUIM MEC TOTAL APP/RL": {"rendimento_hh": 0.0, "rendimento_hm": 0.95, "tipo": "Mecanizada"},
    "CAPINA QUIM MEC TOTAL Manut. APP/RL": {"rendimento_hh": 0.0, "rendimento_hm": 0.95, "tipo": "Mecanizada"},
    "COMBATE A FORMIGAS Impl. CD APP/RL": {"rendimento_hh": 3.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "COMBATE A FORMIGAS Impl. PL APP/ RL": {"rendimento_hh": 3.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "COMBATE A FORMIGAS MAdap. APP/RL": {"rendimento_hh": 3.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "COMBATE A FORMIGAS Manut. APP/RL": {"rendimento_hh": 3.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "COMBATE INCENDIOS CAM PIPA H 100%": {"rendimento_hh": 0.0, "rendimento_hm": 1.0, "tipo": "Mecanizada"},
    "COMBATE INCENDIOS CAM PIPA H 50%": {"rendimento_hh": 0.0, "rendimento_hm": 1.0, "tipo": "Mecanizada"},
    "COMBATE INCENDIOS CAM PIPA HORA NORMAL": {"rendimento_hh": 0.0, "rendimento_hm": 1.0, "tipo": "Mecanizada"},
    "COMBATE INCENDIOS HH (100%)": {"rendimento_hh": 2.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "COMBATE INCENDIOS HH (50%)": {"rendimento_hh": 1.5, "rendimento_hm": 0.0, "tipo": "Manual"},
    "COMBATE INCENDIOS HH NORMAL": {"rendimento_hh": 1.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "CONDUÇÃO DE REGENERAÇÃO": {"rendimento_hh": 40.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "CONTROLE DE INVASORAS APP/RL I": {"rendimento_hh": 70.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "COVEAM AREA NAO SUBSOL Impl. PL APP/ RL": {"rendimento_hh": 38.2, "rendimento_hm": 0.0, "tipo": "Manual"},
    "COVEAM AREA NAO SUBSOL MAdap. PL APP/ RL": {"rendimento_hh": 39.5, "rendimento_hm": 0.0, "tipo": "Manual"},
    "COVEAMENTO - MOTOCOVEADOR MAdap. APP/RL": {"rendimento_hh": 42.4, "rendimento_hm": 0.0, "tipo": "Manual"},
    "COVEAMENTO - MOTOCOVEADOR Manut. APP/RL": {"rendimento_hh": 40.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "COVEAMENTO - MOTOCOVEADOR PL APP/RL": {"rendimento_hh": 40.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "COVEAMENTO AREA NAO SUBSOL APP/RL": {"rendimento_hh": 38.2, "rendimento_hm": 0.0, "tipo": "Manual"},
    "COVEAMENTO AREA SUBSOL APP/RL": {"rendimento_hh": 10.8, "rendimento_hm": 0.0, "tipo": "Manual"},
    "COVEAMENTO AREA SUBSOL Impl. PL - APP/RL": {"rendimento_hh": 10.8, "rendimento_hm": 0.0, "tipo": "Manual"},
    "COVEAMENTO MOTOCOV. + ADUB. APP/RL": {"rendimento_hh": 45.8, "rendimento_hm": 0.0, "tipo": "Manual"},
    "DIARIA AVULSA - H NORMAL": {"rendimento_hh": 1.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "ELIMINAÇAO DE EXOTICAS Impl. CD - APP/RL": {"rendimento_hh": 72.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "ELIMINAÇAO DE EXOTICAS Impl. PL - APP/RL": {"rendimento_hh": 70.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "ELIMINAÇAO DE INVASORAS Manut. APP/RL": {"rendimento_hh": 75.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "IRRIGACAO INICIAL MAN APP/RL": {"rendimento_hh": 14.6, "rendimento_hm": 0.0, "tipo": "Manual"},
    "IRRIGACAO INICIAL MAN Impl. PL - APP/ RL": {"rendimento_hh": 13.1, "rendimento_hm": 0.0, "tipo": "Manual"},
    "IRRIGACAO INICIAL MAN MAdap. APP/ RL": {"rendimento_hh": 15.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "IRRIGACAO INICIAL SEMIMEC APP/RL": {"rendimento_hh": 4.0, "rendimento_hm": 0.0, "tipo": "SemiMecanizada"},
    "LIMPEZA DE AREA QUIM. Impl. CD APP/RL": {"rendimento_hh": 9.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "LIMPEZA DE AREA QUIM. MAN APP/RL": {"rendimento_hh": 9.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "LIMPEZA DE AREA QUIM. MEC APP/RL": {"rendimento_hh": 0.0, "rendimento_hm": 0.85, "tipo": "Mecanizada"},
    "NUCLEAÇÃO EM FAIXAS APP/RL": {"rendimento_hh": 10.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "PLANTIO CONSORCIADO NAT/EUC APP/RL": {"rendimento_hh": 13.9, "rendimento_hm": 0.0, "tipo": "Manual"},
    "PLANTIO DE ADENSAMENTO APP/RL": {"rendimento_hh": 15.8, "rendimento_hm": 0.0, "tipo": "Manual"},
    "PLANTIO ENRIQUECIMENTO APP/RL": {"rendimento_hh": 10.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "PLANTIO MANUAL APP/RL": {"rendimento_hh": 15.8, "rendimento_hm": 0.0, "tipo": "Manual"},
    "PREPARO DE SOLO MEC C/ ADUB APP/RL": {"rendimento_hh": 0.0, "rendimento_hm": 2.0, "tipo": "Mecanizada"},
    "PREPARO DE SOLO MEC C/ GRADE APP/RL": {"rendimento_hh": 0.0, "rendimento_hm": 2.0, "tipo": "Mecanizada"},
    "PREPARO DE SOLO MEC S/ ADUB APP/RL": {"rendimento_hh": 0.0, "rendimento_hm": 1.5, "tipo": "Mecanizada"},
    "PREPARO SOLO MEC CABEC COV C/ADUB APP/RL": {"rendimento_hh": 0.0, "rendimento_hm": 12.2, "tipo": "Mecanizada"},
    "RECUPERAÇAO DE AREAS DEGRADADAS": {"rendimento_hh": 45.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "REPLANTIO APP/RL I": {"rendimento_hh": 6.2, "rendimento_hm": 0.0, "tipo": "Manual"},
    "REPLANTIO APP/RL II": {"rendimento_hh": 7.2, "rendimento_hm": 0.0, "tipo": "Manual"},
    "REPLANTIO APP/RL III": {"rendimento_hh": 10.1, "rendimento_hm": 0.0, "tipo": "Manual"},
    "REPLANTIO APP/RL IV": {"rendimento_hh": 11.5, "rendimento_hm": 0.0, "tipo": "Manual"},
    "REPLANTIO APP/RL V": {"rendimento_hh": 13.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "RESTAURAÇAO PASSIVA APP/RL": {"rendimento_hh": 0.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "ROÇADA MANUAL Impl. CD APP/RL I": {"rendimento_hh": 8.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "ROÇADA MANUAL Impl. CD APP/RL II": {"rendimento_hh": 11.2, "rendimento_hm": 0.0, "tipo": "Manual"},
    "ROÇADA MANUAL Impl. CD APP/RL III": {"rendimento_hh": 15.6, "rendimento_hm": 0.0, "tipo": "Manual"},
    "ROÇADA MANUAL Impl. CD APP/RL IV": {"rendimento_hh": 22.4, "rendimento_hm": 0.0, "tipo": "Manual"},
    "ROÇADA MANUAL Impl. CD APP/RL V": {"rendimento_hh": 44.5, "rendimento_hm": 0.0, "tipo": "Manual"},
    "ROÇADA MANUAL Impl. PL APP/RL I": {"rendimento_hh": 8.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "ROÇADA MANUAL Impl. PL APP/RL II": {"rendimento_hh": 11.2, "rendimento_hm": 0.0, "tipo": "Manual"},
    "ROÇADA MANUAL Impl. PL APP/RL III": {"rendimento_hh": 15.6, "rendimento_hm": 0.0, "tipo": "Manual"},
    "ROÇADA MANUAL Impl. PL APP/RL IV": {"rendimento_hh": 22.4, "rendimento_hm": 0.0, "tipo": "Manual"},
    "ROÇADA MANUAL Impl. PL APP/RL V": {"rendimento_hh": 44.5, "rendimento_hm": 0.0, "tipo": "Manual"},
    "ROÇADA MANUAL MAdap. APP/RL I": {"rendimento_hh": 8.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "ROÇADA MANUAL MAdap. APP/RL II": {"rendimento_hh": 11.2, "rendimento_hm": 0.0, "tipo": "Manual"},
    "ROÇADA MANUAL MAdap. APP/RL III": {"rendimento_hh": 15.6, "rendimento_hm": 0.0, "tipo": "Manual"},
    "ROÇADA MANUAL MAdap. APP/RL IV": {"rendimento_hh": 22.4, "rendimento_hm": 0.0, "tipo": "Manual"},
    "ROÇADA MANUAL MAdap. APP/RL V": {"rendimento_hh": 44.5, "rendimento_hm": 0.0, "tipo": "Manual"},
    "ROÇADA MANUAL Manut. APP/RL I": {"rendimento_hh": 8.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "ROÇADA MANUAL Manut. APP/RL II": {"rendimento_hh": 11.2, "rendimento_hm": 0.0, "tipo": "Manual"},
    "ROÇADA MANUAL Manut. APP/RL III": {"rendimento_hh": 15.6, "rendimento_hm": 0.0, "tipo": "Manual"},
    "ROÇADA MANUAL Manut. APP/RL IV": {"rendimento_hh": 22.4, "rendimento_hm": 0.0, "tipo": "Manual"},
    "ROÇADA MANUAL Manut. APP/RL V": {"rendimento_hh": 44.5, "rendimento_hm": 0.0, "tipo": "Manual"},
    "ROÇADA MECANIZADA PL APP/RL": {"rendimento_hh": 0.0, "rendimento_hm": 1.7, "tipo": "Mecanizada"},
    "SEMEADURA DIRETA APP/RL": {"rendimento_hh": 8.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "COBERTURA DA SEMEADURA APP/RL I": {"rendimento_hh": 10.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "QUILOMETRAGEM EXCEDENTE (KM)": {"rendimento_hh": 1.0, "rendimento_hm": 0.0, "tipo": "Mecanizada"},
    "LIMPEZA DE ÁREA QU. MAN APP/RL": {"rendimento_hh": 9.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "COMBATE À FORMIGAS Impl. CD APP/RL": {"rendimento_hh": 3.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "COMBATE À FORMIGAS Impl. PL APP/ RL": {"rendimento_hh": 3.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "COVEAM ÁREA NÃO SUBSOL Impl. PL APP/ RL": {"rendimento_hh": 38.2, "rendimento_hm": 0.0, "tipo": "Manual"},
    "ADUBAÇÃO QUÍM MAN DE BASE Impl.PL-APP/RL": {"rendimento_hh": 10.0, "rendimento_hm": 0.0, "tipo": "Manual"},
    "IRRIGAÇÃO INICIAL MAN Impl. PL - APP/ RL": {"rendimento_hh": 13.1, "rendimento_hm": 0.0, "tipo": "Manual"},
}


# ═══════════════════════════════════════════════════════════════════════════════
# COMPARATIVO MANUAL vs MECANIZADO — Mapeamento de atividades substituíveis
# ═══════════════════════════════════════════════════════════════════════════════
# Dicionário de pares: Atividade Manual → Atividade Mecanizada equivalente
# Use para comparar cenários de mão-de-obra humana vs máquina
COMPARATIVO_MANUAL_MEC = {
    # ROÇADA
    "ROÇADA MANUAL Impl. CD APP/RL I": "ROÇADA MECANIZADA PL APP/RL",
    "ROÇADA MANUAL Impl. CD APP/RL II": "ROÇADA MECANIZADA PL APP/RL",
    "ROÇADA MANUAL Impl. CD APP/RL III": "ROÇADA MECANIZADA PL APP/RL",
    "ROÇADA MANUAL Impl. CD APP/RL IV": "ROÇADA MECANIZADA PL APP/RL",
    "ROÇADA MANUAL Impl. CD APP/RL V": "ROÇADA MECANIZADA PL APP/RL",
    "ROÇADA MANUAL Impl. PL APP/RL I": "ROÇADA MECANIZADA PL APP/RL",
    "ROÇADA MANUAL Impl. PL APP/RL II": "ROÇADA MECANIZADA PL APP/RL",
    "ROÇADA MANUAL Impl. PL APP/RL III": "ROÇADA MECANIZADA PL APP/RL",
    "ROÇADA MANUAL Impl. PL APP/RL IV": "ROÇADA MECANIZADA PL APP/RL",
    "ROÇADA MANUAL Impl. PL APP/RL V": "ROÇADA MECANIZADA PL APP/RL",
    "ROÇADA MANUAL MAdap. APP/RL I": "ROÇADA MECANIZADA PL APP/RL",
    "ROÇADA MANUAL MAdap. APP/RL II": "ROÇADA MECANIZADA PL APP/RL",
    "ROÇADA MANUAL MAdap. APP/RL III": "ROÇADA MECANIZADA PL APP/RL",
    "ROÇADA MANUAL MAdap. APP/RL IV": "ROÇADA MECANIZADA PL APP/RL",
    "ROÇADA MANUAL MAdap. APP/RL V": "ROÇADA MECANIZADA PL APP/RL",
    "ROÇADA MANUAL Manut. APP/RL I": "ROÇADA MECANIZADA PL APP/RL",
    "ROÇADA MANUAL Manut. APP/RL II": "ROÇADA MECANIZADA PL APP/RL",
    "ROÇADA MANUAL Manut. APP/RL III": "ROÇADA MECANIZADA PL APP/RL",
    "ROÇADA MANUAL Manut. APP/RL IV": "ROÇADA MECANIZADA PL APP/RL",
    "ROÇADA MANUAL Manut. APP/RL V": "ROÇADA MECANIZADA PL APP/RL",
    # CAPINA
    "CAPINA MANUAL COROA Impl. CD APP/ RL I": "CAPINA QUIM MEC TOTAL APP/RL",
    "CAPINA MANUAL COROA Impl. CD APP/ RL II": "CAPINA QUIM MEC TOTAL APP/RL",
    "CAPINA MANUAL COROA Impl. CD APP/ RL III": "CAPINA QUIM MEC TOTAL APP/RL",
    "CAPINA MANUAL COROA Impl. CD APP/ RL IV": "CAPINA QUIM MEC TOTAL APP/RL",
    "CAPINA MANUAL COROA Impl. CD APP/ RL V": "CAPINA QUIM MEC TOTAL APP/RL",
    "CAPINA MANUAL COROA Impl. PL - APP/ RL I": "CAPINA QUIM MEC TOTAL APP/RL",
    "CAPINA MANUAL COROA Impl. PL - APP/ RL II": "CAPINA QUIM MEC TOTAL APP/RL",
    "CAPINA MANUAL COROA Impl. PL - APP/ RL III": "CAPINA QUIM MEC TOTAL APP/RL",
    "CAPINA MANUAL COROA Impl. PL - APP/ RL IV": "CAPINA QUIM MEC TOTAL APP/RL",
    "CAPINA MANUAL COROA Impl. PL - APP/ RL V": "CAPINA QUIM MEC TOTAL APP/RL",
    "CAPINA MANUAL COROAMENTO MAdap. APP/RL I": "CAPINA QUIM MEC TOTAL APP/RL",
    "CAPINA MANUAL COROAMENTO MAdap. APP/RL II": "CAPINA QUIM MEC TOTAL APP/RL",
    "CAPINA MANUAL COROAMENTO MAdap. APP/RL III": "CAPINA QUIM MEC TOTAL APP/RL",
    "CAPINA MANUAL COROAMENTO MAdap. APP/RL IV": "CAPINA QUIM MEC TOTAL APP/RL",
    "CAPINA MANUAL COROAMENTO MAdap. APP/RL V": "CAPINA QUIM MEC TOTAL APP/RL",
    "CAPINA MANUAL COROAMENTO Manut. APP/RL I": "CAPINA QUIM MEC TOTAL APP/RL",
    "CAPINA MANUAL COROAMENTO Manut. APP/RL II": "CAPINA QUIM MEC TOTAL APP/RL",
    "CAPINA MANUAL COROAMENTO Manut. APP/RL III": "CAPINA QUIM MEC TOTAL APP/RL",
    "CAPINA MANUAL COROAMENTO Manut. APP/RL IV": "CAPINA QUIM MEC TOTAL APP/RL",
    "CAPINA MANUAL COROAMENTO Manut. APP/RL V": "CAPINA QUIM MEC TOTAL APP/RL",
    # LIMPEZA DE ÁREA
    "LIMPEZA DE AREA QUIM. Impl. CD APP/RL": "LIMPEZA DE AREA QUIM. MEC APP/RL",
    "LIMPEZA DE AREA QUIM. MAN APP/RL": "LIMPEZA DE AREA QUIM. MEC APP/RL",
    "LIMPEZA DE ÁREA QU. MAN APP/RL": "LIMPEZA DE AREA QUIM. MEC APP/RL",
    # COVEAMENTO
    "COVEAM AREA NAO SUBSOL Impl. PL APP/ RL": "COVEAMENTO - MOTOCOVEADOR PL APP/RL",
    "COVEAM ÁREA NÃO SUBSOL Impl. PL APP/ RL": "COVEAMENTO - MOTOCOVEADOR PL APP/RL",
    "COVEAM AREA NAO SUBSOL MAdap. PL APP/ RL": "COVEAMENTO - MOTOCOVEADOR MAdap. APP/RL",
    "COVEAMENTO AREA NAO SUBSOL APP/RL": "COVEAMENTO - MOTOCOVEADOR PL APP/RL",
    "COVEAMENTO - MOTOCOVEADOR MAdap. APP/RL": "COVEAMENTO - MOTOCOVEADOR PL APP/RL",
    "COVEAMENTO - MOTOCOVEADOR Manut. APP/RL": "COVEAMENTO - MOTOCOVEADOR PL APP/RL",
    # APLICAÇÃO QUÍMICA
    "APLIC QUIM (LIQ) C/ DRONE (EXTRA)": "APLIC QUIM (LIQ) C/ DRONE TOT APP/RL",
}


def _atividades_com_mecanizado_disponivel(atividades_reais):
    """
    Retorna lista de atividades que têm equivalente mecanizado.
    """
    pares = []
    for atv in atividades_reais:
        if atv in COMPARATIVO_MANUAL_MEC:
            mec = COMPARATIVO_MANUAL_MEC[atv]
            pares.append((atv, mec))
    return pares


def _substituir_por_mecanizado(df_faz, substituicoes):
    """
    Substitui atividades manuais pelas mecanizadas equivalentes em um dataframe.
    
    Args:
        df_faz: DataFrame com as atividades
        substituicoes: dict {atividade_manual: atividade_mecanizada}
    
    Returns:
        DataFrame modificado com atividades substituídas
    """
    df_mec = df_faz.copy()
    for manual, mecanizada in substituicoes.items():
        if isinstance(mecanizada, dict):
            alvo = str(
                mecanizada.get("atividade_mecanizada")
                or mecanizada.get("nome")
                or mecanizada.get("recurso")
                or ""
            ).strip()
        else:
            alvo = str(mecanizada).strip()
        if not alvo:
            continue
        mask = df_mec["atividade"] == manual
        if mask.any():
            df_mec.loc[mask, "atividade"] = alvo
    return df_mec


def _formatar_substituicao_comparativo(valor):
    if isinstance(valor, dict):
        nome = str(
            valor.get("atividade_mecanizada")
            or valor.get("nome")
            or valor.get("recurso")
            or ""
        ).strip()
        hm = float(valor.get("rendimento_hm", valor.get("hm", 0)) or 0)
        custo = float(valor.get("custo_h", 0) or 0)
        origem = str(valor.get("origem", "custom") or "custom")
        if nome:
            return f"{nome} [HM={hm:.2f}, R$ {custo:.2f}/h, {origem}]"
        return f"[HM={hm:.2f}, R$ {custo:.2f}/h, {origem}]"
    return str(valor)


def _clonar_cfg_comparativo_mecanizado(cfg, substituicoes):
    """Cria uma copia isolada do cfg e injeta recursos mecanizados customizados temporarios."""
    cfg_var = copy.deepcopy(cfg or {})
    tarifas = cfg_var.setdefault("tarifas", {})
    de_para = cfg_var.setdefault("de_para", {})

    for manual, mecanizada in (substituicoes or {}).items():
        if not isinstance(mecanizada, dict):
            continue
        nome = str(
            mecanizada.get("atividade_mecanizada")
            or mecanizada.get("nome")
            or mecanizada.get("recurso")
            or ""
        ).strip()
        if not nome:
            continue

        hm = float(mecanizada.get("rendimento_hm", mecanizada.get("hm", 0)) or 0)
        custo_h = float(mecanizada.get("custo_h", mecanizada.get("custo", 0)) or 0)
        preco_ha = float(
            mecanizada.get("preco_ha", mecanizada.get("preco_unit", 0)) or 0
        )
        tipo = str(mecanizada.get("tipo") or "Mecanizada").strip() or "Mecanizada"

        row = dict(tarifas.get(nome, {}) or {})
        row.update(
            {
                "rendimento_hh": 0.0,
                "rendimento_hm": hm,
                "preco_ha": preco_ha,
                "preco_unit": preco_ha,
                "custo_hora": custo_h,
                "tipo": tipo,
                "recurso": "maquina",
                "origem": "comparativo_custom",
            }
        )
        tarifas[nome] = row
        de_para[manual] = nome

    return cfg_var


def _cadastrar_recurso_mecanizado_externo(manual_sugestao=""):
    """Coleta um recurso mecanizado externo para uso apenas na comparacao."""
    sub()
    print(C + BL + "  RECURSO MECANIZADO EXTERNO (comparativo)" + RS)
    nome = prompt(
        "Nome do recurso/modelo externo",
        manual_sugestao or "Navu",
    )
    nome = str(nome).strip()
    if not nome:
        aviso("Nome vazio. Recurso externo cancelado.")
        return None
    hm = pedir_float("HM/ha do recurso externo", 1.0)
    custo_h = pedir_float("Custo R$/h do recurso externo", 0.0, allow_zero=True)
    preco_ha = pedir_float("Preco R$/ha (opcional)", 0.0, allow_zero=True)
    return {
        "atividade_mecanizada": nome,
        "rendimento_hm": float(hm or 0.0),
        "custo_h": float(custo_h or 0.0),
        "preco_ha": float(preco_ha or 0.0),
        "preco_unit": float(preco_ha or 0.0),
        "tipo": "Mecanizada",
        "origem": "externo",
    }


def _to_float_json(v, default=0.0):
    if v is None:
        return float(default)
    if isinstance(v, (int, float)):
        try:
            return float(v)
        except Exception:
            return float(default)
    s = str(v).strip()
    if not s:
        return float(default)
    s = s.replace("R$", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return float(default)


def _candidatos_preco_final_json(cfg=None):
    out = []
    if isinstance(cfg, dict):
        p_cfg = str(cfg.get("preco_final_json_path", "") or "").strip()
        if p_cfg:
            p_cfg = os.path.expanduser(p_cfg)
            if not os.path.isabs(p_cfg):
                p_cfg = os.path.join(INPUT_DIR, p_cfg)
            out.append(p_cfg)
        out.append(os.path.join(INPUT_DIR, PRECO_FINAL_JSON_DEFAULT))
    out.append(PRECO_FINAL_JSON_DOWNLOADS)
    uniq = []
    seen = set()
    for p in out:
        ap = os.path.abspath(os.path.expanduser(str(p)))
        if ap in seen:
            continue
        seen.add(ap)
        uniq.append(ap)
    return uniq


def _score_payload_preco(payload):
    hh = float(payload.get("rendimento_hh", 0) or 0)
    hm = float(payload.get("rendimento_hm", 0) or 0)
    pr = float(payload.get("preco_ha", 0) or payload.get("preco_unit", 0) or 0)
    return (1 if hh > 0 else 0, hh, 1 if pr > 0 else 0, pr, hm)


def _carregar_mapa_preco_final_json(cfg=None):
    global _PRECO_FINAL_JSON_CACHE

    caminho = ""
    for c in _candidatos_preco_final_json(cfg):
        if os.path.isfile(c):
            caminho = c
            break
    if not caminho:
        return {}

    try:
        mtime = os.path.getmtime(caminho)
    except Exception:
        return {}

    cache = _PRECO_FINAL_JSON_CACHE or {}
    if (
        cache.get("path") == caminho
        and cache.get("mtime") == mtime
        and isinstance(cache.get("mapa"), dict)
    ):
        return dict(cache.get("mapa") or {})

    try:
        with open(caminho, "r", encoding="utf-8-sig") as f:
            raw = json.load(f)
    except Exception:
        return {}

    rows = []
    if isinstance(raw, list):
        rows = raw
    elif isinstance(raw, dict):
        for k in ("data", "rows", "tarifas", "atividades", "preco_final"):
            if isinstance(raw.get(k), list):
                rows = raw.get(k)
                break
        if not rows:
            # Aceita formato dict "nome" -> payload.
            if raw and all(isinstance(v, dict) for v in raw.values()):
                rows = []
                for k, v in raw.items():
                    item = dict(v)
                    item.setdefault("operacao", k)
                    rows.append(item)

    mapa = {}
    idx_norm = {}
    for item in rows:
        if not isinstance(item, dict):
            continue
        nome = str(
            item.get("operacao")
            or item.get("atividade")
            or item.get("nome")
            or item.get("descricao")
            or ""
        ).strip()
        if not nome:
            continue

        hh = _to_float_json(item.get("rendimento_hh_ha", item.get("rendimento_hh")), 0.0)
        hm = _to_float_json(
            item.get("rendimento_maquina_ha", item.get("rendimento_hm")), 0.0
        )
        preco = _to_float_json(
            item.get(
                "preco_rs",
                item.get("preco_ha", item.get("preco_unit", item.get("preco"))),
            ),
            0.0,
        )
        if hh > 0:
            hm = 0.0
        tipo = str(item.get("tipo") or "").strip()
        if not tipo:
            tipo = "Mecanizada" if hm > 0 and hh <= 0 else "Manual"

        payload = {
            "rendimento_hh": float(hh or 0.0),
            "rendimento_hm": float(hm or 0.0),
            "preco_ha": float(preco or 0.0),
            "preco_unit": float(preco or 0.0),
            "tipo": tipo,
        }
        if (
            payload["rendimento_hh"] <= 0
            and payload["rendimento_hm"] <= 0
            and payload["preco_ha"] <= 0
        ):
            continue

        nk = normalizar_chave(nome)
        if not nk:
            continue
        prev_nome = idx_norm.get(nk)
        if prev_nome is None:
            mapa[nome] = payload
            idx_norm[nk] = nome
        else:
            prev_payload = mapa.get(prev_nome, {})
            if _score_payload_preco(payload) >= _score_payload_preco(prev_payload):
                if prev_nome != nome:
                    mapa.pop(prev_nome, None)
                mapa[nome] = payload
                idx_norm[nk] = nome

    _PRECO_FINAL_JSON_CACHE = {"path": caminho, "mtime": mtime, "mapa": mapa}
    return dict(mapa)


def _aplicar_mapa_preco_final_em_tarifas(tarifas, mapa_json):
    if not isinstance(tarifas, dict) or not mapa_json:
        return 0
    alterados = 0
    idx_norm = {
        normalizar_chave(k): k
        for k in tarifas.keys()
        if isinstance(k, str) and normalizar_chave(k)
    }

    for nome, base in mapa_json.items():
        hh = float(base.get("rendimento_hh", 0) or 0)
        hm = float(base.get("rendimento_hm", 0) or 0)
        preco = float(base.get("preco_ha", 0) or base.get("preco_unit", 0) or 0)
        tipo = str(base.get("tipo") or "").strip()
        if hh > 0:
            hm = 0.0

        nk = normalizar_chave(nome)
        key = nome if nome in tarifas else idx_norm.get(nk)
        if key is None:
            key = nome
            tarifas[key] = {}
            if nk:
                idx_norm[nk] = key

        row = tarifas.get(key)
        if not isinstance(row, dict):
            row = {}
            tarifas[key] = row

        mudou = False

        if hh > 0:
            if float(row.get("rendimento_hh", 0) or 0) != hh:
                row["rendimento_hh"] = hh
                mudou = True
            if float(row.get("rendimento_hm", 0) or 0) != 0.0:
                row["rendimento_hm"] = 0.0
                mudou = True
            if tipo and str(row.get("tipo") or "").strip() != tipo:
                row["tipo"] = tipo
                mudou = True
            elif not row.get("tipo"):
                row["tipo"] = "Manual"
                mudou = True
        elif hm > 0 and float(row.get("rendimento_hh", 0) or 0) <= 0:
            if float(row.get("rendimento_hm", 0) or 0) != hm:
                row["rendimento_hm"] = hm
                mudou = True
            if tipo and str(row.get("tipo") or "").strip() != tipo:
                row["tipo"] = tipo
                mudou = True
            elif not row.get("tipo"):
                row["tipo"] = "Mecanizada"
                mudou = True

        if preco > 0 and float(row.get("preco_ha", 0) or 0) != preco:
            row["preco_ha"] = preco
            mudou = True
        if preco > 0 and float(row.get("preco_unit", 0) or 0) != preco:
            row["preco_unit"] = preco
            mudou = True

        hh_row = float(row.get("rendimento_hh", 0) or 0)
        hm_row = float(row.get("rendimento_hm", 0) or 0)
        if hh_row > 0 and hm_row > 0:
            row["rendimento_hm"] = 0.0
            hm_row = 0.0
            mudou = True

        recurso = "homem" if hh_row > 0 else ("maquina" if hm_row > 0 else "homem")
        if row.get("recurso") != recurso:
            row["recurso"] = recurso
            mudou = True
        if "eficiencia" not in row:
            row["eficiencia"] = 1.0
            mudou = True

        if mudou:
            alterados += 1
    return alterados


def _aplicar_mapa_preco_final_em_rows_by_name(rows_by_name, mapa_json, fonte_tag):
    if not isinstance(rows_by_name, dict) or not mapa_json:
        return 0
    alterados = 0
    idx_norm = {
        normalizar_chave(k): k
        for k in rows_by_name.keys()
        if isinstance(k, str) and normalizar_chave(k)
    }

    for nome, base in mapa_json.items():
        hh = float(base.get("rendimento_hh", 0) or 0)
        hm = float(base.get("rendimento_hm", 0) or 0)
        preco = float(base.get("preco_ha", 0) or base.get("preco_unit", 0) or 0)
        tipo = str(base.get("tipo") or "").strip()
        if hh > 0:
            hm = 0.0
        if not tipo:
            tipo = "Mecanizada" if hm > 0 and hh <= 0 else "Manual"

        nk = normalizar_chave(nome)
        key = nome if nome in rows_by_name else idx_norm.get(nk)
        if key is None:
            rows_by_name[nome] = {
                "atividade": nome,
                "tipo": tipo,
                "rendimento_hh": hh,
                "rendimento_hm": hm,
                "preco_ha": preco,
                "custo_hora": 0.0,
                "custo_ha": 0.0,
                "fonte_aba": fonte_tag,
            }
            if nk:
                idx_norm[nk] = nome
            alterados += 1
            continue

        row = rows_by_name.get(key)
        if not isinstance(row, dict):
            row = {}
            rows_by_name[key] = row

        mudou = False
        if str(row.get("atividade") or "").strip() != key:
            row["atividade"] = key
            mudou = True

        if hh > 0:
            if float(row.get("rendimento_hh", 0) or 0) != hh:
                row["rendimento_hh"] = hh
                mudou = True
            if float(row.get("rendimento_hm", 0) or 0) != 0.0:
                row["rendimento_hm"] = 0.0
                mudou = True
            if str(row.get("tipo") or "").strip() != (tipo or "Manual"):
                row["tipo"] = tipo or "Manual"
                mudou = True
        elif hm > 0 and float(row.get("rendimento_hh", 0) or 0) <= 0:
            if float(row.get("rendimento_hm", 0) or 0) != hm:
                row["rendimento_hm"] = hm
                mudou = True
            if tipo and str(row.get("tipo") or "").strip() != tipo:
                row["tipo"] = tipo
                mudou = True

        if preco > 0 and float(row.get("preco_ha", 0) or 0) != preco:
            row["preco_ha"] = preco
            mudou = True

        if float(row.get("rendimento_hh", 0) or 0) > 0 and float(
            row.get("rendimento_hm", 0) or 0
        ) > 0:
            row["rendimento_hm"] = 0.0
            mudou = True

        if mudou:
            row["fonte_aba"] = fonte_tag
            alterados += 1
    return alterados


def _depara_heuristico_exame_ct317(kn, tarifas):
    """
    Fallback heurístico foi DELIBERADAMENTE DESATIVADO para não impor horas/homem erradas
    com nomes legados de planilhas antigas. Agora a ferramenta utilizará os exatos nomes
    da CT317 presente na pasta. Se não encontrar um nome igual, a planilha pedirá mapeamento manual.
    """
    return None



def _find_preco_final_sheet(xls):
    for s in xls.sheet_names:
        ns = remover_acentos(s).replace(" ", "").replace("_", "").replace("-", "")
        if "precofinal" in ns:
            return s
    return None


def normalizar_ct313(caminho_ct):
    """
    Le CT_313 bruta e gera CT_313_NORMALIZADA.xlsx com aba STG_TARIFAS.
    Retorna (caminho_stg, n_linhas, custo_hora_tf).
    """
    xls = pd.ExcelFile(caminho_ct)
    pf = _find_preco_final_sheet(xls)
    if not pf:
        return None, 0, 0.0

    # Layout antigo (indices fixos) e layout CT317 real (cabecalho na linha 1)
    # coexistem. Aqui tentamos primeiro por cabecalho real; se falhar, caimos para
    # o parser legado por indice.
    dfh = pd.read_excel(caminho_ct, sheet_name=pf)

    custo_hora_tf = 0.0
    rows_by_name = {}

    def col_by_tokens(df, token_sets):
        cols = list(df.columns)
        for c in cols:
            nc = normalizar_chave(c)
            for toks in token_sets:
                if all(t in nc for t in toks):
                    return c
        return None

    col_nome = col_by_tokens(
        dfh,
        [["operac"], ["atividade"], ["descricao"], ["servico"]],
    )
    col_tipo = col_by_tokens(dfh, [["tipo"]])
    col_hh = col_by_tokens(
        dfh,
        [["rendimento", "hh"], ["homem", "hora"], ["hh", "ha"]],
    )
    col_hm = col_by_tokens(
        dfh,
        [["rendimento", "maq"], ["rendimento", "maquina"], ["hm"], ["maquina", "ha"]],
    )
    col_preco = col_by_tokens(dfh, [["preco"], ["tarifa"], ["valor"]])
    col_custo_h = col_by_tokens(dfh, [["custo", "hora"], ["r", "h"]])

    # Parser moderno (CT317 real)
    if col_nome and (col_hh or col_hm):
        for _, r in dfh.iterrows():
            nome = str(r.get(col_nome, "")).strip()
            if not nome or nome.lower() in {"nan", "none"}:
                continue
            if normalizar_chave(nome) in {"operacoes", "operacao", "atividade"}:
                continue
            hh = _to_float_any(r.get(col_hh)) if col_hh else 0.0
            hm = _to_float_any(r.get(col_hm)) if col_hm else 0.0
            preco = _to_float_any(r.get(col_preco)) if col_preco else 0.0
            tipo = str(r.get(col_tipo, "")).strip() if col_tipo else ""
            custo_h = _to_float_any(r.get(col_custo_h)) if col_custo_h else 0.0
            hh = float(hh or 0.0)
            hm = float(hm or 0.0)
            preco = float(preco or 0.0)
            custo_h = float(custo_h or 0.0)
            if hh <= 0 and hm <= 0 and preco <= 0:
                continue
            prev = rows_by_name.get(nome)
            payload = {
                "atividade": nome,
                "tipo": tipo or ("Mecanizada" if hm > 0 else "Manual"),
                "rendimento_hh": hh,
                "rendimento_hm": hm,
                "preco_ha": preco,
                "custo_hora": custo_h,
                "custo_ha": (hh * custo_h) if custo_h > 0 else 0.0,
                "fonte_aba": pf,
            }
            if prev is None:
                rows_by_name[nome] = payload
            else:
                # Mantem a linha mais informativa (HH/HM/preco/custo maiores).
                prev_score = (
                    float(prev.get("rendimento_hh", 0) or 0)
                    + float(prev.get("rendimento_hm", 0) or 0)
                    + float(prev.get("preco_ha", 0) or 0)
                    + float(prev.get("custo_hora", 0) or 0)
                )
                cur_score = hh + hm + preco + custo_h
                if cur_score >= prev_score:
                    rows_by_name[nome] = payload

    # Fallback legado (layout por indice fixo)
    if len(rows_by_name) < 20:
        df = pd.read_excel(caminho_ct, sheet_name=pf, header=None)
        for i in range(5, len(df)):
            r = df.iloc[i]
            nome = str(r[2]).strip() if pd.notna(r[2]) else ""
            if not nome or nome == "0":
                continue
            tipo = str(r[4]).strip() if pd.notna(r[4]) else ""
            try:
                hh = float(r[5]) if pd.notna(r[5]) else 0.0
            except (TypeError, ValueError):
                hh = 0.0
            try:
                hm = float(r[6]) if pd.notna(r[6]) else 0.0
            except (TypeError, ValueError):
                hm = 0.0
            try:
                preco = float(r[7]) if pd.notna(r[7]) else 0.0
            except (TypeError, ValueError):
                preco = 0.0
            if hh <= 0 and hm <= 0 and preco <= 0:
                continue
            rows_by_name[nome] = {
                "atividade": nome,
                "tipo": tipo or ("Mecanizada" if hm > 0 else "Manual"),
                "rendimento_hh": hh,
                "rendimento_hm": hm,
                "preco_ha": preco,
                "custo_hora": 0.0,
                "custo_ha": 0.0,
                "fonte_aba": pf,
            }

    # Fonte oficial adicional: preco_final.json (projeto/Downloads) com HH/HM/preco.
    # Regra de negocio: quando HH existir, HM e zerado.
    mapa_json = _carregar_mapa_preco_final_json()
    if mapa_json:
        _aplicar_mapa_preco_final_em_rows_by_name(
            rows_by_name, mapa_json, f"{pf}|preco_final_json"
        )

    # Fallback hardcoded de HH para operacao basica quando o parser nao conseguir
    # popular uma tarifa essencial do fluxo ATM.
    for nome, base in CT317_HARDCODE_HH_BASE.items():
        cur = rows_by_name.get(nome)
        if cur is None:
            cur = {
                "atividade": nome,
                "tipo": base.get("tipo", "Manual"),
                "rendimento_hh": float(base.get("rendimento_hh", 0) or 0),
                "rendimento_hm": float(base.get("rendimento_hm", 0) or 0),
                "preco_ha": 0.0,
                "custo_hora": 0.0,
                "custo_ha": 0.0,
                "fonte_aba": f"{pf}|hardcoded_hh",
            }
            rows_by_name[nome] = cur
            continue
        hh_cur = float(cur.get("rendimento_hh", 0) or 0)
        hm_cur = float(cur.get("rendimento_hm", 0) or 0)
        hh_base = float(base.get("rendimento_hh", 0) or 0)
        hm_base = float(base.get("rendimento_hm", 0) or 0)
        if hh_base > 0 and hh_cur <= 0:
            cur["rendimento_hh"] = hh_base
            cur["rendimento_hm"] = 0.0
            cur["tipo"] = base.get("tipo", "Manual")
            cur["fonte_aba"] = f"{pf}|hardcoded_hh"
        elif hh_cur <= 0 and hm_cur <= 0:
            cur["rendimento_hh"] = hh_base
            cur["rendimento_hm"] = hm_base
            cur["tipo"] = cur.get("tipo") or base.get("tipo", "Manual")
            cur["fonte_aba"] = f"{pf}|hardcoded_hh"
        if float(cur.get("rendimento_hh", 0) or 0) > 0 and float(
            cur.get("rendimento_hm", 0) or 0
        ) > 0:
            cur["rendimento_hm"] = 0.0

    rows = list(rows_by_name.values())

    # Custo hora base: mediana dos custos hora lidos da planilha.
    custos_h_validos = [float(r.get("custo_hora", 0) or 0) for r in rows if float(r.get("custo_hora", 0) or 0) > 0]
    if custos_h_validos:
        custo_hora_tf = float(median(custos_h_validos))

    # Se algum item de HH nao tiver custo_hora, aplica custo_hora_tf para manter custo_ha coerente.
    for r in rows:
        hh = float(r.get("rendimento_hh", 0) or 0)
        ch = float(r.get("custo_hora", 0) or 0)
        if hh > 0 and ch <= 0 and custo_hora_tf > 0:
            ch = custo_hora_tf
            r["custo_hora"] = ch
        r["custo_ha"] = (hh * ch) if ch > 0 else 0.0

    df_stg = pd.DataFrame(rows)
    meta = pd.DataFrame(
        [
            {
                "gerado_em": datetime.datetime.now().isoformat(),
                "arquivo_origem": os.path.basename(caminho_ct),
                "linhas_validas": len(rows),
                "custo_hora_tf": round(custo_hora_tf, 4),
            }
        ]
    )

    stg_path = os.path.join(INPUT_DIR, STG_FILENAME)
    with pd.ExcelWriter(stg_path, engine="openpyxl") as w:
        df_stg.to_excel(w, sheet_name="STG_TARIFAS", index=False)
        meta.to_excel(w, sheet_name="STG_METADATA", index=False)

    return stg_path, len(rows), custo_hora_tf


def carregar_stg_tarifas(stg_path):
    """Le STG_TARIFAS e retorna dict {atividade: {rendimento_hh, preco_ha, custo_hora, custo_ha, tipo}}."""
    df = pd.read_excel(stg_path, sheet_name="STG_TARIFAS")
    t = {}
    for _, r in df.iterrows():
        nome = str(r.get("atividade", "")).strip()
        if not nome:
            continue
        hh = float(r.get("rendimento_hh") or 0)
        hm = float(r.get("rendimento_hm") or 0)
        t[nome] = {
            "rendimento_hh": hh,
            "rendimento_hm": hm,
            "preco_ha": float(r.get("preco_ha") or 0),
            "custo_hora": float(r.get("custo_hora") or 0),
            "custo_ha": float(r.get("custo_ha") or 0),
            "tipo": str(r.get("tipo") or ""),
            "preco_unit": float(r.get("preco_ha") or 0),
            "recurso": "homem" if hh > 0 else "maquina",
            "eficiencia": 1.0,
        }
    return t


def modulo_normalizar_ct(cfg):
    """Menu: selecionar CT bruta, gerar STG, integrar em config.tarifas."""
    dashboard_header()
    subcabecalho("NORMALIZAR CT (CT317 REAL) -> STG_TARIFAS")
    caminho = selecionar_arquivo("CT BRUTA/REAL (.xlsm ou .xlsx)")
    if not caminho:
        return

    print(DM + "  Processando... pode demorar alguns segundos." + RS)
    stg_path, n, custo_h = normalizar_ct313(caminho)
    if not stg_path:
        erro("Aba 'Preco Final' nao encontrada neste arquivo.")
        input(DM + "\n  [ENTER] " + RS)
        return

    if MODO_SOMENTE_HH:
        ok(f"Gerado {STG_FILENAME}: {n} atividades (modo somente HH).")
    else:
        ok(f"Gerado {STG_FILENAME}: {n} atividades | custo/hora TF = R${custo_h:.2f}")

    if confirmar(
        "Integrar STG_TARIFAS em config.json (substitui tarifas existentes)?",
        default=True,
    ):
        tarifas = carregar_stg_tarifas(stg_path)
        cfg["tarifas"] = tarifas
        cfg["custo_hora_tf"] = round(custo_h, 4)
        salvar_config(cfg)
        ok(f"{len(tarifas)} tarifas integradas no config.")
    input(DM + "\n  [ENTER para voltar] " + RS)


def _to_float_any(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    s = s.replace("R$", "").replace(" ", "")
    if "," in s and "." in s:
        # Ex.: 1.234,56 -> 1234.56
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def resolver_chave_tarifa(cfg, tarifas, atv):
    """
    Resolve a chave de tarifa para uma atividade do micro.
    Prioridade:
    1) de_para[atividade] quando existir e estiver em tarifas;
    2) nome original da atividade quando existir em tarifas;
    3) fallback para a chave mapeada (mesmo ausente) para manter diagnostico claro.
    """
    de_para = cfg.get("de_para", {}) or {}
    tarifas = tarifas or {}

    def _find_key_norm(target, keys):
        nt = normalizar_chave(target)
        if not nt:
            return None
        for k in keys:
            if normalizar_chave(k) == nt:
                return k
        for k in keys:
            nk = normalizar_chave(k)
            if nt in nk or nk in nt:
                return k
        return None

    # 1) de_para exato
    t_map = de_para.get(atv)
    # 1b) de_para por chave normalizada (evita mismatch de acento/espacos/caixa)
    if not t_map:
        natv = normalizar_chave(atv)
        for km, vm in de_para.items():
            if normalizar_chave(km) == natv:
                t_map = vm
                break
    t_map = str(t_map or atv)

    # 2) busca exata em tarifas
    if t_map in tarifas:
        return t_map
    if atv in tarifas:
        return atv

    # 3) busca normalizada nas tarifas
    k_norm = _find_key_norm(t_map, tarifas.keys())
    if k_norm:
        return k_norm
    a_norm = _find_key_norm(atv, tarifas.keys())
    if a_norm:
        return a_norm

    # 4) fallback diagnostico
    return t_map


def modulo_mapeamentos_de_para(cfg, df_micro=None):
    """CRUD de_para: nome no microplanejamento -> nome da tarifa em config.tarifas."""
    tarifas = cfg.get("tarifas", {})
    nomes_tarifa = sorted(tarifas.keys(), key=lambda x: str(x))
    atividades_micro = []
    if (
        df_micro is not None
        and getattr(df_micro, "columns", None) is not None
        and "atividade" in df_micro.columns
    ):
        atividades_micro = sorted(
            df_micro["atividade"].dropna().unique().tolist(), key=str
        )

    while True:
        dashboard_header()
        subcabecalho("MAPEAMENTOS de_para (micro -> tarifa)")
        d = cfg.get("de_para", {})
        pairs = [(k, v) for k, v in d.items() if not str(k).startswith("_")]
        if not pairs:
            print(
                DM
                + "  Nenhum par (o sistema usa nome micro = nome na tarifa, ou default 8 h/ha)."
                + RS
            )
        else:
            for k, v in sorted(pairs, key=lambda x: str(x[0]))[:35]:
                print(G + f"  {str(k)[:36]:36} -> " + C + f"{str(v)[:36]}" + RS)
            if len(pairs) > 35:
                print(DM + f"  ... +{len(pairs) - 35} pares no arquivo" + RS)
        sub()
        print(DM + "  [1] Incluir ou alterar par" + RS)
        print(DM + "  [2] Remover par" + RS)
        print(DM + "  [3] Listar catalogo de TARIFAS (nomes em config)" + RS)
        print(DM + "  [0] Voltar" + RS)
        op = prompt("Opcao").strip()
        if op == "0":
            return
        if op == "1":
            chave_micro = ""
            if atividades_micro and confirmar(
                "Escolher atividade da planilha carregada?", default=True
            ):
                idx = selecionar_paginado(
                    "ATIVIDADE no micro", atividades_micro, page_size=8
                )
                if idx >= 0:
                    chave_micro = atividades_micro[idx]
            if not chave_micro:
                chave_micro = prompt("Nome EXATO da atividade no microplanejamento", "")
            if not chave_micro:
                aviso("Nome vazio.")
                continue
            val_tarifa = ""
            if nomes_tarifa and confirmar(
                "Escolher tarifa na lista importada?", default=True
            ):
                idx = selecionar_paginado(
                    "TARIFA (orcamento)", nomes_tarifa, page_size=8
                )
                if idx >= 0:
                    val_tarifa = nomes_tarifa[idx]
            if not val_tarifa:
                val_tarifa = prompt("Nome da TARIFA (chave em tarifas)", "")
            if not val_tarifa:
                aviso("Tarifa vazio.")
                continue
            if val_tarifa not in tarifas:
                if not confirmar(
                    f"  '{str(val_tarifa)[:42]}' nao esta em tarifas. Gravar mesmo assim?",
                    default=False,
                ):
                    continue
            cfg.setdefault("de_para", {})
            cfg["de_para"][chave_micro] = val_tarifa
            salvar_config(cfg)
            ok("Mapeamento salvo em config.json.")
        elif op == "2":
            keys = sorted([k for k in d.keys() if not str(k).startswith("_")], key=str)
            if not keys:
                aviso("Nada para remover.")
                continue
            idx = selecionar_paginado("REMOVER mapeamento", [str(k) for k in keys])
            if idx >= 0:
                del cfg["de_para"][keys[idx]]
                salvar_config(cfg)
                ok("Removido.")
        elif op == "3":
            if not nomes_tarifa:
                aviso("Nenhuma tarifa em config. Use menu [2] Importar.")
            else:
                for i, n in enumerate(nomes_tarifa[:60], 1):
                    print(DM + f"  {i:3}. {str(n)[:58]}" + RS)
                if len(nomes_tarifa) > 60:
                    print(DM + f"  ... +{len(nomes_tarifa) - 60}" + RS)
            input(DM + "\n  [ENTER] " + RS)
        else:
            aviso("Opcao invalida.")


# ──────────────────────────────────────────────
#  COLUMN MAPPING: KNOWN FIRST, THEN FALLBACK
# ──────────────────────────────────────────────
def encontrar_coluna(cols, campo):
    """Try KNOWN_COLUMNS exact matches first, then fuzzy fallback."""
    known = KNOWN_COLUMNS.get(campo, [])
    for k in known:
        if k in cols:
            return k
    # Fuzzy fallback
    cn_map = {remover_acentos(c): c for c in cols}
    for k in known:
        kn = remover_acentos(k)
        for cn, original in cn_map.items():
            if kn in cn or cn in kn:
                return original
    return None


# ──────────────────────────────────────────────
#  FILE SELECTOR
# ──────────────────────────────────────────────
def buscar_arquivos_excel():
    return [
        f
        for f in os.listdir(INPUT_DIR)
        if not f.startswith("~")
        and any(f.endswith(e) for e in [".xlsx", ".xls", ".csv", ".xlsm"])
    ]


def _find_default_micro_path(cfg=None):
    """
    Prioridade: config.arquivo_micro (se existir no disco) > lista fixa > primeiro .xlsx com
    'inovesa'/'consolidado'/'exame'/'micro' no nome.
    """
    if cfg:
        pref = cfg.get("arquivo_micro")
        if pref and isinstance(pref, str):
            p = os.path.join(INPUT_DIR, pref.strip())
            if os.path.exists(p):
                return p
        candidatos = [
            "MICROPLANEJAMENTO_CONSOLIDADO_INOVESA 1.xlsx",
            "exame.xlsx",
            "EXAME.xlsx",
            "Exame.xlsx",
            "microplanejamento.xlsx",
            "MICROPLANEJAMENTO.xlsx",
        ]
        for c in candidatos:
            p = os.path.join(INPUT_DIR, c)
            if os.path.exists(p):
                return p
        for f in buscar_arquivos_excel():
            n = remover_acentos(f)
            if "inovesa" in n or "consolidado" in n:
                return os.path.join(INPUT_DIR, f)
        for f in buscar_arquivos_excel():
            n = remover_acentos(f)
            if "exame" in n or "micro" in n:
                return os.path.join(INPUT_DIR, f)
        return None


def _prefer_micro_sheet(abas):
    skip_prefixes = ("planilha", "previs")
    for a in abas:
        if "microplanejamento_abril_junho" in remover_acentos(a).replace(" ", ""):
            return a
    for a in abas:
        na = remover_acentos(a).lower()
        if any(na.startswith(p) for p in skip_prefixes):
            continue
        if "micropl" in na and "_v5" in na:
            return a
    for a in abas:
        na = remover_acentos(a).lower()
        if any(na.startswith(p) for p in skip_prefixes):
            continue
        if "micropl" in na and "impl" in na and ("_v3" in na or "_v4" in na):
            continue
        if "microplanejamento" in na:
            return a
    for a in abas:
        if "inovesa" in remover_acentos(a) or "consolidado" in remover_acentos(a):
            return a
    for a in abas:
        na = remover_acentos(a).lower()
        if any(na.startswith(p) for p in skip_prefixes):
            continue
        if "micropl" in na:
            return a
    return abas[0] if abas else None


def _find_default_ct_path():
    preferido = os.path.join(INPUT_DIR, CT_REAL_FILENAME)
    if os.path.exists(preferido):
        return preferido

    for f in buscar_arquivos_excel():
        if f == STG_FILENAME:
            continue
        n = remover_acentos(f)
        if "ct317" in n or ("ct" in n and "317" in n):
            return os.path.join(INPUT_DIR, f)

    for f in buscar_arquivos_excel():
        if f == STG_FILENAME:
            continue
        n = remover_acentos(f)
        if "ct_313" in n or ("ct" in n and "313" in n):
            return os.path.join(INPUT_DIR, f)
    return None


def selecionar_arquivo(titulo="Selecione um Arquivo"):
    """
    Seletor com navegacao de pastas.
    - [DIR] .. : sobe um nivel
    - [DIR] nome: entra na pasta
    - [ARQ] nome: seleciona arquivo
    """
    exts = (".xlsx", ".xls", ".xlsm", ".csv")
    cwd = INPUT_DIR
    while True:
        try:
            nomes = sorted(os.listdir(cwd), key=lambda s: normalizar_chave(str(s)))
        except Exception as ex:
            erro(f"Nao foi possivel abrir pasta: {ex}")
            return None

        entries = []
        parent = os.path.dirname(cwd.rstrip("\\/"))
        if parent and parent != cwd:
            entries.append(("dir_up", "..", parent))

        for n in nomes:
            p = os.path.join(cwd, n)
            if os.path.isdir(p):
                if n.startswith("."):
                    continue
                entries.append(("dir", n, p))
        for n in nomes:
            p = os.path.join(cwd, n)
            if os.path.isfile(p):
                ln = n.lower()
                if n.startswith("~") or n.startswith("~$"):
                    continue
                if ln.endswith(exts):
                    entries.append(("file", n, p))

        if not entries:
            aviso(f"Nada encontrado em: {cwd}")
            if cwd == INPUT_DIR:
                return None
            cwd = parent if parent else INPUT_DIR
            continue

        labels = []
        for kind, name, _ in entries:
            if kind == "dir_up":
                labels.append("[DIR] ..")
            elif kind == "dir":
                labels.append(f"[DIR] {name}")
            else:
                labels.append(f"[ARQ] {name}")

        subcabecalho(titulo)
        print(DM + f"  Pasta atual: {cwd}" + RS)
        sub()
        idx = selecionar_paginado(titulo, labels, page_size=12, zero_label="Cancelar")
        if idx < 0:
            return None
        kind, name, path = entries[idx]
        if kind in ("dir_up", "dir"):
            cwd = path
            continue
        return path


# ──────────────────────────────────────────────
#  MICROPLANEJAMENTO
# ──────────────────────────────────────────────
def carregar_planilha_microplanejamento(cfg, caminho=None, modo_auto=False):
    if not caminho:
        caminho = selecionar_arquivo("MICROPLANEJAMENTO (ex. exame.xlsx)")
    if not caminho:
        return None
    try:
        xls = pd.ExcelFile(caminho)
        abas = xls.sheet_names
        if len(abas) == 1:
            aba = abas[0]
        else:
            if modo_auto:
                aba = _prefer_micro_sheet(abas)
            else:
                aba = selecionar("SELECIONE A ABA", abas)
                if aba is None:
                    return None

        df = pd.read_excel(caminho, sheet_name=aba, header=0)
        cols = df.columns.tolist()

        # Mapear com KNOWN_COLUMNS primeiro
        faz_col = encontrar_coluna(cols, "fazenda")
        chv_col = encontrar_coluna(cols, "chave")
        area_col = encontrar_coluna(cols, "area")
        atv_col = encontrar_coluna(cols, "atividade")

        # BETA: planilhas de testes podem nao ter CHAVE POLIGONO; usar NUCLEO como fallback automatico.
        if modo_auto and not chv_col:
            for c in cols:
                if "nucleo" in normalizar_chave(c):
                    chv_col = c
                    break

        # Se algum nao bateu, pedir manual
        mapeados = {
            "Fazenda": faz_col,
            "Talhao": chv_col,
            "Area(ha)": area_col,
            "Atividade": atv_col,
        }
        todos_ok = all(v is not None for v in mapeados.values())

        sub()
        print(G + BL + "  MAPEAMENTO AUTOMATICO:" + RS)
        for label, val in mapeados.items():
            cor = G if val else R
            print(cor + f"  {label:12}: " + C + f"{val or '??? NAO ENCONTRADO'}" + RS)
        sub()

        if todos_ok and (modo_auto or confirmar("Usar este mapeamento?", default=True)):
            pass  # tudo certo
        else:
            if modo_auto:
                erro("Nao foi possivel mapear colunas do micro automaticamente.")
                return None
            print(G + "\n  Selecione manualmente cada coluna:\n" + RS)
            if not faz_col:
                idx = selecionar_paginado("COLUNA DA FAZENDA", cols)
                faz_col = cols[idx] if idx >= 0 else cols[0]
            if not chv_col:
                idx = selecionar_paginado("COLUNA DO TALHAO/CHAVE", cols)
                chv_col = cols[idx] if idx >= 0 else cols[1]
            if not area_col:
                idx = selecionar_paginado("COLUNA DA AREA (HECTARES)", cols)
                area_col = cols[idx] if idx >= 0 else cols[2]
            if not atv_col:
                idx = selecionar_paginado("COLUNA DA ATIVIDADE", cols)
                atv_col = cols[idx] if idx >= 0 else cols[3]

        sel_cols = [faz_col, chv_col, area_col, atv_col]
        sel_names = ["fazenda", "chave", "area_ha", "atividade"]

        equipe_col = None
        for c in cols:
            if normalizar_chave(c) in ("equipe", "nome equipe", "empresa"):
                equipe_col = c
                break
        if equipe_col:
            sel_cols.append(equipe_col)
            sel_names.append("equipe")
            print(G + f"  Coluna EQUIPE detectada: " + C + f"{equipe_col}" + RS)

        metodologia_col = None
        pref_metodologia = str(cfg.get("coluna_metodologia_micro", "") or "").strip()
        if pref_metodologia and pref_metodologia in cols:
            metodologia_col = pref_metodologia
        metodologia_labels = {
            "metodologia",
            "tipo metodologia",
            "tipo de metodologia",
            "metodologia talhao",
            "metodologia atividade",
            "metodo",
        }
        if not metodologia_col:
            for c in cols:
                nc = normalizar_chave(c)
                if not nc:
                    continue
                if nc in metodologia_labels or "metodolog" in nc:
                    metodologia_col = c
                    break

        if not metodologia_col:
            # Fallback heuristico: coluna textual com poucos valores e termos de metodologia.
            chaves_metodo = (
                "implant",
                "conduc",
                "manut",
                "adapt",
                "adap",
                "restaur",
                "passiv",
            )
            for c in cols:
                if c in sel_cols:
                    continue
                serie = df[c].dropna()
                if serie.empty:
                    continue
                amostra = serie.astype(str).str.strip()
                amostra = amostra[~amostra.str.lower().isin(["nan", "none", ""])]
                if amostra.empty:
                    continue
                unicos = sorted(
                    {
                        normalizar_chave(v)
                        for v in amostra.tolist()
                        if normalizar_chave(v)
                    }
                )
                if not unicos or len(unicos) > 30:
                    continue
                hits = sum(
                    1 for v in unicos if any(k in v for k in chaves_metodo)
                )
                if hits >= max(1, len(unicos) // 5):
                    metodologia_col = c
                    break

        if not metodologia_col and not modo_auto:
            if confirmar(
                "Nao detectei a coluna de METODOLOGIA automaticamente. Selecionar manualmente?",
                default=False,
            ):
                cols_disp = [c for c in cols if c not in sel_cols]
                if cols_disp:
                    idx = selecionar_paginado("COLUNA DA METODOLOGIA", cols_disp)
                    if idx >= 0:
                        metodologia_col = cols_disp[idx]

        if metodologia_col:
            sel_cols.append(metodologia_col)
            sel_names.append("metodologia")
            cfg["coluna_metodologia_micro"] = metodologia_col
            salvar_config(cfg)
            print(
                G + f"  Coluna METODOLOGIA detectada: " + C + f"{metodologia_col}" + RS
            )

        df_filtro = df[sel_cols].copy()
        df_filtro.columns = sel_names
        for c_txt in ("fazenda", "chave", "atividade", "equipe", "metodologia"):
            if c_txt in df_filtro.columns:
                df_filtro[c_txt] = (
                    df_filtro[c_txt]
                    .astype(str)
                    .str.replace(r"\s+", " ", regex=True)
                    .str.strip()
                )
                df_filtro.loc[
                    df_filtro[c_txt].str.lower().isin(["nan", "none", ""]), c_txt
                ] = None
        df_filtro = df_filtro.dropna(subset=["atividade", "area_ha"])
        df_filtro["area_ha"] = df_filtro["area_ha"].apply(
            lambda x: _to_float_br(x, default=0.0)
        )

        validos = df_filtro[df_filtro["area_ha"] > 0]

        equipe_vazia = (
            "equipe" not in validos.columns
            or validos["equipe"].dropna().str.strip().replace("", None).dropna().empty
        )
        if equipe_vazia:
            try:
                xls_eq = pd.ExcelFile(caminho)
                eq_sheet = None
                for s in xls_eq.sheet_names:
                    if normalizar_chave(s).startswith("planilha7") or (
                        "logistic" in normalizar_chave(s)
                    ):
                        eq_sheet = s
                        break
                if not eq_sheet:
                    for s in xls_eq.sheet_names:
                        sh = pd.read_excel(xls_eq, sheet_name=s, nrows=2)
                        cols_s = [normalizar_chave(c) for c in sh.columns]
                        if any("equipe" in c for c in cols_s):
                            has_faz = any("fazenda" in c for c in cols_s)
                            if has_faz and len(sh.columns) <= 20:
                                eq_sheet = s
                                break
                if eq_sheet:
                    df_eq = pd.read_excel(xls_eq, sheet_name=eq_sheet)
                    eq_col = None
                    fz_col = None
                    for c in df_eq.columns:
                        if "equipe" in normalizar_chave(c) and not eq_col:
                            eq_col = c
                        if "fazenda" in normalizar_chave(c) and not fz_col:
                            fz_col = c
                    if eq_col and fz_col:
                        map_eq = dict(
                            zip(
                                df_eq[fz_col].astype(str).str.strip(),
                                df_eq[eq_col].astype(str).str.strip(),
                            )
                        )
                        map_eq = {k: v for k, v in map_eq.items() if v and v.lower() not in ("nan", "none", "")}
                        if map_eq:
                            def _match_equipe(faz_name, mapping):
                                fn = str(faz_name).strip()
                                if fn in mapping:
                                    return mapping[fn]
                                fn_n = remover_acentos(fn).lower()
                                for k, v in mapping.items():
                                    if remover_acentos(k).lower() == fn_n:
                                        return v
                                for k, v in mapping.items():
                                    if remover_acentos(k).lower() in fn_n or fn_n in remover_acentos(k).lower():
                                        return v
                                return None

                            if "equipe" not in validos.columns:
                                validos = validos.copy()
                                validos["equipe"] = validos["fazenda"].apply(
                                    lambda f: _match_equipe(f, map_eq)
                                )
                            else:
                                mask_empty = validos["equipe"].isna() | (validos["equipe"].str.strip() == "")
                                validos.loc[mask_empty, "equipe"] = validos.loc[mask_empty, "fazenda"].apply(
                                    lambda f: _match_equipe(f, map_eq)
                                )
                            n_assigned = validos["equipe"].notna().sum()
                            if n_assigned > 0:
                                ok(f"EQUIPE enriquecida a partir de '{eq_sheet}': {n_assigned} registros atribuidos")
            except Exception:
                pass

        ok(f"Carregadas {len(validos)} atividades validas.")
        return validos
    except Exception as e:
        erro(f"Erro ao ler microplanejamento: {e}")
        return None


# ──────────────────────────────────────────────
#  IMPORTADOR CT_313
# ──────────────────────────────────────────────
def modulo_importar_tarifas(cfg):
    dashboard_header()
    subcabecalho("IMPORTAR TARIFAS ORCADAS (CT_313)")
    caminho = selecionar_arquivo("PLANILHA DE ORCAMENTO (CT_313 ou Tarifas)")
    if not caminho:
        return

    try:
        print(DM + "  Carregando arquivo..." + RS)
        xls = pd.ExcelFile(caminho)
        aba = selecionar("SELECIONE A ABA (ex: Preco Final)", xls.sheet_names)
        if aba is None:
            return

        print(DM + f"  Lendo aba '{aba}'..." + RS)
        df = pd.read_excel(caminho, sheet_name=aba, nrows=1000)
        cols_ct = df.columns.tolist()

        # Tentar mapear automaticamente
        col_atv = encontrar_coluna(cols_ct, "atividade")
        sub()
        print(G + BL + "  MAPEAMENTO:" + RS)
        print(G + f"  Atividade: " + C + f"{col_atv or '???'}" + RS)
        sub()

        if not col_atv or not confirmar("Usar este mapeamento?", default=True):
            idx = selecionar_paginado("COLUNA DA ATIVIDADE", cols_ct)
            col_atv = cols_ct[idx] if idx >= 0 else None
            if not col_atv:
                aviso("Atividade obrigatoria.")
                return

        # Para HH e Preco, perguntar diretamente
        print(G + "\n  Selecione as colunas adicionais (0 = ignorar):\n" + RS)
        idx = selecionar_paginado("COLUNA DE HH/HA", cols_ct)
        col_hh = cols_ct[idx] if idx >= 0 else None
        idx = selecionar_paginado("COLUNA DE PRECO UNITARIO", cols_ct)
        col_preco = cols_ct[idx] if idx >= 0 else None

        tarifas = cfg.get("tarifas", {})
        importadas = 0
        for _, row in df.iterrows():
            nome = str(row.get(col_atv, "")).strip()
            if not nome or nome.lower() == "nan":
                continue
            hh = 0 if not col_hh else row.get(col_hh, 0)
            preco = 0 if not col_preco else row.get(col_preco, 0)
            if pd.notna(hh) and str(hh).strip() != "":
                hh_val = float(str(hh).replace(",", "."))
            else:
                hh_val = resolver_rendimento_hh(cfg, tarifas, nome)
            preco_val = float(str(preco).replace(",", ".")) if pd.notna(preco) else 0.0
            tarifas[nome] = {
                "rendimento_hh": hh_val,
                "preco_unit": preco_val,
                "recurso": "homem",
                "eficiencia": 1.0,
            }
            importadas += 1

        cfg["tarifas"] = tarifas
        salvar_config(cfg)
        ok(f"{importadas} tarifas integradas!")
        sem_hh = [
            k for k, v in tarifas.items() if float(v.get("rendimento_hh", 0) or 0) <= 0
        ]
        sem_preco = [
            k for k, v in tarifas.items() if float(v.get("preco_unit", 0) or 0) <= 0
        ]
        if sem_hh:
            print(Y + f"  Pos-import: {len(sem_hh)} tarifa(s) com HH zerado." + RS)
            for x in sem_hh[:5]:
                print(DM + f"    - {str(x)[:55]}" + RS)
        if sem_preco:
            print(
                Y + f"  Pos-import: {len(sem_preco)} tarifa(s) com preco zerado." + RS
            )
            for x in sem_preco[:5]:
                print(DM + f"    - {str(x)[:55]}" + RS)
    except Exception as e:
        erro(f"Erro ao importar: {e}")

    input(DM + "\n  [ENTER para voltar] " + RS)


# ──────────────────────────────────────────────
#  DECLIVIDADE
# ──────────────────────────────────────────────
def aviso_politica_tarifas_planas():
    """Politica comercial-executiva: base CT sempre 'plana' (Classe I) onde o micro nao discrimina."""
    sub()
    print(Y + BL + "  POLITICA DE DECLIVIDADE E ROÇADA MANUAL (CT)" + RS)
    print(
        DM
        + "  Na CT, ROÇADA MANUAL CLASSE I = terreno mais plano (menos HH/ha, menor R$/ha); "
        "CLASSE V = declive maximo (mais HH, mais R$/ha — obra mais cara e precos mais altos)."
        + RS
    )
    print(
        Y
        + "  Padrao deste app: o exame nao informa a classe por talhao — usamos sempre as linhas "
        "EQUIVALENTES AO CENARIO MAIS PLANO (ex.: ROÇADA MANUAL CLASSE I) no de_para fixo."
        + RS
    )
    print(
        DM
        + "  Interpretacao: simulacao conservadora em LUCRO — como se nao houvesse premio de "
        "declividade na mixagem; em campo inclinado real, revise o menu [4] de_para para "
        "Classes II–V conforme a CT." + RS
    )
    sub()


def avaliar_terreno(df_faz):
    print(G + BL + "\n  REFINAMENTO DE DECLIVIDADE\n" + RS)
    print(
        DM
        + "  Isto aplica um fator multiplicativo extra sobre HH/ha (1,0 / 1,15 / 1,30), "
        "independente da classe I–V da CT. Classe I vs V ja esta na linha de preco da CT; "
        "este passo e so para penalizar o cronograma se quiser simular declive geral."
        + RS
    )
    if not confirmar("Aplicar penalidade por declive?", default=False):
        df_faz["penalidade"] = 1.0
        return df_faz
    terrenos = ["Plano (Base x1.0)", "Misto (x1.15)", "Inclinado (x1.30)"]
    t = selecionar("DECLIVIDADE", terrenos)
    if t and "Inclinado" in t:
        df_faz["penalidade"] = 1.3
    elif t and "Misto" in t:
        df_faz["penalidade"] = 1.15
    else:
        df_faz["penalidade"] = 1.0
    return df_faz


def filtrar_atividades_por_texto(atividades, texto):
    """Nomes cuja versao sem acento contem o filtro (substring)."""
    t = remover_acentos(texto)
    if not t:
        return []
    out = []
    for a in atividades:
        if t in remover_acentos(str(a)):
            out.append(a)
    return out


def _to_float_br(v, default=0.0):
    """Converte numero em formato BR/EN para float de forma tolerante."""
    try:
        if v is None:
            return float(default)
        if isinstance(v, (int, float)) and not pd.isna(v):
            return float(v)
        s = str(v).strip()
        if not s or s.lower() == "nan":
            return float(default)
        s = s.replace(" ", "")
        if "," in s and "." in s:
            # Assume separador de milhar + decimal (pt-BR): 1.234,56
            s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            s = s.replace(",", ".")
        return float(s)
    except Exception:
        return float(default)


def fazendas_unicas_micro(df):
    """Nomes unicos da coluna fazenda (micro), ordenados."""
    return sorted(
        {str(x).strip() for x in df["fazenda"].dropna().unique() if str(x).strip()},
        key=lambda s: normalizar_chave(s),
    )


def _resolver_fazenda_demo_ulianopolis(df):
    """
    Fazenda alvo no modo DEMO:
    1) Nome que contenha 'ulianopolis' (ex.: fazenda cadastrada como Ulianopolis SWG);
    2) Senao, a fazenda com mais linhas no micro (planilhas municipio Ulianopolis sem 'Ulianopolis' no nome da fazenda).
    """
    fazendas = fazendas_unicas_micro(df)
    for f in fazendas:
        if "ulianopolis" in normalizar_chave(f):
            return f
    best, nmax = None, -1
    for f in fazendas:
        n = len(df[df["fazenda"] == f])
        if n > nmax:
            nmax, best = n, f
    return best


def garantir_fazenda_ulianopolis_no_ct(cfg, df):
    """
    Modo DEMO: acrescenta em fazendas_ct todas as fazendas presentes no micro demo,
    para o aviso micro-vs-CT nao bloquear (CT pode nao listar municipio).
    Retorna quantos nomes novos foram adicionados.
    """
    idx = _indice_fazendas_ct(cfg)
    ad = 0
    for f in fazendas_unicas_micro(df):
        nk = normalizar_chave(f)
        if not nk:
            continue
        if nk not in idx:
            cfg.setdefault("fazendas_ct", []).append(f)
            idx[nk] = f
            ad += 1
    return ad


def reconstruir_demo_ulianopolis_a_partir_da_fonte():
    """
    Le USEESTAPLANILHAULIANOPOLIS.xlsx (Planilha1), filtra municipio Ulianopolis,
    gera ulianopolisswg.xlsx no formato esperado pelo carregador (NOME FAZENDA, CHAVE POLIGONO, ...).
    Retorna (n_linhas, n_atividades_unicas) ou None se a fonte nao existir.
    """
    src = os.path.join(INPUT_DIR, DEMO_MICRO_SOURCE_FILENAME)
    if not os.path.exists(src):
        return None
    try:
        raw = pd.read_excel(src, sheet_name=0)
    except Exception:
        return None
    if raw.shape[1] < 11:
        return None
    # Colunas fixas pela ordem do export escritorio (USEESTA...)
    muni = raw.iloc[:, 3]
    mask = muni.astype(str).str.contains("ulian", case=False, na=False)
    sub = raw.loc[mask].copy()
    if sub.empty:
        sub = raw.copy()
    cod = sub.iloc[:, 0]
    nome_faz = sub.iloc[:, 1]
    nucleo = sub.iloc[:, 2]
    atividades = sub.iloc[:, 9]
    area = sub.iloc[:, 10]
    chaves = []
    for i in range(len(sub)):
        chaves.append(
            f"{str(cod.iloc[i]).strip()}_{str(nucleo.iloc[i]).strip()}_{i:04d}"
        )
    out = pd.DataFrame(
        {
            "NOME FAZENDA": nome_faz.astype(str).str.strip(),
            "CHAVE POLÍGONO": chaves,
            "ÁREA TRABALHADA ESTIMADA (HECTARE)": pd.to_numeric(area, errors="coerce"),
            "ATIVIDADES": atividades.astype(str).str.strip(),
        }
    )
    out = out.dropna(subset=["ATIVIDADES"])
    out = out[
        pd.to_numeric(
            out["ÁREA TRABALHADA ESTIMADA (HECTARE)"], errors="coerce"
        ).fillna(0)
        > 0
    ]
    dest = os.path.join(INPUT_DIR, DEMO_MICRO_FILENAME)
    out.to_excel(dest, index=False, sheet_name="MICROPLANEJAMENTO_ULIANOPOLIS")
    return len(out), out["ATIVIDADES"].nunique()


def _eh_rocada(atividade):
    """Heuristica robusta para identificar atividades de roçada no nome."""
    n = normalizar_chave(str(atividade or ""))
    if not n:
        return False
    return (
        "rocada" in n
        or "rocagem" in n
        or "roco" in n
        or "roca" in n
    )


def construir_cronograma_humano_sem_rocada(
    cronograma_base, turmas, jornada, executores
):
    """
    Remove atividades de roçada do cronograma humano e recompõe dias por turma
    mantendo ordem original de execução dentro de cada turma.
    """
    turmas_ops = {t["nome"]: int(t["operarios"]) for t in turmas}
    # Mantem ordem original de aparicao no cronograma base
    por_turma = defaultdict(list)
    for c in cronograma_base:
        atv = str(c.get("Atividade", ""))
        if _eh_rocada(atv):
            continue
        por_turma[str(c.get("Turma", ""))].append(dict(c))

    novo = []
    for nm_turma, itens in por_turma.items():
        if not itens:
            continue
        if nm_turma == "Pelotao_Unificado":
            n_ops = int(executores)
        else:
            n_ops = int(turmas_ops.get(nm_turma, 1))
        cap_dia = max(0.01, float(n_ops) * float(jornada))
        dia = 1
        saldo = cap_dia
        for it in itens:
            hh_rest = float(it.get("HH", 0) or 0)
            if hh_rest <= 0.01:
                continue
            while hh_rest > 0.01:
                if saldo <= 0.01:
                    dia += 1
                    saldo = cap_dia
                cons = min(hh_rest, saldo)
                hh_rest -= cons
                saldo -= cons
                row = dict(it)
                row["Dia"] = dia
                row["HH"] = round(cons, 2)
                # custo ja vem por HH na base; reescala proporcionalmente
                c_old = float(it.get("Custo_MO", 0) or 0)
                hh_old = float(it.get("HH", 0) or 0)
                row["Custo_MO"] = (
                    round((cons / hh_old) * c_old, 2) if hh_old > 0.01 else 0.0
                )
                novo.append(row)
    novo = sorted(novo, key=lambda r: (int(r.get("Dia", 0)), str(r.get("Turma", ""))))
    return novo


def construir_cronograma_robo_rocada(
    demandas, fazenda, jornada, prod_ha_h=0.18, custo_h=0.0
):
    """
    Fila dedicada do robô para TODAS as atividades de roçada (paralelo ao humano).
    Capacidade diária em área: prod_ha_h * jornada.
    """
    tarefas = []
    for talhao, ls in demandas.items():
        for t in ls:
            atv = str(t.get("atividade", ""))
            if not _eh_rocada(atv):
                continue
            area = float(t.get("area", 0) or 0)
            if area <= 0.0001:
                continue
            tarefas.append({"Talhao": talhao, "Atividade": atv, "Area_ha": area})

    if not tarefas:
        return []

    cap_area_dia = max(0.0001, float(prod_ha_h) * float(jornada))
    dia = 1
    area_saldo_dia = cap_area_dia
    out = []
    for t in tarefas:
        area_rest = float(t["Area_ha"])
        while area_rest > 0.0001:
            if area_saldo_dia <= 0.0001:
                dia += 1
                area_saldo_dia = cap_area_dia
            area_exec = min(area_rest, area_saldo_dia)
            area_rest -= area_exec
            area_saldo_dia -= area_exec
            hh = area_exec / float(prod_ha_h) if float(prod_ha_h) > 0 else 0.0
            out.append(
                {
                    "Dia": int(dia),
                    "Fazenda": fazenda,
                    "Talhao": t["Talhao"],
                    "Atividade": t["Atividade"],
                    "Turma": "ROBO_ROCADOR",
                    "Operarios": 1,
                    "HH": round(hh, 2),
                    "Modo": "RoboRocador",
                    "Area_ha": round(area_exec, 4),
                }
            )
    return out


def construir_cronograma_mecanizado(demandas, fazenda, jornada, recursos_mec):
    """
    Fila dedicada para cada recurso mecanizado.
    recursos_mec: [{"nome", "prod_ha_h", "custo_h", "atividades": set}]
    """
    todas_filas = []
    for rec in recursos_mec:
        nome_rec = rec["nome"]
        prod = float(rec.get("prod_ha_h") or 0.18)
        tarefas = []
        for talhao, ls in demandas.items():
            for t in ls:
                atv = str(t.get("atividade", ""))
                if atv not in rec.get("atividades", set()):
                    continue
                area = float(t.get("area", 0) or 0)
                if area <= 0.0001:
                    continue
                tarefas.append({"Talhao": talhao, "Atividade": atv, "Area_ha": area})
        if not tarefas:
            continue
        cap_area_dia = max(0.0001, prod * float(jornada))
        dia = 1
        saldo = cap_area_dia
        for t in tarefas:
            rest = float(t["Area_ha"])
            while rest > 0.0001:
                if saldo <= 0.0001:
                    dia += 1
                    saldo = cap_area_dia
                exe = min(rest, saldo)
                rest -= exe
                saldo -= exe
                hh = exe / prod if prod > 0 else 0.0
                todas_filas.append(
                    {
                        "Dia": int(dia),
                        "Fazenda": fazenda,
                        "Talhao": t["Talhao"],
                        "Atividade": t["Atividade"],
                        "Turma": f"MEC_{nome_rec}",
                        "Operarios": 1,
                        "HH": round(hh, 2),
                        "HM": round(hh, 2),
                        "Modo": "Mecanizado",
                        "Origem_Mec": "Cadastro",
                        "Area_ha": round(exe, 4),
                    }
                )
    return sorted(
        todas_filas, key=lambda r: (int(r.get("Dia", 0)), str(r.get("Turma", "")))
    )


def construir_cronograma_mecanizado_auto_hm_tarifa(
    demandas, fazenda, jornada, cfg, tarifas, atividades_alvo=None
):
    """
    Gera cronograma mecanizado automaticamente a partir de HM/ha da tarifa.
    Cada atividade com HM vira uma fila dedicada (paralela), sem cadastro manual de recurso.
    """
    atividades_alvo = set(atividades_alvo or [])
    filas_por_atividade = defaultdict(list)
    for talhao, ls in demandas.items():
        for t in ls:
            atv = str(t.get("atividade", ""))
            if atividades_alvo and atv not in atividades_alvo:
                continue
            area = float(t.get("area", 0) or 0)
            hm_ha = float(t.get("hm_ha", 0) or 0)
            hm_total = float(t.get("hm_total", area * hm_ha) or 0)
            if area <= 0.0001 or hm_ha <= 0 or hm_total <= 0.0001:
                continue
            filas_por_atividade[atv].append(
                {
                    "Talhao": talhao,
                    "Atividade": atv,
                    "Area_ha": area,
                    "HM_ha": hm_ha,
                    "HM_total": hm_total,
                }
            )

    if not filas_por_atividade:
        return [], []

    cronograma = []
    recursos_auto = []
    for atv in sorted(filas_por_atividade.keys(), key=str):
        itens = filas_por_atividade[atv]

        area_total = sum(float(x.get("Area_ha", 0) or 0) for x in itens)
        hm_total_atv = sum(float(x.get("HM_total", 0) or 0) for x in itens)
        hm_ha_medio = (hm_total_atv / area_total) if area_total > 0.0001 else 0.0
        prod_ha_h = (1.0 / hm_ha_medio) if hm_ha_medio > 0.0001 else 0.0
        recursos_auto.append(
            {
                "nome": f"AutoHM_{_slug_ficheiro_seguro(atv, max_len=22)}",
                "prod_ha_h": round(prod_ha_h, 4) if prod_ha_h > 0 else 0.0,
                "atividades": {atv},
            }
        )

        cap_hm_dia = max(0.0001, float(jornada))
        dia = 1
        saldo_hm_dia = cap_hm_dia
        turma_nome = f"MEC_AUTO_{_slug_ficheiro_seguro(atv, max_len=16)}"

        for item in itens:
            hm_rest = float(item.get("HM_total", 0) or 0)
            hm_ha_item = float(item.get("HM_ha", 0) or 0)
            while hm_rest > 0.0001:
                if saldo_hm_dia <= 0.0001:
                    dia += 1
                    saldo_hm_dia = cap_hm_dia
                hm_exec = min(hm_rest, saldo_hm_dia)
                hm_rest -= hm_exec
                saldo_hm_dia -= hm_exec
                area_exec = (hm_exec / hm_ha_item) if hm_ha_item > 0.0001 else 0.0
                cronograma.append(
                    {
                        "Dia": int(dia),
                        "Fazenda": fazenda,
                        "Talhao": item["Talhao"],
                        "Atividade": item["Atividade"],
                        "Turma": turma_nome,
                        "Operarios": 1,
                        "HH": round(hm_exec, 2),
                        "HM": round(hm_exec, 2),
                        "Modo": "MecanizadoAutoHM",
                        "Origem_Mec": "HM_Tarifa",
                        "Area_ha": round(area_exec, 4),
                    }
                )

    cronograma = sorted(
        cronograma, key=lambda r: (int(r.get("Dia", 0)), str(r.get("Turma", "")))
    )
    return cronograma, recursos_auto


def construir_cronograma_humano_sem_mecanizadas(
    cronograma_base, turmas, jornada, executores, atividades_mec
):
    """Remove atividades assumidas por mecanizados do cronograma humano e recompoe dias."""
    turmas_ops = {t["nome"]: int(t["operarios"]) for t in turmas}
    por_turma = defaultdict(list)
    for c in cronograma_base:
        atv = str(c.get("Atividade", ""))
        if atv in atividades_mec:
            continue
        por_turma[str(c.get("Turma", ""))].append(dict(c))
    novo = []
    for nm_turma, itens in por_turma.items():
        if not itens:
            continue
        n_ops = (
            int(executores)
            if nm_turma == "Pelotao_Unificado"
            else int(turmas_ops.get(nm_turma, 1))
        )
        cap_dia = max(0.01, float(n_ops) * float(jornada))
        dia = 1
        saldo = cap_dia
        for it in itens:
            hh_rest = float(it.get("HH", 0) or 0)
            if hh_rest <= 0.01:
                continue
            while hh_rest > 0.01:
                if saldo <= 0.01:
                    dia += 1
                    saldo = cap_dia
                cons = min(hh_rest, saldo)
                hh_rest -= cons
                saldo -= cons
                row = dict(it)
                row["Dia"] = dia
                row["HH"] = round(cons, 2)
                c_old = float(it.get("Custo_MO", 0) or 0)
                hh_old = float(it.get("HH", 0) or 0)
                row["Custo_MO"] = (
                    round((cons / hh_old) * c_old, 2) if hh_old > 0.01 else 0.0
                )
                novo.append(row)
    return sorted(novo, key=lambda r: (int(r.get("Dia", 0)), str(r.get("Turma", ""))))


def _menu_editar_recurso_mecanizado(recursos, pool_catalogo):
    """Permite revisar e alterar atividades/produtividade/custo de recursos mecanizados."""
    if not recursos:
        aviso("Nao ha recursos mecanizados para editar.")
        return recursos

    while True:
        sub()
        print(G + BL + "  EDICAO RETROATIVA — RECURSOS MECANIZADOS" + RS)
        nomes = [
            f"{r['nome']} ({len(r.get('atividades', set()))} atividades)"
            for r in recursos
        ]
        op = selecionar("SELECIONE O RECURSO", nomes + ["Concluir edicao"])
        if not op or op == "Concluir edicao":
            break
        idx = nomes.index(op)
        rec = recursos[idx]

        while True:
            cur = sorted(list(rec.get("atividades", set())), key=str)
            sub()
            print(G + BL + f" RECURSO: {rec['nome']}" + RS)
            print(DM + f" Produtividade: {rec.get('prod_ha_h', 0)} ha/h" + RS)
            print(DM + f" Custo/h: R$ {rec.get('custo_h', 0)}" + RS)
            print(DM + f" Atividades vinculadas: {len(cur)}" + RS)
            acao = selecionar(
                    "ACAO",
                    [
                        "Adicionar atividade",
                        "Remover atividade",
                        "Substituir atividade",
                        "Alterar produtividade",
                        "Alterar custo/h",
                        "Ver lista filtrada (mec/mecanizado/semimec)",
                        "Ver listas completas (vinculadas x catalogo)",
                        "Voltar",
                    ],
                )
            if not acao or acao == "Voltar":
                break

            if acao == "Adicionar atividade":
                disp = [a for a in pool_catalogo if a not in rec.get("atividades", set())]
                if not disp:
                    aviso("Nao ha atividade nova para adicionar.")
                    continue
                idx_add = selecionar_paginado("ADICIONAR ATIVIDADE", disp)
                if idx_add >= 0:
                    rec.setdefault("atividades", set()).add(disp[idx_add])
                    ok("Atividade adicionada.")
                continue

            if acao == "Remover atividade":
                if not cur:
                    aviso("Recurso sem atividades vinculadas.")
                    continue
                idx_rm = selecionar_paginado("REMOVER ATIVIDADE", cur)
                if idx_rm >= 0:
                    rec.setdefault("atividades", set()).discard(cur[idx_rm])
                    ok("Atividade removida.")
                continue

            if acao == "Substituir atividade":
                if not cur:
                    aviso("Recurso sem atividades vinculadas.")
                    continue
                idx_src = selecionar_paginado("ATIVIDADE ORIGEM", cur)
                if idx_src < 0:
                    continue
                src = cur[idx_src]
                disp = [a for a in pool_catalogo if a != src]
                idx_dst = selecionar_paginado("ATIVIDADE DESTINO", disp)
                if idx_dst >= 0:
                    dst = disp[idx_dst]
                    rec.setdefault("atividades", set()).discard(src)
                    rec.setdefault("atividades", set()).add(dst)
                    ok(f"Substituida: '{src[:45]}' -> '{dst[:45]}'.")
                continue

            if acao == "Alterar produtividade":
                rec["prod_ha_h"] = pedir_float(
                    "Nova produtividade (ha/h)",
                    float(rec.get("prod_ha_h") or 0.18),
                )
                ok("Produtividade atualizada.")
                continue

            if acao == "Alterar custo/h":
                rec["custo_h"] = pedir_float(
                    "Novo custo (R$/h)",
                    float(rec.get("custo_h") or 0.0),
                    allow_zero=True,
                )
                ok("Custo/h atualizado.")
                continue

                if acao == "Ver lista filtrada (mec/mecanizado/semimec)":
                    # Filtrar apenas atividades mecanizadas (com HM > 0 ou tipo Mecanizada/SemiMecanizada)
                    atividades_mec = []
                    for atv_nome, dados in CT317_HARDCODE_HH_BASE.items():
                        if dados.get("rendimento_hm", 0) > 0 or dados.get("tipo", "").lower() in ("mecanizada", "semimecanizada"):
                            atividades_mec.append(atv_nome)
                    
                    # Também incluir do catálogo que tenham 'mec', 'mecaniz', 'semimec' no nome
                    for atv in pool_catalogo:
                        atv_norm = normalizar_chave(str(atv))
                        if any(k in atv_norm for k in ["mec", "mecaniz", "semimec", "trator", "robo", "drone"]):
                            if atv not in atividades_mec:
                                atividades_mec.append(atv)
                    
                    atividades_mec.sort(key=str)
                    
                    print(G + BL + "\n LISTA FILTRADA — ATIVIDADES MECANIZADAS" + RS)
                    print(DM + f" (atividades com HM > 0, 'mec' no nome, ou tipo Mecanizada/SemiMecanizada)" + RS)
                    if atividades_mec:
                        for i, a in enumerate(atividades_mec, 1):
                            # Buscar valor HM se disponível
                            hm_val = ""
                            for atv_nome, dados in CT317_HARDCODE_HH_BASE.items():
                                if atv_nome == a:
                                    hm = dados.get("rendimento_hm", 0)
                                    if hm > 0:
                                        hm_val = f" {C}(HM={hm:.2f}){RS}"
                                    break
                            print(f" {Y}{i:2}{RS}. {a}{hm_val}")
                        sub()
                        print(G + f"Total filtrado: {len(atividades_mec)} atividade(s)" + RS)
                    else:
                        print(Y + " (nenhuma atividade mecanizada encontrada)" + RS)
                    
                    # Mostrar também as já vinculadas
                    cur_mec = [a for a in cur if a in atividades_mec]
                    if cur_mec:
                        sub()
                        print(G + " Já vinculadas a este recurso:" + RS)
                        for a in cur_mec:
                            print(f" {C}✓{RS} {a}")
                    
                    input(DM + "\n [ENTER] " + RS)
                    continue

                if acao == "Ver listas completas (vinculadas x catalogo)":
                    _mostrar_catalogo_atividades(cur, pool_catalogo)
                    input(DM + "\n [ENTER] " + RS)

                return recursos


def _cadastrar_recursos_mecanizados_sn(atividades_reais, cfg=None, atividades_catalogo=None):
    """Cadastrar N recursos mecanizados com seleção de atividades via S/N e edição retroativa."""
    pool_catalogo = _catalogo_atividades_completo(
        atividades_reais,
        cfg=cfg,
        atividades_catalogo=atividades_catalogo,
    )
    cand_mec = atividades_candidatas_mecanizado(atividades_reais, cfg)
    pool = list(atividades_reais)
    if cand_mec:
        sub()
        print(G + BL + "  LISTA DE ATIVIDADES (modo mecanizado)" + RS)
        print(
            DM
            + f"  Encontradas {len(cand_mec)} candidata(s) (nome: trator, mec., solo mec, etc.; ou tipo HM na tarifa)."
            + RS
        )
        if confirmar(
            "  Mostrar apenas candidatas a mecanizado na pergunta S/N abaixo?",
            default=True,
        ):
            pool = cand_mec
        else:
            pool = list(pool_catalogo)
    elif cfg:
        aviso("Nenhuma candidata automatica; listando todas as atividades da fazenda.")
        pool = list(pool_catalogo)
    recursos = []
    while True:
        sub()
        print(G + BL + f"  MODO MECANIZADO — recurso #{len(recursos) + 1}" + RS)
        nome = prompt(
            "Nome do recurso (ex: Robo Rocador, Trator X)",
            f"Mecanizado_{len(recursos) + 1}",
        )
        prod = pedir_float("Produtividade (ha/h)", 0.18)
        custo = pedir_float("Custo (R$/h, 0 se placeholder)", 0.0, allow_zero=True)
        print(G + BL + f"\n  Selecionar atividades para '{nome}' (S/N):" + RS)
        print(DM + "  s=sim  n=nao  a=nao e encerrar  ok=sim e encerrar" + RS)
        atvs = set()
        cur_all = sorted(pool, key=str)
        for i, a in enumerate(cur_all, 1):
            v = prompt(f"[{i}/{len(cur_all)}] Vincular '{str(a)[:54]}'? (s/n/a/ok)", "")
            v = str(v).strip().lower()
            if v in ("s", "sim", "y", "yes"):
                atvs.add(a)
            elif v == "a":
                ok("Selecao encerrada (sem vincular esta).")
                break
            elif v == "ok":
                atvs.add(a)
                ok("Selecao encerrada por comando rapido.")
                break
        if not atvs:
            aviso("Nenhuma atividade selecionada para este recurso.")
        else:
            recursos.append(
                {"nome": nome, "prod_ha_h": prod, "custo_h": custo, "atividades": atvs}
            )
            ok(f"Recurso '{nome}': {len(atvs)} atividades, {prod} ha/h, R$ {custo}/h")
        if not confirmar("Adicionar mais um recurso mecanizado?", default=False):
            break

    if recursos and confirmar(
        "Revisar/editar atividades dos recursos mecanizados agora?",
        default=True,
    ):
        recursos = _menu_editar_recurso_mecanizado(recursos, pool_catalogo)

    return recursos


def _parse_lista_numeros(txt, as_int=False):
    out = []
    for p in str(txt).replace(";", ",").split(","):
        s = p.strip()
        if not s:
            continue
        try:
            v = float(s.replace(",", "."))
            if as_int:
                v = int(round(v))
            if v > 0:
                out.append(v)
        except Exception:
            pass
    return sorted(set(out))


def coletar_config_comparativo_multifator(executores_base, jornada_base):
    """Coleta grade de cenarios (jornada/equipe) de forma antecipada e explicita."""
    sub()
    print(C + BL + "  [CENARIOS] CONFIGURAR COMPARATIVO MULTI-FATOR" + RS)
    print(DM + "  O comparativo sera exportado no Excel (COMPARATIVO_CENARIOS)." + RS)
    print(DM + "  Exemplo entradas: jornadas 4.3,5.3,8 | equipes 4,6,8,10" + RS)
    jornadas_txt = prompt("  Jornadas (h/dia) separadas por virgula", f"{jornada_base}")
    equipes_txt = prompt(
        "  Equipes (executores) separadas por virgula", f"{executores_base}"
    )
    jornadas = _parse_lista_numeros(jornadas_txt, as_int=False)
    equipes = _parse_lista_numeros(equipes_txt, as_int=True)
    if not jornadas:
        jornadas = [float(jornada_base)]
    if not equipes:
        equipes = [int(executores_base)]
    ok(
        f"Comparativo configurado: {len(jornadas)} jornada(s) x {len(equipes)} equipe(s)."
    )
    return {"jornadas": jornadas, "equipes": equipes}


def simular_cenarios_multifator(
    total_hh,
    dias_meta,
    executores_base,
    jornada_base,
    jornadas_in=None,
    equipes_in=None,
    interativo=True,
):
    """
    Simulador de cenarios em lote (aproximacao operacional):
    dias ~ HH total / (executores * jornada)
    """
    jornadas = sorted(
        set(float(x) for x in (jornadas_in or [] if not interativo else []))
    )
    equipes = sorted(set(int(x) for x in (equipes_in or [] if not interativo else [])))
    if interativo:
        sub()
        print(C + BL + "  [CENARIOS] COMPARATIVO MULTI-FATOR" + RS)
        print(DM + "  Exemplo entradas: jornadas 4.3,5.3 | equipes 6,8,10" + RS)
        jornadas_txt = prompt(
            "  Jornadas (h/dia) separadas por virgula", f"{jornada_base}"
        )
        equipes_txt = prompt(
            "  Equipes (executores) separadas por virgula", f"{executores_base}"
        )
        jornadas = _parse_lista_numeros(jornadas_txt, as_int=False)
        equipes = _parse_lista_numeros(equipes_txt, as_int=True)
    if not jornadas:
        jornadas = [float(jornada_base)]
    if not equipes:
        equipes = [int(executores_base)]

    rows = []
    for j in jornadas:
        for e in equipes:
            cap = float(e) * float(j)
            dias = int(math.ceil(float(total_hh) / cap)) if cap > 0.01 else 0
            meses = dias / 22.0 if dias > 0 else 0.0
            ganho = int(dias_meta) - int(dias)
            rows.append(
                {
                    "Equipe": int(e),
                    "Jornada_h_dia": float(j),
                    "Dias_Simulados": int(dias),
                    "Meses_Simulados": round(meses, 2),
                    "Ganho_vs_Meta_dias": int(ganho),
                    "HH_Total": round(float(total_hh), 2),
                }
            )
    rows = sorted(
        rows, key=lambda r: (r["Dias_Simulados"], -r["Equipe"], -r["Jornada_h_dia"])
    )
    return rows


def _indice_fazendas_ct(cfg):
    """Chave normalizada -> nome canonico (ultimo da lista)."""
    d = {}
    for x in cfg.get("fazendas_ct") or []:
        s = str(x).strip()
        if not s:
            continue
        k = normalizar_chave(s)
        if k:
            d[k] = s
    return d


def micro_fazendas_ausentes_na_lista_ct(cfg, df):
    """
    Fazendas que aparecem no micro mas nao em config.fazendas_ct (comparacao por normalizar_chave).
    Se fazendas_ct vazia, retorna None (validacao desativada).
    """
    idx = _indice_fazendas_ct(cfg)
    if not idx:
        return None
    out = []
    for m in fazendas_unicas_micro(df):
        if normalizar_chave(m) not in idx:
            out.append(m)
    return out


def aviso_fazendas_micro_sem_cadastro_ct(cfg, df):
    """Aviso no startup / apos recarregar micro."""
    falta = micro_fazendas_ausentes_na_lista_ct(cfg, df)
    if falta is None:
        sub()
        print(
            DM + "  [Fazendas vs CT] Lista `fazendas_ct` vazia no config — "
            "cadastre em menu [6] as fazendas cobertas pelo orcamento CT (ex.: todas menos Ulianopolis)."
            + RS
        )
        return
    if not falta:
        return
    sub()
    print(
        Y
        + "  !  Fazendas no MICRO sem correspondencia na lista `fazendas_ct` (orcamento CT): "
        + RS
    )
    for x in falta:
        print(Y + f"      - {str(x)[:72]}" + RS)
    print(
        DM
        + "  Corrija o CT no escritorio ou adicione excecao em [6] se a fazenda estiver coberta."
        + RS
    )


def modulo_validar_fazendas_ct(cfg, df):
    """
    CRUD de `fazendas_ct`: nomes de fazenda que o orcamento CT considera cadastrados.
    O micro pode ter fazendas a mais (esquecimento no CT) — o scan compara por nome normalizado.
    """
    while True:
        dashboard_header()
        subcabecalho("FAZENDAS — micro vs lista CT (orcamento)")
        micro_list = fazendas_unicas_micro(df)
        ct_list = list(cfg.get("fazendas_ct") or [])
        idx = _indice_fazendas_ct(cfg)
        print(G + f"  No micro agora: {len(micro_list)} fazenda(s)" + RS)
        print(G + f"  Na lista fazendas_ct: {len(ct_list)}" + RS)
        falta = micro_fazendas_ausentes_na_lista_ct(cfg, df)
        if falta is None:
            print(Y + "  Lista CT vazia — nenhuma comparacao possivel." + RS)
        elif falta:
            print(Y + f"  Ausentes na lista CT ({len(falta)}):" + RS)
            for x in falta[:25]:
                print(Y + f"    - {str(x)[:68]}" + RS)
            if len(falta) > 25:
                print(DM + f"    ... +{len(falta) - 25}" + RS)
        else:
            ok("  Todas as fazendas do micro constam em fazendas_ct.")
        sub()
        print(DM + "  [1] Ver / listar fazendas_ct" + RS)
        print(DM + "  [2] Adicionar uma fazenda a fazendas_ct" + RS)
        print(
            DM
            + "  [3] Importar TODAS as fazendas do micro para fazendas_ct (substitui lista)"
            + RS
        )
        print(DM + "  [4] Remover uma fazenda da lista" + RS)
        print(DM + "  [5] Colar varios nomes (virgula ou ponto-e-virgula)" + RS)
        print(DM + "  [6] Limpar lista (fazendas_ct = [])" + RS)
        print(DM + "  [0] Voltar" + RS)
        op = prompt("Opcao").strip()
        if op == "0":
            return
        if op == "1":
            if not ct_list:
                aviso("Lista vazia.")
            else:
                for i, x in enumerate(ct_list, 1):
                    print(G + f"  {i:2}. " + C + str(x)[:68] + RS)
            input(DM + "\n  [ENTER] " + RS)
        elif op == "2":
            nome = prompt("Nome EXATO como no micro ou na CT", "")
            if nome.strip():
                k = normalizar_chave(nome)
                if k in idx:
                    aviso("Ja consta (mesma chave normalizada).")
                else:
                    cfg.setdefault("fazendas_ct", []).append(nome.strip())
                    salvar_config(cfg)
                    ok("Adicionado.")
        elif op == "3":
            if not confirmar(
                "Substituir fazendas_ct pelas "
                f"{len(micro_list)} fazendas unicas do micro? (nao altera a planilha CT .xlsm)",
                default=False,
            ):
                continue
            cfg["fazendas_ct"] = micro_list[:]
            salvar_config(cfg)
            ok(f"fazendas_ct = {len(micro_list)} nomes (copia do micro).")
        elif op == "4":
            if not ct_list:
                aviso("Lista vazia.")
                continue
            idx_r = selecionar_paginado("REMOVER", ct_list, page_size=10)
            if idx_r >= 0:
                cfg["fazendas_ct"].pop(idx_r)
                salvar_config(cfg)
                ok("Removido.")
        elif op == "5":
            txt = prompt("Nomes separados por virgula ou ;", "")
            partes = []
            for chunk in txt.replace(";", ",").split(","):
                s = str(chunk).strip()
                if s:
                    partes.append(s)
            ad = 0
            seen = _indice_fazendas_ct(cfg)
            for s in partes:
                k = normalizar_chave(s)
                if k and k not in seen:
                    cfg.setdefault("fazendas_ct", []).append(s)
                    seen[k] = s
                    ad += 1
            if ad:
                salvar_config(cfg)
                ok(f"+{ad} nome(s) novos.")
            else:
                aviso("Nenhum nome novo.")
        elif op == "6":
            if confirmar("Zerar fazendas_ct?", default=False):
                cfg["fazendas_ct"] = []
                salvar_config(cfg)
                ok("Lista limpa.")
        else:
            aviso("Opcao invalida.")


def _catalogo_atividades_completo(atividades_escopo, cfg=None, atividades_catalogo=None):
    """Unifica atividades do escopo atual, catálogo do micro e catálogo da CT/de_para."""
    out = set()

    for a in atividades_escopo or []:
        s = str(a).strip()
        if s:
            out.add(s)

    for a in atividades_catalogo or []:
        s = str(a).strip()
        if s:
            out.add(s)

    if isinstance(cfg, dict):
        tarifas = cfg.get("tarifas", {}) or {}
        for a in tarifas.keys():
            s = str(a).strip()
            if s:
                out.add(s)

        de_para = cfg.get("de_para", {}) or {}
        for k, v in de_para.items():
            if str(k).startswith("_"):
                continue
            ks = _norm_atv(k)
            vs = str(v).strip()
            if ks:
                out.add(ks)
            if vs:
                out.add(vs)

    return sorted(out, key=lambda x: str(x))


def _mostrar_catalogo_atividades(atividades_escopo, atividades_catalogo):
    """Mostra duas tabelas: escopo atual e catálogo completo disponível."""
    escopo = sorted({str(a).strip() for a in atividades_escopo or [] if str(a).strip()}, key=str)
    catalogo = sorted({str(a).strip() for a in atividades_catalogo or [] if str(a).strip()}, key=str)

    print(G + BL + "\n  LISTA 1 — ATIVIDADES NO ESCOPO ATUAL" + RS)
    if escopo:
        for i, a in enumerate(escopo, 1):
            print(G + f"  [{i:2}] " + C + f"{a}" + RS)
    else:
        print(Y + "  (vazio)" + RS)

    print(G + BL + "\n  LISTA 2 — CATALOGO COMPLETO (MICRO + CT + de_para)" + RS)
    if catalogo:
        for i, a in enumerate(catalogo, 1):
            print(G + f"  [{i:2}] " + C + f"{a}" + RS)
    else:
        print(Y + "  (vazio)" + RS)


def menu_vincular_atividades_turma(turma, atividades_reais, atividades_catalogo=None):
    """
    Vincula atividades a uma turma.
    Padrao: percurso S/N atividade-por-atividade.
    Fallback: filtro/lista/paginacao acessiveis via menu auxiliar.
    """
    atv_set = set(turma["atividades"])

    def _catalogo_all():
        return _catalogo_atividades_completo(
            list(atividades_reais) + list(atv_set),
            cfg=None,
            atividades_catalogo=atividades_catalogo,
        )

    def _percurso_sn():
        cur_all = sorted(atividades_reais, key=lambda x: str(x))
        print(
            G
            + BL
            + f"\n  TURMA '{turma['nome']}' — percurso S/N ({len(cur_all)} atividades)"
            + RS
        )
        print(
            DM
            + "  s=vincular  n=desvincular  a=nao e encerrar  ok=sim e encerrar  ENTER=manter atual"
            + RS
            + "\n"
        )
        for i, a in enumerate(cur_all, 1):
            mk = "X" if a in atv_set else " "
            v = prompt(f"[{i}/{len(cur_all)}] [{mk}] '{str(a)[:54]}' (s/n/a/ok)", "")
            v = str(v).strip().lower()
            if v in ("s", "sim", "y", "yes"):
                atv_set.add(a)
                _emitir_monitor_rendimentos(str(a), True)
            elif v in ("n", "nao", "não", "no"):
                atv_set.discard(a)
                _emitir_monitor_rendimentos(str(a), False)
            elif v == "a":
                ok("Percurso encerrado (sem alterar esta atividade).")
                _emitir_monitor_rendimentos("", False)  # Indicates user aborted
                break
            elif v == "ok":
                atv_set.add(a)
                ok("Percurso encerrado por comando rapido.")
                break
        ok(f"Percurso concluido. Vinculadas: {len(atv_set)}")

    def _assistente_sn_vinculos():
        """
        Revisao guiada S/N das atividades da turma:
        - ENTER: manter
        - n: remover
        - t: trocar por outra atividade
        - a: adicionar nova atividade agora
        - ok: encerrar assistente
        """
        while True:
            cur_all = sorted(atividades_reais, key=lambda x: str(x))
            cat_all = _catalogo_all()
            cur_v = sorted(atv_set, key=lambda x: str(x))
            print(G + BL + f"\n  ASSISTENTE S/N — TURMA '{turma['nome']}'" + RS)
            print(
                DM
                + "  ENTER=manter  n=remover  t=trocar  a=adicionar  ok=encerrar"
                + RS
                + "\n"
            )
            for i, a in enumerate(cur_all, 1):
                if a not in atv_set:
                    continue
                v = (
                    prompt(f"[{i}/{len(cur_all)}] '{str(a)[:54]}' (ENTER/n/t/a/ok)", "")
                    .strip()
                    .lower()
                )
                if not v:
                    continue
                if v in ("ok",):
                    ok("Assistente encerrado.")
                    return
                if v in ("n", "nao", "não", "no"):
                    atv_set.discard(a)
                    continue
                if v in ("a",):
                    disp_add = [x for x in cat_all if x not in atv_set]
                    if not disp_add:
                        aviso("Nao ha atividade disponivel para adicionar.")
                        continue
                    idx_add = selecionar_paginado("ADICIONAR ATIVIDADE", disp_add)
                    if idx_add >= 0:
                        atv_set.add(disp_add[idx_add])
                        ok("Adicionada.")
                    continue
                if v in ("t", "trocar"):
                    disp = [x for x in cat_all if x != a]
                    idxd = selecionar_paginado("DESTINO DA TROCA", disp)
                    if idxd >= 0:
                        dest = disp[idxd]
                        atv_set.discard(a)
                        atv_set.add(dest)
                        ok(f"Troca: '{str(a)[:40]}' -> '{str(dest)[:40]}'.")
                    continue
            if not confirmar("Repassar assistente S/N novamente?", default=False):
                return

    _percurso_sn()

    while True:
        cur = sorted(atv_set, key=lambda x: str(x))
        sub()
        print(
            G
            + BL
            + f"  TURMA: {turma['nome']} ({turma['operarios']} ops) — {len(cur)} atividade(s) vinculadas"
            + RS
        )
        print(DM + "  [1] Refazer percurso S/N" + RS)
        print(DM + "  [2] Adicionar por filtro de texto" + RS)
        print(DM + "  [3] Adicionar por lista/indices (fallback)" + RS)
        print(DM + "  [4] Remover por filtro" + RS)
        print(DM + "  [5] Remover UMA (lista)" + RS)
        print(DM + "  [6] Ver vinculadas" + RS)
        print(DM + "  [7] Trocar atividade (substituir 1:1)" + RS)
        print(DM + "  [8] Assistente inteligente S/N (revisao guiada)" + RS)
        print(DM + "  [9] Ver duas listas (escopo x catalogo completo)" + RS)
        print(DM + "  [0] Concluir esta turma" + RS)
        sub()
        op = prompt("Opcao", "0").strip()
        if op == "0":
            turma["atividades"] = sorted(atv_set, key=lambda x: str(x))
            return
        if op == "1":
            _percurso_sn()
        elif op == "2":
            filtro = prompt("Texto no nome (ex: roçada)", "")
            if not str(filtro).strip():
                aviso("Filtro vazio.")
                continue
            matches = filtrar_atividades_por_texto(atividades_reais, filtro)
            if not matches:
                aviso("Nenhuma atividade bateu com o filtro.")
                continue
            print(G + f"\n  {len(matches)} encontrada(s):" + RS)
            for m in matches[:12]:
                print(DM + f"    - {str(m)[:62]}" + RS)
            if len(matches) > 12:
                print(DM + f"    ... +{len(matches) - 12}" + RS)
            if confirmar("Adicionar TODAS ao vinculo desta turma?", default=True):
                for m in matches:
                    atv_set.add(m)
                ok(f"+{len(matches)} atividades.")
            else:
                for i, m in enumerate(matches, 1):
                    if confirmar(f"  [{i}] {str(m)[:55]}", default=False):
                        atv_set.add(m)
        elif op == "3":
            disp = [a for a in _catalogo_all() if a not in atv_set]
            if not disp:
                aviso("Ja estao todas vinculadas ou lista vazia.")
                continue
            print(
                DM
                + f"\n  Indices de 1 a {len(disp)} (ex.: 1,3,5-8). ENTER = lista paginada"
                + RS
            )
            multi = prompt("Indices", "")
            if str(multi).strip():
                idxs = parse_intervalos_escolha(multi, len(disp))
                if not idxs:
                    aviso("Nenhum indice valido.")
                else:
                    for i in idxs:
                        atv_set.add(disp[i])
                    ok(f"+{len(idxs)} atividades.")
                continue
            idx = selecionar_paginado("ADICIONAR ATIVIDADE", disp)
            if idx >= 0:
                atv_set.add(disp[idx])
                ok("Adicionada.")
        elif op == "4":
            filtro = prompt("Remover cujo nome contem", "")
            if not str(filtro).strip():
                aviso("Filtro vazio.")
                continue
            rem = filtrar_atividades_por_texto(list(atv_set), filtro)
            if not rem:
                aviso("Nenhuma vinculada bateu com o filtro.")
                continue
            if confirmar(
                f"Remover {len(rem)} da turma '{turma['nome']}'?", default=True
            ):
                for r in rem:
                    atv_set.discard(r)
                ok("Removidas.")
        elif op == "5":
            cur2 = sorted(atv_set, key=lambda x: str(x))
            if not cur2:
                aviso("Nada vinculado ainda.")
                continue
            idx = selecionar_paginado("REMOVER ATIVIDADE", cur2)
            if idx >= 0:
                atv_set.discard(cur2[idx])
                ok("Removida.")
        elif op == "6":
            cur2 = sorted(atv_set, key=lambda x: str(x))
            print(G + f"\n  Vinculadas ({len(cur2)}): " + RS)
            for x in cur2[:40]:
                print(DM + f"    - {str(x)[:62]}" + RS)
            if len(cur2) > 40:
                print(DM + f"    ... +{len(cur2) - 40}" + RS)
            input(DM + "\n  [ENTER] " + RS)
        elif op == "7":
            cur2 = sorted(atv_set, key=lambda x: str(x))
            if not cur2:
                aviso("Nada vinculado para trocar.")
                continue
            old = selecionar_paginado("ATIVIDADE ORIGEM (será removida)", cur2)
            if old < 0:
                continue
            origem = cur2[old]
            disp = sorted([a for a in _catalogo_all() if a != origem], key=lambda x: str(x))
            if not disp:
                aviso("Nao ha atividade destino disponivel.")
                continue
            print(
                DM + "  Dica: ENTER para lista paginada ou use filtro por texto." + RS
            )
            filtro = prompt("Filtro do destino (opcional)", "")
            if str(filtro).strip():
                candidatos = filtrar_atividades_por_texto(disp, filtro)
                if not candidatos:
                    aviso("Nenhum destino bateu com o filtro.")
                    continue
                destino = selecionar("ATIVIDADE DESTINO", candidatos)
            else:
                idxd = selecionar_paginado("ATIVIDADE DESTINO", disp)
                destino = disp[idxd] if idxd >= 0 else None
            if not destino:
                continue
            if confirmar(
                f"Trocar '{str(origem)[:48]}' por '{str(destino)[:48]}'?", default=True
            ):
                atv_set.discard(origem)
                atv_set.add(destino)
                ok("Troca aplicada.")
        elif op == "8":
            _assistente_sn_vinculos()
        elif op == "9":
            _mostrar_catalogo_atividades(sorted(atv_set, key=str), _catalogo_all())
            input(DM + "\n  [ENTER] " + RS)
        else:
            aviso("Opcao invalida.")


def resolver_conflitos_e_reatribuir(turmas, atividades_reais):
    """
    Atividades com mais de uma turma: paralelo ou exclusivo.
    Reatribuicao: qualquer atividade -> turma executora (reforco / outra funcao).
    Retorna reatribuicao, paralelo, primaria.
    """
    reatribuicao = {}
    paralelo = {}
    primaria = {}

    def candidatos(atv):
        return [t["nome"] for t in turmas if atv in t["atividades"]]

    for atv in atividades_reais:
        c = candidatos(atv)
        if len(c) <= 1:
            continue
        sub()
        print(Y + f"  Conflito: '{str(atv)[:58]}'" + RS)
        print(DM + f"  Turmas: {', '.join(c)}" + RS)
        if confirmar(
            "  Varias turmas em PARALELO (dividem a mesma demanda no tempo)?",
            default=True,
        ):
            paralelo[atv] = True
        else:
            paralelo[atv] = False
            p = selecionar("  Turma EXCLUSIVA para esta atividade", c)
            if p:
                primaria[atv] = p

    if confirmar(
        "\n  Reatribuir atividades (reforco: outra turma executa, ex. adubacao faz uma roçada)?",
        default=False,
    ):
        nomes_turmas = [t["nome"] for t in turmas]
        while True:
            idx = selecionar_paginado(
                "REATRIBUIR — escolha a ATIVIDADE", atividades_reais, page_size=6
            )
            if idx < 0:
                break
            atv = atividades_reais[idx]
            print(G + f"\n  Atividade: {str(atv)[:62]}" + RS)
            t_alvo = selecionar(
                "  Turma que EXECUTA (capacidade desta turma)", nomes_turmas
            )
            if t_alvo:
                reatribuicao[atv] = t_alvo
                ok(
                    f"Executora: '{t_alvo}' (sobrescreve vinculos anteriores para o cronograma)."
                )

    return reatribuicao, paralelo, primaria


def turmas_que_executam(atv, turmas, reatribuicao, paralelo, primaria):
    """Lista de nomes de turma que trabalham nesta atividade no simulador."""
    if atv in reatribuicao:
        return [reatribuicao[atv]]
    c = [t["nome"] for t in turmas if atv in t["atividades"]]
    if not c:
        return []
    if len(c) == 1:
        return c
    if paralelo.get(atv, True):
        return c
    p = primaria.get(atv)
    return [p] if p else c


def atividades_por_filtro(atividades_reais, filtros_texto):
    """Retorna atividades cujo nome contem algum filtro (sem acento)."""
    filtros = [
        remover_acentos(x).strip() for x in (filtros_texto or []) if str(x).strip()
    ]
    out = set()
    for atv in atividades_reais:
        nome = remover_acentos(str(atv))
        if any(f in nome for f in filtros):
            out.add(atv)
    return sorted(out, key=lambda x: str(x))


_FILTROS_NOME_CANDIDATAS_MECANIZADO = [
    "mecaniz",
    "maquina",
    "máquina",
    "trator",
    "motocovead",
    "motocultor",
    "robo",
    "robô",
    "pulveriz",
    "atomiz",
    "implemento",
    "coveador",
    "solo mec",
    "mec c/",
    "mec s/",
    "esteira",
    "drone",
    "máq",
    "maq.",
]


def atividades_candidatas_mecanizado(atividades_reais, cfg=None):
    """Atividades provavelmente mecanizadas: palavras-chave no nome e/ou tipo HM na tarifa CT."""
    cfg = cfg or {}
    tarifas = cfg.get("tarifas", {}) or {}
    merged = set(
        atividades_por_filtro(atividades_reais, _FILTROS_NOME_CANDIDATAS_MECANIZADO)
    )
    for atv in atividades_reais:
        t_nome = resolver_chave_tarifa(cfg, tarifas, atv)
        row = tarifas.get(t_nome)
        if not isinstance(row, dict):
            continue
        tipo = str(row.get("tipo", "")).lower()
        try:
            hm = float(row.get("rendimento_hm", 0) or 0)
        except (TypeError, ValueError):
            hm = 0.0
        if "mecaniz" in tipo or hm > 0:
            merged.add(atv)
    return sorted(merged, key=str)


def sequencia_manutencao_seco_placeholder(cfg):
    aviso(
        "Modo manutencao_seco: regras de sequencia ainda nao definidas (stub). Cascata desligada nesta execucao."
    )


def sequencia_manutencao_umido_placeholder(cfg):
    aviso(
        "Modo manutencao_umido: regras de sequencia ainda nao definidas (stub). Cascata desligada nesta execucao."
    )


def _match_filtros_fase(nome_atv, filtros, exclusoes=None):
    """True se nome contem algum filtro (normalizar_chave) e nenhuma exclusao."""
    kn = normalizar_chave(nome_atv)
    for ex in exclusoes or []:
        exn = normalizar_chave(ex)
        if exn and exn in kn:
            return False
    for f in filtros or []:
        fn = normalizar_chave(f)
        if fn and fn in kn:
            return True
    return False


def eh_limpeza_quimica_pos_plantio(atv, seq_cfg):
    """Limpeza quimica pos-plantio: todos filtros presentes e nenhuma exclusao."""
    kn = normalizar_chave(atv)
    filtros = seq_cfg.get("limpeza_quimica_filtros") or ["limpeza", "quim"]
    exclusoes = seq_cfg.get("limpeza_quimica_exclusoes") or []
    for f in filtros:
        fn = normalizar_chave(f)
        if not fn or fn not in kn:
            return False
    for ex in exclusoes:
        exn = normalizar_chave(ex)
        if exn and exn in kn:
            return False
    return True


def _fases_ordem_config(seq_cfg, modo):
    if modo == "personalizado" and seq_cfg.get("personalizado_ordem"):
        return seq_cfg["personalizado_ordem"]
    if modo == "manutencao_swg":
        return seq_cfg.get("swg_fases") or []
    return seq_cfg.get("implantacao_fases") or []


def classificar_fase_cascata_valor(
    atv, seq_cfg, modo, atividades_plantio, atividades_irrig
):
    """
    Retorna indice numerico de fase para cascata global.
    manutencao_seco/umido: 0.0 (sem cascata).
    """
    if modo in ("manutencao_seco", "manutencao_umido"):
        return 0.0
    if eh_limpeza_quimica_pos_plantio(atv, seq_cfg):
        return 8.0
    if atv in atividades_plantio or _match_filtros_fase(
        atv, seq_cfg.get("filtros_plantio") or ["plantio"], None
    ):
        return 6.0
    if atv in atividades_irrig or _match_filtros_fase(
        atv, seq_cfg.get("filtros_irrigacao") or ["irrig"], None
    ):
        return 7.0
    fases = _fases_ordem_config(seq_cfg, modo)
    for i, fase in enumerate(fases):
        if _match_filtros_fase(atv, fase.get("filtros") or [], fase.get("exclusoes")):
            return float(i)
    try:
        return float(seq_cfg.get("implantacao_outras_fase", 5.5))
    except (TypeError, ValueError):
        return 5.5


def _demanda_plantio_talhao(talhao, demanda_global, atividades_plantio):
    for p in atividades_plantio:
        if demanda_global.get((talhao, p), 0) > 0.01:
            return True
    return False


def limpeza_permitida_por_talhao(
    talhao,
    dia,
    seq_cfg,
    dia_termino_plantio,
    tem_plantio_previsto_no_talhao,
):
    """Se nao ha plantio no talhao, limpeza pos-plantio nao exige offset."""
    if not tem_plantio_previsto_no_talhao:
        return True
    d = dia_termino_plantio.get(talhao)
    if d is None:
        return False
    off = int(seq_cfg.get("offset_limpeza_quimica_dias", 30) or 30)
    return dia >= d + off


def _min_fase_cascata(
    demanda_global,
    seq_cfg,
    modo,
    usar_cascata,
    usar_bloqueio_global,
    atividades_bloqueadas,
    atividades_plantio,
    atividades_irrig,
    dia,
    dia_termino_plantio,
    tem_plantio_por_talhao,
):
    """Menor fase entre demandas ainda > 0 e elegiveis neste dia."""
    if not usar_cascata or modo in ("manutencao_seco", "manutencao_umido"):
        return None
    bloqueadas = set(atividades_bloqueadas or [])
    vals = []
    for (talhao, atv), hh in demanda_global.items():
        if hh <= 0.01:
            continue
        if usar_bloqueio_global and atv in bloqueadas:
            if _ha_trabalho_nao_bloqueado(demanda_global, bloqueadas):
                continue
        if eh_limpeza_quimica_pos_plantio(atv, seq_cfg):
            if not limpeza_permitida_por_talhao(
                talhao,
                dia,
                seq_cfg,
                dia_termino_plantio,
                tem_plantio_por_talhao.get(talhao, False),
            ):
                continue
        fv = classificar_fase_cascata_valor(
            atv, seq_cfg, modo, atividades_plantio, atividades_irrig
        )
        vals.append(fv)
    if not vals:
        return None
    return min(vals)


def pode_agendar_atividade_cascata(
    talhao,
    atv,
    demanda_global,
    seq_cfg,
    modo,
    usar_cascata,
    usar_bloqueio_global,
    atividades_bloqueadas,
    atividades_plantio,
    atividades_irrig,
    dia,
    dia_termino_plantio,
    tem_plantio_por_talhao,
    min_fase_dia,
):
    """Regras combinadas: cascata, bloqueio global, plantio antes de irrigacao, limpeza pos-plantio."""
    hh = demanda_global.get((talhao, atv), 0)
    if hh <= 0.01:
        return False
    if usar_bloqueio_global and atv in atividades_bloqueadas:
        if _ha_trabalho_nao_bloqueado(demanda_global, atividades_bloqueadas):
            return False
    if eh_limpeza_quimica_pos_plantio(atv, seq_cfg):
        if not limpeza_permitida_por_talhao(
            talhao,
            dia,
            seq_cfg,
            dia_termino_plantio,
            tem_plantio_por_talhao.get(talhao, False),
        ):
            return False
    if atv in atividades_irrig or _match_filtros_fase(
        atv, seq_cfg.get("filtros_irrigacao") or ["irrig"], None
    ):
        if _demanda_plantio_talhao(talhao, demanda_global, atividades_plantio):
            return False
    if usar_cascata and modo not in ("manutencao_seco", "manutencao_umido"):
        fv = classificar_fase_cascata_valor(
            atv, seq_cfg, modo, atividades_plantio, atividades_irrig
        )
        if (
            min_fase_dia is not None
            and abs(fv - min_fase_dia) > 1e-6
            and fv > min_fase_dia + 1e-6
        ):
            return False
    return True


def diagnosticar_sequencia_atividades(atividades_reais, seq_cfg, modo):
    """Avisos: atividades classificadas como 'outras' (fase intermediaria) e lista."""
    if modo in ("manutencao_seco", "manutencao_umido"):
        return
    ap = set(
        atividades_por_filtro(
            atividades_reais, seq_cfg.get("filtros_plantio") or ["plantio"]
        )
    )
    ai = set(
        atividades_por_filtro(
            atividades_reais, seq_cfg.get("filtros_irrigacao") or ["irrig"]
        )
    )
    outras = []
    for atv in atividades_reais:
        if eh_limpeza_quimica_pos_plantio(atv, seq_cfg):
            continue
        if atv in ap or atv in ai:
            continue
        fases = _fases_ordem_config(seq_cfg, modo)
        ok_fase = False
        for fase in fases:
            if _match_filtros_fase(
                atv, fase.get("filtros") or [], fase.get("exclusoes")
            ):
                ok_fase = True
                break
        if not ok_fase:
            outras.append(atv)
    if outras:
        print(
            DM
            + f"\n  Sequencia ({modo}): {len(outras)} atividade(s) em fase generica 'demais (antes plantio)'."
            + RS
        )
        print(
            DM
            + "  Executam antes de plantio, sem fase fixa na cascata. Para priorizar, adicione filtros em config.sequencia."
            + RS
        )
        for a in sorted(outras, key=str)[:15]:
            print(DM + f"    - {str(a)[:70]}" + RS)
        if len(outras) > 15:
            print(DM + f"    ... +{len(outras) - 15}" + RS)
    else:
        ok("Todas as atividades possuem fase explicita na sequencia.")


def _ha_trabalho_nao_bloqueado(demanda_global, atividades_bloqueadas):
    """True se ainda existe demanda >0 para atividade fora do grupo bloqueado."""
    bloqueadas = set(atividades_bloqueadas or [])
    for (_, atv), hh in demanda_global.items():
        if hh > 0.01 and atv not in bloqueadas:
            return True
    return False


def auditar_cadeia_dados(cfg, demandas, atividades_reais, session_hh=None):
    """Auditoria unificada: micro → de_para → tarifa CT → HH/Preço/Custo_hora."""
    tarifas = cfg.get("tarifas", {})
    de_para = cfg.get("de_para", {})
    sem_depara = []
    sem_tarifa = []
    sem_hh = []
    sem_preco = []
    total = 0
    for talhao, lista in demandas.items():
        for t in lista:
            total += 1
            atv = t["atividade"]
            chave = resolver_chave_tarifa(cfg, tarifas, atv)
            if atv not in de_para and atv != chave:
                pass
            elif atv not in de_para and chave == atv and chave not in tarifas:
                sem_depara.append(str(atv)[:55])
            if chave not in tarifas:
                sem_tarifa.append(f"{str(atv)[:40]} → {str(chave)[:40]}")
            else:
                entry = tarifas[chave]
                rh_e = float(entry.get("rendimento_hh", 0) or 0)
                if session_hh and (atv in session_hh or chave in session_hh):
                    rh_e = max(rh_e, 1e-6)
                if rh_e <= 0:
                    sem_hh.append(f"{str(chave)[:55]}")
                if float(entry.get("preco_unit", 0) or 0) <= 0:
                    sem_preco.append(f"{str(chave)[:55]}")
    sub()
    print(G + BL + "  AUDITORIA CADEIA DE DADOS" + RS)
    print(
        G
        + f"  Demandas: {total} | Atividades unicas: {len(atividades_reais)} | de_para: {len(de_para)} | Tarifas CT: {len(tarifas)}"
        + RS
    )
    if session_hh:
        print(
            DM
            + f"  Overrides HH/ha nesta execucao: {len(session_hh)} chave(s) (nao gravados no config)."
            + RS
        )
    if sem_depara:
        u = sorted(set(sem_depara))
        print(Y + f"\n  Sem de_para ({len(u)}) — atividade micro nao mapeada:" + RS)
        for x in u[:10]:
            print(Y + f"    - {x}" + RS)
        if len(u) > 10:
            print(DM + f"    ... +{len(u) - 10}" + RS)
    if sem_tarifa:
        u = sorted(set(sem_tarifa))
        print(
            Y
            + f"\n  Sem tarifa CT ({len(u)}) — chave nao encontrada no orcamento importado:"
            + RS
        )
        for x in u[:10]:
            print(Y + f"    - {x}" + RS)
        if len(u) > 10:
            print(DM + f"    ... +{len(u) - 10}" + RS)
    if sem_hh:
        u = sorted(set(sem_hh))
        print(Y + f"\n  HH zerado ({len(u)}) — rendimento_hh = 0 na tarifa:" + RS)
        for x in u[:10]:
            print(Y + f"    - {x}" + RS)
        if len(u) > 10:
            print(DM + f"    ... +{len(u) - 10}" + RS)
    if sem_preco:
        u = sorted(set(sem_preco))
        print(Y + f"\n  Preco zerado ({len(u)}) — preco_unit = 0 na tarifa:" + RS)
        for x in u[:10]:
            print(Y + f"    - {x}" + RS)
        if len(u) > 10:
            print(DM + f"    ... +{len(u) - 10}" + RS)
    if not sem_depara and not sem_tarifa and not sem_hh and not sem_preco:
        ok("Cadeia de dados completa — nenhuma lacuna detectada.")
    else:
        total_lacunas = (
            len(set(sem_depara))
            + len(set(sem_tarifa))
            + len(set(sem_hh))
            + len(set(sem_preco))
        )
        aviso(
            f"Total de lacunas: {total_lacunas}. Corrija via menu [4] de_para ou [2] importar tarifas."
        )
    sub()


def auto_mapear_de_para(cfg, atividades_reais):
    """
    Mapeia automaticamente atividades do micro para chaves de tarifa por similaridade textual.
    Usa normalizar_chave para comparacao robusta.
    """
    tarifas = cfg.get("tarifas", {})
    if not tarifas:
        return 0
    de_para = cfg.setdefault("de_para", {})
    tarif_norm = {k: normalizar_chave(k) for k in tarifas.keys()}
    novos = 0
    for atv in atividades_reais:
        if atv in de_para:
            continue
        cands = _candidatos_chave_atividade(atv)
        an = cands[-1] if cands else normalizar_chave(atv)
        melhor = None
        melhor_score = 0
        for tk, tn in tarif_norm.items():
            score = 0
            if an == tn:
                score = 1000
            elif an in tn or tn in an:
                score = min(len(an), len(tn))
            else:
                toks_a = set(x for x in an.split() if len(x) > 2)
                toks_t = set(x for x in tn.split() if len(x) > 2)
                inter = len(toks_a & toks_t)
                if inter >= 3:
                    score = inter
            if score > melhor_score:
                melhor_score = score
                melhor = tk
        if melhor and melhor_score >= 3:
            de_para[atv] = melhor
            novos += 1
    if novos > 0:
        salvar_config(cfg)
    return novos


def aplicar_depara_padrao_exame(cfg, atividades_reais):
    """
    Aplica mapeamento fixo (hardcoded) do prototipo EXAME->CT_317.
    1) Dicionario exato por normalizar_chave; 2) heuristica por palavras-chave (APPN, parenteses, etc.).
    """
    tarifas = cfg.get("tarifas", {})
    if not tarifas:
        return 0
    de_para = cfg.setdefault("de_para", {})
    novo = 0
    for atv in atividades_reais:
        alvo = None
        for kn in _candidatos_chave_atividade(atv):
            if kn in DEFAULT_DEPARA_EXAME_CT317:
                alvo = DEFAULT_DEPARA_EXAME_CT317[kn]
                break
            alvo = _depara_heuristico_exame_ct317(kn, tarifas)
            if alvo:
                break
        if alvo and alvo in tarifas and de_para.get(atv) != alvo:
            de_para[atv] = alvo
            novo += 1
    if novo:
        salvar_config(cfg)
    return novo


def _somente_bloqueado_restante(demanda_global, atividades_bloqueadas):
    """True se so resta demanda em atividades bloqueadas (ex. plantio/irrigacao)."""
    bloqueadas = set(atividades_bloqueadas or [])
    tem_nao_bloqueado = False
    tem_bloqueado = False
    for (_, atv), hh in demanda_global.items():
        if hh <= 0.01:
            continue
        if atv in bloqueadas:
            tem_bloqueado = True
        else:
            tem_nao_bloqueado = True
    return tem_bloqueado and not tem_nao_bloqueado


def _mostrar_painel_hh_hm_pre_scheduler(demandas, fazenda, detalhado=False, limite=120):
    """
    Painel rapido pre-simulacao com HH/HM por atividade.
    detalhado=False: consolidado por atividade na fazenda.
    detalhado=True: detalhado por talhao + atividade.
    """
    linhas = []
    for talhao, tarefas in (demandas or {}).items():
        for t in tarefas or []:
            area = float(t.get("area", 0) or 0)
            hh_t = float(t.get("hh_total", 0) or 0)
            hm_t = float(t.get("hm_total", 0) or 0)
            atv = str(t.get("atividade", ""))
            tipo = str(t.get("tipo", ""))
            if area <= 0.0001 and hh_t <= 0.0001 and hm_t <= 0.0001:
                continue
            linhas.append(
                {
                    "talhao": str(talhao),
                    "atividade": atv,
                    "area": area,
                    "hh_total": hh_t,
                    "hm_total": hm_t,
                    "tipo": tipo,
                }
            )

    if not linhas:
        return

    if detalhado:
        table = Table(title=f"Pre-voo HH/HM por Talhao - {fazenda}")
        table.add_column("Talhao", style="cyan")
        table.add_column("Atividade", style="green")
        table.add_column("Area(ha)", justify="right")
        table.add_column("HH/ha", justify="right")
        table.add_column("HM/ha", justify="right")
        table.add_column("HH total", justify="right")
        table.add_column("HM total", justify="right")
        table.add_column("Tipo", justify="center")
        rows = sorted(linhas, key=lambda r: (str(r["talhao"]), str(r["atividade"])))
        total = len(rows)
        for i, r in enumerate(rows):
            if i >= limite:
                break
            area = float(r["area"])
            hh_total = float(r["hh_total"])
            hm_total = float(r["hm_total"])
            hh_ha = (hh_total / area) if area > 0.0001 else 0.0
            hm_ha = (hm_total / area) if area > 0.0001 else 0.0
            table.add_row(
                str(r["talhao"]),
                str(r["atividade"])[:38],
                f"{area:.2f}",
                f"{hh_ha:.2f}",
                f"{hm_ha:.2f}",
                f"{hh_total:.2f}",
                f"{hm_total:.2f}",
                str(r["tipo"] or "-"),
            )
        console.print(table)
        if total > limite:
            print(DM + f"  ... +{total - limite} linha(s) detalhadas no total." + RS)
        return

    agg = defaultdict(
        lambda: {
            "area": 0.0,
            "hh_total": 0.0,
            "hm_total": 0.0,
            "tipo": "",
        }
    )
    for r in linhas:
        k = str(r["atividade"])
        agg[k]["area"] += float(r["area"])
        agg[k]["hh_total"] += float(r["hh_total"])
        agg[k]["hm_total"] += float(r["hm_total"])
        if not agg[k]["tipo"]:
            agg[k]["tipo"] = str(r.get("tipo", ""))

    table = Table(title=f"Pre-voo HH/HM por Atividade - {fazenda}")
    table.add_column("Atividade", style="green")
    table.add_column("Area(ha)", justify="right")
    table.add_column("HH/ha", justify="right")
    table.add_column("HM/ha", justify="right")
    table.add_column("HH total", justify="right")
    table.add_column("HM total", justify="right")
    table.add_column("Tipo", justify="center")

    rows = sorted(agg.items(), key=lambda kv: str(kv[0]))
    total = len(rows)
    for i, (atv, dados) in enumerate(rows):
        if i >= limite:
            break
        area = float(dados["area"])
        hh_total = float(dados["hh_total"])
        hm_total = float(dados["hm_total"])
        hh_ha = (hh_total / area) if area > 0.0001 else 0.0
        hm_ha = (hm_total / area) if area > 0.0001 else 0.0
        table.add_row(
            str(atv)[:42],
            f"{area:.2f}",
            f"{hh_ha:.2f}",
            f"{hm_ha:.2f}",
            f"{hh_total:.2f}",
            f"{hm_total:.2f}",
            str(dados.get("tipo", "") or "-"),
        )
    console.print(table)
    if total > limite:
        print(DM + f"  ... +{total - limite} atividade(s) no total." + RS)


def menu_ajustes_hh_apenas_sessao(atividades_reais, cfg, session_hh):
    """Edita HH/ha por atividade apenas na memoria (nao salva config.json)."""
    if session_hh is None:
        return
    tarifas = cfg.get("tarifas", {})
    strict = cfg.get("orcamento_estrito", True)
    sub()
    print(G + BL + "  AJUSTE DE HH/ha — APENAS ESTA EXECUCAO" + RS)
    print(DM + "  Nao grava em config. ENTER = manter valor atual." + RS + "\n")
    n = 0
    n_skip_mec_hm = 0
    for atv in sorted(set(atividades_reais), key=str):
        t_nome = resolver_chave_tarifa(cfg, tarifas, atv)
        row_tarifa = tarifas.get(t_nome, {})
        if not isinstance(row_tarifa, dict):
            row_tarifa = {}
        tipo = str(row_tarifa.get("tipo", "")).lower()
        try:
            hm = float(row_tarifa.get("rendimento_hm", 0) or 0)
        except (TypeError, ValueError):
            hm = 0.0
        if "mecaniz" in tipo or hm > 0:
            n_skip_mec_hm += 1
            continue
        cur = resolver_rendimento_hh(
            cfg, tarifas, t_nome, strict=strict, session_hh=session_hh, atv_micro=atv
        )
        if cur is None:
            cur = 0.0
        v = prompt(f"  [{str(atv)[:46]}]  CT:{str(t_nome)[:36]}  HH/ha [{cur}]", "")
        vs = str(v).strip()
        if not vs:
            continue
        try:
            nv = float(vs.replace(",", "."))
            session_hh[atv] = nv
            session_hh[t_nome] = nv
            n += 1
        except (TypeError, ValueError):
            aviso("Valor invalido, ignorado.")
    if n:
        ok(f"{n} override(s) HH/ha nesta sessao.")
    else:
        print(DM + "  Nenhum override informado." + RS)
    if n_skip_mec_hm:
        print(
            DM
            + f"  {n_skip_mec_hm} atividade(s) HM-only/mecanizadas foram ocultadas deste ajuste de HH."
            + RS
        )


def validar_e_completar_orcamento(cfg, atividades_reais, session_hh=None):
    """
    Modo orcamento_estrito: toda atividade do micro precisa de chave em tarifas com dados CT.
    Lacunas: escolher tarifa na lista ou informar HH/preco/custo manualmente.
    session_hh: dict opcional; HH informado para linha zerada pode ir apenas para a sessao (nao grava tarifas).
    Retorna False se usuario abortar.
    """
    if not cfg.get("orcamento_estrito", True):
        return True
    tarifas = cfg.setdefault("tarifas", {})
    de_para = cfg.setdefault("de_para", {})
    nomes_tarifa = sorted(tarifas.keys(), key=str)
    if not nomes_tarifa:
        erro("Nenhuma tarifa em config. Normalize a CT [3] ou importe [2].")
        return False

    for atv in sorted(set(atividades_reais), key=str):
        t_nome = resolver_chave_tarifa(cfg, tarifas, atv)
        if t_nome not in tarifas:
            sub()
            print(Y + f"  [ESTRITO] Sem tarifa CT para atividade do micro:" + RS)
            print(Y + f"    {str(atv)[:70]}" + RS)
            print(DM + f"    Chave atual: {t_nome}" + RS)
            if confirmar(
                "  Escolher uma linha existente em tarifas (recomendado)?", default=True
            ):
                idx = selecionar_paginado(
                    "TARIFA CT (orcamento)", nomes_tarifa, page_size=8
                )
                if idx < 0:
                    return False
                de_para[atv] = nomes_tarifa[idx]
                salvar_config(cfg)
            else:
                hh_m = pedir_float("  HH/ha (manual)", 8.0)
                pr_m = pedir_float("  Preco R$/ha (manual)", 0.0, allow_zero=True)
                ch_m = pedir_float(
                    "  Custo R$/h (manual)",
                    float(cfg.get("custo_hora_tf") or 50),
                    allow_zero=True,
                )
                chave = prompt(
                    "  Nome da chave a gravar em tarifas (ex.: alias)", t_nome[:48]
                )
                if not chave:
                    chave = t_nome
                tarifas[chave] = {
                    "rendimento_hh": hh_m,
                    "preco_ha": pr_m,
                    "preco_unit": pr_m,
                    "custo_hora": ch_m,
                    "custo_ha": hh_m * ch_m if ch_m > 0 else 0,
                    "tipo": "Manual",
                    "recurso": "homem",
                    "eficiencia": 1.0,
                }
                de_para[atv] = chave
                salvar_config(cfg)
        t_nome = resolver_chave_tarifa(cfg, tarifas, atv)
        row = tarifas.get(t_nome, {})
        try:
            rh = float(row.get("rendimento_hh", 0))
        except (TypeError, ValueError):
            rh = 0.0
        tipo = str(row.get("tipo", "")).lower()
        hm = float(row.get("rendimento_hm", 0) or 0)
        is_mec = "mecaniz" in tipo or hm > 0
        if rh <= 0.01 and not is_mec:
            sub()
            print(
                Y
                + f"  [ESTRITO] rendimento_hh zero ou invalido na tarifa '{str(t_nome)[:50]}'"
                + RS
            )
            hh_m = pedir_float(
                "  Informe HH/ha para esta linha (ou 0 se so maquina)",
                0.0,
                allow_zero=True,
            )
            if session_hh is not None and confirmar(
                "  Aplicar SO nesta execucao (nao gravar em config.json)?", default=True
            ):
                session_hh[atv] = float(hh_m)
                session_hh[t_nome] = float(hh_m)
            else:
                tarifas[t_nome]["rendimento_hh"] = float(hh_m)
                salvar_config(cfg)
    return True


# ──────────────────────────────────────────────
#  SMART SCHEDULER v2:
#  Filas por TURMA, sequenciais por talhao.
#  Atividades podem ser paralelas (varias turmas) ou exclusivas (uma turma).
#  Reatribuicao: qualquer atividade pode ser executada por qualquer turma.
# ──────────────────────────────────────────────
def dias_uteis_no_periodo(mes_ini, ano_ini, meses):
    dias = 0
    m, a = mes_ini, ano_ini
    for _ in range(int(math.ceil(meses))):
        for sem in calendar.monthcalendar(a, m):
            for i, d in enumerate(sem):
                if d != 0 and i < 5:
                    dias += 1
        m += 1
        if m > 12:
            m = 1
            a += 1
    return dias


_SEQUENCIAS_DISPONIVEIS = [
    (
        "implantacao",
        "Rocada > Formiga > Coroamento > Coveamento > Adubacao > Plantio > Irrigacao (cascata)",
    ),
    (
        "manutencao_swg",
        "Rocada manual > Limpeza de area > Capina de coroa > Formigas > Coveamento > Adubacao > Plantio > Irrigacao (ordem SWG)",
    ),
    (
        "manutencao_seco",
        "[EM PROGRESSO] Manutencao periodo seco — regras ainda nao definidas",
    ),
    (
        "manutencao_umido",
        "[EM PROGRESSO] Manutencao periodo umido — regras ainda nao definidas",
    ),
    ("personalizado", "Ordem livre (sem bloqueio global plantio/irrigacao)"),
]


def _selecionar_sequencia_padrao_sn(cfg, seq_cfg):
    sub()
    print(G + BL + "  SELECIONAR SEQUENCIA PADRAO:" + RS)
    print(DM + "  Responda S para a sequencia desejada (apenas UMA):" + RS + "\n")
    escolhido = None
    for modo_id, descr in _SEQUENCIAS_DISPONIVEIS:
        resp = confirmar(
            f"  {modo_id}: {descr}",
            default=(modo_id == seq_cfg.get("modo", "implantacao")),
        )
        if resp:
            escolhido = modo_id
            break
    if not escolhido:
        aviso("Nenhuma sequencia selecionada. Repetindo...")
        return _selecionar_sequencia_padrao_sn(cfg, seq_cfg)
    seq_cfg["modo"] = escolhido
    ok(f"Sequencia: {escolhido}")
    if confirmar("  Salvar como padrao para proximas execucoes?", default=True):
        cfg["sequencia"] = seq_cfg
        salvar_config(cfg)
    return escolhido


def _norm_atv(x):
    """Normaliza nome de atividade para cruzamento template x micro (NA-safe, str strip)."""
    if x is None:
        return ""
    try:
        if pd.isna(x):
            return ""
    except (TypeError, ValueError):
        pass
    return str(x).strip()


def _slug_ficheiro_seguro(s, max_len=48):
    """Nome seguro para ficheiros Windows (sem acentos problematicos)."""
    t = remover_acentos(str(s).strip()) if s else ""
    t = re.sub(r'[<>:"/\\|?*]+', "_", t)
    t = re.sub(r"\s+", "_", t)
    t = re.sub(r"_+", "_", t).strip("_")
    return t[:max_len] if t else "escopo"


def _distribuir_atividades_faltantes_turmas(turmas, atividades_reais, fazenda):
    """
    Atribui à turma mais numerosa as atividades da fazenda que não entraram no template
    (ex.: equipe só irrigação em micro sem irrigação — evita cronograma vazio).
    """
    if not turmas or not atividades_reais:
        return
    cobertura = set()
    for t in turmas:
        for a in t.get("atividades") or []:
            na = _norm_atv(a)
            if na:
                cobertura.add(na)
    farm_set = set(atividades_reais)
    orfas = [a for a in atividades_reais if a not in cobertura]
    if not orfas:
        return
    tpl_extra = sorted(cobertura - farm_set, key=str)
    if tpl_extra:
        print(
            DM
            + f"  Modelo de turmas: {len(tpl_extra)} atividade(s) sem demanda nesta fazenda "
            + f"('{str(fazenda)[:42]}') — ignoradas no micro atual."
            + RS
        )
        for x in tpl_extra[:5]:
            print(DM + f"    ~ {str(x)[:58]}" + RS)
        if len(tpl_extra) > 5:
            print(DM + f"    ... +{len(tpl_extra) - 5}" + RS)
    alvo = max(turmas, key=lambda t: int(t.get("operarios", 0) or 0))
    cur = {_norm_atv(x) for x in alvo.get("atividades", []) if _norm_atv(x)}
    cur |= set(orfas)
    alvo["atividades"] = sorted(cur, key=str)
    ok(
        f"{len(orfas)} atividade(s) desta fazenda sem turma no modelo foram atribuidas a '{alvo['nome']}' "
        f"(evita cronograma vazio quando o template nao cruza o micro)."
    )


def calcular_cronograma_inteligente(
    cfg,
    df_faz,
    fazenda,
    esperar_enter=True,
    ctx=None,
    escopo_meta=None,
    atividades_catalogo=None,
    modo_comparativo=False,
    substituicoes_comparativo=None,
):
    """
    ctx: optional dict with preconfigured session state for batch mode.
    When ctx is provided, interactive setup questions are skipped.
    """
    _batch = ctx is not None
    comparativo_cfg = None
    contexto_sessao.atualizar_configuracoes(cfg)
    contexto_sessao.atualizar_fazenda(fazenda, df_faz)
    _emitir_monitor_atual()
    dashboard_header()

    if not _batch:
        subcabecalho(f"SMART SCHEDULER - {fazenda}")
        df_faz = avaliar_terreno(df_faz)
        aviso_politica_tarifas_planas()
    else:
        sub()
        print(G + BL + f"  SMART SCHEDULER - {fazenda}" + RS)
        df_faz["penalidade"] = float(ctx.get("penalidade", 1.0))

        _emitir_monitor_state(
            {
                "operacao": {
                    "fazenda_atual": str(fazenda),
                    "modo": "batch" if _batch else "single",
                    "micro_basename": str(cfg.get("arquivo_micro", "")),
                    "status_geral": "iniciando_scheduler",
                    "mensagem_curta": f"Scheduler iniciado para {fazenda}",
                }
            }
        )

    # ── Extrair atividades REAIS da fazenda ──
    df_faz = df_faz.copy()
    df_faz["atividade"] = df_faz["atividade"].map(
        lambda x: _norm_atv(x) if pd.notna(x) else x
    )
    if not _batch and confirmar(
        "Ajustar escopo de atividades (substituir/remover/adicionar) nesta execucao?",
        default=False,
    ):
        df_faz = _menu_ajustar_escopo_atividades(
            df_faz,
            cfg=cfg,
            atividades_catalogo=atividades_catalogo,
        )
    atividades_reais = sorted(
        {a for a in df_faz["atividade"].dropna().unique().tolist() if _norm_atv(a)},
        key=str,
    )
    catalogo_global = _catalogo_atividades_completo(
        atividades_reais,
        cfg=cfg,
        atividades_catalogo=atividades_catalogo,
    )
    talhoes_ordenados = sorted(df_faz["chave"].dropna().unique().tolist())
    escopo_talhoes = []
    if isinstance(escopo_meta, dict):
        escopo_talhoes = list(escopo_meta.get("talhoes") or [])

    if escopo_talhoes:
        contexto_sessao.definir_escopo_talhoes(escopo_talhoes, talhoes_ordenados)
    else:
        contexto_sessao.definir_escopo_talhoes(talhoes_ordenados, talhoes_ordenados)
    contexto_sessao.atualizar_atividades(0, len(atividades_reais))
    _emitir_monitor_atual()

    novos_fixos = aplicar_depara_padrao_exame(cfg, atividades_reais)
    if novos_fixos > 0:
        ok(f"de_para PADRAO aplicado: {novos_fixos} mapeamento(s) fixos EXAME->CT_313.")
    if not cfg.get("orcamento_estrito", True):
        novos_de_para = auto_mapear_de_para(cfg, atividades_reais)
        if novos_de_para > 0:
            ok(f"de_para complementar: {novos_de_para} mapeamento(s) adicionais.")

    if not _batch:
        sub()
        print(
            DM
            + "  Orcamento estrito (sem mediana silenciosa; lacunas pedem input): "
            + C
            + str(cfg.get("orcamento_estrito", True))
            + RS
        )
        if confirmar("  Alternar orcamento_estrito para esta execucao?", default=False):
            cfg["orcamento_estrito"] = not cfg.get("orcamento_estrito", True)
            salvar_config(cfg)
            ok(f"orcamento_estrito = {cfg['orcamento_estrito']}")

    sub()
    print(G + BL + "  ATIVIDADES ENCONTRADAS NESTA FAZENDA:" + RS)
    for i, a in enumerate(atividades_reais, 1):
        print(G + f"  {i:2}. " + C + a + RS)
    print(G + f"\n  Talhoes: " + C + f"{len(talhoes_ordenados)}" + RS)
    if escopo_talhoes:
        n_show = min(8, len(escopo_talhoes))
        base = ", ".join(str(x)[:24] for x in escopo_talhoes[:n_show])
        if len(escopo_talhoes) > n_show:
            base += f", ... (+{len(escopo_talhoes) - n_show})"
            print(DM + f" Escopo talhoes selecionados: {base}" + RS)
    sub()

    # ═══════════════════════════════════════════════════════════════════════════
    # MODO COMPARATIVO: MANUAL vs MECANIZADO
    # ═══════════════════════════════════════════════════════════════════════════
    modo_comparativo = False
    substituicoes_comparativo = {}
    
    seq_cfg = cfg.get("sequencia") or {}

    if not _batch:
        # Verificar se há atividades com equivalente mecanizado
        pares_mecanizaveis = _atividades_com_mecanizado_disponivel(atividades_reais)
        if atividades_reais:
            sub()
            print(C + BL + " MODO COMPARATIVO MANUAL vs MECANIZADO" + RS)
            if pares_mecanizaveis:
                print(
                    DM
                    + f" Detectadas {len(pares_mecanizaveis)} atividade(s) com equivalente mecanizado."
                    + RS
                )
            else:
                print(
                    DM
                    + " Nenhuma sugestao automatica encontrada; use modo manual [2] ou recurso externo [3]."
                    + RS
                )
            
            if confirmar("Deseja executar comparativo MANUAL vs MECANIZADO?", default=False):
                modo_comparativo = True
                # Loop para permitir voltar entre modos
                while True:
                    sub()
                    print(G + " [1] Usar sugestões automáticas (detecção por nome)" + RS)
                    print(G + " [2] Escolher manualmente do catálogo completo" + RS)
                    print(G + " [3] Cadastrar recurso mecanizado externo" + RS)
                    print(R + " [0] Cancelar comparativo" + RS)
                    print(DM + "    (opcão 2 permite escolher QUALQUER atividade mecanizada)" + RS)
                    sub()
                    modo_escolha = input(DM + ">> Opção [1/2/3/0]: " + RS).strip() or "1"
                    
                    if modo_escolha == "0":
                        modo_comparativo = False
                        substituicoes_comparativo = {}
                        aviso("Modo comparativo cancelado. Continuando com modo normal.")
                        break
                    
                    if modo_escolha == "1":
                        # MODO AUTOMÁTICO
                        if not pares_mecanizaveis:
                            aviso("Nao ha sugestoes automaticas para esta fazenda. Use [2] ou [3].")
                            continue
                        sub()
                        print(G + " Atividades detectadas automaticamente:" + RS)
                        print()
                        
                        # Mostrar lista numerada
                        for i, (manual, mec) in enumerate(pares_mecanizaveis, 1):
                            print(f" {Y}{i:2}{RS}. {manual}")
                            print(f" {DM}→{RS} {C}{mec}{RS}")
                            print()
                        
                        sub()
                        print(DM + "Digite os números das atividades para trocar (ex: 1,3,5)" + RS)
                        print(DM + "ou ENTER para TODAS, ou 0 para VOLTAR ao menu anterior" + RS)
                        escolha = input(DM + " >> Escolha: " + RS).strip()
                        
                        # Verificar se quer voltar
                        if escolha == "0":
                            continue  # Volta ao início do while loop
                        
                        indices_trocar = []
                        if escolha:
                            try:
                                # Parse lista de números
                                for parte in escolha.split(","):
                                    idx = int(parte.strip()) - 1
                                    if 0 <= idx < len(pares_mecanizaveis):
                                        indices_trocar.append(idx)
                            except ValueError:
                                aviso("Entrada inválida. Usando TODAS as atividades.")
                                indices_trocar = list(range(len(pares_mecanizaveis)))
                        else:
                            # ENTER = todas
                            indices_trocar = list(range(len(pares_mecanizaveis)))
                        
                        # Construir dicionário de substituições
                        for idx in indices_trocar:
                            manual, mec = pares_mecanizaveis[idx]
                            substituicoes_comparativo[manual] = mec
                        
                        if substituicoes_comparativo:
                            ok(f"Selecionadas {len(substituicoes_comparativo)} substituição(ões) para comparativo.")
                            break  # Sai do loop - seleção concluída
                        else:
                            aviso("Nenhuma atividade selecionada. Voltando ao menu.")
                            continue
                    if modo_escolha == "2":
                        # MODO MANUAL: Mostrar todas as atividades mecanizadas disponíveis
                        modo_manual_ativo = True
                        historico_substituicoes_manual = []
                        while modo_manual_ativo:
                            sub()
                            print(C + BL + " CATÁLOGO DE ATIVIDADES MECANIZADAS DISPONÍVEIS" + RS)
                            print(DM + " (todas as atividades com rendimento HM > 0)" + RS)
                            sub()

                            atividades_mecanizadas = []
                            for atv_nome, dados in CT317_HARDCODE_HH_BASE.items():
                                if dados.get("rendimento_hm", 0) > 0:
                                    atividades_mecanizadas.append((atv_nome, dados.get("rendimento_hm", 0)))

                            atividades_mecanizadas.sort(key=lambda x: x[0])

                            for i, (atv_nome, hm_val) in enumerate(atividades_mecanizadas, 1):
                                print(f" {Y}{i:2}{RS}. {atv_nome[:55]:<55} {C}(HM={hm_val:.2f}){RS}")

                            sub()
                            print(G + f"Total: {len(atividades_mecanizadas)} atividades mecanizadas" + RS)
                            print()
                            print(DM + "Escolha uma atividade mecanizada (ou 0 para voltar):" + RS)
                            print(DM + "Comandos: [L] listar substituições atuais | [U] desfazer última | [A] ver sugestões automáticas" + RS)

                            escolha_mec = input(DM + ">> Número (0 para voltar ao menu): " + RS).strip()
                            cmd_mec = escolha_mec.upper()
                            if cmd_mec == "0":
                                break

                            if cmd_mec == "L":
                                if substituicoes_comparativo:
                                    sub()
                                    print(G + BL + " SUBSTITUIÇÕES ATUAIS" + RS)
                                    for i, (manual, mec) in enumerate(substituicoes_comparativo.items(), 1):
                                        print(f" {Y}{i:2}{RS}. {manual}")
                                        print(f"    {DM}→{RS} {C}{_formatar_substituicao_comparativo(mec)}{RS}")
                                else:
                                    aviso("Nenhuma substituição selecionada até agora.")
                                continue

                            if cmd_mec == "U":
                                if historico_substituicoes_manual:
                                    atividade_desfazer, valor_anterior = historico_substituicoes_manual.pop()
                                    if valor_anterior is None:
                                        valor_removido = substituicoes_comparativo.pop(atividade_desfazer, None)
                                        if valor_removido is not None:
                                            ok(
                                                "Desfeito: "
                                                + f"{atividade_desfazer[:40]}..."
                                                + " removido da lista de substituições."
                                            )
                                        else:
                                            aviso("Nada para desfazer neste item.")
                                    else:
                                        substituicoes_comparativo[atividade_desfazer] = valor_anterior
                                        ok(
                                            "Desfeito: "
                                            + f"{atividade_desfazer[:40]}..."
                                            + " restaurado para a seleção anterior."
                                        )
                                else:
                                    aviso("Nao ha substituições recentes para desfazer.")
                                continue

                            if cmd_mec == "A":
                                sub()
                                print(C + BL + " SUGESTÕES AUTOMÁTICAS (MANUAL -> MECANIZADA)" + RS)
                                if pares_mecanizaveis:
                                    print()
                                    for i, (manual, mec) in enumerate(pares_mecanizaveis, 1):
                                        print(f" {Y}{i:2}{RS}. {manual}")
                                        print(f"    {DM}→{RS} {C}{_formatar_substituicao_comparativo(mec)}{RS}")
                                else:
                                    aviso("Nao ha sugestões automáticas para esta fazenda.")
                                sub()
                                input(DM + " [ENTER para voltar ao catálogo manual] " + RS)
                                continue

                            if not escolha_mec:
                                continue

                            try:
                                idx_mec = int(escolha_mec) - 1
                                if 0 <= idx_mec < len(atividades_mecanizadas):
                                    atividade_mecanizada_escolhida = atividades_mecanizadas[idx_mec][0]

                                    sub()
                                    print(C + " Atividades MANUAIS na fazenda:" + RS)
                                    for i, atv_manual in enumerate(atividades_reais, 1):
                                        print(f" {Y}{i:2}{RS}. {atv_manual}")

                                    print()
                                    print(DM + f"Qual atividade substituir por '{atividade_mecanizada_escolhida}'?" + RS)
                                    escolha_manual = input(DM + ">> Número da atividade manual (0 para cancelar): " + RS).strip()

                                    if escolha_manual == "0":
                                        continue

                                    try:
                                        idx_manual = int(escolha_manual) - 1
                                        if 0 <= idx_manual < len(atividades_reais):
                                            atividade_manual_escolhida = atividades_reais[idx_manual]
                                            valor_anterior = substituicoes_comparativo.get(atividade_manual_escolhida)
                                            if valor_anterior is not None:
                                                print()
                                                print(
                                                    Y
                                                    + "  Esta atividade já possui substituição:"
                                                    + RS
                                                )
                                                print(
                                                    DM
                                                    + f"  {atividade_manual_escolhida}"
                                                    + RS
                                                )
                                                print(
                                                    DM
                                                    + f"  atual: {_formatar_substituicao_comparativo(valor_anterior)}"
                                                    + RS
                                                )
                                                print(
                                                    DM
                                                    + f"  nova : {_formatar_substituicao_comparativo(atividade_mecanizada_escolhida)}"
                                                    + RS
                                                )
                                                if not confirmar("Substituir mapeamento existente?", default=True):
                                                    continue

                                            historico_substituicoes_manual.append(
                                                (atividade_manual_escolhida, valor_anterior)
                                            )
                                            substituicoes_comparativo[atividade_manual_escolhida] = atividade_mecanizada_escolhida
                                            ok(f"Adicionado: {atividade_manual_escolhida[:40]}... → {atividade_mecanizada_escolhida[:40]}...")
                                        else:
                                            aviso("Número inválido.")
                                    except ValueError:
                                        aviso("Entrada inválida.")

                                    print()
                                    if not confirmar("Adicionar outra substituição manual?", default=False):
                                        modo_manual_ativo = False
                                else:
                                    aviso("Número inválido.")
                            except ValueError:
                                aviso("Entrada inválida.")

                        if substituicoes_comparativo:
                            ok(f"Selecionadas {len(substituicoes_comparativo)} substituição(ões) para comparativo.")
                            sub()
                            print(G + " Resumo das substituições:" + RS)
                            for manual, mec in substituicoes_comparativo.items():
                                print(f" • {manual}")
                                print(f" → {C}{_formatar_substituicao_comparativo(mec)}{RS}")
                            sub()
                            input(DM + " [ENTER para continuar] " + RS)
                            break

                        aviso("Nenhuma substituição selecionada. Voltando ao menu.")
                        continue

                    if modo_escolha == "3":
                        # MODO EXTERNO: cadastrar recurso mecanizado fora do CT e ligar em uma atividade manual.
                        modo_externo_ativo = True
                        while modo_externo_ativo:
                            sub()
                            print(C + BL + " CADASTRAR RECURSO MECANIZADO EXTERNO" + RS)
                            print(DM + " Ex.: Navu, trator alugado, drone, serviço terceirizado." + RS)
                            sub()
                            idx_manual = selecionar_paginado("ATIVIDADE MANUAL A SUBSTITUIR", atividades_reais)
                            if idx_manual < 0:
                                break
                            atividade_manual_escolhida = atividades_reais[idx_manual]
                            recurso_custom = _cadastrar_recurso_mecanizado_externo(
                                atividade_manual_escolhida
                            )
                            if recurso_custom:
                                substituicoes_comparativo[atividade_manual_escolhida] = recurso_custom
                                ok(
                                    f"Adicionado: {atividade_manual_escolhida[:40]}... → {recurso_custom['atividade_mecanizada'][:40]}..."
                                )
                            if not confirmar("Adicionar outro recurso externo?", default=False):
                                modo_externo_ativo = False

                        if substituicoes_comparativo:
                            ok(f"Selecionadas {len(substituicoes_comparativo)} substituição(ões) para comparativo.")
                            sub()
                            print(G + " Resumo das substituições:" + RS)
                            for manual, mec in substituicoes_comparativo.items():
                                print(f" • {manual[:50]}...")
                                print(f" → {C}{_formatar_substituicao_comparativo(mec)}{RS}")
                            sub()
                            input(DM + " [ENTER para continuar] " + RS)
                            break

                        aviso("Nenhum recurso externo foi vinculado. Voltando ao menu.")
                        continue

                    aviso("Opcao invalida. Use 1, 2, 3 ou 0.")
                    continue
    _merge_sequencia_defaults(seq_cfg)
    cfg["sequencia"] = seq_cfg

    if _batch:
        modo_seq = ctx["modo_seq"]
    else:
        modo_seq = _selecionar_sequencia_padrao_sn(cfg, seq_cfg)

    modo_ctx = f"seq:{modo_seq}"
    modo_existente = contexto_sessao.modo_atual
    if modo_existente:
        if modo_ctx not in str(modo_existente):
            contexto_sessao.atualizar_modo(f"{modo_existente} | {modo_ctx}")
    else:
        contexto_sessao.atualizar_modo(modo_ctx)

    if modo_seq == "manutencao_seco":
        sequencia_manutencao_seco_placeholder(cfg)
    elif modo_seq == "manutencao_umido":
        sequencia_manutencao_umido_placeholder(cfg)
    usar_cascata = modo_seq in ("implantacao", "manutencao_swg", "personalizado")
    diagnosticar_sequencia_atividades(atividades_reais, seq_cfg, modo_seq)

    if _batch:
        usar_bloqueio_global = ctx.get("usar_bloqueio_global", False)
        atividades_bloqueadas = set()
        if usar_bloqueio_global:
            filtros_bloqueio = cfg.get("filtros_bloqueio_global", ["plantio", "irrig"])
            atividades_bloqueadas = set(
                atividades_por_filtro(atividades_reais, filtros_bloqueio)
            )
        usar_reforco_automatico = ctx.get("usar_reforco_automatico", True)
        usar_pool_pos_bloqueio = ctx.get("usar_pool_pos_bloqueio", False)
    else:
        filtros_bloqueio = cfg.get("filtros_bloqueio_global", ["plantio", "irrig"])
        candidatas_bloqueio = atividades_por_filtro(atividades_reais, filtros_bloqueio)
        usar_bloqueio_global = False
        atividades_bloqueadas = set()
        if modo_seq == "personalizado":
            print(
                DM
                + "  Modo PERSONALIZADO: bloqueio global plantio/irrigacao DESLIGADO."
                + RS
            )
        elif candidatas_bloqueio:
            usar_bloqueio_global = confirmar(
                "Aplicar BLOQUEIO GLOBAL (plantio/irrigacao so iniciam quando TODO o resto zerar na fazenda)?",
                default=True,
            )
            if usar_bloqueio_global:
                atividades_bloqueadas = set(candidatas_bloqueio)
                print(
                    Y
                    + f"\n  BLOQUEADAS ATE LIBERACAO GLOBAL ({len(atividades_bloqueadas)}):"
                    + RS
                )
                for a in sorted(atividades_bloqueadas, key=lambda x: str(x))[:20]:
                    print(Y + f"    - {str(a)[:58]}" + RS)
                if len(atividades_bloqueadas) > 20:
                    print(DM + f"    ... +{len(atividades_bloqueadas) - 20}" + RS)
                if confirmar(
                    "Salvar estes filtros de bloqueio no config para proximas execucoes?",
                    default=True,
                ):
                    cfg["filtros_bloqueio_global"] = filtros_bloqueio
                    salvar_config(cfg)
        usar_reforco_automatico = confirmar(
            "Ativar REFORCO AUTOMATICO (turma ociosa ajuda outras atividades nao bloqueadas)?",
            default=True,
        )
        usar_pool_pos_bloqueio = False
        if usar_bloqueio_global:
            usar_pool_pos_bloqueio = confirmar(
                "Usar PELOTAO UNIFICADO (todos os executores) so em plantio/irrigacao apos liberacao global?",
                default=True,
            )

    if _batch:
        prazo_meses = ctx["prazo_meses"]
        mes_ref = ctx["mes_ref"]
        ano_ref = ctx["ano_ref"]
        data_inicio_txt = ctx.get("data_inicio_txt")
        data_fim_txt = ctx.get("data_fim_txt")
        if data_inicio_txt or data_fim_txt:
            contexto_sessao.definir_datas(data_inicio_txt, data_fim_txt)
            # Não chamar dashboard_header() aqui para evitar flickering
        jornada = ctx["jornada"]
        executores = ctx["executores"]
        comparativo_cfg = ctx.get("comparativo_cfg")
        turmas = []
        for t in ctx["turmas"]:
            turmas.append(
                {
                    "nome": t["nome"],
                    "operarios": t["operarios"],
                    "atividades": [
                        _norm_atv(a)
                        for a in (t.get("atividades") or [])
                        if _norm_atv(a)
                    ],
                }
            )
    else:
        # ── Config equipe ──
        print(G + BL + "\n  CONFIGURACAO DO PROJETO" + RS + "\n")

        prazo_meses = pedir_float("Prazo META para conclusao (meses)", 6.0)
        hoje = datetime.datetime.now()
        print(
            DM
            + "  Referencia do calendario para DIAS UTEIS da meta (meses corridos a partir de): "
            + RS
        )
        mes_ref = pedir_int("Mes inicial (1-12)", hoje.month)
        mes_ref = max(1, min(12, int(mes_ref)))
        ano_ref = pedir_int("Ano inicial", hoje.year)
        dia_max = calendar.monthrange(ano_ref, mes_ref)[1]
        dia_ref = pedir_int(f"Dia inicial (1-{dia_max})", min(hoje.day, dia_max))
        dia_ref = max(1, min(dia_max, int(dia_ref)))

        data_inicio_txt = _formatar_data_dia(dia_ref, mes_ref, ano_ref)
        data_fim_txt = None
        if confirmar("Informar dia final manualmente?", default=False):
            mes_fim = pedir_int("Mes final (1-12)", mes_ref)
            mes_fim = max(1, min(12, int(mes_fim)))
            ano_fim = pedir_int("Ano final", ano_ref)
            dia_max_fim = calendar.monthrange(ano_fim, mes_fim)[1]
            dia_fim = pedir_int(
                f"Dia final (1-{dia_max_fim})", min(dia_ref, dia_max_fim)
            )
            dia_fim = max(1, min(dia_max_fim, int(dia_fim)))
            data_fim_txt = _formatar_data_dia(dia_fim, mes_fim, ano_fim)
        else:
            fim_calc = _calcular_data_fim_por_meses(
                dia_ref, mes_ref, ano_ref, prazo_meses
            )
            if fim_calc:
                data_fim_txt = _formatar_data_dia(fim_calc[0], fim_calc[1], fim_calc[2])

        contexto_sessao.definir_datas(data_inicio_txt, data_fim_txt)
        # Não chamar dashboard_header() aqui para evitar flickering

        j_def = float(cfg.get("jornada_horas") or 4.6)
        if j_def <= 0:
            j_def = 4.6
        executores = pedir_int(
            "Operarios totais (quem realmente trabalha)",
            9,
        )
        jornada = pedir_jornada(
            "Jornada efetiva diaria (ex: 6.5 ou 6:30 = 6h30)", round(j_def, 2)
        )
        cfg["jornada_horas"] = jornada
        salvar_config(cfg)

        if executores <= 0:
            erro("Precisa de pelo menos 1 executor.")
            return
        print(
            G + f"\n Equipe Operacional: {executores} operarios @ {jornada}h/dia" + RS
        )
        if confirmar(
            "Configurar COMPARATIVO MULTI-FATOR agora (para exportar no Excel)?",
            default=False,
        ):
            comparativo_cfg = coletar_config_comparativo_multifator(executores, jornada)

        # ──────────────────────────────────────────
        #  ETAPA 1: CRIAR TURMAS
        # ──────────────────────────────────────────
        sub()
        print(G + BL + "  ETAPA 1: CRIAR TURMAS / FUNCOES" + RS)
        print(
            DM + "  Defina grupos de trabalho (ex: Rocadores, Adubadores, Geral)." + RS
        )
        print(
            DM + "  Depois voce vinculara quais atividades cada turma executa.\n" + RS
        )

        turmas = []
        restantes = executores

        while restantes > 0:
            print(G + f"  Operarios disponiveis: {restantes}" + RS)
            nome_turma = prompt(
                "Nome da turma (ex: Rocadores)", f"Turma {len(turmas) + 1}"
            )
            def_pad = min(restantes, max(1, restantes // 2 or restantes))
            qtd = pedir_int(f"  Quantos operarios na turma '{nome_turma}'", def_pad)
            if qtd > restantes:
                aviso(f"Maximo disponivel: {restantes}. Ajustando.")
                qtd = restantes
            turmas.append({"nome": nome_turma, "operarios": qtd, "atividades": []})
            restantes -= qtd
            if restantes > 0:
                if not confirmar(
                    f"Criar outra turma? ({restantes} restantes)", default=True
                ):
                    turmas.append(
                        {"nome": "Geral", "operarios": restantes, "atividades": []}
                    )
                    restantes = 0

        sub()
        print(G + BL + "  TURMAS CRIADAS:" + RS)
        for t in turmas:
            print(G + f"  - {t['nome']}: " + C + f"{t['operarios']} operarios" + RS)
        sub()

        # ──────────────────────────────────────────
        #  ETAPA 2: VINCULAR ATIVIDADES AS TURMAS
        # ──────────────────────────────────────────
        print(G + BL + "\n  ETAPA 2: VINCULAR ATIVIDADES AS TURMAS" + RS)
        print(
            DM
            + "  Use FILTRO por texto para ligar varias de uma vez (ex: todas com 'roçada')."
            + RS
        )
        print(
            DM
            + "  Depois: conflitos (paralelo vs uma turma) e opcao de REATRIBUIR a outra turma.\n"
            + RS
        )

    atividade_remap = {}
    for manual, destino in (cfg.get("de_para", {}) or {}).items():
        manual_n = _norm_atv(manual)
        destino_n = _norm_atv(destino)
        if manual_n and destino_n:
            atividade_remap[manual_n] = destino_n
    if _batch and isinstance(ctx.get("substituicoes_template"), dict):
        for manual, destino in (ctx.get("substituicoes_template") or {}).items():
            manual_n = _norm_atv(manual)
            if isinstance(destino, dict):
                destino_nome = str(
                    destino.get("atividade_mecanizada")
                    or destino.get("nome")
                    or destino.get("recurso")
                    or ""
                ).strip()
            else:
                destino_nome = str(destino).strip()
            destino_n = _norm_atv(destino_nome)
            if manual_n and destino_n:
                atividade_remap[manual_n] = destino_n

    atividades_reais_set = set(atividades_reais)

    for turma in turmas:
        if not _batch:
            menu_vincular_atividades_turma(
                turma,
                atividades_reais,
                atividades_catalogo=catalogo_global,
            )
        else:
            existing = {
                _norm_atv(a) for a in (turma.get("atividades") or []) if _norm_atv(a)
            }
            remapeadas = set(existing)
            for atv in list(existing):
                alvo = atividade_remap.get(atv)
                if alvo:
                    remapeadas.add(alvo)
            matched = remapeadas & atividades_reais_set
            turma["atividades"] = sorted(matched, key=str)

    def _cobertura_atual_turmas():
        s = set()
        for t in turmas:
            for a in t.get("atividades") or []:
                na = _norm_atv(a)
                if na:
                    s.add(na)
        return s

    cob_pre = _cobertura_atual_turmas()
    orfas_pre = [a for a in atividades_reais if a not in cob_pre]
    preencher_orfas = False
    if orfas_pre:
        if _batch:
            preencher_orfas = bool(ctx.get("preencher_orfas_template", False))
        elif modo_seq == "personalizado":
            preencher_orfas = confirmar(
                "Esta fazenda tem demandas sem turma no modelo. Preencher na turma com mais operarios? "
                "(N = equipe especializada; HH dessas atividades nao entram no cronograma)",
                default=False,
            )
    if preencher_orfas:
        _distribuir_atividades_faltantes_turmas(turmas, atividades_reais, fazenda)

    for turma in turmas:
        if not turma["atividades"]:
            aviso(f"Turma '{turma['nome']}' ficou sem atividades!")

    def coletar_vinculadas():
        s = set()
        for t in turmas:
            s.update(t["atividades"])
        return s

    atividades_vinculadas = coletar_vinculadas()
    orfas = [a for a in atividades_reais if a not in atividades_vinculadas]
    if orfas:
        print(Y + f"\n  ATENCAO: {len(orfas)} atividades sem turma vinculada:" + RS)
        for o in orfas:
            print(Y + f"    - {str(o)[:55]}" + RS)
        vincular_orfas = False
        turma_alvo = None
        if _batch:
            vincular_orfas = bool(ctx.get("preencher_orfas_template", False))
            if vincular_orfas and turmas:
                turma_alvo = max(
                    turmas, key=lambda t: int(t.get("operarios", 0) or 0)
                ).get("nome")
        else:
            vincular_orfas = confirmar(
                "Vincular todas as orfas a uma turma existente?", default=True
            )
            if vincular_orfas:
                nomes = [t["nome"] for t in turmas]
                turma_alvo = selecionar("TURMA PARA ORFAS", nomes)

        if vincular_orfas and turma_alvo:
            for t in turmas:
                if t["nome"] == turma_alvo:
                    t["atividades"] = sorted(
                        set(t["atividades"]) | set(orfas), key=lambda x: str(x)
                    )
                    ok(f"{len(orfas)} atividades vinculadas a '{turma_alvo}'.")
                    break
        elif _batch and vincular_orfas:
            aviso("Modo batch: sem turma destino valida para atividades orfas.")

    atividades_vinculadas = coletar_vinculadas()
    contexto_sessao.atualizar_atividades(
        len(atividades_vinculadas), len(atividades_reais)
    )
    # Não chamar dashboard_header() aqui para evitar flickering

    # ──────────────────────────────────────────
    #  ETAPA 3: Conflitos (paralelo / exclusivo) + reatribuicao opcional
    # ──────────────────────────────────────────
    print(G + BL + "\n  ETAPA 3: CONFLITOS E REATRIBUICAO" + RS)
    if _batch:
        reatribuicao_tpl = dict((ctx.get("reatribuicao_template") if ctx else {}) or {})
        paralelo_tpl = dict((ctx.get("paralelo_template") if ctx else {}) or {})
        primaria_tpl = dict((ctx.get("primaria_template") if ctx else {}) or {})

        reatribuicao = {}
        paralelo = {}
        primaria = {}
        for atv, turma_nome in reatribuicao_tpl.items():
            atv_n = _norm_atv(atv)
            atv_n = atividade_remap.get(atv_n, atv_n)
            if atv_n in atividades_reais_set and turma_nome:
                reatribuicao[atv_n] = turma_nome
        for atv, em_paralelo in paralelo_tpl.items():
            atv_n = _norm_atv(atv)
            atv_n = atividade_remap.get(atv_n, atv_n)
            if atv_n in atividades_reais_set:
                paralelo[atv_n] = bool(em_paralelo)
        for atv, turma_nome in primaria_tpl.items():
            atv_n = _norm_atv(atv)
            atv_n = atividade_remap.get(atv_n, atv_n)
            if atv_n in atividades_reais_set and turma_nome:
                primaria[atv_n] = turma_nome
    else:
        reatribuicao, paralelo, primaria = resolver_conflitos_e_reatribuir(
            turmas, atividades_reais
        )

    session_hh = {}
    if ctx and isinstance(ctx.get("session_hh"), dict):
        session_hh.update(ctx["session_hh"])

    def _recalcular_apos_ajuste_escopo():
        nonlocal df_faz, atividades_reais, talhoes_ordenados, catalogo_global
        atividades_reais = sorted(
            {
                a
                for a in df_faz["atividade"].dropna().unique().tolist()
                if _norm_atv(a)
            },
            key=str,
        )
        talhoes_ordenados = sorted(df_faz["chave"].dropna().unique().tolist())
        catalogo_global = _catalogo_atividades_completo(
            atividades_reais,
            cfg=cfg,
            atividades_catalogo=atividades_catalogo,
        )
        for t in turmas:
            cur = [a for a in (t.get("atividades") or []) if a in catalogo_global]
            t["atividades"] = sorted(set(cur), key=str)

    if not _batch:
        if confirmar(
            "Ajustar HH/ha por atividade APENAS nesta execucao (nao grava config)?",
            default=False,
        ):
            menu_ajustes_hh_apenas_sessao(atividades_reais, cfg, session_hh)

    while True:
        sub()
        print(G + BL + " CHECKPOINT RETROATIVO" + RS)
        op_cp = selecionar(
            "O QUE DESEJA REVISAR?",
            [
                "Editar atividades de uma turma",
                "Reprocessar conflitos/reatribuicao",
                "Ajustar HH/ha desta sessao",
                "Ajustar escopo de atividades desta execucao",
                "Revisar jornada/equipe",
                "Voltar ao seletor de fazenda/escopo",
                "Continuar para simulacao",
            ],
        )

        if not op_cp or op_cp == "Continuar para simulacao":
            break

        if op_cp == "Voltar ao seletor de fazenda/escopo":
            return {"acao": "retroceder_escopo"}

        if op_cp == "Revisar jornada/equipe":
            print(DM + f"\n  Atual: {executores} operarios @ {jornada}h/dia" + RS)
            if confirmar("Alterar jornada?", default=False):
                jornada = pedir_jornada(
                    "Nova jornada (ex: 6.5 ou 6:30 = 6h30)", round(jornada, 2)
                )
                cfg["jornada_horas"] = jornada
                salvar_config(cfg)
                ok(f"Jornada atualizada: {jornada}h/dia")
            if confirmar("Alterar operarios?", default=False):
                executores = pedir_int("Operarios totais", executores)
            print(
                G + f" Equipe: {executores} operarios @ {jornada}h/dia = {executores * jornada:.1f} HH/dia" + RS
            )
            continue

            if op_cp == "Editar atividades de uma turma":
                nomes_t = [t["nome"] for t in turmas]
                nm = selecionar("TURMA", nomes_t)
                if nm:
                    for t in turmas:
                        if t["nome"] == nm:
                            menu_vincular_atividades_turma(
                                t,
                                atividades_reais,
                                atividades_catalogo=catalogo_global,
                            )
                            break
                continue

            if op_cp == "Reprocessar conflitos/reatribuicao":
                reatribuicao, paralelo, primaria = resolver_conflitos_e_reatribuir(
                    turmas, atividades_reais
                )
                continue

            if op_cp == "Ajustar HH/ha desta sessao":
                menu_ajustes_hh_apenas_sessao(atividades_reais, cfg, session_hh)
                continue

            if op_cp == "Ajustar escopo de atividades desta execucao":
                df_faz = _menu_ajustar_escopo_atividades(
                    df_faz,
                    cfg=cfg,
                    atividades_catalogo=catalogo_global,
                )
                _recalcular_apos_ajuste_escopo()
                reatribuicao, paralelo, primaria = resolver_conflitos_e_reatribuir(
                    turmas, atividades_reais
                )
                continue

    # ── Validacao orcamento estrito (antes das demandas) ──
    if not validar_e_completar_orcamento(cfg, atividades_reais, session_hh=session_hh):
        if not _batch:
            input(DM + "\n  [ENTER para voltar] " + RS)
            return
        aviso("Modo batch: validacao de orcamento falhou; cenario cancelado.")
        return {"acao": "orcamento_invalido"}

    tarifas = cfg.get("tarifas", {})
    de_para = cfg.get("de_para", {})
    strict = cfg.get("orcamento_estrito", True)

    # ── Construir demandas por talhao ──
    demandas = OrderedDict()  # {talhao: [{atividade, area, hh_total}, ...]}
    total_hh = 0.0
    total_hm = 0.0
    hm_only_atividades = set()
    fallback_hh_items = []

    for talhao in talhoes_ordenados:
        df_t = df_faz[df_faz["chave"] == talhao]
        tarefas = []
        for _, row in df_t.iterrows():
            atv = row["atividade"]
            area = float(row["area_ha"])
            pen = float(row["penalidade"])

            t_nome = resolver_chave_tarifa(cfg, tarifas, atv)
            rend_base = resolver_rendimento_hh(
                cfg,
                tarifas,
                t_nome,
                strict=strict,
                session_hh=session_hh,
                atv_micro=atv,
            )
            hm_base = resolver_rendimento_hm(cfg, tarifas, t_nome, strict=strict)
            if rend_base is None:
                if float(hm_base or 0) > 0.0:
                    # Atividade mecanizada HM-only: HH pode ficar zerado sem abortar o cronograma.
                    rend_base = 0.0
                    fallback_hh_items.append((str(atv), str(t_nome), "HM-only => HH=0"))
                else:
                    rend_fb = resolver_rendimento_hh(
                        cfg,
                        tarifas,
                        t_nome,
                        strict=False,
                        session_hh=session_hh,
                        atv_micro=atv,
                    )
                    if rend_fb is None:
                        rend_fb = 0.0
                    rend_base = float(rend_fb)
                    fallback_hh_items.append(
                        (str(atv), str(t_nome), f"fallback HH={float(rend_base):.2f}")
                    )
            rend_hh_ha = float(rend_base) * pen
            hm_ha = float(hm_base) * pen
            in_tarifa = t_nome in tarifas

            horas = area * rend_hh_ha
            hm_horas = area * hm_ha
            total_hh += horas
            total_hm += hm_horas

            tarifa_row = tarifas.get(t_nome, {})
            tipo_tarifa = str(tarifa_row.get("tipo", "")).lower()
            is_mec = "mecaniz" in tipo_tarifa or (hm_ha > 0 and rend_hh_ha <= 0)
            if is_mec and rend_hh_ha <= 0 and hm_ha > 0:
                hm_only_atividades.add(str(atv))

            if strict:
                origem_linha = "CT"
                rfonte = "CT"
            else:
                origem_linha = "tarifa" if in_tarifa else "fallback"
                rfonte = "CT" if in_tarifa else "estimado"
            tarefas.append(
                {
                    "atividade": atv,
                    "area": area,
                    "hh_total": horas,
                    "hm_ha": hm_ha,
                    "hm_total": hm_horas,
                    "chave_tarifa": t_nome,
                    "origem": origem_linha,
                    "rendimento_fonte": rfonte,
                    "tipo": "Mecanizada" if is_mec else "Manual",
                }
            )
        demandas[talhao] = tarefas

    print(DM + f"\n  Total HH da fazenda (bruto): {total_hh:.1f} horas-homem" + RS)
    print(DM + f"  Total HM da fazenda (bruto): {total_hm:.1f} horas-maquina" + RS)
    if total_hm > 0.01:
        print(
            DM
            + "  Regra de fluxo HM-only: atividades mecanizadas rodam em paralelo e a equipe humana"
            + " continua o fluxo sem esperar 100% da maquina."
            + RS
        )
    if fallback_hh_items:
        print(
            Y
            + f"  Fallback HH aplicado em {len(fallback_hh_items)} item(ns) do escopo."
            + RS
        )
        for atv_fb, t_fb, motivo_fb in fallback_hh_items[:5]:
            print(
                Y
                + f"    - {str(atv_fb)[:44]} | CT:{str(t_fb)[:30]} | {str(motivo_fb)[:24]}"
                + RS
            )
        if len(fallback_hh_items) > 5:
            print(Y + f"    ... +{len(fallback_hh_items) - 5}" + RS)

    auditar_cadeia_dados(cfg, demandas, atividades_reais, session_hh=session_hh)

    sem_tarifa = []
    for talhao, tarefas in demandas.items():
        for t in tarefas:
            atv = t["atividade"]
            t_nome = resolver_chave_tarifa(cfg, tarifas, atv)
            if t_nome not in tarifas:
                sem_tarifa.append((str(atv)[:50], str(t_nome)[:50]))
    if not strict and sem_tarifa:
        est_fb = resolver_rendimento_hh(
            cfg, tarifas, "!__chave_inexistente__!", strict=False
        )
        print(
            Y
            + "\n  !  Chave de tarifa NAO encontrada no orcamento importado (desencontro de nome)."
            + RS
        )
        print(
            Y
            + f"     Rendimento estimado aplicado: ~{est_fb:.2f} h/ha (mediana/config; ver doc)."
            + RS
        )
        visto = set()
        for a, tn in sem_tarifa:
            key = (a, tn)
            if key in visto:
                continue
            visto.add(key)
            print(Y + f"    micro: {a}  ->  chave buscada: {tn}" + RS)
        print(
            DM
            + "    Correcao: menu [4] de_para ou importe tarifas [2] — no orcamento o homem/ha existe."
            + RS
        )

    sub()
    print(C + BL + "  PRE-CHECAGEM HH/HM ANTES DO CRONOGRAMA" + RS)
    if modo_comparativo and substituicoes_comparativo:
        print(
            DM
            + "  [COMPARATIVO] Esta pre-checagem e o cronograma abaixo representam o CENARIO BASELINE (manual atual)."
            + RS
        )
    _mostrar_painel_hh_hm_pre_scheduler(demandas, fazenda, detalhado=False)
    if (not _batch) and confirmar("Exibir HH/HM detalhado por talhao?", default=False):
        _mostrar_painel_hh_hm_pre_scheduler(demandas, fazenda, detalhado=True)

    sem_executor = []
    for talhao, tarefas in demandas.items():
        for t in tarefas:
            if t["hh_total"] < 0.01:
                continue
            atv = t["atividade"]
            if not turmas_que_executam(atv, turmas, reatribuicao, paralelo, primaria):
                sem_executor.append(atv)
    if sem_executor:
        print(R + "\n  X  Atividades com demanda mas SEM turma executora:" + RS)
        for a in sorted(set(str(x) for x in sem_executor))[:15]:
            print(R + f"    - {a[:58]}" + RS)
        if len(set(sem_executor)) > 15:
            print(DM + f"    ... +{len(set(sem_executor)) - 15}" + RS)
        continuar_sem_executor = True
        if _batch:
            aviso("Modo batch: HH sem turma executora serao zeradas automaticamente.")
        else:
            continuar_sem_executor = confirmar(
                "  Continuar mesmo assim (essas HH nao serao agendadas)?", default=False
            )
        if not continuar_sem_executor:
            input(DM + "\n  [ENTER para voltar] " + RS)
            return
        for talhao, tarefas in demandas.items():
            for t in tarefas:
                atv = t["atividade"]
                if t["hh_total"] > 0.01 and not turmas_que_executam(
                    atv, turmas, reatribuicao, paralelo, primaria
                ):
                    t["hh_total"] = 0.0
        total_hh = sum(t["hh_total"] for tarefas in demandas.values() for t in tarefas)
        aviso("HH sem executora foram zeradas no cronograma.")
        print(DM + f"  Total HH agendavel: {total_hh:.1f} horas-homem" + RS)

    sub()
    print(G + BL + "  GERANDO CRONOGRAMA (talhao a talhao)..." + RS + "\n")

    # ──────────────────────────────────────────
    #  SCHEDULER: Filas por TURMA, sequenciais por talhao.
    #  Cada turma tem uma fila de trabalho construida a partir
    #  das atividades vinculadas, na ordem dos talhoes.
    #  Quando a turma termina no talhao 1, avanca pro 2.
    # ──────────────────────────────────────────

    # Build per-turma work queue: list of {talhao, atividade, hh_rest}
    turma_filas = {}
    for turma in turmas:
        fila = []
        for talhao in talhoes_ordenados:
            for tarefa in demandas.get(talhao, []):
                atv = tarefa["atividade"]
                if tarefa["hh_total"] > 0.01 and turma["nome"] in turmas_que_executam(
                    atv, turmas, reatribuicao, paralelo, primaria
                ):
                    fila.append(
                        {
                            "talhao": talhao,
                            "atividade": atv,
                            "hh_rest": tarefa["hh_total"],
                        }
                    )
        turma_filas[turma["nome"]] = fila

    # Uma entrada por (talhao, atividade): todas as turmas autorizadas
    # consomem o mesmo saldo (paralelo) ou so uma turma (exclusivo/reatribuido).
    demanda_global = {}  # key=(talhao,atividade) -> remaining hh
    for talhao, tarefas in demandas.items():
        for t in tarefas:
            demanda_global[(talhao, t["atividade"])] = t["hh_total"]

    atividades_plantio = set(
        atividades_por_filtro(
            atividades_reais, seq_cfg.get("filtros_plantio") or ["plantio"]
        )
    )
    atividades_irrig = set(
        atividades_por_filtro(
            atividades_reais, seq_cfg.get("filtros_irrigacao") or ["irrig"]
        )
    )
    tem_plantio_por_talhao = {}
    for th in talhoes_ordenados:
        tem_plantio_por_talhao[th] = any(
            t["atividade"] in atividades_plantio and t["hh_total"] > 0.01
            for t in demandas.get(th, [])
        )
    dia_termino_plantio = {}

    if usar_cascata:
        for _tn, fila in turma_filas.items():
            fila.sort(
                key=lambda x: (
                    classificar_fase_cascata_valor(
                        x["atividade"],
                        seq_cfg,
                        modo_seq,
                        atividades_plantio,
                        atividades_irrig,
                    ),
                    str(x["talhao"]),
                    str(x["atividade"]),
                )
            )

    cronograma = []
    dia = 0
    MAX_DIAS = 10000

    def _registrar_fim_plantio_talhao(th, dia_atual):
        if dia_termino_plantio.get(th) is not None:
            return
        if not _demanda_plantio_talhao(th, demanda_global, atividades_plantio):
            dia_termino_plantio[th] = dia_atual

    while dia < MAX_DIAS:
        tem_trabalho = any(v > 0.01 for v in demanda_global.values())
        if not tem_trabalho:
            break

        dia += 1
        pool_only = (
            usar_bloqueio_global
            and usar_pool_pos_bloqueio
            and _somente_bloqueado_restante(demanda_global, atividades_bloqueadas)
        )
        if pool_only:
            cap_pool = float(executores) * float(jornada)
            while cap_pool > 0.01:
                fez = False
                min_fase_dia = _min_fase_cascata(
                    demanda_global,
                    seq_cfg,
                    modo_seq,
                    usar_cascata,
                    usar_bloqueio_global,
                    atividades_bloqueadas,
                    atividades_plantio,
                    atividades_irrig,
                    dia,
                    dia_termino_plantio,
                    tem_plantio_por_talhao,
                )
                for talhao in talhoes_ordenados:
                    tlist = list(demandas.get(talhao, []))
                    tlist.sort(
                        key=lambda t: (
                            0
                            if t["atividade"] in atividades_plantio
                            else (1 if t["atividade"] in atividades_irrig else 2),
                            str(t["atividade"]),
                        )
                    )
                    for t in tlist:
                        atv = t["atividade"]
                        if atv not in atividades_bloqueadas:
                            continue
                        key = (talhao, atv)
                        rest = demanda_global.get(key, 0.0)
                        if rest <= 0.01:
                            continue
                        if not pode_agendar_atividade_cascata(
                            talhao,
                            atv,
                            demanda_global,
                            seq_cfg,
                            modo_seq,
                            usar_cascata,
                            usar_bloqueio_global,
                            atividades_bloqueadas,
                            atividades_plantio,
                            atividades_irrig,
                            dia,
                            dia_termino_plantio,
                            tem_plantio_por_talhao,
                            min_fase_dia,
                        ):
                            continue
                        consumo = min(rest, cap_pool)
                        demanda_global[key] -= consumo
                        cap_pool -= consumo
                        fez = True
                        _registrar_fim_plantio_talhao(talhao, dia)
                        cronograma.append(
                            {
                                "Dia": dia,
                                "Fazenda": fazenda,
                                "Talhao": talhao,
                                "Atividade": atv,
                                "Turma": "Pelotao_Unificado",
                                "Operarios": executores,
                                "HH": round(consumo, 2),
                                "Modo": "PoolPosBloqueio",
                            }
                        )
                        if cap_pool <= 0.01:
                            break
                    if cap_pool <= 0.01:
                        break
                if not fez:
                    break
            for turma in turmas:
                fila = turma_filas[turma["nome"]]
                while (
                    fila
                    and demanda_global.get((fila[0]["talhao"], fila[0]["atividade"]), 0)
                    < 0.01
                ):
                    fila.pop(0)
            continue

        for turma in turmas:
            fila = turma_filas[turma["nome"]]
            n_ops = turma["operarios"]
            cap_dia = n_ops * jornada

            # Process items in queue order
            idx = 0
            while cap_dia > 0.01 and idx < len(fila):
                min_fase_dia = _min_fase_cascata(
                    demanda_global,
                    seq_cfg,
                    modo_seq,
                    usar_cascata,
                    usar_bloqueio_global,
                    atividades_bloqueadas,
                    atividades_plantio,
                    atividades_irrig,
                    dia,
                    dia_termino_plantio,
                    tem_plantio_por_talhao,
                )
                item = fila[idx]
                key = (item["talhao"], item["atividade"])
                rest = demanda_global.get(key, 0)

                if rest < 0.01:
                    idx += 1  # Already done (by another turma perhaps)
                    continue

                if not pode_agendar_atividade_cascata(
                    item["talhao"],
                    item["atividade"],
                    demanda_global,
                    seq_cfg,
                    modo_seq,
                    usar_cascata,
                    usar_bloqueio_global,
                    atividades_bloqueadas,
                    atividades_plantio,
                    atividades_irrig,
                    dia,
                    dia_termino_plantio,
                    tem_plantio_por_talhao,
                    min_fase_dia,
                ):
                    idx += 1
                    continue

                consumo = min(rest, cap_dia)
                demanda_global[key] -= consumo
                cap_dia -= consumo
                _registrar_fim_plantio_talhao(item["talhao"], dia)

                cronograma.append(
                    {
                        "Dia": dia,
                        "Fazenda": fazenda,
                        "Talhao": item["talhao"],
                        "Atividade": item["atividade"],
                        "Turma": turma["nome"],
                        "Operarios": n_ops,
                        "HH": round(consumo, 2),
                    }
                )

                if demanda_global[key] < 0.01:
                    idx += 1  # Move to next item in queue
                # else stay on same item (partially done today)

            # Clean up completed items from front of queue
            while (
                fila
                and demanda_global.get((fila[0]["talhao"], fila[0]["atividade"]), 0)
                < 0.01
            ):
                fila.pop(0)

            # Mutirao/realloc automatico:
            # se ainda sobrou capacidade no dia, ajuda demanda de outras atividades nao bloqueadas.
            if usar_reforco_automatico and cap_dia > 0.01:
                for talhao in talhoes_ordenados:
                    if cap_dia <= 0.01:
                        break
                    tarefas_t = list(demandas.get(talhao, []))
                    if usar_cascata:
                        tarefas_t.sort(
                            key=lambda t: (
                                classificar_fase_cascata_valor(
                                    t["atividade"],
                                    seq_cfg,
                                    modo_seq,
                                    atividades_plantio,
                                    atividades_irrig,
                                ),
                                str(t["atividade"]),
                            )
                        )
                    for t in tarefas_t:
                        min_fase_dia = _min_fase_cascata(
                            demanda_global,
                            seq_cfg,
                            modo_seq,
                            usar_cascata,
                            usar_bloqueio_global,
                            atividades_bloqueadas,
                            atividades_plantio,
                            atividades_irrig,
                            dia,
                            dia_termino_plantio,
                            tem_plantio_por_talhao,
                        )
                        atv = t["atividade"]
                        key_ref = (talhao, atv)
                        rest_ref = demanda_global.get(key_ref, 0.0)
                        if rest_ref <= 0.01:
                            continue
                        if not pode_agendar_atividade_cascata(
                            talhao,
                            atv,
                            demanda_global,
                            seq_cfg,
                            modo_seq,
                            usar_cascata,
                            usar_bloqueio_global,
                            atividades_bloqueadas,
                            atividades_plantio,
                            atividades_irrig,
                            dia,
                            dia_termino_plantio,
                            tem_plantio_por_talhao,
                            min_fase_dia,
                        ):
                            continue
                        consumo_ref = min(rest_ref, cap_dia)
                        if consumo_ref <= 0.01:
                            continue
                        demanda_global[key_ref] -= consumo_ref
                        cap_dia -= consumo_ref
                        _registrar_fim_plantio_talhao(talhao, dia)
                        cronograma.append(
                            {
                                "Dia": dia,
                                "Fazenda": fazenda,
                                "Talhao": talhao,
                                "Atividade": atv,
                                "Turma": turma["nome"],
                                "Operarios": n_ops,
                                "HH": round(consumo_ref, 2),
                                "Modo": "Reforco",
                            }
                        )

    # Baseline mecanizado: HM-only do orcamento entra automaticamente no cronograma base.
    hm_only_list = sorted(hm_only_atividades, key=str)
    cronograma_mec_base = []
    recursos_mec_base = []
    if hm_only_list:
        cronograma_mec_base, recursos_mec_base = construir_cronograma_mecanizado_auto_hm_tarifa(
            demandas,
            fazenda,
            jornada,
            cfg,
            tarifas,
            atividades_alvo=hm_only_list,
        )
        if cronograma_mec_base:
            ok(
                f"Cronograma base incluiu {len(cronograma_mec_base)} linha(s) mecanizadas (HM do orcamento)."
            )

    cronograma_base = sorted(
        cronograma + cronograma_mec_base,
        key=lambda r: (int(r.get("Dia", 0)), str(r.get("Turma", ""))),
    )

    dias_simulado_hum = dia
    d_mec_base = max([int(x.get("Dia", 0)) for x in cronograma_mec_base], default=0)
    dias_simulado = max(dias_simulado_hum, d_mec_base)

    # ── Diagnostico ──
    dias_meta = dias_uteis_no_periodo(mes_ref, ano_ref, prazo_meses)
    exec_teoricos = (
        math.ceil(total_hh / (dias_meta * jornada)) if (dias_meta * jornada) > 0 else 1
    )
    meses_simulado = dias_simulado / 22.0 if dias_simulado > 0 else 0

    # ── Tabela semanal ──
    table = Table(title=f"Cronograma - {fazenda} ({executores} Exec.)")
    table.add_column("Semana", justify="center", style="cyan")
    table.add_column("Dias", justify="center")
    table.add_column("Talhoes / Atividades", style="green")

    semanas = defaultdict(lambda: {"dias": set(), "acoes": set()})
    for c in cronograma_base:
        sem = math.ceil(c["Dia"] / 5)
        semanas[sem]["dias"].add(c["Dia"])
        semanas[sem]["acoes"].add(f"[{c['Talhao']}] {c['Atividade'][:18]}")

    for sem in sorted(semanas.keys())[:8]:
        d = semanas[sem]
        dias_str = f"Dia {min(d['dias'])} a {max(d['dias'])}"
        acoes = ", ".join(list(d["acoes"])[:3])
        if len(d["acoes"]) > 3:
            acoes += " (+)"
        table.add_row(f"Sem {sem}", dias_str, acoes)

    console.print(table)
    if len(semanas) > 8:
        print(DM + f"  ... e mais {len(semanas) - 8} semanas no Excel." + RS)

    # ── Metricas operacionais ──
    hh_por_turma = defaultdict(float)
    for c in cronograma:
        hh_por_turma[c["Turma"]] += float(c["HH"])

    hh_agendada_total = sum(hh_por_turma.values())
    n_demandas = sum(1 for tarefas in demandas.values() for t in tarefas)
    n_fb = sum(
        1
        for tarefas in demandas.values()
        for t in tarefas
        if t.get("origem") == "fallback"
    )
    pct_fallback = (100.0 * n_fb / n_demandas) if n_demandas > 0 else 0.0

    recursos_mec = []
    cronograma_mec = []
    cronograma_com_mec = []
    atividades_mec_set = set()
    default_ativar_mec = False
    if _batch:
        sub()
        print(
            DM
            + "  Modo batch: pulando 'modo mecanizado opcional' (sem prompts interativos)."
            + RS
        )
    elif modo_comparativo and substituicoes_comparativo:
        sub()
        print(
            DM
            + "  Comparativo MANUAL vs MECANIZADO ativo: pulando 'modo mecanizado opcional' para evitar duplicidade de cenarios."
            + RS
        )
    else:
        sub()
        print(C + BL + "  ATIVAR MODO MECANIZADO" + RS)
        print(
            DM
            + "  Cenario opcional: cadastrar recurso extra para adicionar/substituir atividades."
            + RS
        )
        if cronograma_mec_base:
            print(
                DM
                + "  As atividades HM do orcamento ja foram contabilizadas automaticamente no cronograma base."
                + RS
            )
            for a in hm_only_list[:5]:
                print(DM + f"    - {str(a)[:58]}" + RS)
            if len(hm_only_list) > 5:
                print(DM + f"    ... +{len(hm_only_list) - 5}" + RS)
        if hm_only_list:
            print(
                DM
                + f"  HM-only (HH=0) detectadas: {len(hm_only_list)} atividade(s)."
                + RS
            )
        if confirmar("  Ativar modo mecanizado opcional?", default=default_ativar_mec):
            recursos_mec = _cadastrar_recursos_mecanizados_sn(
                atividades_reais,
                cfg,
                atividades_catalogo=catalogo_global,
            )
            for rec in recursos_mec:
                atividades_mec_set.update(rec.get("atividades", set()))
            if recursos_mec and atividades_mec_set:
                cronograma_mec = construir_cronograma_mecanizado(
                    demandas, fazenda, jornada, recursos_mec
                )

            if cronograma_mec and atividades_mec_set:
                regra_implantacao_mec = "substituir_total"
                if confirmar(
                    "Regra de implantacao mecanizado: manter humano em PARALELO nas atividades mecanizadas?",
                    default=False,
                ):
                    regra_implantacao_mec = "paralelo"
                if regra_implantacao_mec == "paralelo":
                    crono_hum_sem_mec = [dict(x) for x in cronograma_base]
                else:
                    crono_hum_sem_mec_h = construir_cronograma_humano_sem_mecanizadas(
                        cronograma, turmas, jornada, executores, atividades_mec_set
                    )
                    crono_hum_sem_mec = sorted(
                        crono_hum_sem_mec_h + cronograma_mec_base,
                        key=lambda r: (int(r.get("Dia", 0)), str(r.get("Turma", ""))),
                    )
                cronograma_com_mec = sorted(
                    crono_hum_sem_mec + cronograma_mec,
                    key=lambda r: (int(r.get("Dia", 0)), str(r.get("Turma", ""))),
                )
                d_hum = max(
                    [int(x.get("Dia", 0)) for x in crono_hum_sem_mec], default=0
                )
                d_mec = max([int(x.get("Dia", 0)) for x in cronograma_mec], default=0)
                d_comb = max(d_hum, d_mec)
                t_mec = Table(title="Comparativo Operacional - Modo Mecanizado")
                t_mec.add_column("Metrica", style="cyan")
                t_mec.add_column("Valor", justify="right")
                t_mec.add_row("Dias baseline (cronograma base)", str(dias_simulado))
                t_mec.add_row("Dias base sem atividades opcionais", str(d_hum))
                t_mec.add_row("Dias recursos mecanizados (filas dedicadas)", str(d_mec))
                t_mec.add_row(
                    "Dias cenario combinado (humano || mecanizado)", str(d_comb)
                )
                t_mec.add_row(
                    "Ganho de prazo (dias)", f"{int(dias_simulado) - int(d_comb):+d}"
                )
                t_mec.add_row("Regra mecanizada", regra_implantacao_mec)
                for rec in recursos_mec:
                    t_mec.add_row(
                        f"  Recurso: {rec['nome']}",
                          f"{rec['prod_ha_h']} ha/h",
                    )
                    t_mec.add_row(
                        f"  Atividades ({rec['nome']})",
                        str(len(rec.get("atividades", set()))),
                    )
                hm_mec_total = sum(
                    float(x.get("HM", x.get("HH", 0)) or 0) for x in cronograma_mec
                )
                t_mec.add_row("Horas mecanizadas (HM)", f"{hm_mec_total:.1f}")
                console.print(t_mec)

                t_alt = Table(title="Cronograma Alternativo (Humano + Mecanizado)")
                t_alt.add_column("Semana", justify="center", style="cyan")
                t_alt.add_column("Dias", justify="center")
                t_alt.add_column("Acoes", style="green")
                sem_alt = defaultdict(lambda: {"dias": set(), "acoes": set()})
                for c in cronograma_com_mec:
                    s = (
                        int(math.ceil(float(c.get("Dia", 0)) / 5.0))
                        if c.get("Dia")
                        else 0
                    )
                    if s <= 0:
                        continue
                    sem_alt[s]["dias"].add(int(c["Dia"]))
                    txt = f"[{str(c.get('Talhao', ''))[:18]}] {str(c.get('Atividade', ''))[:18]} ({c.get('Turma', '')})"
                    sem_alt[s]["acoes"].add(txt)
                for s in sorted(sem_alt.keys())[:8]:
                    d = sem_alt[s]
                    dias_str = f"Dia {min(d['dias'])} a {max(d['dias'])}"
                    acoes = ", ".join(list(d["acoes"])[:3])
                    if len(d["acoes"]) > 3:
                        acoes += " (+)"
                    t_alt.add_row(f"Sem {s}", dias_str, acoes)
                console.print(t_alt)

    # ── Tabela ocupacao ──
    sub()
    print(G + BL + "  OCUPACAO POR TURMA" + RS)
    t_occ = Table()
    t_occ.add_column("Turma", style="cyan")
    t_occ.add_column("HH", justify="right")
    t_occ.add_column("Cap. max", justify="right")
    t_occ.add_column("Uso %", justify="right")
    crit_nm, crit_pct = "", 0.0
    for turma in turmas:
        nm = turma["nome"]
        cap = float(dias_simulado_hum) * float(turma["operarios"]) * float(jornada)
        us = hh_por_turma.get(nm, 0.0)
        pct = (100.0 * us / cap) if cap > 0.01 else 0.0
        if pct > crit_pct:
            crit_pct, crit_nm = pct, nm
        t_occ.add_row(
            nm,
            f"{us:.1f}",
            f"{cap:.1f}",
            f"{pct:.0f}%",
        )
    if hh_por_turma.get("Pelotao_Unificado", 0) > 0.01:
        d_pool = len(
            set(c["Dia"] for c in cronograma if c.get("Turma") == "Pelotao_Unificado")
        )
        pu = hh_por_turma["Pelotao_Unificado"]
        cap_p = float(d_pool) * float(executores) * float(jornada)
        pct_p = (100.0 * pu / cap_p) if cap_p > 0.01 else 0.0
        t_occ.add_row(
            "Pelotao_Unificado",
            f"{pu:.1f}",
            f"{cap_p:.1f}",
            f"{pct_p:.0f}%",
        )
    console.print(t_occ)
    print(
        DM
        + "  Uso % = HH no cronograma com o nome da turma / (dias simulados x operarios x jornada)."
        + RS
    )
    print(
        DM
        + "  Reforco nao aumenta n_ops; bloqueio global impede reforco em plantio/irrigacao ate liberar tudo."
        + RS
    )
    if usar_pool_pos_bloqueio and usar_bloqueio_global:
        print(
            DM
            + "  Pelotao_Unificado: plantio/irrigacao apos liberacao usam todos os executores num so pelotao."
            + RS
        )
    if crit_nm:
        print(
            DM
            + f"  Heuristica caminho critico (maior Uso %): turma '{crit_nm}' (~{crit_pct:.0f}%)."
            + RS
        )
    if n_fb > 0:
        print(
            DM
            + f"  Cobertura CT no escopo: {100 - pct_fallback:.0f}% (fallback em {n_fb}/{n_demandas} item(ns))."
            + RS
        )

    def _render_tabela_cenarios(rows, label):
        if not rows:
            return
        t_sc = Table(title=f"Comparativo de Cenários (Equipe x Jornada) - {label}")
        t_sc.add_column("Equipe", justify="right")
        t_sc.add_column("Jornada", justify="right")
        t_sc.add_column("Dias", justify="right")
        t_sc.add_column("Meses", justify="right")
        t_sc.add_column("Ganho vs Meta", justify="right")
        for r in rows[:40]:
            t_sc.add_row(
                str(r["Equipe"]),
                f"{r['Jornada_h_dia']:.2f}",
                str(r["Dias_Simulados"]),
                f"{r['Meses_Simulados']:.2f}",
                f"{r['Ganho_vs_Meta_dias']:+d}",
            )
        console.print(t_sc)

    cenarios_rows = []
    if comparativo_cfg is not None and isinstance(comparativo_cfg, dict):
        hh_base_multi = float(total_hh)
        lbl_base_multi = "Sem mecanizado"
        if (not _batch) and recursos_mec and cronograma_com_mec:
            hh_hum_pos_mec = sum(
                float(x.get("HH", 0) or 0)
                for x in cronograma_com_mec
                if not str(x.get("Turma", "")).startswith("MEC_")
            )
            base_opt = selecionar(
                "BASE DO COMPARATIVO MULTI-FATOR",
                [
                    "Sem mecanizado (HH total atual)",
                    "Com mecanizado (HH humano remanescente)",
                ],
            )
            if base_opt and base_opt.startswith("Com mecanizado"):
                hh_base_multi = float(hh_hum_pos_mec)
                lbl_base_multi = "Com mecanizado"
        print(
            DM + f"  Base selecionada: {lbl_base_multi} | HH={hh_base_multi:.1f}" + RS
        )
        cenarios_rows = simular_cenarios_multifator(
            total_hh=hh_base_multi,
            dias_meta=dias_meta,
            executores_base=executores,
            jornada_base=jornada,
            jornadas_in=comparativo_cfg.get("jornadas"),
            equipes_in=comparativo_cfg.get("equipes"),
            interativo=False,
        )
        _render_tabela_cenarios(cenarios_rows, lbl_base_multi)

    if not _batch:
        while confirmar(
            "Recalcular comparativo multi-fator com novos valores agora?", default=False
        ):
            hh_base_multi = float(total_hh)
            lbl_base_multi = "Sem mecanizado"
            if recursos_mec and cronograma_com_mec:
                hh_hum_pos_mec = sum(
                    float(x.get("HH", 0) or 0)
                    for x in cronograma_com_mec
                    if not str(x.get("Turma", "")).startswith("MEC_")
                )
                base_opt = selecionar(
                    "BASE DO COMPARATIVO MULTI-FATOR",
                    [
                        "Sem mecanizado (HH total atual)",
                        "Com mecanizado (HH humano remanescente)",
                    ],
                )
                if base_opt and base_opt.startswith("Com mecanizado"):
                    hh_base_multi = float(hh_hum_pos_mec)
                    lbl_base_multi = "Com mecanizado"
            print(
                DM + f"  Base selecionada: {lbl_base_multi} | HH={hh_base_multi:.1f}" + RS
            )
            cenarios_rows = simular_cenarios_multifator(
                total_hh=hh_base_multi,
                dias_meta=dias_meta,
                executores_base=executores,
                jornada_base=jornada,
                jornadas_in=comparativo_cfg.get("jornadas") if isinstance(comparativo_cfg, dict) else None,
                equipes_in=comparativo_cfg.get("equipes") if isinstance(comparativo_cfg, dict) else None,
                interativo=True,
            )
            _render_tabela_cenarios(cenarios_rows, lbl_base_multi)

    atividades_escopo = sorted(
        {
            str(a).strip()
            for a in df_faz["atividade"].dropna().tolist()
            if str(a).strip()
        },
        key=str,
    )
    escopo_set = set(atividades_escopo)
    ag_hum_set = set()
    ag_mec_set = set()
    cronograma_ref = cronograma_com_mec if cronograma_com_mec else cronograma_base
    for item in cronograma_ref or []:
        atividade_item = str(item.get("Atividade", "") or "").strip()
        if not atividade_item:
            continue
        hh_item = float(item.get("HH", 0) or 0)
        hm_item = float(item.get("HM", 0) or 0)
        turma_item = str(item.get("Turma", "") or "")
        if turma_item.startswith("MEC_") or (hm_item > 0 and hh_item <= 0):
            ag_mec_set.add(atividade_item)
        else:
            ag_hum_set.add(atividade_item)
    faltantes_set = escopo_set - (ag_hum_set | ag_mec_set)

    hh_por_atividade = defaultdict(float)
    for tarefas in demandas.values():
        for t in tarefas:
            atividade_t = str(t.get("atividade", "") or "").strip()
            if not atividade_t:
                continue
            hh_por_atividade[atividade_t] += float(t.get("hh_total", 0) or 0)

    rows_audit = []
    for a in atividades_escopo:
        if a in ag_hum_set:
            status = "agendada_humana"
        elif a in ag_mec_set:
            status = "agendada_mecanizada"
        else:
            status = "nao_agendada"

        motivo = ""
        if status == "nao_agendada":
            if a in atividades_mec_set and not recursos_mec:
                motivo = "atividade mecanizada sem recurso cadastrado"
            else:
                motivo = "sem alocacao no cronograma"

        rows_audit.append(
            {
                "Atividade": a,
                "HH_Escopo": round(float(hh_por_atividade.get(a, 0) or 0), 2),
                "Status": status,
                "Motivo": motivo,
            }
        )
    df_audit = pd.DataFrame(rows_audit)
    sub()
    print(G + BL + "  AUDITORIA DO ESCOPO (ANTES DA EXPORTACAO)" + RS)
    print(DM + f"  Atividades no escopo: {len(atividades_escopo)}" + RS)
    print(DM + f"  Agendadas no humano: {len(ag_hum_set & escopo_set)}" + RS)
    print(DM + f"  Agendadas no mecanizado: {len(ag_mec_set & escopo_set)}" + RS)
    print(DM + f"  Nao agendadas: {len(faltantes_set)}" + RS)
    rocadas_escopo = [a for a in atividades_escopo if "rocada" in normalizar_chave(a)]
    if rocadas_escopo:
        for rcv in rocadas_escopo:
            if rcv in ag_hum_set:
                st = "agendada_humana"
            elif rcv in ag_mec_set:
                st = "agendada_mecanizada"
            else:
                st = "nao_agendada"
            print(DM + f"    rocada: {rcv[:56]} -> {st}" + RS)

    # ── Export Dossier Excel (somente operacional) ──
    if cronograma_base:
        try:

            def _slug_nome(v):
                return str(v).replace("/", "_").replace(" ", "_")

            scope_tag = "__FAZENDA_TODOS"
            if isinstance(escopo_meta, dict):
                modo_th = str(escopo_meta.get("modo_talhao") or "")
                ths = [
                    str(x) for x in (escopo_meta.get("talhoes") or []) if str(x).strip()
                ]
                if modo_th in ("unico", "parcial") and len(ths) == 1:
                    scope_tag = f"__TH_{_slug_nome(ths[0])}"
                elif modo_th == "parcial" and len(ths) > 1:
                    scope_tag = f"__TH_MULTI_{len(ths)}"
                elif modo_th in ("todos", "fallback_todos"):
                    scope_tag = "__FAZENDA_TODOS"

            nome_base = f"Dossier_{_slug_nome(fazenda)}{scope_tag}"
            nome_op = f"{nome_base}_OPERACIONAL.xlsx"
            pasta_dossier = OUTPUT_DIR
            os.makedirs(pasta_dossier, exist_ok=True)
            nome_op, caminho_op = _proximo_caminho_livre(pasta_dossier, nome_op)

            df_crono = pd.DataFrame(cronograma_base)
            if "Dia" in df_crono.columns:
                df_crono["Semana"] = df_crono["Dia"].apply(
                    lambda d: int(math.ceil(float(d) / 5.0)) if pd.notna(d) else ""
                )

            rows_op = [
                {"Metrica": "Fazenda", "Valor": fazenda},
                {"Metrica": "Arquivo operacional", "Valor": nome_op},
                {"Metrica": "Executores", "Valor": executores},
                {"Metrica": "Jornada (h/dia)", "Valor": jornada},
                {"Metrica": "Prazo Meta (meses)", "Valor": prazo_meses},
                {"Metrica": "Dias Uteis Meta", "Valor": dias_meta},
                {"Metrica": "Duracao Simulada (dias uteis)", "Valor": dias_simulado},
                {
                    "Metrica": "Duracao Simulada (meses)",
                    "Valor": f"{meses_simulado:.1f}",
                },
                {"Metrica": "HH Total Simulado", "Valor": f"{total_hh:,.1f}"},
                {
                    "Metrica": "Fonte dos dados",
                    "Valor": "100% CT"
                    if pct_fallback < 0.01
                    else f"{100 - pct_fallback:.0f}% CT ({n_fb} fallbacks)",
                },
                {"Metrica": "", "Valor": ""},
                {"Metrica": "Atividades no escopo", "Valor": len(atividades_escopo)},
                {
                    "Metrica": "Agendadas (humano)",
                    "Valor": len(ag_hum_set & escopo_set),
                },
                {
                    "Metrica": "Agendadas (mecanizado)",
                    "Valor": len(ag_mec_set & escopo_set),
                },
                {"Metrica": "Nao agendadas", "Valor": len(faltantes_set)},
            ]
            if isinstance(escopo_meta, dict):
                rows_op.append(
                    {
                        "Metrica": "Escopo talhoes",
                        "Valor": ", ".join(
                            str(x) for x in (escopo_meta.get("talhoes") or [])
                        )
                        or "todos",
                    }
                )
            if recursos_mec:
                rows_op += [{"Metrica": "", "Valor": ""}]
                for rec in recursos_mec:
                    rows_op.append(
                        {
                            "Metrica": f"Mecanizado: {rec['nome']}",
                            "Valor": f"{rec['prod_ha_h']} ha/h",
                        }
                    )
                    rows_op.append(
                        {
                            "Metrica": f"  Atividades ({rec['nome']})",
                            "Valor": str(len(rec.get("atividades", set()))),
                        }
                    )
            resumo_op = pd.DataFrame(rows_op)

            df_cascata = _gerar_aba_cascata_explicada(
                cronograma_base, jornada, dia_ref, mes_ref, ano_ref
            )
            df_ocupacao = _gerar_aba_ocupacao_turmas(
                cronograma, turmas, jornada, dias_simulado_hum, dia_ref, mes_ref, ano_ref
            )
            df_crono_op = _df_crono_operacional(df_crono, dia_ref, mes_ref, ano_ref)

            with pd.ExcelWriter(caminho_op, engine="openpyxl") as writer_op:
                resumo_op.to_excel(
                    writer_op, sheet_name="RESUMO_OPERACIONAL", index=False
                )
                df_crono_op.to_excel(
                    writer_op, sheet_name="CRONOGRAMA_DETALHADO", index=False
                )
                if not df_cascata.empty:
                    df_cascata.to_excel(
                        writer_op, sheet_name="CASCATA_EXPLICADA", index=False
                    )
                if not df_ocupacao.empty:
                    df_ocupacao.to_excel(
                        writer_op, sheet_name="OCUPACAO_TURMAS_DIA", index=False
                    )
                if recursos_mec and cronograma_mec:
                    df_mec_crono = _df_crono_operacional(pd.DataFrame(cronograma_mec))
                    df_mec_crono.to_excel(
                        writer_op, sheet_name="CRONOGRAMA_MECANIZADO", index=False
                    )
                    df_combinado = _df_crono_operacional(pd.DataFrame(cronograma_com_mec))
                    df_combinado.to_excel(
                        writer_op, sheet_name="CRONOGRAMA_COMBINADO", index=False
                    )
                if cronograma_mec_base:
                    df_mec_base = _df_crono_operacional(pd.DataFrame(cronograma_mec_base))
                    df_mec_base.to_excel(
                        writer_op, sheet_name="CRONOGRAMA_MEC_BASE", index=False
                    )
                if not df_audit.empty:
                    df_audit.to_excel(
                        writer_op, sheet_name="AUDITORIA_ESCOPO", index=False
                    )
                wb_op = writer_op.book
                _aplicar_cores_ocupacao_excel(wb_op, "OCUPACAO_TURMAS_DIA")
                try:
                    from srf_excel_format import aplicar_formatacao_operacional

                    aplicar_formatacao_operacional(wb_op, dias_simulado, cronograma_base)
                except Exception:
                    pass

            ok(f"Dossier operacional exportado: {nome_op}")

            if cenarios_rows:
                nome_xlsx_cmp = f"Dossier_{fazenda.replace('/', '_').replace(' ', '_')}_COMPARATIVO_CENARIOS.xlsx"
                nome_xlsx_cmp, caminho_xlsx_cmp = _proximo_caminho_livre(
                    pasta_dossier, nome_xlsx_cmp
                )
                with pd.ExcelWriter(caminho_xlsx_cmp, engine="openpyxl") as writer3:
                    pd.DataFrame(cenarios_rows).to_excel(
                        writer3, sheet_name="COMPARATIVO_CENARIOS", index=False
                    )
                ok(f"Dossier comparativo de cenarios exportado: {nome_xlsx_cmp}")

            if recursos_mec and cronograma_com_mec:
                nome_mec_op = f"{nome_base}_COM_MECANIZADO_OPERACIONAL.xlsx"
                nome_mec_op, caminho_mec_op = _proximo_caminho_livre(
                    pasta_dossier, nome_mec_op
                )
                df_mec_full = pd.DataFrame(cronograma_com_mec)
                if "Dia" in df_mec_full.columns:
                    df_mec_full["Semana"] = df_mec_full["Dia"].apply(
                        lambda d: int(math.ceil(float(d) / 5.0)) if pd.notna(d) else ""
                    )
                d_comb = max(
                    [int(x.get("Dia", 0)) for x in cronograma_com_mec], default=0
                )
                rows_mec_op = [
                    {"Metrica": "Fazenda", "Valor": fazenda},
                    {"Metrica": "Arquivo operacional", "Valor": nome_mec_op},
                    {"Metrica": "Cenario", "Valor": "Humano + Mecanizado"},
                    {"Metrica": "Dias baseline (humano)", "Valor": dias_simulado},
                    {"Metrica": "Dias cenario combinado", "Valor": d_comb},
                    {
                        "Metrica": "Ganho de prazo (dias)",
                        "Valor": int(dias_simulado) - int(d_comb),
                    },
                ]
                for rec in recursos_mec:
                    rows_mec_op.append(
                        {
                            "Metrica": f"Recurso: {rec['nome']}",
                            "Valor": f"{rec['prod_ha_h']} ha/h",
                        }
                    )

                df_cascata_mec = _gerar_aba_cascata_explicada(
                    cronograma_com_mec, jornada, dia_ref, mes_ref, ano_ref
                )
                df_mec_op = _df_crono_operacional(
                    df_mec_full, dia_ref, mes_ref, ano_ref
                )

                with pd.ExcelWriter(caminho_mec_op, engine="openpyxl") as writer_mo:
                    pd.DataFrame(rows_mec_op).to_excel(
                        writer_mo, sheet_name="RESUMO_OPERACIONAL", index=False
                    )
                    df_mec_op.to_excel(
                        writer_mo, sheet_name="CRONOGRAMA_DETALHADO", index=False
                    )
                    if not df_cascata_mec.empty:
                        df_cascata_mec.to_excel(
                            writer_mo, sheet_name="CASCATA_EXPLICADA", index=False
                        )
                    wb_mo = writer_mo.book
                    try:
                        from srf_excel_format import aplicar_formatacao_operacional

                        aplicar_formatacao_operacional(wb_mo, d_comb, cronograma_com_mec)
                    except Exception:
                        pass

                ok(f"Dossier cenario mecanizado (operacional): {nome_mec_op}")
        except Exception as ex:
            aviso(f"Nao foi possivel salvar Dossier: {ex}")

    # ── Diagnostico final ──
    linha()
    print(G + BL + "  DIAGNOSTICO DE PRAZO" + RS)
    sub()
    print(
        G
        + f"  Meta informada             : {prazo_meses} meses ({dias_meta} dias uteis a partir de {mes_ref:02d}/{ano_ref})"
        + RS
    )
    print(
        G
        + f"  Duracao simulada           : {dias_simulado} dias ({meses_simulado:.1f} meses)"
        + RS
    )
    if recursos_mec and cronograma_com_mec:
        d_mc = max([int(x.get("Dia", 0)) for x in cronograma_com_mec], default=0)
        m_mc = d_mc / 22.0 if d_mc > 0 else 0.0
        print(C + f"  Duracao cenario mecanizado : {d_mc} dias ({m_mc:.1f} meses)" + RS)
        print(
            C
            + f"  Ganho operacional estimado : {int(dias_simulado) - int(d_mc):+d} dias"
            + RS
        )
    sub()

    if meses_simulado <= prazo_meses:
        print(G + BL + "  STATUS: DENTRO DO PRAZO" + RS)
        print(G + f"  Equipe de {executores} executores conclui antes da meta." + RS)
    else:
        print(Y + BL + "  STATUS: PRAZO EXCEDIDO" + RS)
        print(
            Y
            + f"  Equipe atual levara {meses_simulado:.1f} meses (meta: {prazo_meses})."
            + RS
        )
        print(
            C
            + f"  [SUGESTAO] ~{exec_teoricos} executores @ {jornada}h/dia cumpririam a meta."
            + RS
        )
        if dias_meta > 0 and total_hh > 0.01:
            ex5 = math.ceil(total_hh / (dias_meta * 5.0))
            ex6 = math.ceil(total_hh / (dias_meta * 6.0))
            print(
                DM
                + f"  [DICA] Com a mesma jornada na meta, ~{ex5} executores @ 5h/dia ou ~{ex6} @ 6h/dia "
                f"(aprox.: HH total / {dias_meta} dias uteis / jornada)." + RS
            )

    linha()
    
    # ═══════════════════════════════════════════════════════════════════════════
    # MODO COMPARATIVO: Executar segunda simulação com atividades mecanizadas
    # ═══════════════════════════════════════════════════════════════════════════
    resultado_mecanizado = None
    resultado_mecanizado_valido = False
    if modo_comparativo and substituicoes_comparativo:
        sub()
        print(C + BL + " EXECUTANDO CENÁRIO MECANIZADO (Comparativo)" + RS)
        print(DM + " Preparando substituições de atividades..." + RS)

        comparativo_cfg = cfg.get("comparativo", {}) if isinstance(cfg, dict) else {}
        execucao_compacta = bool(comparativo_cfg.get("execucao_compacta", True))
        if execucao_compacta:
            print(
                DM
                + " Cenário mecanizado em modo compacto: detalhes intermediários suprimidos."
                + RS
            )
        
        # Criar cópia do dataframe com atividades substituídas
        df_mec = _substituir_por_mecanizado(df_faz, substituicoes_comparativo)
        cfg_mec = _clonar_cfg_comparativo_mecanizado(cfg, substituicoes_comparativo)
        
        # Contar quantas atividades foram trocadas
        n_substituicoes = 0
        for manual, mec in substituicoes_comparativo.items():
            if (df_faz["atividade"] == manual).any():
                n_substituicoes += (df_faz["atividade"] == manual).sum()
        
        ok(f"{n_substituicoes} registro(s) serão executados com versão mecanizada.")
        
        # Executar segundo cenário (mecanizado) em modo batch para não interagir
        ctx_mec = {
            "modo_seq": modo_seq,
            "usar_bloqueio_global": usar_bloqueio_global,
            "usar_reforco_automatico": usar_reforco_automatico,
            "usar_pool_pos_bloqueio": usar_pool_pos_bloqueio,
            "prazo_meses": prazo_meses,
            "mes_ref": mes_ref,
            "ano_ref": ano_ref,
            "data_inicio_txt": data_inicio_txt,
            "data_fim_txt": data_fim_txt,
            "jornada": jornada,
            "executores": executores,
            "turmas": turmas,
            "preencher_orfas_template": preencher_orfas,
            "substituicoes_template": substituicoes_comparativo,
            "reatribuicao_template": reatribuicao,
            "paralelo_template": paralelo,
            "primaria_template": primaria,
            "session_hh": session_hh,
        }
        
        if execucao_compacta:
            _buf_cmp = io.StringIO()
            try:
                with redirect_stdout(_buf_cmp), redirect_stderr(_buf_cmp):
                    resultado_mecanizado = calcular_cronograma_inteligente(
                        cfg_mec,
                        df_mec,
                        fazenda + " (MECANIZADO)",
                        esperar_enter=False,  # Não esperar enter no segundo cenário
                        ctx=ctx_mec,
                        escopo_meta=escopo_meta,
                        atividades_catalogo=atividades_catalogo,
                        modo_comparativo=False,  # Evitar recursão infinita
                        substituicoes_comparativo=None,
                    )
            except Exception as e:
                resultado_mecanizado = None
                aviso(f"Cenario mecanizado falhou em modo compacto: {e}")
        else:
            resultado_mecanizado = calcular_cronograma_inteligente(
                cfg_mec,
                df_mec,
                fazenda + " (MECANIZADO)",
                esperar_enter=False,  # Não esperar enter no segundo cenário
                ctx=ctx_mec,
                escopo_meta=escopo_meta,
                atividades_catalogo=atividades_catalogo,
                modo_comparativo=False,  # Evitar recursão infinita
                substituicoes_comparativo=None,
            )
        
        # Verificar se houve retrocesso
        if isinstance(resultado_mecanizado, dict) and resultado_mecanizado.get("acao") == "retroceder_escopo":
            aviso("Cenário mecanizado cancelado.")
            resultado_mecanizado = None
        elif isinstance(resultado_mecanizado, dict) and resultado_mecanizado.get("acao"):
            aviso(
                f"Cenario mecanizado finalizou com acao '{resultado_mecanizado.get('acao')}'."
            )
            resultado_mecanizado = None
        elif not isinstance(resultado_mecanizado, dict):
            aviso("Cenario mecanizado nao retornou resultado valido.")
            resultado_mecanizado = None
        else:
            chaves_obrigatorias = (
                "dias_simulado",
                "total_hh",
            )
            faltantes = [k for k in chaves_obrigatorias if k not in resultado_mecanizado]
            if faltantes:
                aviso(
                    "Cenario mecanizado retornou resultado incompleto; comparativo nao sera exibido."
                )
                resultado_mecanizado = None
            else:
                resultado_mecanizado_valido = True
                ok("Cenário mecanizado concluído!")
    
    # ═══════════════════════════════════════════════════════════════════════════
    # EXIBIR COMPARATIVO (se modo comparativo ativo)
    # ═══════════════════════════════════════════════════════════════════════════
    if resultado_mecanizado_valido:
        sub()
        print(G + BL + "═══════════════════════════════════════════════════════════════════" + RS)
        print(G + BL + "       COMPARATIVO: MANUAL vs MECANIZADO" + RS)
        print(G + BL + "═══════════════════════════════════════════════════════════════════" + RS)
        print()
        
        # Preparar dados
        d_manual = float(dias_simulado)
        d_mec = float(resultado_mecanizado.get("dias_simulado") or 0)
        hh_manual = float(total_hh)
        hh_mec = float(resultado_mecanizado.get("total_hh") or 0)
        hm_manual = float(total_hm)
        hm_mec = float(resultado_mecanizado.get("total_hm") or 0)
        
        # Calcular economia
        economia_dias = int(d_manual - d_mec)
        economia_hh = hh_manual - hh_mec
        economia_hm = hm_mec - hm_manual
        cap_hh_dia = float(executores) * float(jornada)
        dias_eq_hh_manual = (hh_manual / cap_hh_dia) if cap_hh_dia > 0.01 else 0.0
        dias_eq_hh_mec = (hh_mec / cap_hh_dia) if cap_hh_dia > 0.01 else 0.0
        delta_dias_eq_hh = dias_eq_hh_manual - dias_eq_hh_mec
        cronograma_mec_ref = resultado_mecanizado.get("cronograma") or []
        turmas_mec_comp = sorted(
            {
                str(x.get("Turma", ""))
                for x in cronograma_mec_ref
                if str(x.get("Turma", "")).startswith("MEC_")
            },
            key=str,
        )
        
        # Tabela de comparação
        print(f"  {C}Métrica{RS}                    {C}Manual{RS}          {C}Mecanizado{RS}      {C}Diferença{RS}")
        print(f"  {DM}{'─' * 70}{RS}")
        print(f"  {'Dias necessários':<25} {d_manual:>10.0f}      {d_mec:>10.0f}      {Y}{economia_dias:>+10.0f}{RS}")
        print(f"  {'HH totais':<25} {hh_manual:>10.1f}      {hh_mec:>10.1f}      {Y}{economia_hh:>+10.1f}{RS}")
        print(f"  {'HM totais':<25} {hm_manual:>10.1f}      {hm_mec:>10.1f}      {Y}{economia_hm:>+10.1f}{RS}")
        print(f"  {'Dias eq. via HH/cap':<25} {dias_eq_hh_manual:>10.2f}      {dias_eq_hh_mec:>10.2f}      {Y}{delta_dias_eq_hh:>+10.2f}{RS}")
        print()
        if turmas_mec_comp:
            print(G + BL + "  TURMAS MECANIZADAS NO CENARIO:" + RS)
            for nm_turma in turmas_mec_comp:
                print(DM + f"    - {nm_turma}" + RS)
            print()
        if substituicoes_comparativo:
            print(G + BL + "  SUBSTITUICOES APLICADAS:" + RS)
            for manual, mec in substituicoes_comparativo.items():
                print(f"  • {manual[:50]} → {C}{_formatar_substituicao_comparativo(mec)}{RS}")
            print()
        
        # Destaques
        print(G + BL + "  DESTAQUES:" + RS)
        if economia_dias > 0:
            print(f"  {G}✓{RS} Redução de {G}{economia_dias}{RS} dias com mecanização")
        if economia_hh > 0:
            print(f"  {G}✓{RS} Economia de {G}{economia_hh:.1f}{RS} HH (mão de obra humana)")
        if economia_dias <= 0 and economia_hh > 0 and cap_hh_dia > 0.01:
            print(
                DM
                + f"  Nota: a reducao de HH equivale a ~{delta_dias_eq_hh:.2f} dia(s), "
                + "mas o cronograma fecha por dias inteiros e caminho critico; por isso pode manter o mesmo total de dias."
                + RS
            )
        print()
        print(G + BL + "═══════════════════════════════════════════════════════════════════" + RS)
        sub()
    
    if esperar_enter:
        input(DM + "\n [ENTER para voltar ao menu] " + RS)
    d_mc = (
        max([int(x.get("Dia", 0)) for x in cronograma_com_mec], default=0)
        if (recursos_mec and cronograma_com_mec)
        else None
    )
    ganho_mc = (int(dias_simulado) - int(d_mc)) if d_mc is not None else 0
    rendimentos_feed = []
    if callable(_monitor_build_rendimentos):
        try:
            rendimentos_feed = _monitor_build_rendimentos(demandas)
        except Exception:
            rendimentos_feed = []
    _emitir_monitor_state(
        {
            "operacao": {
                "fazenda_atual": str(fazenda),
                "status_geral": "concluido",
                "mensagem_curta": f"{dias_simulado} dia(s) simulados | HH {total_hh:.1f}",
            },
            "lote": {
                "dias_meta": int(dias_meta),
                "dias_consumidos": int(dias_simulado),
                "saldo_dias": int(max(0, int(dias_meta) - int(dias_simulado))),
                "status_meta_continuo": "OK"
                if meses_simulado <= prazo_meses
                else "EXCEDIDO",
                "prazo_absoluto": True,
            },
            "rendimentos_sessao": rendimentos_feed,
        }
    )
    resumo_monitor = [
        f"Fazenda: {fazenda}",
        f"Dias simulados: {int(dias_simulado)}",
        f"HH total: {float(total_hh):.1f}",
    ]
    _emitir_monitor_relatorio(f"Resumo {fazenda}", "\n".join(resumo_monitor))
    
    resultado_final = {
        "fazenda": fazenda,
        "dias_simulado": int(dias_simulado),
        "meses_simulado": float(meses_simulado),
        "dias_mecanizado": d_mc,
        "ganho_mecanizado_dias": int(ganho_mc),
        "total_hh": float(total_hh),
        "total_hm": float(total_hm),
        "cronograma": cronograma_base,
        "turmas_snapshot": [
            {"nome": t["nome"], "operarios": t["operarios"]} for t in turmas
        ],
    }
    
    # Incluir resultados comparativos se disponíveis
    if resultado_mecanizado_valido:
        resultado_final["comparativo_mecanizado"] = {
            "dias_simulado": resultado_mecanizado.get("dias_simulado"),
            "total_hh": resultado_mecanizado.get("total_hh"),
            "total_hm": resultado_mecanizado.get("total_hm"),
            "substituicoes_aplicadas": [
                {
                    "manual": manual,
                    "mecanizado": _formatar_substituicao_comparativo(mec),
                }
                for manual, mec in (substituicoes_comparativo or {}).items()
            ],
        }
    
    return resultado_final


# ──────────────────────────────────────────────
# V6: ABAS EXCEL TIMELINE + OCUPACAO + PERFIS
# ──────────────────────────────────────────────

_FASE_CORES_HEX = {
    "rocada": "4472C4",
    "formiga": "ED7D31",
    "coroamento": "70AD47",
    "coveamento": "FFC000",
    "adubacao_quimica": "9B59B6",
    "plantio": "2ECC71",
    "irrigacao": "3498DB",
    "limpeza_quimica": "95A5A6",
    "demais": "BDC3C7",
    "reforco": "D5DBDB",
    "pool": "1ABC9C",
}


def _fase_nome_pt(fase_id):
    m = {
        "rocada": "Rocada",
        "formiga": "Formiga",
        "coroamento": "Coroamento",
        "coveamento": "Coveamento",
        "adubacao_quimica": "Adubacao quim.",
        "plantio": "Plantio",
        "irrigacao": "Irrigacao",
        "limpeza_quimica": "Limpeza quim.",
        "demais": "Demais",
        "reforco": "Reforco",
        "pool": "Pelotao unif.",
    }
    return m.get(fase_id, str(fase_id).capitalize())


def _classificar_fase_nome(atv, seq_cfg, modo, atvs_plantio, atvs_irrig):
    """Retorna (fase_id, fase_valor) para rotulagem no Excel."""
    from collections import OrderedDict as _OD

    if eh_limpeza_quimica_pos_plantio(atv, seq_cfg):
        return "limpeza_quimica", 8.0
    if atv in atvs_plantio or _match_filtros_fase(
        atv, seq_cfg.get("filtros_plantio") or ["plantio"], None
    ):
        return "plantio", 6.0
    if atv in atvs_irrig or _match_filtros_fase(
        atv, seq_cfg.get("filtros_irrigacao") or ["irrig"], None
    ):
        return "irrigacao", 7.0
    fases = _fases_ordem_config(seq_cfg, modo)
    for i, fase in enumerate(fases):
        if _match_filtros_fase(atv, fase.get("filtros") or [], fase.get("exclusoes")):
            return fase.get("id", f"fase_{i}"), float(i)
    return "demais", 5.5


def _gerar_aba_timeline(cronograma, seq_cfg, modo_seq, atividades_reais, fazenda, dia_ref=None, mes_ref=None, ano_ref=None):
    """Retorna DataFrame para aba TIMELINE_CASCATA com colunas de visualização.
    
    Se dia_ref, mes_ref, ano_ref forem fornecidos, adiciona colunas de data real.
    """
    atvs_plantio = set(
        atividades_por_filtro(
            atividades_reais, seq_cfg.get("filtros_plantio") or ["plantio"]
        )
    )
    atvs_irrig = set(
        atividades_por_filtro(
            atividades_reais, seq_cfg.get("filtros_irrigacao") or ["irrig"]
        )
    )
    rows = []
    for c in cronograma:
        atv = c.get("Atividade", "")
        fase_id, fase_val = _classificar_fase_nome(
            atv, seq_cfg, modo_seq, atvs_plantio, atvs_irrig
        )
        modo_exec = c.get("Modo", "Normal")
        if modo_exec == "Reforco":
            fase_id_display = "reforco"
        elif modo_exec == "PoolPosBloqueio":
            fase_id_display = "pool"
        else:
            fase_id_display = fase_id
        
        dia_simulado = c.get("Dia", 1)
        
        # Calcular data real se parametros fornecidos
        data_real = None
        dia_semana = ""
        if dia_ref and mes_ref and ano_ref:
            data_tuple = _converter_dia_simulado_para_data(
                dia_simulado, dia_ref, mes_ref, ano_ref
            )
            if data_tuple:
                data_real = data_tuple[0]  # "20/04/2025"
                dia_semana = data_tuple[1]  # "Seg"
        
        row = {
            "Dia": dia_simulado,
            "Semana": int(math.ceil(float(dia_simulado) / 5.0)),
            "Fazenda": fazenda,
            "Talhao": c.get("Talhao", ""),
            "Atividade": atv,
            "Fase": _fase_nome_pt(fase_id),
            "Fase_ID": fase_id,
            "Fase_Ordem": fase_val,
            "Turma": c.get("Turma", ""),
            "Operarios": c.get("Operarios", 0),
            "HH": c.get("HH", 0),
            "Custo_MO": c.get("Custo_MO", 0),
            "Modo": modo_exec,
            "Cor_Hex": _FASE_CORES_HEX.get(fase_id_display, "BDC3C7"),
        }
        
        # Adicionar colunas de data real se calculadas
        if data_real:
            row["Data"] = data_real
            row["Dia_Semana"] = dia_semana
            
        rows.append(row)
    
    df = pd.DataFrame(rows) if rows else pd.DataFrame()
    
    # Reordenar colunas: Data, Dia_Semana primeiro se existirem
    if not df.empty and "Data" in df.columns:
        cols = ["Data", "Dia_Semana"] + [c for c in df.columns if c not in ["Data", "Dia_Semana"]]
        df = df[cols]
    
    return df


def _gerar_aba_cascata_explicada(cronograma, jornada, dia_ref=None, mes_ref=None, ano_ref=None):
    """
    Trilha explicativa da cascata por dia/turma/atividade.
    Mostra capacidade, consumo, saldo e pendencia (carry-over) de forma didatica.
    Se dia_ref, mes_ref, ano_ref forem fornecidos, adiciona colunas de data real.
    """
    if not cronograma:
        return pd.DataFrame()

    rows_src = []
    for i, c in enumerate(cronograma):
        try:
            dia = int(c.get("Dia", 0) or 0)
            hh = float(c.get("HH", 0) or 0.0)
            ops = float(c.get("Operarios", 0) or 0.0)
        except (TypeError, ValueError):
            continue
        if dia <= 0 or hh <= 0.0:
            continue
        rows_src.append(
            {
                "_ord": i,
                "Dia": dia,
                "Semana": int(math.ceil(float(dia) / 5.0)),
                "Fazenda": c.get("Fazenda", ""),
                "Talhao": c.get("Talhao", ""),
                "Atividade": c.get("Atividade", ""),
                "Turma": c.get("Turma", ""),
                "Operarios": ops,
                "HH": hh,
            }
        )
    if not rows_src:
        return pd.DataFrame()

    # demanda total por atividade/talhao/turma para calcular pendencia durante o consumo
    demanda_total = defaultdict(float)
    for r in rows_src:
        k = (str(r["Talhao"]), str(r["Atividade"]), str(r["Turma"]))
        demanda_total[k] += float(r["HH"])

    df_rows = (
        pd.DataFrame(rows_src)
        .sort_values(["Dia", "Turma", "_ord"])
        .reset_index(drop=True)
    )
    out = []
    consumido_atividade = defaultdict(float)

    for (dia, turma), grp in df_rows.groupby(["Dia", "Turma"], sort=True):
        ops_dia = max(float(x) for x in grp["Operarios"].tolist()) if len(grp) else 0.0
        cap_dia = max(0.0, float(ops_dia) * float(jornada))
        usado_dia = 0.0
        
        # Calcular data real se parametros fornecidos
        data_real = None
        dia_semana = ""
        if dia_ref and mes_ref and ano_ref:
            data_tuple = _converter_dia_simulado_para_data(
                dia, dia_ref, mes_ref, ano_ref
            )
            if data_tuple:
                data_real = data_tuple[0]
                dia_semana = data_tuple[1]
        
        for _, r in grp.iterrows():
            hh_inicio = max(0.0, cap_dia - usado_dia)
            hh_cons = float(r["HH"])
            usado_dia += hh_cons
            hh_saldo = max(0.0, cap_dia - usado_dia)
            k = (str(r["Talhao"]), str(r["Atividade"]), str(turma))
            consumido_atividade[k] += hh_cons
            pend = max(0.0, float(demanda_total[k]) - float(consumido_atividade[k]))
            op = float(r["Operarios"] or 0.0)
            hh_equiv_op = (hh_cons / op) if op > 0.01 else 0.0
            
            row = {
                "Tipo_Linha": "ATIVIDADE",
                "Dia": int(dia),
                "Semana": int(r["Semana"]),
                "Fazenda": r["Fazenda"],
                "Talhao": r["Talhao"],
                "Atividade": r["Atividade"],
                "Turma": turma,
                "Operarios": int(round(op)) if op > 0 else 0,
                "Jornada_h_dia": round(float(jornada), 2),
                "Capacidade_Dia_HH": round(cap_dia, 2),
                "HH_Disponivel_Inicio_Dia": round(hh_inicio, 2),
                "HH_Atividade_Demandado": round(float(demanda_total[k]), 2),
                "HH_Atividade_Consumido": round(hh_cons, 2),
                "HH_Consumido_Por_Operador_Equiv": round(hh_equiv_op, 3),
                "HH_Saldo_Apos_Atividade": round(hh_saldo, 2),
                "HH_Pendente_Atividade": round(pend, 2),
                "Fechou_Dia": "S" if hh_saldo <= 0.01 else "N",
                "Calculo_Dia": f"{cap_dia:.2f} - {usado_dia - hh_cons:.2f} - {hh_cons:.2f} = {hh_saldo:.2f}",
            }
            
            # Adicionar data real se calculada
            if data_real:
                row["Data"] = data_real
                row["Dia_Semana"] = dia_semana
                
            out.append(row)
            
        # Resumo do dia
        resumo_row = {
            "Tipo_Linha": "RESUMO_DIA",
            "Dia": int(dia),
            "Semana": int(math.ceil(float(dia) / 5.0)),
            "Fazenda": "",
            "Talhao": "",
            "Atividade": "__RESUMO_DIA__",
            "Turma": turma,
            "Operarios": int(round(ops_dia)) if ops_dia > 0 else 0,
            "Jornada_h_dia": round(float(jornada), 2),
            "Capacidade_Dia_HH": round(cap_dia, 2),
            "HH_Disponivel_Inicio_Dia": round(cap_dia, 2),
            "HH_Atividade_Demandado": "",
            "HH_Atividade_Consumido": round(usado_dia, 2),
            "HH_Consumido_Por_Operador_Equiv": round((usado_dia / ops_dia), 3)
            if ops_dia > 0.01
            else 0.0,
            "HH_Saldo_Apos_Atividade": round(max(0.0, cap_dia - usado_dia), 2),
            "HH_Pendente_Atividade": "",
            "Fechou_Dia": "S" if max(0.0, cap_dia - usado_dia) <= 0.01 else "N",
            "Calculo_Dia": f"{cap_dia:.2f} - {usado_dia:.2f} = {max(0.0, cap_dia - usado_dia):.2f}",
        }
        
        # Adicionar data real no resumo
        if data_real:
            resumo_row["Data"] = data_real
            resumo_row["Dia_Semana"] = dia_semana
            
        out.append(resumo_row)
        
    df = pd.DataFrame(out)
    
    # Reordenar colunas: Data, Dia_Semana primeiro se existirem
    if not df.empty and "Data" in df.columns:
        cols = ["Data", "Dia_Semana"] + [c for c in df.columns if c not in ["Data", "Dia_Semana"]]
        df = df[cols]
        
    return df


def _gerar_aba_ocupacao_turmas(cronograma, turmas, jornada, dias_simulado, dia_ref=None, mes_ref=None, ano_ref=None):
    """Retorna DataFrame pivot: dia x turma com HH, Cap, Uso%, Status.
    Se dia_ref, mes_ref, ano_ref forem fornecidos, adiciona colunas de data real.
    """
    if not cronograma or dias_simulado < 1:
        return pd.DataFrame()
    turma_nomes = sorted(set(t["nome"] for t in turmas))
    turma_ops = {t["nome"]: t["operarios"] for t in turmas}
    hh_dia_turma = defaultdict(lambda: defaultdict(float))
    for c in cronograma:
        hh_dia_turma[c["Dia"]][c.get("Turma", "")] += float(c.get("HH", 0))
    rows = []
    for dia in range(1, dias_simulado + 1):
        # Calcular data real se parametros fornecidos
        data_real = None
        dia_semana = ""
        if dia_ref and mes_ref and ano_ref:
            data_tuple = _converter_dia_simulado_para_data(
                dia, dia_ref, mes_ref, ano_ref
            )
            if data_tuple:
                data_real = data_tuple[0]
                dia_semana = data_tuple[1]
        
        row = {"Dia": dia, "Semana": int(math.ceil(dia / 5.0))}
        
        # Adicionar data real se calculada
        if data_real:
            row["Data"] = data_real
            row["Dia_Semana"] = dia_semana
        hh_total_dia = 0.0
        cap_total_dia = 0.0
        for tn in turma_nomes:
            hh = hh_dia_turma[dia].get(tn, 0.0)
            cap = turma_ops.get(tn, 0) * jornada
            pct = (hh / cap * 100) if cap > 0.01 else 0.0
            row[f"{tn}_HH"] = round(hh, 2)
            row[f"{tn}_Cap"] = round(cap, 2)
            row[f"{tn}_Uso%"] = round(pct, 1)
            if pct >= 90:
                row[f"{tn}_Status"] = "ALTO"
            elif pct >= 50:
                row[f"{tn}_Status"] = "MEDIO"
            elif pct > 0.01:
                row[f"{tn}_Status"] = "BAIXO"
            else:
                row[f"{tn}_Status"] = "OCIOSO"
            hh_total_dia += hh
            cap_total_dia += cap
        row["Total_HH"] = round(hh_total_dia, 2)
        row["Total_Cap"] = round(cap_total_dia, 2)
        row["Total_Uso%"] = round(
            (hh_total_dia / cap_total_dia * 100) if cap_total_dia > 0.01 else 0.0, 1
        )
        rows.append(row)
    return pd.DataFrame(rows)


def _df_crono_operacional(df_crono, dia_ref=None, mes_ref=None, ano_ref=None):
    """Remove colunas monetarias do cronograma para export operacional.
    Se dia_ref, mes_ref, ano_ref forem fornecidos, adiciona colunas de data real.
    """
    drop = [c for c in ("Custo_MO",) if c in df_crono.columns]
    df = df_crono.drop(columns=drop, errors="ignore")
    
    # Adicionar colunas de data real se parametros fornecidos
    if dia_ref and mes_ref and ano_ref and "Dia" in df.columns:
        datas_reais = []
        dias_semana = []
        for _, row in df.iterrows():
            dia_simulado = row.get("Dia", 1)
            data_tuple = _converter_dia_simulado_para_data(
                dia_simulado, dia_ref, mes_ref, ano_ref
            )
            if data_tuple:
                datas_reais.append(data_tuple[0])
                dias_semana.append(data_tuple[1])
            else:
                datas_reais.append(f"Dia_{dia_simulado}")
                dias_semana.append("")
        
        # Inserir colunas no inicio
        df.insert(0, "Dia_Semana", dias_semana)
        df.insert(0, "Data", datas_reais)
    
    return df


def _escrever_cronograma_e_cascata(
    writer, df_crono_op, df_timeline, sheet_name="CRONOGRAMA_E_CASCATA"
):
    """
    Cronograma e timeline na mesma folha (linha em branco entre blocos).
    Retorna a linha 1-based do cabecalho da timeline (para colorir Fase), ou None.
    """
    df_crono_op.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
    if df_timeline is None or getattr(df_timeline, "empty", True):
        return None
    start = len(df_crono_op) + 2
    df_timeline.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start)
    return start + 1


def _aplicar_cores_timeline_excel(wb, sheet_name="TIMELINE_CASCATA", header_row=1):
    """Colorir coluna Cor_Hex como fill real na coluna da Fase (header_row = linha do cabecalho da timeline, 1-based)."""
    try:
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    header = [cell.value for cell in ws[header_row]]
    if "Cor_Hex" not in header or "Fase" not in header:
        return
    idx_cor = header.index("Cor_Hex") + 1
    idx_fase = header.index("Fase") + 1
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        hex_val = str(row[idx_cor - 1].value or "BDC3C7")
        if len(hex_val) == 6:
            fill = PatternFill(
                start_color=hex_val, end_color=hex_val, fill_type="solid"
            )
            row[idx_fase - 1].fill = fill
            row[idx_fase - 1].font = Font(color="FFFFFF", bold=True)


def _aplicar_cores_ocupacao_excel(wb, sheet_name="OCUPACAO_TURMAS_DIA"):
    """Colorir Status (ALTO/MEDIO/BAIXO/OCIOSO) na aba de ocupação."""
    try:
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    header = [cell.value for cell in ws[1]]
    status_cols = [i for i, h in enumerate(header) if h and str(h).endswith("_Status")]
    fills = {
        "ALTO": PatternFill(
            start_color="E74C3C", end_color="E74C3C", fill_type="solid"
        ),
        "MEDIO": PatternFill(
            start_color="F39C12", end_color="F39C12", fill_type="solid"
        ),
        "BAIXO": PatternFill(
            start_color="3498DB", end_color="3498DB", fill_type="solid"
        ),
        "OCIOSO": PatternFill(
            start_color="95A5A6", end_color="95A5A6", fill_type="solid"
        ),
    }
    font_w = Font(color="FFFFFF", bold=True)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for ci in status_cols:
            val = str(row[ci].value or "")
            if val in fills:
                row[ci].fill = fills[val]
                row[ci].font = font_w


PERFIS_DIR = PROFILES_DIR


def _salvar_perfil_equipe(turmas, executores, jornada, nome_perfil):
    os.makedirs(PERFIS_DIR, exist_ok=True)
    dados = {
        "nome": nome_perfil,
        "executores": executores,
        "jornada": jornada,
        "turmas": [
            {
                "nome": t["nome"],
                "operarios": t["operarios"],
                "atividades": list(t.get("atividades") or []),
            }
            for t in turmas
        ],
    }
    caminho = os.path.join(PERFIS_DIR, f"{_slug_ficheiro_seguro(nome_perfil)}.json")
    with open(caminho, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)
    return caminho


def _listar_perfis_equipe():
    if not os.path.isdir(PERFIS_DIR):
        return []
    out = []
    for fn in sorted(os.listdir(PERFIS_DIR)):
        if fn.endswith(".json"):
            try:
                with open(os.path.join(PERFIS_DIR, fn), "r", encoding="utf-8") as f:
                    d = json.load(f)
                out.append(d)
            except Exception:
                pass
    return out


def _carregar_perfil_equipe_menu():
    perfis = _listar_perfis_equipe()
    if not perfis:
        aviso("Nenhum perfil de equipe salvo ainda.")
        return None
    nomes = [p.get("nome", "?") for p in perfis]
    sel = selecionar("PERFIL DE EQUIPE", nomes)
    if not sel:
        return None
    for p in perfis:
        if p.get("nome") == sel:
            return p
    return None


# ──────────────────────────────────────────────
#  EQUIPE PADRAO + BATCH ORCHESTRATOR
# ──────────────────────────────────────────────


def _checkpoint_editar_template(turmas, atividades_reais_global):
    """Checkpoint before each farm: allow quick edits to team template."""
    sub()
    print(G + BL + "  CHECKPOINT — Equipe Padrao" + RS)
    for t in turmas:
        print(
            G
            + f"  - {t['nome']}: "
            + C
            + f"{t['operarios']} ops, {len(t.get('atividades', []))} atividades"
            + RS
        )
    print(DM + "  [0] Continuar sem alterar" + RS)
    print(DM + "  [1] Editar operarios de uma turma" + RS)
    print(DM + "  [2] Adicionar nova turma" + RS)
    print(DM + "  [3] Redistribuir atividades (S/N) de uma turma" + RS)
    sub()
    op = prompt("Opcao", "0").strip()
    if op == "1":
        nomes = [t["nome"] for t in turmas]
        nm = selecionar("TURMA PARA EDITAR", nomes)
        if nm:
            for t in turmas:
                if t["nome"] == nm:
                    t["operarios"] = pedir_int(
                        f"  Novos operarios para '{nm}'", t["operarios"]
                    )
    elif op == "2":
        nome = prompt("Nome da nova turma", f"Turma {len(turmas) + 1}")
        qtd = pedir_int("Quantos operarios", 1)
        nova = {"nome": nome, "operarios": qtd, "atividades": []}
        if atividades_reais_global:
            menu_vincular_atividades_turma(nova, atividades_reais_global)
        turmas.append(nova)
        ok(f"Turma '{nome}' adicionada.")
    elif op == "3":
        nomes = [t["nome"] for t in turmas]
        nm = selecionar("TURMA PARA REDISTRIBUIR", nomes)
        if nm and atividades_reais_global:
            for t in turmas:
                if t["nome"] == nm:
                    menu_vincular_atividades_turma(t, atividades_reais_global)
    return turmas


def _recomendar_equipes_padrao(
    total_hh, dias_meta, cap_ep_dia, jornada, prazo_absoluto
):
    """Compute how many standard-team sets are needed and return recommendation dict."""
    if cap_ep_dia <= 0.01 or dias_meta <= 0:
        return None
    hh_capacidade_ep = cap_ep_dia * dias_meta
    if hh_capacidade_ep >= total_hh:
        return {
            "status": "suficiente",
            "ep_necessarias": 1,
            "ep_extras": 0,
            "trabalhadores_extras": 0,
        }
    ep_necessarias = math.ceil(total_hh / hh_capacidade_ep)
    ep_extras = ep_necessarias - 1
    trab_necessarios = math.ceil(total_hh / (dias_meta * jornada))
    trab_extras = max(0, trab_necessarios - int(cap_ep_dia / jornada))
    return {
        "status": "insuficiente",
        "ep_necessarias": ep_necessarias,
        "ep_extras": ep_extras,
        "trabalhadores_extras": trab_extras,
        "trab_total_necessario": trab_necessarios,
    }


def _imprimir_recomendacao_ep(rec, fazenda, prazo_absoluto):
    """Print equipe padrao recommendation for one farm."""
    if not rec:
        return
    if rec["status"] == "suficiente":
        print(G + f"  Equipe padrao SUFICIENTE para '{fazenda}'." + RS)
    else:
        print(Y + f"  Equipe padrao INSUFICIENTE para '{fazenda}'." + RS)
        if prazo_absoluto:
            print(
                C
                + f"  [SUGESTAO] +{rec['ep_extras']} equipe(s) padrao (total {rec['ep_necessarias']}) cumpririam a meta."
                + RS
            )
            print(
                C
                + f"  [ALTERNATIVA] +{rec['trabalhadores_extras']} trabalhador(es) extra(s) (total {rec['trab_total_necessario']})."
                + RS
            )


def _exportar_excel_consolidado_lote(
    resultados, empresa_filtro=None, nome_arquivo_micro="", extras=None
):
    """Exporta workbook consolidado do lote com cascata inter-fazendas e timeline unificada."""
    if not resultados:
        return
    extras = extras or {}
    try:
        pasta = OUTPUT_DIR
        os.makedirs(pasta, exist_ok=True)
        emp_slug = (
            _slug_ficheiro_seguro(empresa_filtro)
            if empresa_filtro
            else "Todas_empresas"
        )
        nome_xlsx = f"Consolidado_SmartScheduler_{emp_slug}.xlsx"
        caminho = os.path.join(pasta, nome_xlsx)
        meta_rows = [
            {"Campo": "Empresa_filtro_EQUIPE", "Valor": empresa_filtro or "(todas)"},
            {
                "Campo": "Microplanejamento",
                "Valor": os.path.basename(nome_arquivo_micro)
                if nome_arquivo_micro
                else "",
            },
        ]
        for k, v in extras.items():
            meta_rows.append({"Campo": str(k), "Valor": str(v)})

        dias_acum_total = max(
            (int(x.get("dia_fim_acumulado", 0)) for x in resultados), default=0
        )
        dias_meta_val = int(extras.get("Dias_meta", 0) or 0)
        resumo_rows = [
            {"Metrica": "Fazendas processadas", "Valor": len(resultados)},
            {
                "Metrica": "HH total (soma)",
                "Valor": round(sum(float(x.get("total_hh", 0)) for x in resultados), 1),
            },
            {
                "Metrica": "Dias acumulados lote continuo",
                "Valor": dias_acum_total,
            },
            {
                "Metrica": "Dias meta",
                "Valor": dias_meta_val,
            },
            {
                "Metrica": "Saldo meta (dias)",
                "Valor": max(0, dias_meta_val - dias_acum_total),
            },
            {
                "Metrica": "Status meta global",
                "Valor": "DENTRO" if dias_acum_total <= dias_meta_val else "EXCEDIDO",
            },
        ]
        d_mec_vals = [
            int(x.get("dias_mecanizado") or 0)
            for x in resultados
            if x.get("dias_mecanizado")
        ]
        if d_mec_vals:
            resumo_rows.append(
                {"Metrica": "Dias cenario mecanizado (max)", "Valor": max(d_mec_vals)}
            )

        rows_faz = []
        for x in resultados:
            rec = x.get("rec_ep") or {}
            row_faz = {
                "Fazenda": x.get("fazenda"),
                "Dias_simulado": x.get("dias_simulado"),
                "Dia_inicio_acum": x.get("dia_inicio_acumulado"),
                "Dia_fim_acum": x.get("dia_fim_acumulado"),
                "Meta_consumida_%": x.get("pct_meta_consumida"),
                "Saldo_meta_dias": x.get("saldo_meta_apos"),
                "Status_meta": x.get("status_meta_continuo"),
                "Total_HH": x.get("total_hh"),
            }
            rows_faz.append(row_faz)

        curva_rows = []
        for x in resultados:
            curva_rows.append(
                {
                    "Fazenda": x.get("fazenda"),
                    "Dia_fim_acumulado": x.get("dia_fim_acumulado", 0),
                    "Meta_dias": dias_meta_val,
                    "Consumido_%": x.get("pct_meta_consumida", 0),
                    "HH_acumulado": round(
                        sum(
                            float(r.get("total_hh", 0))
                            for r in resultados[: resultados.index(x) + 1]
                        ),
                        1,
                    ),
                }
            )

        crono_all_rows = []
        for x in resultados:
            offset = int(x.get("dia_inicio_acumulado", 1)) - 1
            for c in x.get("cronograma") or []:
                row = dict(c)
                row["Dia_Lote"] = int(c.get("Dia", 0)) + offset
                row["Semana_Lote"] = int(math.ceil(row["Dia_Lote"] / 5.0))
                crono_all_rows.append(row)

        with pd.ExcelWriter(caminho, engine="openpyxl") as w:
            pd.DataFrame(meta_rows).to_excel(w, sheet_name="METADADOS", index=False)
            pd.DataFrame(resumo_rows).to_excel(w, sheet_name="RESUMO", index=False)
            pd.DataFrame(rows_faz).to_excel(
                w, sheet_name="CASCATA_FAZENDAS", index=False
            )
            pd.DataFrame(curva_rows).to_excel(
                w, sheet_name="CURVA_CONSUMO_META", index=False
            )
            if crono_all_rows:
                pd.DataFrame(crono_all_rows).to_excel(
                    w, sheet_name="CRONOGRAMA_LOTE", index=False
                )
            try:
                wb = w.book
                from openpyxl.styles import Font, PatternFill

                if "CASCATA_FAZENDAS" in wb.sheetnames:
                    ws = wb["CASCATA_FAZENDAS"]
                    header = [cell.value for cell in ws[1]]
                    if "Status_meta" in header:
                        idx_st = header.index("Status_meta") + 1
                        fills_st = {
                            "OK": PatternFill(
                                start_color="27AE60",
                                end_color="27AE60",
                                fill_type="solid",
                            ),
                            "RISCO": PatternFill(
                                start_color="F39C12",
                                end_color="F39C12",
                                fill_type="solid",
                            ),
                            "EXCEDIDO": PatternFill(
                                start_color="E74C3C",
                                end_color="E74C3C",
                                fill_type="solid",
                            ),
                        }
                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                            val = str(row[idx_st - 1].value or "")
                            if val in fills_st:
                                row[idx_st - 1].fill = fills_st[val]
                                row[idx_st - 1].font = Font(color="FFFFFF", bold=True)
            except Exception:
                pass
        ok(f"Consolidado Excel exportado: {nome_xlsx}")
    except Exception as ex:
        aviso(f"Nao foi possivel exportar consolidado Excel: {ex}")


def _executar_lote_fazendas(
    cfg, df_scope, fazendas, empresa_filtro=None, nome_arquivo_micro=""
):
    """Orchestrate all-farms batch: one-time setup, per-farm checkpoint, consolidated report."""

    # ── One-time global setup ──
    dashboard_header()
    subcabecalho("CONFIGURACAO GLOBAL — TODAS AS FAZENDAS")

    todas_atvs = sorted(
        {_norm_atv(x) for x in df_scope["atividade"].dropna().unique() if _norm_atv(x)},
        key=str,
    )

    # Sequence selection
    seq_cfg = cfg.get("sequencia") or {}
    _merge_sequencia_defaults(seq_cfg)
    cfg["sequencia"] = seq_cfg
    modo_seq = _selecionar_sequencia_padrao_sn(cfg, seq_cfg)

    # Bloqueio / reforco / pool
    usar_bloqueio_global = False
    if modo_seq != "personalizado":
        usar_bloqueio_global = confirmar(
            "Aplicar BLOQUEIO GLOBAL (plantio/irrigacao so iniciam quando TODO o resto zerar)?",
            default=True,
        )
    usar_reforco_automatico = confirmar("Ativar REFORCO AUTOMATICO?", default=True)
    usar_pool_pos_bloqueio = False
    if usar_bloqueio_global:
        usar_pool_pos_bloqueio = confirmar(
            "Usar PELOTAO UNIFICADO apos liberacao global?", default=True
        )

    # Deadline
    prazo_meses = pedir_float("Prazo META para conclusao (meses)", 6.0)
    prazo_absoluto = confirmar(
        f"  {prazo_meses} meses e o periodo ABSOLUTO? Se sim, havera sugestoes se necessario",
        default=True,
    )
    hoje = datetime.datetime.now()
    mes_ref = pedir_int("Mes inicial (1-12)", hoje.month)
    mes_ref = max(1, min(12, int(mes_ref)))
    ano_ref = pedir_int("Ano inicial", hoje.year)
    dia_max = calendar.monthrange(ano_ref, mes_ref)[1]
    dia_ref = pedir_int(f"Dia inicial (1-{dia_max})", min(hoje.day, dia_max))
    dia_ref = max(1, min(dia_max, int(dia_ref)))

    data_inicio_txt = _formatar_data_dia(dia_ref, mes_ref, ano_ref)
    data_fim_txt = None
    if confirmar("Informar dia final manualmente para o lote?", default=False):
        mes_fim = pedir_int("Mes final (1-12)", mes_ref)
        mes_fim = max(1, min(12, int(mes_fim)))
        ano_fim = pedir_int("Ano final", ano_ref)
        dia_max_fim = calendar.monthrange(ano_fim, mes_fim)[1]
        dia_fim = pedir_int(f"Dia final (1-{dia_max_fim})", min(dia_ref, dia_max_fim))
        dia_fim = max(1, min(dia_max_fim, int(dia_fim)))
        data_fim_txt = _formatar_data_dia(dia_fim, mes_fim, ano_fim)
    else:
        fim_calc = _calcular_data_fim_por_meses(dia_ref, mes_ref, ano_ref, prazo_meses)
        if fim_calc:
            data_fim_txt = _formatar_data_dia(fim_calc[0], fim_calc[1], fim_calc[2])

    contexto_sessao.definir_datas(data_inicio_txt, data_fim_txt)
    j_def = float(cfg.get("jornada_horas") or 4.6)
    if j_def <= 0:
        j_def = 4.6
    jornada = pedir_jornada("Jornada efetiva diaria (ex: 6.5 ou 6:30 = 6h30)", round(j_def, 2))
    cfg["jornada_horas"] = jornada
    salvar_config(cfg)

    # Team template — carregar perfil ou criar
    sub()
    print(G + BL + "  CONFIGURAR EQUIPE PADRAO" + RS)
    print(DM + "  Defina as turmas que serao reutilizadas em todas as fazendas." + RS)
    print(DM + "  Voce podera ajustar antes de cada fazenda no checkpoint.\n" + RS)

    perfil_carregado = None
    perfis_existentes = _listar_perfis_equipe()
    if perfis_existentes:
        if confirmar("Carregar perfil de equipe salvo anteriormente?", default=False):
            perfil_carregado = _carregar_perfil_equipe_menu()

    if perfil_carregado:
        turmas = [
            {
                "nome": t["nome"],
                "operarios": t["operarios"],
                "atividades": list(t.get("atividades") or []),
            }
            for t in perfil_carregado.get("turmas", [])
        ]
        executores = perfil_carregado.get(
            "executores", sum(t["operarios"] for t in turmas)
        )
        ok(
            f"Perfil '{perfil_carregado['nome']}' carregado: {executores} executores, {len(turmas)} turma(s)."
        )
        for t in turmas:
            print(
                G
                + f"  - {t['nome']}: "
                + C
                + f"{t['operarios']} ops, {len(t.get('atividades', []))} atividades"
                + RS
            )
        if confirmar("Editar este perfil antes de usar?", default=False):
            for turma in turmas:
                menu_vincular_atividades_turma(turma, todas_atvs)
    else:
        executores = pedir_int(
            "Operarios totais da equipe padrao (quem realmente trabalha)",
            9,
        )
        if executores <= 0:
            erro("Precisa de pelo menos 1 executor.")
            return

        turmas = []
        restantes = executores
        while restantes > 0:
            print(G + f"  Operarios disponiveis: {restantes}" + RS)
            nome_turma = prompt("Nome da turma", f"Turma {len(turmas) + 1}")
            def_pad = min(restantes, max(1, restantes // 2 or restantes))
            qtd = pedir_int(f"  Quantos operarios na turma '{nome_turma}'", def_pad)
            if qtd > restantes:
                qtd = restantes
            turmas.append({"nome": nome_turma, "operarios": qtd, "atividades": []})
            restantes -= qtd
            if restantes > 0:
                if not confirmar(
                    f"Criar outra turma? ({restantes} restantes)", default=True
                ):
                    turmas.append(
                        {"nome": "Geral", "operarios": restantes, "atividades": []}
                    )
                    restantes = 0

        sub()
        print(
            G
            + BL
            + "  VINCULAR ATIVIDADES (usa todas as atividades do escopo)"
            + RS
            + "\n"
        )
        for turma in turmas:
            menu_vincular_atividades_turma(turma, todas_atvs)

    if confirmar("Salvar este perfil de equipe para reusar depois?", default=False):
        nome_p = prompt("Nome do perfil", "padrao")
        cam_p = _salvar_perfil_equipe(turmas, executores, jornada, nome_p)
        ok(f"Perfil salvo: {cam_p}")

    sub()
    print(G + BL + "  LOTE: TEMPLATE vs MICRO (lacunas)" + RS)
    print(
        DM
        + "  Template estreito (ex. so irrigacao): outras demandas podem ficar sem turma."
        + RS
    )
    print(
        DM
        + "  Com N, a turma especializada nao recebe tarefas que voce nao vinculou no modelo."
        + RS
    )
    preencher_orfas_template = confirmar(
        "  Por fazenda: distribuir automaticamente demandas sem turma para a turma com mais operarios?",
        default=False,
    )

    cap_ep_dia = float(executores) * float(jornada)
    dias_meta = dias_uteis_no_periodo(mes_ref, ano_ref, prazo_meses)

    ctx_base = {
        "modo_seq": modo_seq,
        "usar_bloqueio_global": usar_bloqueio_global,
        "usar_reforco_automatico": usar_reforco_automatico,
        "usar_pool_pos_bloqueio": usar_pool_pos_bloqueio,
        "prazo_meses": prazo_meses,
        "mes_ref": mes_ref,
        "ano_ref": ano_ref,
        "data_inicio_txt": data_inicio_txt,
        "data_fim_txt": data_fim_txt,
        "jornada": jornada,
        "executores": executores,
        "turmas": turmas,
        "penalidade": 1.0,
        "preencher_orfas_template": preencher_orfas_template,
    }

    # ── Per-farm loop (lote continuo) ──
    resultados = []
    dias_acumulados = 0
    for i_f, fz in enumerate(fazendas, 1):
        linha()
        print(C + BL + f"  [{i_f}/{len(fazendas)}] FAZENDA: {fz}" + RS)
        if prazo_absoluto:
            saldo_pre = dias_meta - dias_acumulados
            pct_consumido = (
                (dias_acumulados / dias_meta * 100) if dias_meta > 0 else 0.0
            )
            print(
                DM + f"  Meta: {dias_meta} dias | Consumido: {dias_acumulados} dias "
                f"({pct_consumido:.0f}%) | Saldo: {saldo_pre} dias" + RS
            )
            if pct_consumido >= 100:
                print(
                    Y + BL + "  !! META GLOBAL JA EXCEDIDA antes desta fazenda !!" + RS
                )
            elif pct_consumido >= 80:
                print(
                    Y + f"  ! Atencao: {pct_consumido:.0f}% da meta ja consumida." + RS
                )
        linha()

        if i_f > 1:
            turmas = _checkpoint_editar_template(turmas, todas_atvs)
            ctx_base["turmas"] = turmas
            ctx_base["executores"] = sum(t["operarios"] for t in turmas)

        r = calcular_cronograma_inteligente(
            cfg,
            df_scope[df_scope["fazenda"] == fz].copy(),
            fz,
            esperar_enter=False,
            ctx=dict(ctx_base),
        )
        if r:
            dias_faz = int(r.get("dias_simulado", 0))
            dia_inicio_acum = dias_acumulados + 1
            dias_acumulados += dias_faz
            r["dia_inicio_acumulado"] = dia_inicio_acum
            r["dia_fim_acumulado"] = dias_acumulados
            r["saldo_meta_apos"] = max(0, dias_meta - dias_acumulados)
            r["pct_meta_consumida"] = round(
                (dias_acumulados / dias_meta * 100) if dias_meta > 0 else 0.0, 1
            )
            if dias_acumulados > dias_meta:
                r["status_meta_continuo"] = "EXCEDIDO"
            elif dias_acumulados >= dias_meta * 0.8:
                r["status_meta_continuo"] = "RISCO"
            else:
                r["status_meta_continuo"] = "OK"

            hh_faz = float(r.get("total_hh", 0))
            rec = _recomendar_equipes_padrao(
                hh_faz, dias_meta, cap_ep_dia, jornada, prazo_absoluto
            )
            r["rec_ep"] = rec
            if rec and prazo_absoluto:
                _imprimir_recomendacao_ep(rec, fz, prazo_absoluto)

            if prazo_absoluto:
                st_lbl = r["status_meta_continuo"]
                cor_st = G if st_lbl == "OK" else (Y if st_lbl == "RISCO" else R)
                sub()
                print(cor_st + BL + f"  LOTE CONTINUO — apos '{fz}':" + RS)
                print(
                    cor_st + f"  Dia {dia_inicio_acum} a {dias_acumulados} | "
                    f"Saldo: {r['saldo_meta_apos']} dias | "
                    f"Consumo: {r['pct_meta_consumida']:.0f}% | "
                    f"Status: {st_lbl}" + RS
                )
            resultados.append(r)

    # ── Consolidated final report ──
    if not resultados:
        return
    linha()
    print(G + BL + "  CONSOLIDADO FINAL (TODAS AS FAZENDAS)" + RS)
    tit_cons = (
        f"Consolidado — {empresa_filtro}"
        if empresa_filtro
        else "Consolidado — todas as empresas (sem filtro EQUIPE)"
    )
    t_all = Table(title=tit_cons)
    t_all.add_column("Metrica", style="cyan")
    t_all.add_column("Valor", justify="right")
    t_all.add_row("Fazendas processadas", str(len(resultados)))
    t_all.add_row(
        "HH total (soma)",
        f"{sum(float(x.get('total_hh', 0)) for x in resultados):,.1f}",
    )
    dias_max_isolado = max(int(x.get("dias_simulado", 0)) for x in resultados)
    t_all.add_row("Dias simulados (maior fazenda isolada)", str(dias_max_isolado))
    t_all.add_row("Dias acumulados lote continuo", str(dias_acumulados))
    t_all.add_row("Meta (dias uteis)", str(dias_meta))
    if dias_meta > 0:
        saldo_final = max(0, dias_meta - dias_acumulados)
        st_final = "DENTRO" if dias_acumulados <= dias_meta else "EXCEDIDO"
        cor_final = "[green]" if st_final == "DENTRO" else "[red]"
        t_all.add_row(
            "Saldo apos todas as fazendas", f"{cor_final}{saldo_final} dias[/]"
        )
        t_all.add_row("Status meta global", f"{cor_final}{st_final}[/]")
    d_mec_vals = [
        int(x.get("dias_mecanizado") or 0)
        for x in resultados
        if x.get("dias_mecanizado")
    ]
    if d_mec_vals:
        t_all.add_row("Dias cenario mecanizado (max)", str(max(d_mec_vals)))
        t_all.add_row(
            "Ganho mecanizado total (dias)",
            f"{sum(int(x.get('ganho_mecanizado_dias', 0)) for x in resultados):+d}",
        )
    console.print(t_all)

    if prazo_absoluto:
        sub()
        print(G + BL + "  ANALISE EQUIPE PADRAO — CONSOLIDADO" + RS)
        ep_cap = sum(t["operarios"] for t in turmas)
        print(
            G
            + f"  Equipe padrao: {ep_cap} executores @ {jornada}h/dia = {cap_ep_dia:.1f} HH/dia"
            + RS
        )
        print(
            G + f"  Meta: {prazo_meses} meses = {dias_meta} dias uteis (ABSOLUTO)" + RS
        )

        t_ep = Table(title=f"Cascata de execucao — {tit_cons}")
        t_ep.add_column("Fazenda", style="cyan")
        t_ep.add_column("HH", justify="right")
        t_ep.add_column("Dias", justify="right")
        t_ep.add_column("Inicio", justify="right")
        t_ep.add_column("Fim", justify="right")
        t_ep.add_column("Meta consumida", justify="right")
        t_ep.add_column("Saldo", justify="right")
        t_ep.add_column("Status", justify="center")
        for r in resultados:
            pct = r.get("pct_meta_consumida", 0)
            st = r.get("status_meta_continuo", "?")
            if st == "OK":
                cor_st = "[green]"
            elif st == "RISCO":
                cor_st = "[yellow]"
            else:
                cor_st = "[red]"
            t_ep.add_row(
                str(r["fazenda"])[:28],
                f"{float(r.get('total_hh', 0)):,.1f}",
                str(r.get("dias_simulado", 0)),
                f"Dia {r.get('dia_inicio_acumulado', '?')}",
                f"Dia {r.get('dia_fim_acumulado', '?')}",
                f"{pct:.0f}%",
                f"{r.get('saldo_meta_apos', '?')} dias",
                f"{cor_st}{st}[/]",
            )
        hh_total_all = sum(float(x.get("total_hh", 0)) for x in resultados)
        st_global = "OK" if dias_acumulados <= dias_meta else "EXCEDIDO"
        cor_g = "[green]" if st_global == "OK" else "[red]"
        t_ep.add_row(
            "TOTAL",
            f"{hh_total_all:,.1f}",
            str(dias_acumulados),
            "Dia 1",
            f"Dia {dias_acumulados}",
            f"{(dias_acumulados / dias_meta * 100) if dias_meta > 0 else 0:.0f}%",
            f"{max(0, dias_meta - dias_acumulados)} dias",
            f"{cor_g}{st_global}[/]",
        )
        console.print(t_ep)

    _exportar_excel_consolidado_lote(
        resultados,
        empresa_filtro=empresa_filtro,
        nome_arquivo_micro=nome_arquivo_micro,
        extras={
            "Prazo_meses": prazo_meses,
            "Meta_absoluta": prazo_absoluto,
            "Modo_sequencia": modo_seq,
            "Jornada_h": jornada,
            "Executores_equipe_padrao": sum(t["operarios"] for t in turmas),
            "Preencher_orfas_auto": preencher_orfas_template,
            "Dias_meta": dias_meta,
            "Dias_acumulados_lote": dias_acumulados,
            "Data_inicio": data_inicio_txt,
            "Data_termino": data_fim_txt,
        },
    )

    linha()
    input(DM + "\n  [ENTER para voltar ao menu] " + RS)


# ──────────────────────────────────────────────
#  V6: MODO MULTI-EQUIPES
# ──────────────────────────────────────────────


def _executar_multi_equipes(
    cfg, df_scope, fazendas, empresa_filtro=None, nome_arquivo_micro=""
):
    """Modo avançado: N equipes independentes, cada uma com carteira de fazendas e meta própria."""
    dashboard_header()
    subcabecalho("MODO MULTI-EQUIPES")
    print(
        DM
        + "  Cada equipe tera sua propria configuracao, meta e carteira de fazendas."
        + RS
    )
    print(
        DM
        + "  Ao final, um consolidado comparativo mostra a situacao de cada equipe.\n"
        + RS
    )

    n_equipes = pedir_int("Quantas equipes independentes?", 2)
    if n_equipes < 1:
        aviso("Precisa de pelo menos 1 equipe.")
        return

    todas_atvs = sorted(
        {_norm_atv(x) for x in df_scope["atividade"].dropna().unique() if _norm_atv(x)},
        key=str,
    )

    seq_cfg = cfg.get("sequencia") or {}
    _merge_sequencia_defaults(seq_cfg)
    cfg["sequencia"] = seq_cfg
    modo_seq = _selecionar_sequencia_padrao_sn(cfg, seq_cfg)

    hoje = datetime.datetime.now()
    mes_ref = pedir_int("Mes inicial (1-12)", hoje.month)
    mes_ref = max(1, min(12, int(mes_ref)))
    ano_ref = pedir_int("Ano inicial", hoje.year)
    dia_max = calendar.monthrange(ano_ref, mes_ref)[1]
    dia_ref = pedir_int(f"Dia inicial (1-{dia_max})", min(hoje.day, dia_max))
    dia_ref = max(1, min(dia_max, int(dia_ref)))

    data_inicio_txt = _formatar_data_dia(dia_ref, mes_ref, ano_ref)

    # ──────────────────────────────────────────────
    # MODO TERRITÓRIO AUTOMÁTICO (V6 NOVO)
    # ──────────────────────────────────────────────
    usar_modo_territorio = False
    distribuicao_territorio = None
    config_territorio = None

    if confirmar(
        "Usar modo automatico de distribuicao por territorio/cidade?",
        default=False,
    ):
        dashboard_header()
        subcabecalho("DISTRIBUICAO POR TERRITORIO")
        print(DM + " Analisando fazendas e distribuindo por cidade..." + RS)

        distribuicao, nao_id = _distribuir_fazendas_por_territorio(fazendas)
        config_territorio = _sugerir_config_territorio(fazendas)

        # Mostrar distribuição
        print(G + BL + "\n Distribuicao detectada:" + RS)
        for cidade, fazs in distribuicao.items():
            if fazs:
                cfg_eq = _calcular_equipes_territorio(cidade)
                if cfg_eq:
                    print(
                        G
                        + f"  [{cfg_eq['nome_cidade']}]: "
                        + C
                        + f"{len(fazs)} fazenda(s), "
                        + f"{cfg_eq['n_equipes']} equipe(s) "
                        + f"({cfg_eq['total_operarios']} operarios)"
                        + RS
                    )
                    for f in fazs:
                        print(DM + f"      - {f}" + RS)

        if nao_id:
            print(Y + f"\n  Fazendas nao identificadas ({len(nao_id)}):" + RS)
            for f in nao_id[:5]:
                print(Y + f"      - {f}" + RS)
            if len(nao_id) > 5:
                print(Y + f"      ... e mais {len(nao_id) - 5}" + RS)

        print(G + BL + f"\n Total: {config_territorio['total_equipes']} equipes, " + f"{config_territorio['total_operarios']} operarios" + RS)

        if confirmar("Aceitar esta distribuicao automatica?", default=True):
            usar_modo_territorio = True
            n_equipes = config_territorio["total_equipes"]
            ok(f"Modo territorio ativado: {n_equipes} equipes automaticas.")
        else:
            aviso("Modo automatico cancelado. Prossiga com configuracao manual.")

        sub()
        input(DM + " [ENTER para continuar] " + RS)

    equipes_config = []
    fazendas_restantes = list(fazendas)

    for ie in range(1, n_equipes + 1):
        sub()
        print(G + BL + f" EQUIPE {ie}/{n_equipes}" + RS)

        # ──────────────────────────────────────────────
        # CONFIGURAÇÃO AUTOMÁTICA POR TERRITÓRIO (V6 NOVO)
        # ──────────────────────────────────────────────
        if usar_modo_territorio and config_territorio:
            # Encontrar a configuração de território para esta equipe
            cfg_territorio_eq = None
            equipe_idx_atual = ie - 1
            acum_equipes = 0

            for sug in config_territorio["sugestoes"]:
                n_eq_cidade = sug["n_equipes"]
                if equipe_idx_atual < acum_equipes + n_eq_cidade:
                    # Esta equipe pertence a esta cidade
                    cidade_eq = sug["cidade"]
                    n_cidade = equipe_idx_atual - acum_equipes + 1
                    nome_eq = f"{sug['nome_empresa']} {sug['nome_cidade']} Eq{n_cidade}"
                    j_eq = 4.3
                    exec_eq = sug["operarios_por_equipe"]
                    turmas_eq = [
                        {
                            "nome": f"{sug['nome_empresa']} {sug['nome_cidade']}",
                            "operarios": exec_eq,
                            "atividades": [],
                        }
                    ]
                    # Distribuir fazendas desta cidade entre as equipes dela
                    fazs_cidade = sug["fazendas"]
                    n_por_eq = max(1, len(fazs_cidade) // n_eq_cidade)
                    inicio_idx = n_cidade - 1
                    faz_eq = fazs_cidade[inicio_idx : inicio_idx + n_por_eq]

                    ok(f"Configuracao automatica: {nome_eq}")
                    print(G + f"  Cidade: {sug['nome_cidade']}" + RS)
                    print(G + f"  Empresa: {sug['nome_empresa']}" + RS)
                    print(G + f"  Operarios: {exec_eq}" + RS)
                    print(G + f"  Fazendas: {len(faz_eq)}" + RS)
                    cfg_territorio_eq = {
                        "nome": nome_eq,
                        "jornada": j_eq,
                        "executores": exec_eq,
                        "turmas": turmas_eq,
                        "fazendas": faz_eq,
                    }
                    break
                acum_equipes += n_eq_cidade

            if cfg_territorio_eq:
                nome_eq = cfg_territorio_eq["nome"]
                j_eq = cfg_territorio_eq["jornada"]
                exec_eq = cfg_territorio_eq["executores"]
                turmas_eq = cfg_territorio_eq["turmas"]
                faz_eq = cfg_territorio_eq["fazendas"]
                prazo_eq = pedir_float(f"Prazo meta para '{nome_eq}' (meses)", 3.0)

                data_fim_txt = None
                if confirmar(
                    f"Informar dia final manualmente para '{nome_eq}'?", default=False
                ):
                    mes_fim = pedir_int("Mes final (1-12)", mes_ref)
                    mes_fim = max(1, min(12, int(mes_fim)))
                    ano_fim = pedir_int("Ano final", ano_ref)
                    dia_max_fim = calendar.monthrange(ano_fim, mes_fim)[1]
                    dia_fim = pedir_int(
                        f"Dia final (1-{dia_max_fim})", min(dia_ref, dia_max_fim)
                    )
                    dia_fim = max(1, min(dia_max_fim, int(dia_fim)))
                    data_fim_txt = _formatar_data_dia(dia_fim, mes_fim, ano_fim)
                else:
                    fim_calc = _calcular_data_fim_por_meses(
                        dia_ref, mes_ref, ano_ref, prazo_eq
                    )
                    if fim_calc:
                        data_fim_txt = _formatar_data_dia(
                            fim_calc[0], fim_calc[1], fim_calc[2]
                        )
        else:
            # ──────────────────────────────────────────────
            # CONFIGURAÇÃO MANUAL (modo tradicional)
            # ──────────────────────────────────────────────
            nome_eq = prompt(f"Nome da equipe {ie}", f"Equipe {ie}")
            prazo_eq = pedir_float(f"Prazo meta para '{nome_eq}' (meses)", 3.0)
            j_eq = pedir_float(f"Jornada diaria '{nome_eq}' (horas)", 4.3)
            exec_eq = pedir_int(f"Executores '{nome_eq}'", 10)

            data_fim_txt = None
            if confirmar(
                f"Informar dia final manualmente para '{nome_eq}'?", default=False
            ):
                mes_fim = pedir_int("Mes final (1-12)", mes_ref)
                mes_fim = max(1, min(12, int(mes_fim)))
                ano_fim = pedir_int("Ano final", ano_ref)
                dia_max_fim = calendar.monthrange(ano_fim, mes_fim)[1]
                dia_fim = pedir_int(
                    f"Dia final (1-{dia_max_fim})", min(dia_ref, dia_max_fim)
                )
                dia_fim = max(1, min(dia_max_fim, int(dia_fim)))
                data_fim_txt = _formatar_data_dia(dia_fim, mes_fim, ano_fim)
            else:
                fim_calc = _calcular_data_fim_por_meses(dia_ref, mes_ref, ano_ref, prazo_eq)
                if fim_calc:
                    data_fim_txt = _formatar_data_dia(fim_calc[0], fim_calc[1], fim_calc[2])

            perfil_carregado = None
            perfis = _listar_perfis_equipe()
            if perfis and confirmar(
                f"Carregar perfil de equipe para '{nome_eq}'?", default=False
            ):
                perfil_carregado = _carregar_perfil_equipe_menu()

            if perfil_carregado:
                turmas_eq = [
                    {
                        "nome": t["nome"],
                        "operarios": t["operarios"],
                        "atividades": list(t.get("atividades") or []),
                    }
                    for t in perfil_carregado.get("turmas", [])
                ]
                exec_eq = sum(t["operarios"] for t in turmas_eq)
                ok(
                    f"Perfil carregado: {len(turmas_eq)} turma(s), {exec_eq} executores."
                )
            else:
                turmas_eq = [{"nome": nome_eq, "operarios": exec_eq, "atividades": []}]
                menu_vincular_atividades_turma(turmas_eq[0], todas_atvs)

        if not fazendas_restantes:
            aviso("Todas as fazendas ja foram atribuidas. Esta equipe ficara vazia.")
            faz_eq = []
        elif ie == n_equipes:
            faz_eq = list(fazendas_restantes)
            ok(f"Restantes ({len(faz_eq)}) atribuidas a '{nome_eq}'.")
        else:
            print(G + f"\n  Fazendas disponiveis ({len(fazendas_restantes)}):" + RS)
            for idx_f, f in enumerate(fazendas_restantes, 1):
                print(G + f"  {idx_f:3}. " + C + f + RS)
            sel_txt = prompt(
                f"Indices das fazendas para '{nome_eq}' (ex: 1,3,5-7) ou ENTER=todas restantes",
                "",
            )
            if not sel_txt.strip():
                faz_eq = list(fazendas_restantes)
            else:
                idxs = parse_intervalos_escolha(sel_txt, len(fazendas_restantes))
                faz_eq = [fazendas_restantes[i] for i in idxs]
            for f in faz_eq:
                if f in fazendas_restantes:
                    fazendas_restantes.remove(f)
            ok(f"{len(faz_eq)} fazenda(s) para '{nome_eq}'.")

        equipes_config.append(
            {
                "nome": nome_eq,
                "prazo_meses": prazo_eq,
                "jornada": j_eq,
                "executores": exec_eq,
                "turmas": turmas_eq,
                "fazendas": faz_eq,
                "modo_seq": modo_seq,
                "mes_ref": mes_ref,
                "ano_ref": ano_ref,
                "data_inicio_txt": data_inicio_txt,
                "data_fim_txt": data_fim_txt,
            }
        )

    all_eq_results = []
    for ec in equipes_config:
        linha()
        print(
            G
            + BL
            + f"  PROCESSANDO EQUIPE: {ec['nome']} ({len(ec['fazendas'])} fazendas)"
            + RS
        )
        linha()

        dias_meta_eq = dias_uteis_no_periodo(
            ec["mes_ref"], ec["ano_ref"], ec["prazo_meses"]
        )
        cap_eq_dia = float(ec["executores"]) * float(ec["jornada"])
        eq_resultados = []
        dias_acum_eq = 0

        ctx_eq = {
            "modo_seq": ec["modo_seq"],
            "usar_bloqueio_global": False,
            "usar_reforco_automatico": True,
            "usar_pool_pos_bloqueio": False,
            "prazo_meses": ec["prazo_meses"],
            "mes_ref": ec["mes_ref"],
            "ano_ref": ec["ano_ref"],
            "data_inicio_txt": ec.get("data_inicio_txt"),
            "data_fim_txt": ec.get("data_fim_txt"),
            "jornada": ec["jornada"],
            "executores": ec["executores"],
            "turmas": ec["turmas"],
            "penalidade": 1.0,
            "preencher_orfas_template": True,
        }
        contexto_sessao.atualizar_modo("multi_equipes")
        contexto_sessao.atualizar_equipe(ec["nome"])
        contexto_sessao.definir_datas(ec.get("data_inicio_txt"), ec.get("data_fim_txt"))
        _emitir_monitor_atual()
        _emitir_monitor_state(
            {
                "operacao": {
                    "modo": "multi_equipes",
                    "equipe_atual": str(ec["nome"]),
                    "status_geral": "processando_equipe",
                    "mensagem_curta": f"Equipe {ec['nome']} ({len(ec['fazendas'])} fazendas)",
                },
                "lote": {
                    "dias_meta": int(dias_meta_eq),
                    "dias_consumidos": int(dias_acum_eq),
                    "saldo_dias": int(max(0, int(dias_meta_eq) - int(dias_acum_eq))),
                    "fazenda_indice": 0,
                    "n_fazendas": int(len(ec["fazendas"])),
                    "status_meta_continuo": "OK",
                    "prazo_absoluto": True,
                },
            }
        )

        for fz in ec["fazendas"]:
            r = calcular_cronograma_inteligente(
                cfg,
                df_scope[df_scope["fazenda"] == fz].copy(),
                fz,
                esperar_enter=False,
                ctx=dict(ctx_eq),
            )
            if r:
                dias_faz = int(r.get("dias_simulado", 0))
                r["dia_inicio_acumulado"] = dias_acum_eq + 1
                dias_acum_eq += dias_faz
                r["dia_fim_acumulado"] = dias_acum_eq
                r["saldo_meta_apos"] = max(0, dias_meta_eq - dias_acum_eq)
                r["pct_meta_consumida"] = round(
                    (dias_acum_eq / dias_meta_eq * 100) if dias_meta_eq > 0 else 0.0, 1
                )
                r["status_meta_continuo"] = (
                    "EXCEDIDO"
                    if dias_acum_eq > dias_meta_eq
                    else ("RISCO" if dias_acum_eq >= dias_meta_eq * 0.8 else "OK")
                )
                eq_resultados.append(r)

        all_eq_results.append(
            {
                "equipe": ec["nome"],
                "executores": ec["executores"],
                "jornada": ec["jornada"],
                "prazo_meses": ec["prazo_meses"],
                "data_inicio_txt": ec.get("data_inicio_txt"),
                "data_fim_txt": ec.get("data_fim_txt"),
                "dias_meta": dias_meta_eq,
                "dias_acumulados": dias_acum_eq,
                "hh_total": sum(float(x.get("total_hh", 0)) for x in eq_resultados),
                "n_fazendas": len(ec["fazendas"]),
                "status": "DENTRO" if dias_acum_eq <= dias_meta_eq else "EXCEDIDO",
                "resultados_fazendas": eq_resultados,
            }
        )

    linha()
    print(G + BL + "  CONSOLIDADO MULTI-EQUIPES" + RS)
    t_meq = Table(title="Comparativo entre equipes")
    t_meq.add_column("Equipe", style="cyan")
    t_meq.add_column("Exec.", justify="right")
    t_meq.add_column("Fazendas", justify="right")
    t_meq.add_column("HH", justify="right")
    t_meq.add_column("Dias acum.", justify="right")
    t_meq.add_column("Meta (dias)", justify="right")
    t_meq.add_column("Saldo", justify="right")
    t_meq.add_column("Status", justify="center")
    for eq in all_eq_results:
        saldo = max(0, eq["dias_meta"] - eq["dias_acumulados"])
        st = eq["status"]
        cor = "[green]" if st == "DENTRO" else "[red]"
        t_meq.add_row(
            eq["equipe"],
            str(eq["executores"]),
            str(eq["n_fazendas"]),
            f"{eq['hh_total']:,.1f}",
            str(eq["dias_acumulados"]),
            str(eq["dias_meta"]),
            f"{saldo} dias",
            f"{cor}{st}[/]",
        )
    console.print(t_meq)

    try:
        pasta = OUTPUT_DIR
        os.makedirs(pasta, exist_ok=True)
        emp_slug = _slug_ficheiro_seguro(empresa_filtro) if empresa_filtro else "Todas"
        nome_xlsx = f"MultiEquipes_{emp_slug}.xlsx"
        caminho = os.path.join(pasta, nome_xlsx)
        rows_eq = []
        for eq in all_eq_results:
            for r in eq["resultados_fazendas"]:
                  row_eq = {
                      "Equipe": eq["equipe"],
                      "Data_inicio": eq.get("data_inicio_txt"),
                      "Data_termino": eq.get("data_fim_txt"),
                      "Fazenda": r.get("fazenda"),
                      "Dias": r.get("dias_simulado"),
                      "Dia_inicio_acum": r.get("dia_inicio_acumulado"),
                      "Dia_fim_acum": r.get("dia_fim_acumulado"),
                      "Meta_consumida_%": r.get("pct_meta_consumida"),
                      "Saldo": r.get("saldo_meta_apos"),
                      "Status": r.get("status_meta_continuo"),
                      "HH": r.get("total_hh"),
                  }
                  rows_eq.append(row_eq)
        rows_sumario = []
        for eq in all_eq_results:
            rows_sumario.append(
                {
                    "Equipe": eq["equipe"],
                    "Data_inicio": eq.get("data_inicio_txt"),
                    "Data_termino": eq.get("data_fim_txt"),
                    "Executores": eq["executores"],
                    "Jornada": eq["jornada"],
                    "Fazendas": eq["n_fazendas"],
                    "HH_total": eq["hh_total"],
                    "Dias_acumulados": eq["dias_acumulados"],
                    "Meta_dias": eq["dias_meta"],
                    "Status": eq["status"],
                }
            )
        with pd.ExcelWriter(caminho, engine="openpyxl") as w:
            pd.DataFrame(rows_sumario).to_excel(
                w, sheet_name="SUMARIO_EQUIPES", index=False
            )
            pd.DataFrame(rows_eq).to_excel(
                w, sheet_name="DETALHE_POR_FAZENDA", index=False
            )
        ok(f"Multi-equipes exportado: {nome_xlsx}")
    except Exception as ex:
        aviso(f"Erro ao exportar multi-equipes: {ex}")

    linha()
    input(DM + "\n  [ENTER para voltar ao menu] " + RS)


# ──────────────────────────────────────────────
#  MENU PRINCIPAL
# ──────────────────────────────────────────────
def _aplicar_filtro_empresa_e_escopo(df):
    """Filtro por EQUIPE (empresa) e escopo (uma fazenda / todas). Retorna (df_filtrado, empresa ou None)."""
    tem_equipe = "equipe" in df.columns
    df_filt = df.copy()
    empresa_filtro = None
    if tem_equipe:
        raw_eq = [
            str(x).strip() for x in df["equipe"].dropna().tolist() if str(x).strip()
        ]
        norm_to_raw = {}
        for e in raw_eq:
            nk = normalizar_chave(e)
            if nk and nk not in norm_to_raw:
                norm_to_raw[nk] = e
        equipes = sorted(norm_to_raw.values(), key=str)
        if equipes:
            print(G + BL + "\n  FILTRO POR EMPRESA (EQUIPE)" + RS)
            print(DM + f"  {len(equipes)} empresa(s) encontrada(s) no micro." + RS)
            equipes_disp = ["TODAS"] + equipes
            eq = selecionar("EMPRESA / EQUIPE", equipes_disp)
            if eq and eq != "TODAS":
                empresa_filtro = eq
                nk_sel = normalizar_chave(eq)
                sem_eq = df_filt["equipe"].isna() | (
                    df_filt["equipe"].astype(str).str.strip() == ""
                )
                n_sem = int(sem_eq.sum())
                if n_sem:
                    print(
                        DM
                        + f"  Excluindo {n_sem} linha(s) sem EQUIPE preenchida (nao entram no filtro por empresa)."
                        + RS
                    )
                df_filt = df_filt[~sem_eq]
                df_filt = df_filt[
                    df_filt["equipe"]
                    .astype(str)
                    .apply(lambda x: normalizar_chave(x.strip()) == nk_sel)
                ]
                ok(
                    f"Filtrado por equipe: {eq} ({len(df_filt)} registros, "
                    f"{df_filt['atividade'].nunique()} atividade(s), {df_filt['fazenda'].nunique()} fazenda(s))"
                )
    if empresa_filtro:
        contexto_sessao.atualizar_equipe(empresa_filtro)
    else:
        contexto_sessao.atualizar_equipe(None)
    _emitir_monitor_atual()
    return df_filt, empresa_filtro


def _selecionar_talhoes_fazenda(df_faz, fazenda):
    """Permite recorte por metodologia e talhao dentro da fazenda selecionada."""
    if df_faz is None or df_faz.empty:
        contexto_sessao.definir_escopo_talhoes([], [])
        return df_faz, {
            "fazenda": fazenda,
            "modo_metodologia": "vazio",
            "metodologias": [],
            "modo_talhao": "vazio",
            "talhoes": [],
        }
    if "chave" not in df_faz.columns:
        contexto_sessao.definir_escopo_talhoes([], [])
        return df_faz, {
            "fazenda": fazenda,
            "modo_metodologia": "sem_coluna",
            "metodologias": [],
            "modo_talhao": "sem_coluna",
            "talhoes": [],
        }

    df_work = df_faz.copy()
    meta = {
        "fazenda": fazenda,
        "modo_metodologia": "sem_coluna",
        "metodologias": [],
    }

    if "metodologia" in df_work.columns:
        metodologias = sorted(
            {
                str(x).strip()
                for x in df_work["metodologia"].dropna().tolist()
                if str(x).strip()
            },
            key=str,
        )
        print(G + BL + "\n  ESCOPO POR METODOLOGIA" + RS)
        print(DM + f"  Fazenda: {fazenda}" + RS)
        if metodologias:
            print(DM + f"  Metodologias presentes ({len(metodologias)}):" + RS)
            for i, m in enumerate(metodologias, 1):
                print(G + f"  [{i:2}] " + C + str(m) + RS)
        else:
            print(DM + "  Nenhuma metodologia preenchida no escopo atual." + RS)
        if not metodologias:
            meta["modo_metodologia"] = "sem_valores"
            meta["metodologias"] = []
        elif len(metodologias) == 1:
            meta["modo_metodologia"] = "unica"
            meta["metodologias"] = metodologias[:]
            ok(f"Metodologia unica no escopo: {metodologias[0]}")
        else:
            op_met = selecionar(
                "ESCOPO DAS METODOLOGIAS",
                [
                    "TODAS AS METODOLOGIAS",
                    "SELECIONAR METODOLOGIAS POR LISTA",
                    "SELECIONAR METODOLOGIAS POR NOME",
                    "FILTRAR METODOLOGIAS POR TEXTO",
                ],
            )

            selecionadas_met = []
            if not op_met or op_met == "TODAS AS METODOLOGIAS":
                meta["modo_metodologia"] = "todas"
                meta["metodologias"] = metodologias[:]
            elif op_met == "SELECIONAR METODOLOGIAS POR LISTA":
                print(
                    DM
                    + "  Digite indices separados por virgula (ex.: 1,2,4) ou intervalo (ex.: 1-3)."
                    + RS
                )
                for i, m in enumerate(metodologias, 1):
                    print(G + f"  [{i:2}] " + C + str(m) + RS)
                raw = prompt("Metodologias", "")
                idxs = parse_intervalos_escolha(raw, len(metodologias))
                selecionadas_met = [metodologias[i] for i in idxs]
            elif op_met == "SELECIONAR METODOLOGIAS POR NOME":
                raw = prompt(
                    "Nomes das metodologias (separados por virgula)",
                    "",
                )
                partes = [
                    normalizar_chave(p)
                    for p in str(raw).replace(";", ",").split(",")
                    if normalizar_chave(p)
                ]
                if partes:
                    for m in metodologias:
                        nm = normalizar_chave(m)
                        if any(p == nm or p in nm for p in partes):
                            selecionadas_met.append(m)
            else:
                filtro = normalizar_chave(prompt("Texto para filtrar metodologia", ""))
                if filtro:
                    selecionadas_met = [
                        m for m in metodologias if filtro in normalizar_chave(m)
                    ]

            if op_met and op_met != "TODAS AS METODOLOGIAS":
                selecionadas_met = sorted(set(selecionadas_met), key=str)
                if not selecionadas_met:
                    aviso(
                        "Nenhuma metodologia selecionada; mantendo TODAS as metodologias da fazenda."
                    )
                    meta["modo_metodologia"] = "fallback_todas"
                    meta["metodologias"] = metodologias[:]
                else:
                    sel_norm = {normalizar_chave(x) for x in selecionadas_met}
                    df_filtrado = df_work[
                        df_work["metodologia"].astype(str).apply(
                            lambda x: normalizar_chave(str(x).strip()) in sel_norm
                        )
                    ].copy()
                    if df_filtrado.empty:
                        aviso(
                            "Filtro de metodologia nao retornou linhas; mantendo TODAS as metodologias da fazenda."
                        )
                        meta["modo_metodologia"] = "fallback_todas"
                        meta["metodologias"] = metodologias[:]
                    else:
                        df_work = df_filtrado
                        meta["modo_metodologia"] = "parcial"
                        meta["metodologias"] = selecionadas_met
                        ok(
                            f"Escopo por metodologia aplicado: {len(selecionadas_met)} selecionada(s), "
                            f"{len(df_work)} linha(s) no micro."
                        )

    talhoes = sorted(
        {
            str(x).strip()
            for x in df_work["chave"].dropna().tolist()
            if str(x).strip()
        },
        key=str,
    )
    if not talhoes:
        contexto_sessao.definir_escopo_talhoes([], [])
        out_meta = dict(meta)
        out_meta.update({"modo_talhao": "sem_talhoes", "talhoes": []})
        _emitir_monitor_atual()
        return df_work, out_meta
    if len(talhoes) == 1:
        ok(f"Talhao unico na fazenda: {talhoes[0]}")
        contexto_sessao.definir_escopo_talhoes(talhoes[:], talhoes[:])
        out_meta = dict(meta)
        out_meta.update({"modo_talhao": "unico", "talhoes": talhoes[:]})
        _emitir_monitor_atual()
        return df_work, out_meta

    print(G + BL + "\n  ESCOPO POR TALHAO" + RS)
    print(DM + f"  Fazenda: {fazenda}" + RS)
    print(DM + f"  {len(talhoes)} talhao(oes) disponivel(is)." + RS)
    op = selecionar(
        "ESCOPO DOS TALHOES",
        [
            "TODOS OS TALHOES",
            "SELECIONAR TALHOES POR LISTA",
            "FILTRAR TALHOES POR TEXTO",
        ],
    )
    if not op or op == "TODOS OS TALHOES":
        contexto_sessao.definir_escopo_talhoes(talhoes[:], talhoes[:])
        out_meta = dict(meta)
        out_meta.update({"modo_talhao": "todos", "talhoes": talhoes[:]})
        _emitir_monitor_atual()
        return df_work, out_meta

    selecionados = []
    if op == "SELECIONAR TALHOES POR LISTA":
        print(
            DM
            + "  Digite numeros separados por virgula (ex.: 1,3,7) ou intervalo (ex.: 1-4)."
            + RS
        )
        for i, t in enumerate(talhoes, 1):
            print(G + f"  [{i:2}] " + C + str(t) + RS)
        raw = prompt("Talhoes", "")
        idxs = parse_intervalos_escolha(raw, len(talhoes))
        selecionados = [talhoes[i] for i in idxs]
    else:
        filtro = normalizar_chave(prompt("Texto para filtrar talhoes", ""))
        if filtro:
            selecionados = [t for t in talhoes if filtro in normalizar_chave(t)]

    if not selecionados:
        aviso("Nenhum talhao selecionado; mantendo TODOS os talhoes da fazenda.")
        contexto_sessao.definir_escopo_talhoes(talhoes[:], talhoes[:])
        out_meta = dict(meta)
        out_meta.update({"modo_talhao": "fallback_todos", "talhoes": talhoes[:]})
        _emitir_monitor_atual()
        return df_work, out_meta

    df_sel = df_work[df_work["chave"].astype(str).isin(set(selecionados))].copy()
    ok(
        f"Escopo por talhao aplicado: {len(selecionados)} selecionado(s), {len(df_sel)} linha(s) no micro."
    )
    contexto_sessao.definir_escopo_talhoes(selecionados, talhoes[:])
    out_meta = dict(meta)
    out_meta.update({"modo_talhao": "parcial", "talhoes": selecionados})
    _emitir_monitor_atual()
    return df_sel, out_meta


def _metodologias_presentes(df_faz):
    if df_faz is None or df_faz.empty or "metodologia" not in df_faz.columns:
        return []
    return sorted(
        {
            str(x).strip()
            for x in df_faz["metodologia"].dropna().tolist()
            if str(x).strip()
        },
        key=str,
    )


def _prompt_proximas_metodologias(df_faz_base, meta_escopo, metodologias_executadas):
    """
    Mostra prompt para continuar com metodologias restantes na mesma fazenda.
    Retorna True quando o usuario deseja seguir para as proximas metodologias.
    """
    todas = _metodologias_presentes(df_faz_base)
    if not todas:
        return False

    meta_escopo = meta_escopo or {}
    usadas = list(meta_escopo.get("metodologias") or [])
    for m in usadas:
        nm = normalizar_chave(m)
        if nm:
            metodologias_executadas.add(nm)

    restantes = [m for m in todas if normalizar_chave(m) not in metodologias_executadas]
    if not restantes:
        return False

    sub()
    print(C + BL + "  PROXIMAS METODOLOGIAS DISPONIVEIS" + RS)
    print(DM + f"  Restantes: {len(restantes)}" + RS)
    for i, m in enumerate(restantes[:12], 1):
        print(G + f"  [{i:2}] " + C + str(m) + RS)
    if len(restantes) > 12:
        print(DM + f"  ... +{len(restantes) - 12} metodologia(s)" + RS)
    return confirmar(
        "Executar outra metodologia desta fazenda agora?",
        default=True,
    )


def _executar_scheduler_fazenda_interativo(cfg, df_scope, faz, catalogo_scope):
    metodologias_executadas = set()
    while True:
        df_faz_base = df_scope[df_scope["fazenda"] == faz].copy()
        if df_faz_base.empty:
            aviso("Sem linhas para a fazenda selecionada neste escopo.")
            return

        contexto_sessao.atualizar_fazenda(faz, df_faz_base)
        df_faz, meta_escopo = _selecionar_talhoes_fazenda(df_faz_base, faz)
        resultado = calcular_cronograma_inteligente(
            cfg,
            df_faz,
            faz,
            escopo_meta=meta_escopo,
            atividades_catalogo=catalogo_scope,
        )
        if isinstance(resultado, dict) and resultado.get("acao") == "retroceder_escopo":
            aviso("Checkpoint retroativo: reselecione fazenda/escopo.")
            continue

        if _prompt_proximas_metodologias(
            df_faz_base,
            meta_escopo,
            metodologias_executadas,
        ):
            aviso("Abrindo selecao para as proximas metodologias desta fazenda.")
            continue
        break


def _menu_ajustar_escopo_atividades(df_faz, cfg=None, atividades_catalogo=None):
    """
    Ajustes por execucao no escopo atual:
      - substituir atividade
      - remover atividade
      - adicionar atividade em talhao(es)
    """
    if df_faz is None or df_faz.empty:
        return df_faz

    if "atividade" not in df_faz.columns or "chave" not in df_faz.columns:
        aviso("Escopo sem colunas necessarias para ajuste de atividade.")
        return df_faz

    out = df_faz.copy()
    if "area_ha" not in out.columns:
        out["area_ha"] = 0.0
    if "penalidade" not in out.columns:
        out["penalidade"] = 1.0

    def _atividades():
        return sorted(
            {
                str(x).strip()
                for x in out["atividade"].dropna().tolist()
                if str(x).strip()
            },
            key=str,
        )

    def _talhoes():
        return sorted(
            {str(x).strip() for x in out["chave"].dropna().tolist() if str(x).strip()},
            key=str,
        )

    while True:
        atvs = _atividades()
        catalogo_all = _catalogo_atividades_completo(
            atvs,
            cfg=cfg,
            atividades_catalogo=atividades_catalogo,
        )
        tls = _talhoes()
        sub()
        print(G + BL + "  AJUSTE DE ATIVIDADES (APENAS NESTA EXECUCAO)" + RS)
        print(
            DM
            + f"  Atividades no escopo: {len(atvs)} | Catalogo completo: {len(catalogo_all)} | Talhoes no escopo: {len(tls)}"
            + RS
        )
        op = selecionar(
            "OPERACAO DE AJUSTE",
            [
                "Substituir atividade",
                "Remover atividade",
                "Adicionar atividade",
                "Ver listas completas (escopo x catalogo)",
                "Concluir ajustes",
            ],
        )
        if not op or op == "Concluir ajustes":
            break

        if op == "Ver listas completas (escopo x catalogo)":
            _mostrar_catalogo_atividades(atvs, catalogo_all)
            input(DM + "\n  [ENTER] " + RS)
            continue

        if op == "Substituir atividade":
            if not atvs:
                aviso("Sem atividades para substituir.")
                continue
            print(DM + " Origens (ex: 1,3,5-8 ou ENTER para selecionar uma)" + RS)
            for i, a in enumerate(atvs, 1):
                print(G + f" [{i:2}] " + C + str(a)[:70] + RS)
            raw = prompt("Origens", "")
            if str(raw).strip():
                idxs = parse_intervalos_escolha(raw, len(atvs))
                if not idxs:
                    aviso("Nenhum indice valido.")
                    continue
                srcs = [atvs[i] for i in idxs]
            else:
                src = selecionar("ATIVIDADE ORIGEM", atvs)
                if not src:
                    continue
                srcs = [src]
            destinos = [x for x in catalogo_all if x not in srcs]
            dst_opt = selecionar(
                "DESTINO", destinos + ["[DIGITAR NOVA ATIVIDADE]"]
            )
            if not dst_opt:
                continue
            dst = (
                prompt("Nova atividade", "")
                if dst_opt == "[DIGITAR NOVA ATIVIDADE]"
                else dst_opt
            )
            dst = _norm_atv(dst)
            if not dst:
                aviso("Destino invalido.")
                continue
            n_sub = 0
            for src in srcs:
                mask = out["atividade"].astype(str) == str(src)
                n_sub += int(mask.sum())
                out.loc[mask, "atividade"] = dst
            ok(f"{len(srcs)} atividade(s) substituida(s) -> '{dst}' ({n_sub} linhas).")
            continue

        if op == "Remover atividade":
            if not atvs:
                aviso("Sem atividades para remover.")
                continue
            print(DM + " Remover (ex: 1,3,5-8 ou ENTER para selecionar uma)" + RS)
            for i, a in enumerate(atvs, 1):
                print(G + f" [{i:2}] " + C + str(a)[:70] + RS)
            raw = prompt("Remover", "")
            if str(raw).strip():
                idxs = parse_intervalos_escolha(raw, len(atvs))
                if not idxs:
                    aviso("Nenhum indice valido.")
                    continue
                rms = [atvs[i] for i in idxs]
            else:
                rm = selecionar("ATIVIDADE PARA REMOVER", atvs)
                if not rm:
                    continue
                rms = [rm]
            n0 = len(out)
            mask = out["atividade"].astype(str).isin([str(r) for r in rms])
            out = out[~mask].copy()
            ok(f"{len(rms)} atividade(s) removida(s) ({n0 - len(out)} linha(s)).")
            continue

        if op == "Adicionar atividade":
            print(DM + " Atividades (ex: 1,3,5-8 ou ENTER para selecionar uma)" + RS)
            for i, a in enumerate(catalogo_all, 1):
                print(G + f" [{i:2}] " + C + str(a)[:70] + RS)
            print(G + f" [{len(catalogo_all)+1:2}] " + C + "[DIGITAR NOVA ATIVIDADE]" + RS)
            raw = prompt("Atividades", "")
            if str(raw).strip():
                idxs = parse_intervalos_escolha(raw, len(catalogo_all) + 1)
                novas = []
                for i in idxs:
                    if i < len(catalogo_all):
                        novas.append(catalogo_all[i])
                    else:
                        t = prompt("Nome da atividade", "")
                        t = _norm_atv(t)
                        if t:
                            novas.append(t)
                if not novas:
                    aviso("Nenhuma atividade valida.")
                    continue
            else:
                base_opt = selecionar(
                    "NOVA ATIVIDADE",
                    catalogo_all + ["[DIGITAR NOVA ATIVIDADE]"],
                )
                if not base_opt:
                    continue
                nova = (
                    prompt("Nome da atividade", "")
                    if base_opt == "[DIGITAR NOVA ATIVIDADE]"
                    else base_opt
                )
                nova = _norm_atv(nova)
                if not nova:
                    aviso("Atividade invalida.")
                    continue
                novas = [nova]
            op_t = selecionar(
                "APLICAR EM",
                [
                    "Todos os talhoes do escopo",
                    "Talhoes por lista",
                    "Talhoes por texto",
                ],
            )
            if not op_t:
                continue
            sel_talhoes = tls[:]
            if op_t == "Talhoes por lista":
                print(DM + "  Digite numeros separados por virgula (ex.: 1,3,7)." + RS)
                for i, t in enumerate(tls, 1):
                    print(G + f"  [{i:2}] " + C + str(t) + RS)
                raw = prompt("Talhoes", "")
                idxs = []
                for p in str(raw).replace(";", ",").split(","):
                    p = p.strip()
                    if p.isdigit():
                        iv = int(p)
                        if 1 <= iv <= len(tls):
                            idxs.append(iv - 1)
                idxs = sorted(set(idxs))
                sel_talhoes = [tls[i] for i in idxs]
            elif op_t == "Talhoes por texto":
                fx = normalizar_chave(prompt("Texto no talhao", ""))
                sel_talhoes = [t for t in tls if fx and fx in normalizar_chave(t)]

            if not sel_talhoes:
                aviso("Nenhum talhao selecionado para adicionar atividade.")
                continue

        area_def = float(out["area_ha"].median() or 1.0)
        area_nova = pedir_float(
            "Area/ha para nova atividade (por talhao)", round(area_def, 2),
            allow_zero=True,
        )
        pen_def = float(out["penalidade"].median() or 1.0)
        pen_nova = pedir_float(
            "Penalidade de terreno da nova atividade", round(pen_def, 2),
            allow_zero=False,
        )

        add_rows = []
        for nova in novas:
            for th in sel_talhoes:
                ja = out[
                    (out["chave"].astype(str) == str(th))
                    & (out["atividade"].astype(str) == str(nova))
                ]
                if not ja.empty:
                    continue
                ref = out[out["chave"].astype(str) == str(th)].head(1)
                row = ref.iloc[0].to_dict() if not ref.empty else {}
                row["chave"] = th
                row["atividade"] = nova
                row["area_ha"] = float(area_nova)
                row["penalidade"] = float(pen_nova)
                add_rows.append(row)
        if add_rows:
            out = pd.concat([out, pd.DataFrame(add_rows)], ignore_index=True)
            ok(f"{len(novas)} atividade(s) adicionada(s) em {len(add_rows)} linha(s).")

    return out


def _proximo_caminho_livre(pasta, nome_arquivo):
    """
    Retorna nome/caminho sem colisao:
    - se nao existe, usa o nome original
    - se existe, gera sufixo _v2, _v3...
    """
    base, ext = os.path.splitext(str(nome_arquivo))
    candidato_nome = str(nome_arquivo)
    candidato_path = os.path.join(pasta, candidato_nome)
    if not os.path.exists(candidato_path):
        return candidato_nome, candidato_path
    i = 2
    while True:
        candidato_nome = f"{base}_v{i}{ext}"
        candidato_path = os.path.join(pasta, candidato_nome)
        if not os.path.exists(candidato_path):
            return candidato_nome, candidato_path
        i += 1


def menu_principal(cfg, df, nome_arquivo_micro="", demo_mode=False):
    opcoes = [
        ("1", "Smart Scheduler (Operacional HH/HM)"),
        ("2", "Importar Tarifas (CT real/manual)"),
        ("3", "Normalizar CT317/CT -> STG (auto)"),
        ("4", "Mapeamentos de_para (micro -> tarifa)"),
        ("5", "Trocar planilha de microplanejamento (.xlsx)"),
        ("6", "Fazendas micro vs CT (lista fazendas_ct)"),
        ("M", "Abrir Monitor em Janela Separada"),
        ("0", "Sair"),
    ]
    while True:
        dashboard_header()
        subcabecalho()
        contexto_sessao.atualizar_configuracoes(cfg)
        nf = df["fazenda"].nunique()
        nu = df["chave"].nunique()
        na = df["atividade"].nunique()
        stg_existe = os.path.exists(os.path.join(INPUT_DIR, STG_FILENAME))
        nt = len(cfg.get("tarifas", {}))
        print(
            G
            + f"  Base: "
            + C
            + f"{nf} fazendas  |  {nu} talhoes  |  {na} atividades"
            + RS
        )
        print(
            G
            + f"  Tarifas: "
            + C
            + f"{nt} carregadas"
            + G
            + f"  |  STG: "
            + C
            + f"{'Sim' if stg_existe else 'Nao'}"
            + RS
        )
        print(
            G
            + f"  Orcamento estrito: "
            + C
            + ("Sim" if cfg.get("orcamento_estrito", True) else "Nao")
            + RS
        )
        if "equipe" in df.columns:
            eq_list = sorted(df["equipe"].dropna().unique().tolist(), key=str)
            print(
                G
                + f"  Empresas (EQUIPE): "
                + C
                + f"{len(eq_list)} ({', '.join(str(e)[:20] for e in eq_list[:5])}{'...' if len(eq_list) > 5 else ''})"
                + RS
            )
        if nome_arquivo_micro:
            print(
                G
                + f"  Microplanejamento: "
                + C
                + os.path.basename(nome_arquivo_micro)
                + RS
            )
        if demo_mode and _is_demo_micro_path(nome_arquivo_micro):
            print(
                Y
                + f"  DEMO: opcao [1] = maior fazenda do micro (municipio Ulianopolis), tarifas = CT 313."
                + RS
            )
        sub()
        for cod, desc in opcoes:
            print(G + f"  [{cod}] " + C + desc + RS)
        sub()
        v = prompt("Opcao").strip()
        if v == "1":
            contexto_sessao.atualizar_modo("single")
            if demo_mode and _is_demo_micro_path(nome_arquivo_micro):
                faz = _resolver_fazenda_demo_ulianopolis(df)
                if not faz:
                    aviso(
                        "DEMO: nenhuma fazenda com 'Ulianópolis' na coluna fazenda do micro."
                    )
                else:
                    ok(f"DEMO: fazenda {faz}")
                    df_faz = df[df["fazenda"] == faz].copy()
                    contexto_sessao.atualizar_fazenda(faz, df_faz)
                    resultado = calcular_cronograma_inteligente(
                        cfg,
                        df_faz,
                        faz,
                        atividades_catalogo=sorted(
                            {
                                str(x).strip()
                                for x in df["atividade"].dropna().unique().tolist()
                                if str(x).strip()
                            },
                            key=str,
                        ),
                    )
                    if isinstance(resultado, dict) and resultado.get("acao") == "retroceder_escopo":
                        aviso("Retornando ao seletor de fazenda/escopo.")
            else:
                df_scope, empresa_filtro = _aplicar_filtro_empresa_e_escopo(df)
                if df_scope is None or df_scope.empty:
                    aviso("Nenhum dado apos filtros.")
                    continue
                catalogo_scope = sorted(
                    {
                        str(x).strip()
                        for x in df_scope["atividade"].dropna().unique().tolist()
                        if str(x).strip()
                    },
                    key=str,
                )
                fazendas = sorted(df_scope["fazenda"].unique().tolist())
                if len(fazendas) == 1:
                    faz = fazendas[0]
                    ok(f"Fazenda unica no escopo: {faz}")
                    _executar_scheduler_fazenda_interativo(
                        cfg,
                        df_scope,
                        faz,
                        catalogo_scope,
                    )
                else:
                    op_faz = [
                        "TODAS AS FAZENDAS (equipe unica)",
                        "MULTI-EQUIPES (carteiras separadas)",
                    ] + fazendas
                    faz = selecionar("SELECIONE A FAZENDA OU MODO", op_faz)
                    if faz == "TODAS AS FAZENDAS (equipe unica)":
                        contexto_sessao.atualizar_modo("lote")
                        _executar_lote_fazendas(
                            cfg,
                            df_scope,
                            fazendas,
                            empresa_filtro=empresa_filtro,
                            nome_arquivo_micro=nome_arquivo_micro,
                        )
                    elif faz == "MULTI-EQUIPES (carteiras separadas)":
                        contexto_sessao.atualizar_modo("multi_equipes")
                        _executar_multi_equipes(
                            cfg,
                            df_scope,
                            fazendas,
                            empresa_filtro=empresa_filtro,
                            nome_arquivo_micro=nome_arquivo_micro,
                        )
                    elif faz:
                        _executar_scheduler_fazenda_interativo(
                            cfg,
                            df_scope,
                            faz,
                            catalogo_scope,
                        )
        elif v == "2":
            modulo_importar_tarifas(cfg)
        elif v == "3":
            modulo_normalizar_ct(cfg)
        elif v == "4":
            modulo_mapeamentos_de_para(cfg, df)
        elif v == "5":
            p = selecionar_arquivo("NOVO MICROPLANEJAMENTO (.xlsx)")
            if p:
                ndf = carregar_planilha_microplanejamento(
                    cfg, caminho=p, modo_auto=True
                )
                if ndf is None:
                    aviso(
                        "Nao foi possivel carregar automaticamente. Tente de novo sem modo_auto (o app pedira colunas)."
                    )
                    ndf = carregar_planilha_microplanejamento(
                        cfg, caminho=p, modo_auto=False
                    )
                if ndf is not None:
                    df = ndf
                    nome_arquivo_micro = p
                    cfg["arquivo_micro"] = os.path.basename(p)
                    salvar_config(cfg)
                    atividades_reais = sorted(
                        str(x).strip()
                        for x in df["atividade"].dropna().unique()
                        if str(x).strip()
                    )
                    novos = aplicar_depara_padrao_exame(cfg, atividades_reais)
                    ok(
                        f"Micro atualizado: {os.path.basename(p)} | {len(df)} registros | "
                        f"{df['fazenda'].nunique()} fazendas | de_para +{novos} novos mapeamentos."
                    )
                    if demo_mode and _is_demo_micro_path(p):
                        n = garantir_fazenda_ulianopolis_no_ct(cfg, df)
                        if n:
                            salvar_config(cfg)
                            ok(f"DEMO: +{n} fazenda(s) em fazendas_ct.")
                    aviso_fazendas_micro_sem_cadastro_ct(cfg, df)
                    input(DM + "  [ENTER] " + RS)
        elif v == "6":
            modulo_validar_fazendas_ct(cfg, df)
        elif v.upper() == "M":
            dashboard_header()
            subcabecalho("ABRIR MONITOR EXTERNO")
            print(G + BL + " Tipo de Feed:" + RS)
            print(DM + " [1] meta - Operacao e metas" + RS)
            print(DM + " [2] rendimentos - HH/ha por atividade" + RS)
            print(DM + " [3] relatorios - Buffer de relatorios" + RS)
            sub()
            feed_op = prompt("Opcao", "1").strip()
            feed_map = {"1": "meta", "2": "rendimentos", "3": "relatorios"}
            feed_escolhido = feed_map.get(feed_op, "meta")
            ok(f"Abrindo monitor com feed '{feed_escolhido}'...")
            _abrir_monitor_janela(feed=feed_escolhido)
            input(DM + "\n [ENTER para voltar] " + RS)
        elif v == "0":
            print(G + "\n Sistema encerrado.\n" + RS)
            break
        else:
            aviso("Opcao invalida.")


def main():
    demo = _is_demo_mode()
    beta = _is_beta_mode()
    legacy = _is_legacy_mode()
    sub_titulo = (
        f"DEMO Ulianópolis ({DEMO_MICRO_SOURCE_FILENAME} -> {DEMO_MICRO_FILENAME} + CT317)"
        if demo
        else ""
    )
    if MODO_SOMENTE_HH:
        sub_titulo = (sub_titulo + " | " if sub_titulo else "") + "MODO SOMENTE HH"
    if beta:
        if sub_titulo:
            sub_titulo += " | PADRAO"
        else:
            sub_titulo = "PADRAO - carga robusta micro + comparativos operacionais"
    if legacy:
        if sub_titulo:
            sub_titulo += " | LEGACY"
        else:
            sub_titulo = "LEGACY - comportamento anterior sem comparativos padrao"
    cabecalho(sub_titulo)
    print(DM + "  Inicializando sistema...\n" + RS)
    cfg = carregar_config()
    salvar_config(cfg)

    if demo:
        rebuilt = reconstruir_demo_ulianopolis_a_partir_da_fonte()
        if rebuilt:
            ok(
                f"DEMO: {DEMO_MICRO_FILENAME} atualizado a partir de {DEMO_MICRO_SOURCE_FILENAME} "
                f"({rebuilt[0]} linhas, {rebuilt[1]} atividades unicas)."
            )
        micro_padrao = os.path.join(INPUT_DIR, DEMO_MICRO_FILENAME)
        if not os.path.exists(micro_padrao):
            erro(
                f"Modo DEMO: coloque {DEMO_MICRO_SOURCE_FILENAME} (gera {DEMO_MICRO_FILENAME}) "
                f"ou o proprio {DEMO_MICRO_FILENAME} em:\n {INPUT_DIR}"
            )
            sys.exit(1)
    else:
        micro_padrao = _find_default_micro_path(cfg)
    ct_padrao = _find_default_ct_path()
    if ct_padrao:
        try:
            stg_path, n, custo_h = normalizar_ct313(ct_padrao)
            if stg_path and n > 0:
                cfg["tarifas"] = carregar_stg_tarifas(stg_path)
                cfg["custo_hora_tf"] = round(custo_h, 4)
                salvar_config(cfg)
                ok(
                    f"CT auto: {os.path.basename(ct_padrao)} -> {n} atividades (modo operacional)"
                )
        except Exception as ex:
            aviso(f"Falha no auto-carregamento CT: {ex}")

    if micro_padrao:
        df = carregar_planilha_microplanejamento(
            cfg, caminho=micro_padrao, modo_auto=True
        )
        if df is None:
            aviso("Falha no auto-carregamento do micro padrao; abrindo modo manual.")
            df = carregar_planilha_microplanejamento(cfg)
    else:
        df = carregar_planilha_microplanejamento(cfg)

    if df is not None:
        if micro_padrao:
            cfg["arquivo_micro"] = os.path.basename(micro_padrao)
            salvar_config(cfg)
        atividades_reais = sorted(
            str(x).strip() for x in df["atividade"].dropna().unique() if str(x).strip()
        )
        novos = aplicar_depara_padrao_exame(cfg, atividades_reais)
        if demo and micro_padrao and _is_demo_micro_path(micro_padrao):
            n = garantir_fazenda_ulianopolis_no_ct(cfg, df)
            if n:
                salvar_config(cfg)
                ok(f"DEMO: +{n} fazenda(s) em fazendas_ct.")
        aviso_fazendas_micro_sem_cadastro_ct(cfg, df)
        dp = {
            k: v
            for k, v in cfg.get("de_para", {}).items()
            if not str(k).startswith("_")
        }
        ok(
            f"{len(df)} registros | "
            f"{df['fazenda'].nunique()} fazendas | {df['chave'].nunique()} talhoes | "
            f"{len(dp)} de_para mapeados ({novos} novos)"
        )
        input(DM + "  [ENTER para continuar] " + RS)
        menu_principal(cfg, df, micro_padrao or "", demo_mode=demo)
    else:
        aviso("Nenhuma planilha selecionada.")


def _cleanup_estado_sessao():
    """Remove arquivos estado_sessao_*.json antigos ao sair."""
    try:
        import glob
        pattern = os.path.join(DIR, "estado_sessao_*.json")
        files = glob.glob(pattern)
        removed = 0
        for f in files:
            try:
                os.remove(f)
                removed += 1
            except Exception:
                pass
        if removed > 0:
            print(DM + f" [Limpo: {removed} arquivo(s) de estado removidos]" + RS)
    except Exception:
        pass


# Registrar cleanup ao sair
atexit.register(_cleanup_estado_sessao)


if __name__ == "__main__":
    main()
