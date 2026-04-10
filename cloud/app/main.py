import json
import os
import unicodedata
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from fastapi import FastAPI, WebSocket, WebSocketDisconnect, Request, UploadFile, File, Form, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

from .auth import parse_users, verify_password, create_token, decode_token
from .session import SessionManager
from .storage import Storage


APP_DIR = Path(__file__).resolve().parent
CLOUD_DIR = APP_DIR.parent
REPO_DIR = Path(os.environ.get("SRF_BASE_DIR", str(CLOUD_DIR.parent))).resolve()
SESSIONS_DIR = Path(os.environ.get("SRF_SESSIONS_DIR", str(CLOUD_DIR / "sessions"))).resolve()
DEFAULT_MODE = os.environ.get("SRF_DEFAULT_MODE", "standard").strip().lower()


# region agent log
def _dbg_log(hypothesis_id: str, location: str, message: str, data: dict):
    try:
        payload = {
            "sessionId": "09cd54",
            "runId": "run1",
            "hypothesisId": hypothesis_id,
            "location": location,
            "message": message,
            "data": data,
            "timestamp": int(time.time() * 1000),
        }
        with open("debug-09cd54.log", "a", encoding="utf-8") as f:
            f.write(json.dumps(payload, ensure_ascii=False) + "\n")
    except Exception:
        pass
# endregion

_dbg_log(
    "H1_static_path_missing",
    "app/main.py:init",
    "Startup path snapshot",
    {
        "APP_DIR": str(APP_DIR),
        "CLOUD_DIR": str(CLOUD_DIR),
        "REPO_DIR": str(REPO_DIR),
        "frozen": bool(getattr(__import__("sys"), "frozen", False)),
        "meipass": str(getattr(__import__("sys"), "_MEIPASS", "")),
        "app_static_exists": (APP_DIR / "static").is_dir(),
    },
)

app = FastAPI(title="SRF Cloud Pilot")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://127.0.0.1:8000", "http://localhost:8000", "http://127.0.0.1:8001", "http://127.0.0.1:8002", "http://localhost:8001", "http://localhost:8002"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
templates = Jinja2Templates(directory=str(APP_DIR / "templates"))
_dbg_log(
    "H1_static_path_missing",
    "app/main.py:mount_static",
    "Before static mount",
    {"static_dir": str(APP_DIR / "static"), "exists": (APP_DIR / "static").is_dir()},
)
app.mount("/static", StaticFiles(directory=str(APP_DIR / "static")), name="static")

users = parse_users()
storage = Storage()
manager = SessionManager(str(REPO_DIR), str(SESSIONS_DIR))


def _is_bundled_sample_rel(rel_posix: str) -> bool:
    """Pastas copiadas do repo para o workspace (exemplos); não contam como upload do utilizador."""
    parts = rel_posix.replace("\\", "/").strip("/").split("/")
    if not parts or parts[0] in (".",):
        return False
    return parts[0].lower() in ("testes", "tutorial")


def _semantic_progress_from_events(s) -> dict[str, Any]:
    events = list(getattr(s, "semantic_events", []) or [])
    if not events:
        return {"progress_pct": 0, "stage": "idle", "events": []}
    last = events[-1]
    return {
        "progress_pct": int(last.get("progress") or 0),
        "stage": str(last.get("stage") or "running"),
        "events": events[-25:],
    }


def _append_events_from_log(s, chunks: list[str]) -> None:
    if not chunks:
        return
    txt = "".join(chunks[-40:]).lower()
    markers = [
        ("upload_received", ["upload", "xlsx", "micro atualizado"], 20, "Entrada recebida"),
        ("sequence_selected", ["sequencia", "selecionar sequencia padrao"], 30, "Sequencia definida"),
        ("team_config", ["configuracao do projeto", "turmas criadas"], 40, "Equipe configurada"),
        ("scheduler_running", ["gerando cronograma", "smart scheduler"], 60, "Scheduler em execucao"),
        ("audit_scope", ["auditoria do escopo", "nao agendadas"], 75, "Auditoria de escopo"),
        ("dossier_export", ["dossier financeiro exportado", "dossier operacional exportado"], 90, "Dossier gerado"),
    ]
    exists = {str(e.get("stage")) for e in (getattr(s, "semantic_events", []) or [])}
    for stage, keys, pct, detail in markers:
        if stage in exists:
            continue
        if any(k in txt for k in keys):
            s.add_event(stage, detail, progress=pct)


def get_user_from_token(token: str):
    # Auth temporariamente silenciado para simplificar uso local zero-config
    return "admin"
    # payload = decode_token(token)
    # if not payload:
    #     return None
    # return payload.get("sub")


@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


@app.get("/ui", response_class=HTMLResponse)
def ui_prototype():
    """Interface visual (protótipo HTML de referência — ver aparencia/prototipos/)."""
    p = REPO_DIR / "aparencia" / "prototipos" / "app_srf_5_telas.html"
    if not p.is_file():
        raise HTTPException(status_code=404, detail="UI prototype not found")
    return HTMLResponse(p.read_text(encoding="utf-8"))


@app.get("/api/health")
def api_health():
    return {"ok": True}


@app.post("/api/auto_session")
async def api_auto_session():
    """Zero-config: cria sessão sem necessidade de login separado (uso local)."""
    username = "admin"
    token = create_token(username)
    mode = DEFAULT_MODE if DEFAULT_MODE in ("standard", "legacy") else "standard"
    s = manager.create_session(username=username, mode=mode)
    s.start()
    return {"token": token, "session_id": s.session_id, "mode": mode}


@app.get("/api/sessions/{session_id}/ping")
def api_session_ping(session_id: str):
    """Verifica existência rápida da sessão (sem auth, para autobootstrap)."""
    s = manager.get(session_id)
    if not s:
        raise HTTPException(status_code=404, detail="Sessao expirada ou inexistente")
    return {"ok": True, "session_id": session_id, "running": s.running}


# --- API UI (sessão + workspace real) ---


def _workspace_summary(root: Path) -> dict:
    """Deriva contagens a partir do disco (sem motor Python ainda)."""
    xlsx_root = [
        p
        for p in root.rglob("*.xlsx")
        if p.is_file() and not _is_bundled_sample_rel(p.relative_to(root).as_posix())
    ]
    doss = root / "dossiês"
    doss_xlsx = []
    if doss.is_dir():
        doss_xlsx = [p for p in doss.glob("*.xlsx") if p.is_file()]
    cfg = root / "config.json"
    micro_label = root.name
    return {
        "workspace_ready": True,
        "micro_nome": micro_label,
        "ficheiros_xlsx_total": len(xlsx_root),
        "relatorios_em_dossies": len(doss_xlsx),
        "tem_config_json": cfg.is_file(),
        "fazendas": len(xlsx_root) if xlsx_root else 0,
        "equipes": 0,
        "atividades": 0,
    }


def _norm_txt(v: Any) -> str:
    s = str(v or "").strip().lower()
    s = "".join(ch for ch in unicodedata.normalize("NFD", s) if unicodedata.category(ch) != "Mn")
    return " ".join("".join(ch if ch.isalnum() else " " for ch in s).split())


def _header_pick_activity(norm: list[str]) -> int | None:
    for i, h in enumerate(norm):
        if any(
            k in h
            for k in (
                "atividade",
                "atividades",
                "servico",
                "descricao",
                "tarefa",
                "operacao",
            )
        ):
            return i
    return None


def _header_pick_farm(norm: list[str]) -> int | None:
    for i, h in enumerate(norm):
        if h in ("fazenda", "nome fazenda", "codigo fazenda", "codigo da fazenda",
                 "propriedade", "propriedades") or (
            "fazenda" in h and "nome" in h
        ):
            return i
    return None


def _header_pick_plot(norm: list[str]) -> int | None:
    for i, h in enumerate(norm):
        if h in ("talhao", "chave", "codigo talhao", "gleba", "poligono", "chave poligono"):
            return i
        if "talhao" in h:
            return i
        if "chave" in h and ("polig" in h or "talhao" in h or "polygon" in h):
            return i
    return None


def _extract_catalog_from_xlsx(path: Path) -> dict:
    out = {"farms": set(), "plots": set(), "activities": set()}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(path), read_only=True, data_only=True)
    except Exception:
        return out
    try:
        for sn in wb.sheetnames[:12]:
            ws = wb[sn]
            rows = ws.iter_rows(min_row=1, max_row=min(int(ws.max_row or 1), 5000), values_only=True)
            first = next(rows, None)
            if not first:
                continue
            headers = [str(x or "").strip() for x in first]
            norm = [_norm_txt(h) for h in headers]
            idx_farm = _header_pick_farm(norm)
            idx_plot = _header_pick_plot(norm)
            idx_atv = _header_pick_activity(norm)
            if idx_farm is None:
                idx_farm = next((i for i, h in enumerate(norm) if h in ("fazenda", "codigo fazenda")), None)
            if idx_plot is None:
                idx_plot = next((i for i, h in enumerate(norm) if h in ("chave", "codigo talhao")), None)
            if idx_atv is None:
                idx_atv = next((i for i, h in enumerate(norm) if "atividade" in h), None)
            if idx_farm is None and idx_plot is None and idx_atv is None:
                continue
            for r in rows:
                if idx_farm is not None and idx_farm < len(r):
                    v = str(r[idx_farm] or "").strip()
                    if v and len(v) < 500:
                        out["farms"].add(v)
                if idx_plot is not None and idx_plot < len(r):
                    v = str(r[idx_plot] or "").strip()
                    if v and len(v) < 500:
                        out["plots"].add(v)
                if idx_atv is not None and idx_atv < len(r):
                    v = str(r[idx_atv] or "").strip()
                    if v and len(v) < 500:
                        out["activities"].add(v)
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return out


def _xlsx_for_catalog(root: Path) -> list[Path]:
    """Prioriza ficheiros na raiz e fora de testes/tutorial; inclui mais ficheiros até preencher catálogo."""
    all_xlsx = [p for p in root.rglob("*.xlsx") if p.is_file() and "dossi" not in str(p).lower()]
    user_files: list[Path] = []
    sample_files: list[Path] = []
    for p in all_xlsx:
        rel = p.relative_to(root).as_posix()
        (sample_files if _is_bundled_sample_rel(rel) else user_files).append(p)

    def sort_key(p: Path) -> tuple[int, str]:
        rel = p.relative_to(root).as_posix()
        depth = rel.count("/")
        return (depth, rel.lower())

    user_files.sort(key=sort_key)
    sample_files.sort(key=sort_key)
    return user_files + sample_files


@app.get("/api/micro/summary")
def api_micro_summary(
    session_id: str | None = Query(None, description="Sessão do pilot"),
    token: str | None = Query(None, description="JWT do login"),
):
    """Resumo derivado do workspace da sessão; sem sessão devolve workspace_ready=false."""
    if not session_id or not token:
        return {
            "workspace_ready": False,
            "message": "Passe session_id e token (query) após login no pilot.",
        }
    username = get_user_from_token(token)
    if not username:
        raise HTTPException(status_code=401, detail="Token invalido")
    s = manager.get(session_id)
    if not s or s.username != username:
        raise HTTPException(status_code=404, detail="Sessao nao encontrada")
    root = Path(s.workspace_dir)
    return _workspace_summary(root)


@app.get("/api/sessions/{session_id}/execution_catalog")
def api_execution_catalog(session_id: str, token: str = Query(...)):
    """Catálogo mínimo (fazendas/talhões/atividades) para wizard da UI."""
    username = get_user_from_token(token)
    if not username:
        raise HTTPException(status_code=401, detail="Token invalido")
    s = manager.get(session_id)
    if not s or s.username != username:
        raise HTTPException(status_code=404, detail="Sessao nao encontrada")
    root = Path(s.workspace_dir)
    ordered = _xlsx_for_catalog(root)
    farms, plots, activities = set(), set(), set()
    # Escanear ficheiros do utilizador primeiro (até 32); só depois exemplos se catálogo ainda vazio.
    for p in ordered[:32]:
        cat = _extract_catalog_from_xlsx(p)
        farms |= cat["farms"]
        plots |= cat["plots"]
        activities |= cat["activities"]
    return {
        "farms": sorted(farms, key=str)[:500],
        "plots": sorted(plots, key=str)[:2000],
        "activities": sorted(activities, key=str)[:2000],
        "sources_scanned": min(len(ordered), 32),
    }


@app.get("/api/sessions/{session_id}/reports")
def api_session_reports(session_id: str, token: str = Query(...)):
    """Ficheiros .xlsx em dossiês/ (relatórios gerados)."""
    username = get_user_from_token(token)
    if not username:
        raise HTTPException(status_code=401, detail="Token invalido")
    s = manager.get(session_id)
    if not s or s.username != username:
        raise HTTPException(status_code=404, detail="Sessao nao encontrada")
    root = Path(s.workspace_dir)
    doss = root / "dossiês"
    out = []
    if doss.is_dir():
        for p in sorted(doss.glob("*.xlsx")):
            if p.is_file():
                st = p.stat()
                meta_p = p.with_suffix(p.suffix + ".meta.json")
                meta = {}
                if meta_p.is_file():
                    try:
                        meta = json.loads(meta_p.read_text(encoding="utf-8"))
                    except Exception:
                        meta = {}
                out.append(
                    {
                        "nome": p.name,
                        "path": p.relative_to(root).as_posix(),
                        "tamanho_bytes": st.st_size,
                        "mtime_iso": datetime.fromtimestamp(st.st_mtime, tz=timezone.utc).isoformat(),
                        "metadata": meta,
                    }
                )
    return {"reports": out}


@app.post("/api/sessions/{session_id}/report_metadata")
async def api_save_report_metadata(session_id: str, request: Request, token: str = Query(...)):
    """Associa metadados operacionais a um dossier exportado."""
    username = get_user_from_token(token)
    if not username:
        raise HTTPException(status_code=401, detail="Token invalido")
    s = manager.get(session_id)
    if not s or s.username != username:
        raise HTTPException(status_code=404, detail="Sessao nao encontrada")
    data = await request.json()
    path = str((data or {}).get("path") or "").strip()
    metadata = dict((data or {}).get("metadata") or {})
    if not path:
        raise HTTPException(status_code=400, detail="Path obrigatorio")
    root = Path(s.workspace_dir).resolve()
    p = (root / path).resolve()
    if not str(p).startswith(str(root)):
        raise HTTPException(status_code=400, detail="Path invalido")
    if not p.exists() or not p.is_file():
        raise HTTPException(status_code=404, detail="Arquivo nao encontrado")
    meta_p = p.with_suffix(p.suffix + ".meta.json")
    payload = {
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "execution_config": getattr(s, "execution_config", {}) or {},
        "notes_count": len(getattr(s, "execution_notes", []) or []),
        **metadata,
    }
    meta_p.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    s.add_event("report_metadata_saved", f"Metadados salvos para {p.name}", progress=92)
    return {"ok": True, "metadata": payload}


@app.get("/api/sessions/{session_id}/report_preview")
def api_session_report_preview(session_id: str, token: str = Query(...), path: str = Query(...)):
    """Metadados básicos do XLSX para leitura rápida no app (sem carregar tudo em memória)."""
    username = get_user_from_token(token)
    if not username:
        raise HTTPException(status_code=401, detail="Token invalido")
    s = manager.get(session_id)
    if not s or s.username != username:
        raise HTTPException(status_code=404, detail="Sessao nao encontrada")

    root = Path(s.workspace_dir).resolve()
    p = (root / path).resolve()
    if not str(p).startswith(str(root)):
        raise HTTPException(status_code=400, detail="Path invalido")
    if not p.exists() or not p.is_file():
        raise HTTPException(status_code=404, detail="Arquivo nao encontrado")
    if p.suffix.lower() != ".xlsx":
        raise HTTPException(status_code=400, detail="Apenas .xlsx suportado")

    try:
        from openpyxl import load_workbook
    except Exception as ex:
        raise HTTPException(status_code=500, detail=f"openpyxl indisponivel: {ex}")

    wb = load_workbook(str(p), read_only=True, data_only=True)
    sheets = []
    total_rows = 0
    total_cols = 0
    for name in wb.sheetnames:
        ws = wb[name]
        mr = int(ws.max_row or 0)
        mc = int(ws.max_column or 0)
        total_rows += mr
        total_cols = max(total_cols, mc)
        sheets.append({"name": name, "rows": mr, "cols": mc})
    wb.close()

    sheets_sorted = sorted(sheets, key=lambda x: (x["rows"], x["cols"]), reverse=True)
    top_sheets = sheets_sorted[:8]
    return {
        "file_name": p.name,
        "path": path,
        "sheet_count": len(sheets),
        "total_rows": total_rows,
        "max_cols": total_cols,
        "top_sheets": top_sheets,
        "all_sheets": sheets,
    }


@app.get("/api/sessions/{session_id}/status")
def api_session_status(session_id: str, token: str = Query(...)):
    """Estado real da sessão CLI: running, done, error, idle."""
    username = get_user_from_token(token)
    if not username:
        raise HTTPException(status_code=401, detail="Token invalido")
    s = manager.get(session_id)
    if not s or s.username != username:
        raise HTTPException(status_code=404, detail="Sessao nao encontrada")

    root = Path(s.workspace_dir)
    doss = root / "dossiês"
    n_dossies = len(list(doss.glob("*.xlsx"))) if doss.is_dir() else 0
    n_files = len(
        [
            p
            for p in root.rglob("*.xlsx")
            if p.is_file() and not _is_bundled_sample_rel(p.relative_to(root).as_posix())
        ]
    )

    if s.running:
        status = "running"
    elif s.proc and s.proc.poll() is not None:
        status = "done" if s.proc.returncode == 0 else "error"
    else:
        status = "idle"

    logs = s.read_output_non_block()
    _append_events_from_log(s, logs)
    recent_log = "".join(logs[-20:]) if logs else ""

    progress = 0
    if status == "done":
        progress = 100
    elif n_dossies > 0:
        progress = 85
    elif n_files > 0:
        progress = 50
    elif s.running:
        progress = 15

    sem = _semantic_progress_from_events(s)
    progress = max(progress, int(sem.get("progress_pct") or 0))
    return {
        "session_id": session_id,
        "status": status,
        "progress_pct": progress,
        "stage": sem.get("stage"),
        "semantic_events": sem.get("events", []),
        "files_count": n_files,
        "dossies_count": n_dossies,
        "recent_log": recent_log[-2000:] if recent_log else "",
        "running": s.running,
        "execution_config": getattr(s, "execution_config", {}) or {},
    }


@app.get("/api/sessions/{session_id}/report_analysis")
def api_report_analysis(session_id: str, token: str = Query(...), path: str = Query(...)):
    """Análise completa de um dossier: parsing + insights + Ollama opcional."""
    username = get_user_from_token(token)
    if not username:
        raise HTTPException(status_code=401, detail="Token invalido")
    s = manager.get(session_id)
    if not s or s.username != username:
        raise HTTPException(status_code=404, detail="Sessao nao encontrada")

    root = Path(s.workspace_dir).resolve()
    p = (root / path).resolve()
    if not str(p).startswith(str(root)):
        raise HTTPException(status_code=400, detail="Path invalido")
    if not p.exists() or not p.is_file():
        raise HTTPException(status_code=404, detail="Arquivo nao encontrado")

    from .report_parser import parse_dossier
    from .rules_engine import analyze
    from . import ollama_bridge

    parsed = parse_dossier(str(p))
    insights = analyze(parsed)

    ollama_summary = None
    ollama_status = ollama_bridge.status()
    if ollama_bridge.is_available():
        try:
            ollama_summary = ollama_bridge.analyze_report(parsed, insights)
        except Exception:
            pass

    return {
        "parsed": parsed,
        "insights": insights,
        "ollama_summary": ollama_summary,
        "ollama_status": ollama_status,
    }


@app.get("/api/ollama/status")
def api_ollama_status():
    """Verifica se o adaptador Ollama está ativo e acessível."""
    from . import ollama_bridge
    return ollama_bridge.status()


@app.post("/api/login")
async def api_login(request: Request):
    data = await request.json()
    username = str(data.get("username", "")).strip() or "admin"
    password = str(data.get("password", "")).strip()
    
    # Auth silenciado para modo local zero-config
    # if not verify_password(username, password, users):
    #     raise HTTPException(status_code=401, detail="Credenciais invalidas")
        
    token = create_token(username)
    return {"token": token, "username": username}


@app.post("/api/sessions")
async def api_create_session(token: str = Form(...), mode: str = Form(DEFAULT_MODE)):
    username = get_user_from_token(token)
    if not username:
        raise HTTPException(status_code=401, detail="Token invalido")
    mode = "legacy" if str(mode).strip().lower() == "legacy" else "standard"
    s = manager.create_session(username=username, mode=mode)
    s.start()
    return {"session_id": s.session_id, "mode": mode}


@app.post("/api/sessions/{session_id}/preflight")
async def api_preflight_execution(session_id: str, request: Request, token: str = Query(...)):
    """Valida contrato de execução da UI sem iniciar processamento."""
    username = get_user_from_token(token)
    if not username:
        raise HTTPException(status_code=401, detail="Token invalido")
    s = manager.get(session_id)
    if not s or s.username != username:
        raise HTTPException(status_code=404, detail="Sessao nao encontrada")
    data = await request.json()
    cfg = dict(data or {})
    issues = []
    warnings = []

    scope = cfg.get("scope") or {}
    teams = cfg.get("teams") or []
    params = cfg.get("params") or {}
    compare = cfg.get("comparative") or {}

    if not scope.get("farm"):
        issues.append({"field": "scope.farm", "message": "Fazenda obrigatoria."})
    if scope.get("mode") in ("single_plot", "multi_plot") and not (scope.get("plots") or []):
        issues.append({"field": "scope.plots", "message": "Selecione ao menos um talhao para esse modo."})
    if not teams:
        issues.append({"field": "teams", "message": "Informe pelo menos uma turma."})
    else:
        for i, t in enumerate(teams):
            if int(t.get("workers") or 0) <= 0:
                issues.append({"field": f"teams[{i}].workers", "message": "Operarios deve ser > 0."})
    if float(params.get("jornada_h_dia") or 0) <= 0:
        issues.append({"field": "params.jornada_h_dia", "message": "Jornada deve ser > 0."})
    if compare.get("enabled"):
        if not (compare.get("jornadas") or []):
            warnings.append({"field": "comparative.jornadas", "message": "Sem jornadas; sera usado valor base."})
        if not (compare.get("equipes") or []):
            warnings.append({"field": "comparative.equipes", "message": "Sem equipes; sera usado valor base."})

    return {
        "ok": len(issues) == 0,
        "issues": issues,
        "warnings": warnings,
        "normalized": {
            "scope_mode": scope.get("mode") or "farm_all",
            "teams_count": len(teams),
            "comparative_enabled": bool(compare.get("enabled")),
        },
    }


@app.post("/api/sessions/{session_id}/execution_config")
async def api_set_execution_config(session_id: str, request: Request, token: str = Query(...)):
    """Persistencia leve da configuracao de execucao da UI na sessao atual."""
    username = get_user_from_token(token)
    if not username:
        raise HTTPException(status_code=401, detail="Token invalido")
    s = manager.get(session_id)
    if not s or s.username != username:
        raise HTTPException(status_code=404, detail="Sessao nao encontrada")
    data = await request.json()
    s.execution_config = dict(data or {})
    s.add_event("execution_config_saved", "Configuracao da UI salva na sessao", progress=10)
    return {"ok": True, "execution_config": s.execution_config}


@app.get("/api/sessions/{session_id}/execution_config")
def api_get_execution_config(session_id: str, token: str = Query(...)):
    username = get_user_from_token(token)
    if not username:
        raise HTTPException(status_code=401, detail="Token invalido")
    s = manager.get(session_id)
    if not s or s.username != username:
        raise HTTPException(status_code=404, detail="Sessao nao encontrada")
    return {"execution_config": getattr(s, "execution_config", {}) or {}}


@app.post("/api/sessions/{session_id}/notes")
async def api_add_execution_note(session_id: str, request: Request, token: str = Query(...)):
    username = get_user_from_token(token)
    if not username:
        raise HTTPException(status_code=401, detail="Token invalido")
    s = manager.get(session_id)
    if not s or s.username != username:
        raise HTTPException(status_code=404, detail="Sessao nao encontrada")
    data = await request.json()
    txt = str((data or {}).get("text") or "").strip()
    if not txt:
        raise HTTPException(status_code=400, detail="Nota vazia")
    note = {"ts": datetime.now(timezone.utc).isoformat(), "text": txt}
    s.execution_notes.append(note)
    if len(s.execution_notes) > 200:
        s.execution_notes = s.execution_notes[-200:]
    s.add_event("note_added", "Nota operacional registrada", progress=15)
    return {"ok": True, "note": note}


@app.get("/api/sessions/{session_id}/notes")
def api_list_execution_notes(session_id: str, token: str = Query(...)):
    username = get_user_from_token(token)
    if not username:
        raise HTTPException(status_code=401, detail="Token invalido")
    s = manager.get(session_id)
    if not s or s.username != username:
        raise HTTPException(status_code=404, detail="Sessao nao encontrada")
    return {"notes": list(getattr(s, "execution_notes", []) or [])}


@app.post("/api/sessions/{session_id}/upload")
async def api_upload(session_id: str, token: str = Form(...), file: UploadFile = File(...)):
    username = get_user_from_token(token)
    if not username:
        raise HTTPException(status_code=401, detail="Token invalido")
    s = manager.get(session_id)
    if not s or s.username != username:
        raise HTTPException(status_code=404, detail="Sessao nao encontrada")

    target = Path(s.workspace_dir) / Path(file.filename).name
    content = await file.read()
    target.write_bytes(content)
    if storage.enabled_blob():
        storage.upload_if_enabled(str(target), f"{s.username}/{s.session_id}/uploads/{target.name}")
    return {"ok": True, "filename": target.name}


@app.delete("/api/sessions/{session_id}/workspace_file")
def api_delete_workspace_file(session_id: str, token: str = Query(...), path: str = Query(...)):
    """Remove um ficheiro dentro do workspace da sessão (ex.: .xlsx enviado por engano)."""
    username = get_user_from_token(token)
    if not username:
        raise HTTPException(status_code=401, detail="Token invalido")
    s = manager.get(session_id)
    if not s or s.username != username:
        raise HTTPException(status_code=404, detail="Sessao nao encontrada")
    rel = str(path or "").strip().replace("\\", "/")
    if not rel or ".." in rel or rel.startswith("/"):
        raise HTTPException(status_code=400, detail="Path invalido")
    if _is_bundled_sample_rel(rel):
        raise HTTPException(status_code=403, detail="Nao pode remover ficheiros de exemplo (testes/tutorial)")
    root = Path(s.workspace_dir).resolve()
    p = (root / rel).resolve()
    if not str(p).startswith(str(root)):
        raise HTTPException(status_code=400, detail="Path invalido")
    if not p.is_file():
        raise HTTPException(status_code=404, detail="Arquivo nao encontrado")
    try:
        p.unlink()
    except OSError as e:
        raise HTTPException(status_code=500, detail=str(e)) from e
    s.add_event("workspace_file_removed", f"Removido do workspace: {p.name}", progress=4, level="info")
    return {"ok": True, "removed": rel}


@app.get("/api/sessions/{session_id}/files")
def api_list_files(session_id: str, token: str):
    username = get_user_from_token(token)
    if not username:
        raise HTTPException(status_code=401, detail="Token invalido")
    s = manager.get(session_id)
    if not s or s.username != username:
        raise HTTPException(status_code=404, detail="Sessao nao encontrada")
    root = Path(s.workspace_dir)
    out = []
    out_all = []
    for p in sorted(root.rglob("*")):
        if p.is_file():
            rel = p.relative_to(root).as_posix()
            out_all.append(rel)
            if not _is_bundled_sample_rel(rel):
                out.append(rel)
    return {"files": out, "files_include_samples": out_all}


@app.get("/api/sessions/{session_id}/download")
def api_download(session_id: str, token: str, path: str):
    username = get_user_from_token(token)
    if not username:
        raise HTTPException(status_code=401, detail="Token invalido")
    s = manager.get(session_id)
    if not s or s.username != username:
        raise HTTPException(status_code=404, detail="Sessao nao encontrada")
    root = Path(s.workspace_dir).resolve()
    p = (root / path).resolve()
    if not str(p).startswith(str(root)):
        raise HTTPException(status_code=400, detail="Path invalido")
    if not p.exists() or not p.is_file():
        raise HTTPException(status_code=404, detail="Arquivo nao encontrado")
    return FileResponse(str(p), filename=p.name)


@app.websocket("/ws/{session_id}")
async def ws_session(websocket: WebSocket, session_id: str, token: str):
    username = get_user_from_token(token)
    if not username:
        await websocket.close(code=4401)
        return
    s = manager.get(session_id)
    if not s or s.username != username:
        await websocket.close(code=4404)
        return
    await websocket.accept()
    try:
        while True:
            for chunk in s.read_output_non_block():
                await websocket.send_text(chunk)
            try:
                msg = await websocket.receive_text()
                if msg == "__PING__":
                    await websocket.send_text("")
                    continue
                s.send_input(msg)
            except WebSocketDisconnect:
                break
    finally:
        # sessao pode continuar viva para download/retomada curta
        pass

