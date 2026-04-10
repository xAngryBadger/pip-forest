"""
Formatacao executiva do Dossier Excel (openpyxl).
Cabecalhos navy, zebra no cronograma, aba DASHBOARD, GANTT simples semanal.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F2F47", end_color="1F2F47", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ZEBRA_A = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
ZEBRA_B = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
THIN = Side(style="thin", color="CCCCCC")


def _style_header_row(ws, row_idx=1):
    for cell in ws[row_idx]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _zebra_data_rows(ws, header_row=1):
    max_row = ws.max_row
    for r in range(header_row + 1, max_row + 1):
        fill = ZEBRA_A if (r - header_row) % 2 == 1 else ZEBRA_B
        for cell in ws[r]:
            cell.fill = fill


def _auto_width(ws, max_w=48):
    for col_idx, col in enumerate(ws.columns, 1):
        m = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_w, m + 2)


def format_standard_sheet(ws, zebra=True):
    _style_header_row(ws, 1)
    if zebra:
        _zebra_data_rows(ws, 1)
    ws.freeze_panes = "A2"
    _auto_width(ws)
    if ws.max_row >= 1 and ws.dimensions:
        ws.auto_filter.ref = ws.dimensions


def _highlight_cascata_explicada(ws):
    """
    Destaque visual para leitura didatica:
    - RESUMO_DIA em azul claro com fonte bold
    - HH_Pendente_Atividade > 0 em amarelo
    - Fechou_Dia = N em laranja suave
    """
    if ws.max_row < 2:
        return
    header = [str(c.value or "").strip() for c in ws[1]]
    idx_tipo = header.index("Tipo_Linha") + 1 if "Tipo_Linha" in header else None
    idx_pend = header.index("HH_Pendente_Atividade") + 1 if "HH_Pendente_Atividade" in header else None
    idx_fechou = header.index("Fechou_Dia") + 1 if "Fechou_Dia" in header else None
    if not (idx_tipo or idx_pend or idx_fechou):
        return

    fill_resumo = PatternFill(start_color="DCEBFF", end_color="DCEBFF", fill_type="solid")
    fill_pend = PatternFill(start_color="FFF3B0", end_color="FFF3B0", fill_type="solid")
    fill_aberto = PatternFill(start_color="FFD9B3", end_color="FFD9B3", fill_type="solid")
    bold = Font(bold=True)

    for r in range(2, ws.max_row + 1):
        is_resumo = False
        if idx_tipo:
            v_tipo = str(ws.cell(row=r, column=idx_tipo).value or "").strip().upper()
            is_resumo = v_tipo == "RESUMO_DIA"
        if is_resumo:
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = fill_resumo
                cell.font = bold
            continue

        if idx_pend:
            try:
                v_pend = float(ws.cell(row=r, column=idx_pend).value or 0)
            except (TypeError, ValueError):
                v_pend = 0.0
            if v_pend > 0.01:
                ws.cell(row=r, column=idx_pend).fill = fill_pend
                ws.cell(row=r, column=idx_pend).font = bold

        if idx_fechou:
            v_f = str(ws.cell(row=r, column=idx_fechou).value or "").strip().upper()
            if v_f == "N":
                ws.cell(row=r, column=idx_fechou).fill = fill_aberto
                ws.cell(row=r, column=idx_fechou).font = bold


def add_dashboard_sheet(wb, titulo_fazenda, lucro, margem_pct, dias_simulado, receita, custo_mo_total):
    """Aba DASHBOARD com KPIs em destaque."""
    if "DASHBOARD" in wb.sheetnames:
        del wb["DASHBOARD"]
    ws = wb.create_sheet("DASHBOARD", 0)
    ws.merge_cells("A1:F3")
    c = ws["A1"]
    c.value = titulo_fazenda
    c.font = Font(size=18, bold=True, color="1F2F47")
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A5:B5")
    ws["A5"] = "KPIs (simulacao)"
    ws["A5"].font = Font(size=14, bold=True, color="1F2F47")

    rows = [
        ("Lucro estimado (R$)", f"R$ {lucro:,.2f}"),
        ("Margem bruta (%)", f"{margem_pct:.1f} %"),
        ("Duracao simulada (dias)", f"{dias_simulado}"),
        ("Receita bruta orcada (R$)", f"R$ {receita:,.2f}"),
        ("Custo MO total (R$)", f"R$ {custo_mo_total:,.2f}"),
    ]
    r0 = 6
    for i, (lab, val) in enumerate(rows):
        r = r0 + i * 2
        ws.merge_cells(f"A{r}:B{r}")
        cell_l = ws[f"A{r}"]
        cell_l.value = lab
        cell_l.font = Font(size=11, bold=True)
        ws.merge_cells(f"C{r}:F{r+1}")
        cell_v = ws[f"C{r}"]
        cell_v.value = val
        cell_v.font = Font(size=20 if i < 3 else 14, bold=True, color="1F2F47")
        cell_v.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(f"A{r0 + len(rows)*2 + 1}:F{r0 + len(rows)*2 + 2}")
    note = ws[f"A{r0 + len(rows)*2 + 1}"]
    note.value = (
        "Nota: 'Data final' civil ainda nao aplicada — dias sao sequenciais da simulacao. "
        "Proxima versao: calendario com feriados."
    )
    note.font = Font(size=9, italic=True, color="666666")
    note.alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["C"].width = 40


def add_gantt_simple_sheet(wb, cronograma_rows, dias_simulado):
    """
    Gantt semanal: linhas = Talhao, colunas = Semana 1..N, celula = HH (cor por intensidade).
    cronograma_rows: lista de dicts com Dia, Talhao, HH
    """
    if not cronograma_rows:
        return
    if "GANTT_SIMPLES" in wb.sheetnames:
        del wb["GANTT_SIMPLES"]
    from collections import defaultdict

    sem_max = max(1, int((dias_simulado + 4) // 5))
    agg = defaultdict(lambda: defaultdict(float))
    for row in cronograma_rows:
        try:
            d = int(row.get("Dia", 0))
            th = str(row.get("Talhao", ""))
            hh = float(row.get("HH", 0))
        except (TypeError, ValueError):
            continue
        sem = max(1, (d + 4) // 5)
        if sem > sem_max:
            sem_max = sem
        agg[th][sem] += hh

    talhoes = sorted(agg.keys(), key=str)
    ws = wb.create_sheet("GANTT_SIMPLES")
    ws.append(["Talhao"] + [f"Sem {s}" for s in range(1, sem_max + 1)])
    _style_header_row(ws, 1)

    max_h = 0.01
    for th in talhoes:
        for s in range(1, sem_max + 1):
            max_h = max(max_h, agg[th].get(s, 0))

    def heat(hh, mh):
        if mh <= 0 or hh <= 0.01:
            return "E8E8E8"
        t = min(1.0, hh / mh)
        r = int(200 - t * 100)
        g = int(220 - t * 70)
        b = int(255 - t * 40)
        return f"{r:02X}{g:02X}{b:02X}"

    r = 2
    for th in talhoes:
        ws.cell(row=r, column=1, value=th)
        for s in range(1, sem_max + 1):
            v = agg[th].get(s, 0.0)
            col = s + 1
            if v > 0.01:
                cell = ws.cell(row=r, column=col, value=round(v, 1))
                hx = heat(v, max_h)
                cell.fill = PatternFill(start_color=hx, end_color=hx, fill_type="solid")
            else:
                ws.cell(row=r, column=col, value="")
        r += 1
    ws.freeze_panes = "B2"
    _auto_width(ws)


def aplicar_formatacao_financeiro(wb, lucro, margem_pct, dias_simulado, fazenda, receita, custo_mo_total):
    """Apenas abas financeiras + DASHBOARD (sem GANTT)."""
    for name in ("RESUMO_FINANCEIRO", "CUSTO_POR_ATIVIDADE", "COMPARATIVO_ROBO", "COMPARATIVO_CENARIOS"):
        if name in wb.sheetnames:
            format_standard_sheet(wb[name], zebra=(name == "CUSTO_POR_ATIVIDADE"))
    add_dashboard_sheet(wb, fazenda, lucro, margem_pct, dias_simulado, receita, custo_mo_total)


def aplicar_formatacao_operacional(wb, dias_simulado, cronograma_rows):
    """Abas de execucao + GANTT (sem DASHBOARD financeiro)."""
    for name in (
        "RESUMO_OPERACIONAL",
        "CRONOGRAMA_DETALHADO",
        "CASCATA_EXPLICADA",
        "CRONOGRAMA_E_CASCATA",
        "TIMELINE_CASCATA",
        "OCUPACAO_TURMAS_DIA",
        "CRONOGRAMA_MECANIZADO",
        "CRONOGRAMA_COMBINADO",
        "AUDITORIA_ESCOPO",
    ):
        if name in wb.sheetnames:
            format_standard_sheet(
                wb[name],
                zebra=name in ("CRONOGRAMA_DETALHADO", "CASCATA_EXPLICADA", "CRONOGRAMA_E_CASCATA"),
            )
    if "CASCATA_EXPLICADA" in wb.sheetnames:
        _highlight_cascata_explicada(wb["CASCATA_EXPLICADA"])
    add_gantt_simple_sheet(wb, cronograma_rows, dias_simulado)


def aplicar_formatacao_dossier(wb, lucro, margem_pct, dias_simulado, fazenda, receita, custo_mo_total, cronograma_rows):
    """Dossie unico legado: formata abas conhecidas + DASHBOARD + GANTT."""
    for name in (
        "RESUMO_FINANCEIRO",
        "RESUMO_OPERACIONAL",
        "CRONOGRAMA_DETALHADO",
        "CRONOGRAMA_E_CASCATA",
        "CUSTO_POR_ATIVIDADE",
        "COMPARATIVO_ROBO",
        "COMPARATIVO_CENARIOS",
    ):
        if name in wb.sheetnames:
            format_standard_sheet(
                wb[name],
                zebra=name in ("CRONOGRAMA_DETALHADO", "CRONOGRAMA_E_CASCATA"),
            )

    add_dashboard_sheet(wb, fazenda, lucro, margem_pct, dias_simulado, receita, custo_mo_total)
    add_gantt_simple_sheet(wb, cronograma_rows, dias_simulado)
