"""
SRF — Sistema de Restauracao Florestal  v5.9  (Logistica & Dossier executivo)
Baseado no ATM v3 de Isaac (Zaza), reescrito com Smart Scheduler
Uso  : python atm_v5.py
       ATM_DEMO=1 python atm_v5.py
       python atm_v5.py --demo
       Modo DEMO: se existir USEESTAPLANILHAULIANOPOLIS.xlsx, gera/atualiza ulianopolisswg.xlsx;
       tarifas CT 313 como no fluxo normal; [1] usa a fazenda com mais linhas (micro municipio Ulianopolis).
"""

import os, sys, re, json, math, datetime, calendar, unicodedata, hashlib
import pandas as pd
from collections import defaultdict, OrderedDict
from statistics import median

try:
    from rich.console import Console
    from rich.table import Table
except ImportError:
    print("Instale: pip install rich pandas openpyxl")
    sys.exit(1)

# ──────────────────────────────────────────────
#  CORES & UI (estilo ATM v3)
# ──────────────────────────────────────────────
try:
    import colorama; colorama.init()
    G="\033[92m"; Y="\033[93m"; R="\033[91m"; C="\033[96m"
    DM="\033[2m"; BL="\033[1m"; RS="\033[0m"
except ImportError:
    G=Y=R=C=DM=BL=RS=""

W = 66
console = Console()

ASCII_ART = r"""
        ,@@@@@@@,
    ,,,.   ,@@@@@@/@@,  .oo8888o.
 ,&%%&%&&%,@@@@@/@@@@@@,8888\88/8o
,%&\%&&%&&%,@@@\@@@/@@@88\88888/88'
%&&%&%&/%&&%@@\@@/ /@@@88888\88888'
%&&%/ %&%%&&@@\ V /@@' `88\8 `/88'
`&%\ ` /%&'    |.|        \ '|8'
    |o|        | |         | |
    |.|        | |         | |
 \\/ ._\//_/__/  ,\_//__\\/.  \_//__/_
"""

VERSION = "6.0"
APP_NAME = "SRF - Sistema de Restauracao Florestal"
DIR  = os.path.dirname(os.path.abspath(__file__))
CFGP = os.path.join(DIR, "config.json")
DOSSIER_DIRNAME = "dossiês"

# Modo DEMO (Ulianópolis): ATM_DEMO=1 ou --demo
# Fonte de verdade para reconstruir o demo: USEESTAPLANILHAULIANOPOLIS.xlsx (municipio Ulianopolis)
DEMO_MICRO_FILENAME = "ulianopolisswg.xlsx"
DEMO_MICRO_SOURCE_FILENAME = "USEESTAPLANILHAULIANOPOLIS.xlsx"

def _is_demo_mode():
    v = os.environ.get("ATM_DEMO", "").strip().lower()
    if v in ("1", "true", "yes", "sim", "on"):
        return True
    return "--demo" in sys.argv

def _is_legacy_mode():
    v = os.environ.get("ATM_LEGACY", "").strip().lower()
    if v in ("1", "true", "yes", "sim", "on"):
        return True
    return "--legacy" in sys.argv

def _is_beta_mode():
    # Fluxo beta promovido a padrão; legado fica opt-in por flag/env.
    if _is_legacy_mode():
        return False
    v = os.environ.get("ATM_BETA", "").strip().lower()
    if v in ("0", "false", "no", "off"):
        return False
    return True

def _is_demo_micro_path(path_or_name):
    if not path_or_name:
        return False
    return os.path.basename(str(path_or_name)).lower() == DEMO_MICRO_FILENAME.lower()

# ──────────────────────────────────────────────
#  MAPEAMENTO FIXO DE COLUNAS CONHECIDAS
#  (fallback semantico so se nenhuma bater)
# ──────────────────────────────────────────────
KNOWN_COLUMNS = {
    "fazenda":   ["NOME FAZENDA", "CÓDIGO FAZENDA"],
    "chave":     ["CHAVE POLÍGONO", "CHAVE POLIGONO"],
    "area":      ["ÁREA TRABALHADA ESTIMADA (HECTARE)", "ÁREA POLÍGONO (HECTARE)",
                  "AREA POLIGONO (HECTARE)", "AREA TRABALHADA ESTIMADA (HECTARE)"],
    "atividade": ["ATIVIDADES", "ATIVIDADE"],
}

def linha(c="="): print(G + c*W + RS)
def sub(c="-"):   print(DM + c*W + RS)

def cabecalho(sub_titulo=""):
    os.system("cls" if os.name=="nt" else "clear")
    print(G + ASCII_ART + RS)
    linha()
    print(G+BL + f"  [ SRF ]  {APP_NAME}  v{VERSION}".center(W) + RS)
    if sub_titulo:
        print(DM+G + sub_titulo.center(W) + RS)
    print(DM+G + datetime.datetime.now().strftime("  %d/%m/%Y  %H:%M").center(W) + RS)
    linha()

def aviso(m): print(Y+f"\n  !  {m}"+RS)
def erro(m):  print(R+f"\n  X  {m}"+RS)
def ok(m):    print(G+f"\n  +  {m}"+RS)

def prompt(msg, default=None):
    suf = f" [{default}]" if default is not None else ""
    try:
        v = input(G+"  >> "+C+msg+suf+G+": "+RS).strip()
    except (EOFError, KeyboardInterrupt):
        print(); sys.exit(0)
    return v if v else (str(default) if default is not None else "")

def pedir_float(msg, default, allow_zero=False):
    while True:
        v = prompt(msg, default)
        try:
            f = float(str(v).replace(",","."))
            if f > 0 or (allow_zero and f >= 0): return f
        except ValueError: pass
        aviso("Valor invalido.")

def pedir_int(msg, default, allow_zero=False):
    while True:
        v = prompt(msg, default)
        try:
            i = int(v)
            if i > 0 or (allow_zero and i >= 0): return i
        except ValueError: pass
        aviso("Valor invalido.")

def selecionar(titulo, itens, zero_label="Voltar"):
    print(G+f"\n  -- {titulo} "+"--"*max(0, (W-len(titulo)-6)//2)+RS)
    for i, it in enumerate(itens, 1):
        print(G+f"  [{i:2}] "+C+str(it)+RS)
    print(G+f"  [ 0] "+DM+zero_label+RS)
    while True:
        v = prompt("Escolha").strip()
        if v == "0": return None
        if v.isdigit() and 1 <= int(v) <= len(itens): return itens[int(v)-1]
        aviso("Opcao invalida.")

def selecionar_paginado(titulo, itens, page_size=5, zero_label="Voltar"):
    total = len(itens)
    page = 0
    max_page = math.ceil(total / page_size) - 1
    while True:
        start = page * page_size
        end = min(start + page_size, total)
        print(G+f"\n  -- {titulo} (pag {page+1}/{max_page+1}) "+"--"*max(0,(W-len(titulo)-16)//2)+RS)
        for i in range(start, end):
            print(G+f"  [{i+1:2}] "+C+str(itens[i])+RS)
        nav = []
        if page > 0: nav.append("[-] Anterior")
        if page < max_page: nav.append("[+] Proxima")
        nav.append("[0] " + zero_label)
        print(DM + "  " + "   ".join(nav) + RS)
        v = prompt("Escolha").strip()
        if v == "0": return -1
        if v == "+" and page < max_page: page += 1; continue
        if v == "-" and page > 0: page -= 1; continue
        if v.isdigit() and 1 <= int(v) <= total: return int(v) - 1
        aviso("Opcao invalida.")

def confirmar(msg, default=True):
    s = "S/n" if default else "s/N"
    v = prompt(f"{msg} [{s}]").strip().lower()
    if not v: return default
    return v in ("s", "sim", "y", "yes")

def remover_acentos(texto):
    if not isinstance(texto, str): return ""
    return ''.join(c for c in unicodedata.normalize('NFD', texto)
                   if unicodedata.category(c) != 'Mn').lower().strip()

import re
_RE_PUNCT = re.compile(r'[^a-z0-9 ]+')
_RE_SPACES = re.compile(r'\s+')

def normalizar_chave(texto):
    """remover_acentos + strip punctuation + collapse whitespace. Canonical lookup key."""
    s = remover_acentos(texto)
    s = _RE_PUNCT.sub(' ', s)
    return _RE_SPACES.sub(' ', s).strip()

# ──────────────────────────────────────────────
#  CONFIG
# ──────────────────────────────────────────────
def _default_sequencia_dict():
    return {
        "modo": "implantacao",
        "offset_limpeza_quimica_dias": 30,
        "filtros_plantio": ["plantio"],
        "filtros_irrigacao": ["irrig"],
        "limpeza_quimica_filtros": ["limpeza", "quim"],
        "limpeza_quimica_exclusoes": ["impl", "impl."],
        "implantacao_outras_fase": 5.5,
        "implantacao_fases": [
            {"id": "rocada", "filtros": ["rocada", "roçada"], "exclusoes": []},
            {"id": "formiga", "filtros": ["formiga", "combate a formiga", "combate a formigas"], "exclusoes": []},
            {"id": "coroamento", "filtros": ["coroamento", "coroa"], "exclusoes": []},
            {"id": "coveamento", "filtros": ["coveamento", "coveam"], "exclusoes": []},
            {"id": "adubacao_quimica", "filtros": ["adubacao quim", "adubação quím", "melhora quim"], "exclusoes": []},
        ],
        # Ordem exata solicitada para manutencao SWG (conforme lista de atividades do Excel).
        "swg_fases": [
            {"id": "swg_rocada_manual", "filtros": ["rocada manual", "roçada manual"], "exclusoes": []},
            {"id": "swg_limpeza_area", "filtros": ["limpeza de area", "limpeza de área"], "exclusoes": []},
            {"id": "swg_capina_coroa", "filtros": ["capina manual coroa", "capina manual"], "exclusoes": []},
            {"id": "swg_combate_formigas", "filtros": ["combate a formigas", "combate a formiga", "formigas"], "exclusoes": []},
            {"id": "swg_coveamento", "filtros": ["coveam area nao subsol", "coveam área não subsol", "coveamento"], "exclusoes": []},
            {"id": "swg_adubacao_base", "filtros": ["adubacao quim man de base", "adubação quim man de base", "adubacao"], "exclusoes": []},
            {"id": "swg_plantio_manual", "filtros": ["plantio manual", "plantio"], "exclusoes": []},
            {"id": "swg_irrigacao_inicial", "filtros": ["irrigacao inicial", "irrigação inicial", "irrigacao"], "exclusoes": []},
        ],
        "personalizado_ordem": [],
    }


def _merge_sequencia_defaults(seq):
    """Preenche chaves ausentes em cfg['sequencia'] (muta seq)."""
    d0 = _default_sequencia_dict()
    for k, v in d0.items():
        if k not in seq:
            seq[k] = v
        elif k in ("implantacao_fases", "swg_fases", "personalizado_ordem") and not seq[k]:
            seq[k] = v


def carregar_config():
    if not os.path.exists(CFGP):
        with open(CFGP, "w", encoding="utf-8") as f:
            json.dump({"de_para": {}, "tarifas": {}, "atividades": {}}, f)
    with open(CFGP, "r", encoding="utf-8") as f:
        cfg = json.load(f)
    for k in ("de_para", "tarifas", "atividades"):
        if k not in cfg:
            cfg[k] = {}
    if "orcamento_estrito" not in cfg:
        cfg["orcamento_estrito"] = True
    if "filtros_bloqueio_global" not in cfg:
        cfg["filtros_bloqueio_global"] = ["plantio", "irrig"]
    if "fazendas_ct" not in cfg:
        cfg["fazendas_ct"] = []
    if "metas" not in cfg:
        cfg["metas"] = {
            "lucro_alvo": None,
            "margem_alvo_pct": None,
            "bonus_aa_formula": "",
            "bonus_bb_formula": "",
            "equacao_quadratica": "",
            "nota": "Rotas preparatorias: preencher quando dados oficiais estiverem disponiveis."
        }
    if "precos_contrato" not in cfg:
        cfg["precos_contrato"] = {
            "arquivo": "",
            "sheet_preco_final": "",
            "sheet_custo_direto": "",
            "sheet_custo_indireto": "",
        }
    if "custos_globais" not in cfg:
        cfg["custos_globais"] = {
            "arquivo": "",
            "sheet_custo_direto": "",
            "sheet_custo_indireto": "",
            "valor_direto_total": 0.0,
            "valor_indireto_total": 0.0,
            "criterio": "ultimo_valor_na_linha",
            "itens_direto": [],
            "itens_indireto": [],
        }
    if "sequencia" not in cfg or not isinstance(cfg.get("sequencia"), dict):
        cfg["sequencia"] = {}
    _merge_sequencia_defaults(cfg["sequencia"])
    return cfg


def salvar_config(cfg):
    with open(CFGP, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

def parse_intervalos_escolha(texto, max_n):
    """
    Converte '1,3,5-8' em indices 0-based unicos e ordenados (numeracao 1..max_n).
    Espacos ignorados; intervalos inclusive.
    """
    out = set()
    if not texto or not str(texto).strip() or max_n < 1:
        return []
    for part in str(texto).replace(" ", "").split(","):
        if not part:
            continue
        if "-" in part:
            a, b = part.split("-", 1)
            try:
                lo, hi = int(a), int(b)
                if lo > hi:
                    lo, hi = hi, lo
                for k in range(lo, hi + 1):
                    if 1 <= k <= max_n:
                        out.add(k - 1)
            except ValueError:
                continue
        else:
            try:
                k = int(part)
                if 1 <= k <= max_n:
                    out.add(k - 1)
            except ValueError:
                continue
    return sorted(out)

def mediana_rendimento_hh(tarifas):
    """Mediana dos rendimento_hh > 0 em config.tarifas, ou None."""
    vals = []
    for v in (tarifas or {}).values():
        if not isinstance(v, dict):
            continue
        try:
            x = float(v.get("rendimento_hh", 0))
            if x > 0:
                vals.append(x)
        except (TypeError, ValueError):
            pass
    if not vals:
        return None
    vals.sort()
    n = len(vals)
    mid = n // 2
    if n % 2:
        return vals[mid]
    return (vals[mid - 1] + vals[mid]) / 2.0

def resolver_rendimento_hh(cfg, tarifas, t_nome, strict=False, session_hh=None, atv_micro=None):
    """
    HH/ha para a chave t_nome em tarifas.
    session_hh: dict opcional {nome_micro ou chave_tarifa: hh_ha} valido so na execucao atual (nao grava config).
    Modo strict: sem mediana/8 silenciosos — retorna None se invalido.
    Excecao: atividades mecanizadas (tipo Mecanizada ou HM>0) retornam 0.0 em strict.
    """
    if session_hh:
        try:
            if atv_micro and atv_micro in session_hh:
                return float(session_hh[atv_micro])
            if t_nome and t_nome in session_hh:
                return float(session_hh[t_nome])
        except (TypeError, ValueError):
            pass
    if t_nome in (tarifas or {}):
        row = tarifas[t_nome]
        r = row.get("rendimento_hh")
        try:
            rf = float(r)
            if rf >= 0:
                if strict and rf <= 0:
                    tipo = str(row.get("tipo", "")).lower()
                    hm = float(row.get("rendimento_hm", 0) or 0)
                    if "mecaniz" in tipo or hm > 0:
                        return 0.0
                    return None
                return rf
        except (TypeError, ValueError):
            if strict:
                return None
    if strict:
        return None
    ex = cfg.get("rendimento_hh_fallback")
    if ex is not None:
        try:
            e = float(ex)
            if e > 0:
                return e
        except (TypeError, ValueError):
            pass
    med = mediana_rendimento_hh(tarifas)
    if med is not None and med > 0:
        return med
    return 8.0

# ──────────────────────────────────────────────
#  CT_313 RAW -> STG NORMALIZER
# ──────────────────────────────────────────────
STG_FILENAME = "CT_313_NORMALIZADA.xlsx"

# Padrao de prototipo: mapeamento fixo EXAME -> CT_313 (normalizado sem acentos).
# Fonte: leitura real das abas MICROPLANEJAMENTO_ABRIL_JUNHO e STG_TARIFAS.
#
# Declividade (ROÇADA MANUAL CLASSE I..V na CT): classe I = terreno mais plano (menor HH/ha e
# menor preco/ha); classe V = declive maximo (maior HH e maior preco — mais custo MO mas mais
# receita). O micro nao traz a classe — o padrao SRF usa sempre CLASSE I / limpeza plana onde
# aplicavel (cenario conservador em margem: sem premio de obra em morro). Ver aviso no scheduler.
DEFAULT_DEPARA_EXAME_CT313 = {
    # ADUBAÇÃO
    "adubacao quim man de base impl pl app rl": "ADUBAÇÃO QUÍMICA MANUAL",
    # CAPINA
    "capina manual coroa impl cd app rl i": "CAPINA COROAMENTO MANUAL I",
    "capina manual coroa impl pl app rl i": "CAPINA COROAMENTO MANUAL I",
    "capina quim man total manut app rl": "CAPINA QUÍMICA TOTAL MANUAL PLANO",
    # COMBATE
    "combate a formigas impl cd app rl": "COMBATE DE FORMIGAS MANUAL",
    "combate a formigas impl pl app rl": "COMBATE DE FORMIGAS MANUAL",
    "combate a formigas manut app rl": "CONTROLE DE FORMIGAS MANUAL  (REPASSE)",
    # CONTROLE — com e sem sufixo I
    "controle de invasoras app rl i": "CAPINA QUÍMICA TOTAL MANUAL PLANO",
    "controle de invasoras app rl": "CAPINA QUÍMICA TOTAL MANUAL PLANO",
    # COVEAMENTO — motocoveador, subsol, NÃO subsol
    "coveamento motocoveador pl app rl": "COVEAMENTO SEMI MECANIZADO - 30CM",
    "coveamento area subsol impl pl app rl": "SUBSOLAGEM COM ADUBAÇÃO (TRATOR PNEU)",
    "coveam area nao subsol impl pl app rl": "COVEAMENTO SEMI MECANIZADO - 30CM",
    # IRRIGAÇÃO
    "irrigacao inicial man impl pl app rl": "IRRIGAÇÃO DE PLANTIO MANUAL",
    # LIMPEZA — QUIM. / QU.
    "limpeza de area quim impl cd app rl": "ROÇADA SEMIMECANIZADA ÁREA TOTAL",
    "limpeza de area quim man app rl": "ROÇADA MANUAL CLASSE I",
    "limpeza de area qu man app rl": "ROÇADA MANUAL CLASSE I",
    # PLANTIO
    "plantio manual app rl": "PLANTIO MANUAL SEM GEL",
    # PREPARO DE SOLO
    "preparo de solo mec c grade app rl": "PREPARO DE SOLO COM ADUBAÇÃO DE BASE E MARCAÇÃO DE BACIA ECAVADEIRA",
    "preparo de solo mec s adub app rl": "PREPARO DE SOLO COM MÁQUINA DE ESTEIRA",
    # ROÇADA
    "rocada manual impl cd app rl i": "ROÇADA MANUAL CLASSE I",
    "rocada manual impl pl app rl i": "ROÇADA MANUAL CLASSE I",
    # Atividades sem par explicito nas tarifas operacionais: usa baseline de MO da CT.
    "conducao de regeneracao": "SERVIÇO DE MÃO DE OBRA",
    "eliminacao de exoticas impl cd app rl": "SERVIÇO DE MÃO DE OBRA",
    "nucleacao em faixas app rl": "SERVIÇO DE MÃO DE OBRA",
}


def _depara_heuristico_exame_ct313(kn, tarifas):
    """
    Fallback quando o micro usa outras convencoes de texto (APPN RL/NR, manl., microtrator,
    MELHORA QUIM..., ESTIVA..., parenteses etc.). So aplica se a tarifa alvo existir no CT.
    """
    if not kn or not tarifas:
        return None

    def pick(*names):
        for n in names:
            if n in tarifas:
                return n
        return None

    # Ordem: mais especifico / excecao primeiro
    if "estiva" in kn:
        return pick("SERVIÇO DE MÃO DE OBRA")
    if "subsol" in kn or "area subsol" in kn:
        return pick("SUBSOLAGEM COM ADUBAÇÃO (TRATOR PNEU)")
    if "coveamento" in kn:
        return pick(
            "COVEAMENTO SEMI MECANIZADO - 30CM",
            "COVEAMENTO SEMI MECANIZADO - 40CM",
        )
    if "combate" in kn and "formiga" in kn:
        return pick(
            "COMBATE DE FORMIGAS MANUAL",
            "CONTROLE DE FORMIGAS MANUAL  (REPASSE)",
        )
    # "Melhoria / Mineral" no micro as vezes vem como MELHORA QUIM ... DE BASE
    if ("melhora" in kn or "melhoria" in kn) and ("quim" in kn or "quimica" in kn):
        return pick("ADUBAÇÃO QUÍMICA MANUAL")
    if "adubacao" in kn and ("quim" in kn or "quimica" in kn):
        return pick("ADUBAÇÃO QUÍMICA MANUAL")
    if "limpeza" in kn and "area" in kn:
        if "quim" in kn or "quimica" in kn:
            return pick(
                "ROÇADA SEMIMECANIZADA ÁREA TOTAL",
                "ROÇADA MANUAL CLASSE I",
            )
        return pick("ROÇADA MANUAL CLASSE I")
    if "irrigacao" in kn:
        return pick(
            "IRRIGAÇÃO DE PLANTIO MANUAL",
            "IRRIGAÇÃO DE PLANTIO SEMIMEC",
        )
    if "plantio" in kn and "manual" in kn:
        return pick("PLANTIO MANUAL SEM GEL", "PLANTIO MANUAL COM GEL")
    if "preparo" in kn and "solo" in kn:
        if "rocadeira" in kn or "esteira" in kn or "s adub" in kn or "s adubacao" in kn:
            return pick("PREPARO DE SOLO COM MÁQUINA DE ESTEIRA")
        if "grade" in kn or ("adub" in kn and "s adub" not in kn):
            return pick(
                "PREPARO DE SOLO COM ADUBAÇÃO DE BASE E MARCAÇÃO DE BACIA ECAVADEIRA",
            )
        return pick(
            "PREPARO DE SOLO COM MÁQUINA DE ESTEIRA",
            "PREPARO DE SOLO COM ADUBAÇÃO DE BASE E MARCAÇÃO DE BACIA ECAVADEIRA",
        )
    if "nucleacao" in kn:
        return pick("SERVIÇO DE MÃO DE OBRA")
    if "conducao" in kn:
        return pick("SERVIÇO DE MÃO DE OBRA")
    if "eliminacao" in kn and "exotica" in kn:
        return pick("SERVIÇO DE MÃO DE OBRA")
    if "controle" in kn and "invasora" in kn:
        return pick("CAPINA QUÍMICA TOTAL MANUAL PLANO")
    if "rocada" in kn and "manual" in kn:
        return pick("ROÇADA MANUAL CLASSE I")
    if "capina" in kn:
        if "quim" in kn or "quimica" in kn:
            return pick(
                "CAPINA QUÍMICA TOTAL MANUAL PLANO",
                "CAPINA QUÍMICA TOTAL MANUAL DECLIVIDADE",
            )
        return pick("CAPINA COROAMENTO MANUAL I")
    return None


def _ct_file_hash(path):
    h = hashlib.md5()
    h.update(str(os.path.getmtime(path)).encode())
    h.update(str(os.path.getsize(path)).encode())
    return h.hexdigest()

def _find_preco_final_sheet(xls):
    for s in xls.sheet_names:
        if remover_acentos(s).replace(" ", "") in ("precofinal", "preco final"):
            return s
    return None

def _find_diaria_tf_sheet(xls):
    for s in xls.sheet_names:
        if "diaria_tf" in remover_acentos(s).replace(" ", ""):
            return s
    return None

def normalizar_ct313(caminho_ct):
    """
    Le CT_313 bruta e gera CT_313_NORMALIZADA.xlsx com aba STG_TARIFAS.
    Retorna (caminho_stg, n_linhas, custo_hora_tf).
    """
    xls = pd.ExcelFile(caminho_ct)
    pf = _find_preco_final_sheet(xls)
    if not pf:
        return None, 0, 0.0

    df = pd.read_excel(caminho_ct, sheet_name=pf, header=None)

    custo_hora_tf = 0.0
    tf = _find_diaria_tf_sheet(xls)
    if tf:
        dft = pd.read_excel(caminho_ct, sheet_name=tf, header=None)
        try:
            custo_dia = float(dft.iloc[3][4])
            jornada_tf = float(dft.iloc[4][2])
            if jornada_tf > 0:
                custo_hora_tf = custo_dia / jornada_tf
        except (IndexError, TypeError, ValueError):
            pass

    rows = []
    for i in range(5, len(df)):
        r = df.iloc[i]
        nome = str(r[2]).strip() if pd.notna(r[2]) else ""
        if not nome or nome == "0":
            continue
        tipo = str(r[4]).strip() if pd.notna(r[4]) else ""
        try:
            hh = float(r[5]) if pd.notna(r[5]) else 0.0
        except (TypeError, ValueError):
            hh = 0.0
        try:
            hm = float(r[6]) if pd.notna(r[6]) else 0.0
        except (TypeError, ValueError):
            hm = 0.0
        try:
            preco = float(r[7]) if pd.notna(r[7]) else 0.0
        except (TypeError, ValueError):
            preco = 0.0
        custo_h = custo_hora_tf if hh > 0 and custo_hora_tf > 0 else 0.0
        custo_ha = hh * custo_h if custo_h > 0 else 0.0
        rows.append({
            "atividade": nome,
            "tipo": tipo,
            "rendimento_hh": hh,
            "rendimento_hm": hm,
            "preco_ha": preco,
            "custo_hora": custo_h,
            "custo_ha": custo_ha,
            "fonte_aba": pf,
        })

    df_stg = pd.DataFrame(rows)
    meta = pd.DataFrame([{
        "gerado_em": datetime.datetime.now().isoformat(),
        "arquivo_origem": os.path.basename(caminho_ct),
        "linhas_validas": len(rows),
        "custo_hora_tf": round(custo_hora_tf, 4),
    }])

    stg_path = os.path.join(DIR, STG_FILENAME)
    with pd.ExcelWriter(stg_path, engine="openpyxl") as w:
        df_stg.to_excel(w, sheet_name="STG_TARIFAS", index=False)
        meta.to_excel(w, sheet_name="STG_METADATA", index=False)

    return stg_path, len(rows), custo_hora_tf

def carregar_stg_tarifas(stg_path):
    """Le STG_TARIFAS e retorna dict {atividade: {rendimento_hh, preco_ha, custo_hora, custo_ha, tipo}}."""
    df = pd.read_excel(stg_path, sheet_name="STG_TARIFAS")
    t = {}
    for _, r in df.iterrows():
        nome = str(r.get("atividade", "")).strip()
        if not nome:
            continue
        hh = float(r.get("rendimento_hh") or 0)
        hm = float(r.get("rendimento_hm") or 0)
        t[nome] = {
            "rendimento_hh": hh,
            "rendimento_hm": hm,
            "preco_ha": float(r.get("preco_ha") or 0),
            "custo_hora": float(r.get("custo_hora") or 0),
            "custo_ha": float(r.get("custo_ha") or 0),
            "tipo": str(r.get("tipo") or ""),
            "preco_unit": float(r.get("preco_ha") or 0),
            "recurso": "homem" if hh > 0 else "maquina",
            "eficiencia": 1.0,
        }
    return t

def modulo_normalizar_ct(cfg):
    """Menu: selecionar CT bruta, gerar STG, integrar em config.tarifas."""
    cabecalho("NORMALIZAR CT_313 -> STG_TARIFAS")
    caminho = selecionar_arquivo("CT_313 BRUTA (.xlsm ou .xlsx)")
    if not caminho:
        return

    print(DM+"  Processando... pode demorar alguns segundos."+RS)
    stg_path, n, custo_h = normalizar_ct313(caminho)
    if not stg_path:
        erro("Aba 'Preco Final' nao encontrada neste arquivo.")
        input(DM+"\n  [ENTER] "+RS)
        return

    ok(f"Gerado {STG_FILENAME}: {n} atividades | custo/hora TF = R${custo_h:.2f}")

    if confirmar("Integrar STG_TARIFAS em config.json (substitui tarifas existentes)?", default=True):
        tarifas = carregar_stg_tarifas(stg_path)
        cfg["tarifas"] = tarifas
        cfg["custo_hora_tf"] = round(custo_h, 4)
        salvar_config(cfg)
        ok(f"{len(tarifas)} tarifas integradas no config.")
    input(DM+"\n  [ENTER para voltar] "+RS)

def _guess_sheet(xls, keys):
    for s in xls.sheet_names:
        ns = normalizar_chave(s)
        if all(k in ns for k in keys):
            return s
    return None

def _pick_col(df, required_tokens_sets):
    cols = list(df.columns)
    for c in cols:
        nc = normalizar_chave(c)
        for toks in required_tokens_sets:
            if all(t in nc for t in toks):
                return c
    return None

def _to_float_any(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    s = s.replace("R$", "").replace(" ", "")
    if "," in s and "." in s:
        # Ex.: 1.234,56 -> 1234.56
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None

def _last_non_zero(nums):
    for n in reversed(nums):
        if n is not None and abs(float(n)) > 1e-9:
            return float(n)
    for n in reversed(nums):
        if n is not None:
            return float(n)
    return 0.0

def _is_raw_cost_row_label(lbl):
    n = normalizar_chave(lbl)
    if not n:
        return False
    bad = {
        "indireto",
        "custo indireto pessoal",
        "custo indireto",
        "custo direto",
        "bdi",
        "soma",
    }
    if n in bad:
        return False
    if n in {"d", "premio"}:
        return False
    if n.startswith("previsao "):
        return False
    if n.startswith("resultado "):
        return False
    if n.startswith("desconto "):
        return False
    return True

def _extrair_custos_globais_brutos(caminho, sheet_cd, sheet_ci):
    """
    Extrai custos globais de planilha bruta (layout "torto"):
    - pega coluna A como item
    - consolida cada linha pelo ultimo valor nao nulo/nao-zero das colunas laterais
    - total direto/indireto prioriza linha 'CUSTO DIRETO'/'CUSTO INDIRETO'
      e, se ausente, usa soma dos itens consolidados.
    """
    def parse_sheet(sheet_name):
        df = pd.read_excel(caminho, sheet_name=sheet_name, header=None)
        itens = []
        total_linha = None
        for _, r in df.iterrows():
            label = str(r.iloc[0] if len(r) > 0 else "").strip()
            if not label:
                continue
            nums = []
            for v in list(r.iloc[1:]):
                fv = _to_float_any(v)
                if fv is not None:
                    nums.append(fv)
            nlabel = normalizar_chave(label)
            if nlabel in {"custo direto", "custo indireto"}:
                total_linha = _last_non_zero(nums)
                continue
            if _is_raw_cost_row_label(label):
                val = _last_non_zero(nums)
                if abs(val) > 1e-9:
                    itens.append({"item": label, "valor": round(float(val), 6)})
        total = float(total_linha) if total_linha is not None else float(sum(x["valor"] for x in itens))
        return total, itens

    total_cd, itens_cd = parse_sheet(sheet_cd)
    total_ci, itens_ci = parse_sheet(sheet_ci)
    return {
        "valor_direto_total": round(float(total_cd), 6),
        "valor_indireto_total": round(float(total_ci), 6),
        "itens_direto": itens_cd,
        "itens_indireto": itens_ci,
    }

def modulo_importar_custos_globais_brutos(cfg):
    """
    Rota secundaria:
    importa CUSTO_DIRETO/CUSTO_INDIRETO em formato bruto (coluna A + valores laterais),
    sem rateio por atividade, para uso no fechamento financeiro global.
    """
    cabecalho("IMPORTAR CUSTOS GLOBAIS (BRUTO)")
    caminho = selecionar_arquivo("PLANILHA BRUTA DE CUSTOS (CUSTO_DIRETO/CUSTO_INDIRETO)")
    if not caminho:
        return
    try:
        xls = pd.ExcelFile(caminho)
        cd = _guess_sheet(xls, ["custo", "direto"])
        ci = _guess_sheet(xls, ["custo", "indireto"])
        sub()
        print(G + "  CUSTO_DIRETO   : " + C + f"{cd or '??'}" + RS)
        print(G + "  CUSTO_INDIRETO : " + C + f"{ci or '??'}" + RS)
        if not (cd and ci) or not confirmar("Usar mapeamento automatico de abas?", default=True):
            cd = selecionar("ABA CUSTO_DIRETO", xls.sheet_names)
            if cd is None:
                return
            ci = selecionar("ABA CUSTO_INDIRETO", xls.sheet_names)
            if ci is None:
                return

        ext = _extrair_custos_globais_brutos(caminho, cd, ci)
        cfg["custos_globais"] = {
            "arquivo": os.path.basename(caminho),
            "sheet_custo_direto": cd,
            "sheet_custo_indireto": ci,
            "valor_direto_total": ext["valor_direto_total"],
            "valor_indireto_total": ext["valor_indireto_total"],
            "criterio": "ultimo_valor_na_linha",
            "itens_direto": ext["itens_direto"],
            "itens_indireto": ext["itens_indireto"],
        }
        salvar_config(cfg)
        ok(
            "Custos globais importados: "
            f"Direto R$ {ext['valor_direto_total']:,.2f} | "
            f"Indireto R$ {ext['valor_indireto_total']:,.2f}"
        )
        print(
            DM
            + f"  Itens lidos: direto={len(ext['itens_direto'])} | indireto={len(ext['itens_indireto'])}"
            + RS
        )
    except Exception as ex:
        erro(f"Falha ao importar custos globais brutos: {ex}")
    input(DM + "\n  [ENTER para voltar] " + RS)

def modulo_importar_precos_contrato(cfg):
    """
    Novo importador (substitui CT_313 quando disponivel):
    workbook com abas PRECO_FINAL, CUSTO_DIRETO, CUSTO_INDIRETO.
    Se nomes divergirem, permite mapeamento manual de abas.
    """
    cabecalho("IMPORTAR PLANILHA DE PRECO (CONTRATO)")
    caminho = selecionar_arquivo("PLANILHA DE PRECO (3 abas: PRECO_FINAL/CUSTO_DIRETO/CUSTO_INDIRETO)")
    if not caminho:
        return
    try:
        # Referencia tecnica: HH e custo-hora devem priorizar CT_313.
        # Carrega uma base CT fresca (quando houver arquivo CT no workspace),
        # com fallback para o estado atual de cfg["tarifas"].
        tarifas_ct_ref = {}
        ct_path = _find_default_ct_path()
        if ct_path:
            try:
                stg_path, n_ct, _ = normalizar_ct313(ct_path)
                if stg_path and n_ct > 0:
                    tarifas_ct_ref = carregar_stg_tarifas(stg_path)
            except Exception:
                tarifas_ct_ref = {}
        if not tarifas_ct_ref:
            tarifas_ct_ref = dict(cfg.get("tarifas", {}) or {})
        tarifas_ct_idx = {normalizar_chave(k): v for k, v in tarifas_ct_ref.items()}
        de_para_cfg = cfg.get("de_para", {}) or {}

        xls = pd.ExcelFile(caminho)
        pf = _guess_sheet(xls, ["preco", "final"])
        cd = _guess_sheet(xls, ["custo", "direto"])
        ci = _guess_sheet(xls, ["custo", "indireto"])
        sub()
        print(G+f"  PRECO_FINAL    : "+C+f"{pf or '??'}"+RS)
        print(G+f"  CUSTO_DIRETO   : "+C+f"{cd or '??'}"+RS)
        print(G+f"  CUSTO_INDIRETO : "+C+f"{ci or '??'}"+RS)
        if not (pf and cd and ci) or not confirmar("Usar mapeamento automatico de abas?", default=True):
            pf = selecionar("ABA PRECO_FINAL", xls.sheet_names)
            if pf is None:
                return
            cd = selecionar("ABA CUSTO_DIRETO", xls.sheet_names)
            if cd is None:
                return
            ci = selecionar("ABA CUSTO_INDIRETO", xls.sheet_names)
            if ci is None:
                return

        df_pf = pd.read_excel(caminho, sheet_name=pf)
        df_cd = pd.read_excel(caminho, sheet_name=cd)
        df_ci = pd.read_excel(caminho, sheet_name=ci)

        col_atv_pf = _pick_col(df_pf, [["atividade"], ["servico"], ["descricao"]])
        col_preco = _pick_col(df_pf, [["preco", "final"], ["preco"], ["valor"]])
        col_hh = _pick_col(df_pf, [["hh"], ["homem", "hora"], ["rendimento", "hh"]])
        col_hm = _pick_col(df_pf, [["hm"], ["hora", "maquina"], ["rendimento", "hm"]])
        col_tipo = _pick_col(df_pf, [["tipo"]])

        col_atv_cd = _pick_col(df_cd, [["atividade"], ["servico"], ["descricao"]])
        col_cd = _pick_col(df_cd, [["custo", "direto"], ["direto"], ["valor"]])
        col_atv_ci = _pick_col(df_ci, [["atividade"], ["servico"], ["descricao"]])
        col_ci = _pick_col(df_ci, [["custo", "indireto"], ["indireto"], ["valor"]])

        if not col_atv_pf or not col_preco:
            erro("Nao foi possivel identificar colunas minimas de PRECO_FINAL (atividade/preco).")
            input(DM+"\n  [ENTER] "+RS)
            return

        custo_direto = {}
        if col_atv_cd and col_cd:
            for _, r in df_cd.iterrows():
                atv = str(r.get(col_atv_cd, "")).strip()
                if not atv:
                    continue
                try:
                    custo_direto[normalizar_chave(atv)] = float(str(r.get(col_cd, 0)).replace(",", "."))
                except Exception:
                    pass
        custo_indireto = {}
        if col_atv_ci and col_ci:
            for _, r in df_ci.iterrows():
                atv = str(r.get(col_atv_ci, "")).strip()
                if not atv:
                    continue
                try:
                    custo_indireto[normalizar_chave(atv)] = float(str(r.get(col_ci, 0)).replace(",", "."))
                except Exception:
                    pass

        tarifas = {}
        for _, r in df_pf.iterrows():
            atv = str(r.get(col_atv_pf, "")).strip()
            if not atv:
                continue
            try:
                preco = float(str(r.get(col_preco, 0)).replace(",", "."))
            except Exception:
                preco = 0.0
            try:
                hh_pf = float(str(r.get(col_hh, 0)).replace(",", ".")) if col_hh else 0.0
            except Exception:
                hh_pf = 0.0
            try:
                hm = float(str(r.get(col_hm, 0)).replace(",", ".")) if col_hm else 0.0
            except Exception:
                hm = 0.0
            nk = normalizar_chave(atv)
            chave_ct = str(de_para_cfg.get(atv, atv) or atv).strip()
            nk_ct = normalizar_chave(chave_ct)
            row_ct = tarifas_ct_idx.get(nk_ct, tarifas_ct_idx.get(nk, {}))
            try:
                hh_ct = float(row_ct.get("rendimento_hh", 0) or 0.0)
            except Exception:
                hh_ct = 0.0
            try:
                hm_ct = float(row_ct.get("rendimento_hm", 0) or 0.0)
            except Exception:
                hm_ct = 0.0
            # Regra oficial: HH vem do CT_313. Se CT nao trouxer, usa HH da planilha de preco.
            hh = hh_ct if hh_ct > 0 else hh_pf
            # HM: preserva o maior sinal valido entre CT e planilha de preco
            hm = max(hm, hm_ct)
            tipo = (
                str(r.get(col_tipo, "")).strip()
                if col_tipo and str(r.get(col_tipo, "")).strip()
                else str(row_ct.get("tipo", "")).strip()
            )
            if not tipo:
                tipo = "Mecanizada" if hm > 0 else "Manual"
            cd_v = float(custo_direto.get(nk, 0.0))
            ci_v = float(custo_indireto.get(nk, 0.0))
            try:
                c_h = float(row_ct.get("custo_hora", 0) or 0.0)
            except Exception:
                c_h = 0.0
            if c_h <= 0:
                c_h = float(cfg.get("custo_hora_tf") or 0.0)
            if hh <= 0.01 and hm > 0:
                c_h = 0.0
            payload = {
                "rendimento_hh": hh,
                "rendimento_hm": hm,
                "preco_ha": preco,
                "preco_unit": preco,
                "custo_hora": c_h,
                "custo_ha": (hh * c_h) if c_h > 0 else 0.0,
                "tipo": tipo,
                "recurso": "maquina" if hm > 0 and hh <= 0.01 else "homem",
                "eficiencia": 1.0,
                "custo_direto": cd_v,
                "custo_indireto": ci_v,
            }
            tarifas[atv] = payload
            # Alias tecnico: quando o de_para aponta para chave CT diferente,
            # mantemos a mesma tarifa tambem sob a chave CT para o scheduler.
            if nk_ct and nk_ct != nk:
                tarifas[chave_ct] = dict(payload)

        if not tarifas:
            erro("Nenhuma atividade valida encontrada na planilha de preco.")
            input(DM+"\n  [ENTER] "+RS)
            return

        cfg["tarifas"] = tarifas
        cfg["precos_contrato"] = {
            "arquivo": os.path.basename(caminho),
            "sheet_preco_final": pf,
            "sheet_custo_direto": cd,
            "sheet_custo_indireto": ci,
        }
        salvar_config(cfg)
        ok(f"{len(tarifas)} tarifas importadas da planilha de contrato.")
        sem_hh = [k for k, v in tarifas.items() if float(v.get("rendimento_hh", 0) or 0) <= 0 and float(v.get("rendimento_hm", 0) or 0) <= 0]
        sem_preco = [k for k, v in tarifas.items() if float(v.get("preco_unit", 0) or 0) <= 0]
        if sem_hh:
            print(Y+f"\n  Pos-import: {len(sem_hh)} tarifa(s) sem rendimento (HH e HM zerados):"+RS)
            for x in sem_hh[:5]:
                print(DM+f"    - {str(x)[:55]}"+RS)
            if len(sem_hh) > 5:
                print(DM+f"    ... +{len(sem_hh)-5}"+RS)
        if sem_preco:
            print(Y+f"\n  Pos-import: {len(sem_preco)} tarifa(s) com preco zerado:"+RS)
            for x in sem_preco[:5]:
                print(DM+f"    - {str(x)[:55]}"+RS)
            if len(sem_preco) > 5:
                print(DM+f"    ... +{len(sem_preco)-5}"+RS)
        if not sem_hh and not sem_preco:
            ok("Pos-import: todas as tarifas possuem HH e preco validos.")
        input(DM+"\n  [ENTER para voltar] "+RS)
    except Exception as ex:
        erro(f"Falha ao importar planilha de preco: {ex}")
        input(DM+"\n  [ENTER] "+RS)

def modulo_rotas_metas_bonus(cfg):
    cabecalho("ROTAS DE METAS / BONIFICACAO / EQUACOES (PREPARATORIO)")
    m = cfg.setdefault("metas", {})
    while True:
        sub()
        print(G+f"  Lucro alvo (R$): "+C+f"{m.get('lucro_alvo')}"+RS)
        print(G+f"  Margem alvo (%): "+C+f"{m.get('margem_alvo_pct')}"+RS)
        print(G+f"  Formula bonus AA: "+C+f"{str(m.get('bonus_aa_formula',''))[:56]}"+RS)
        print(G+f"  Formula bonus BB: "+C+f"{str(m.get('bonus_bb_formula',''))[:56]}"+RS)
        print(G+f"  Equacao quadratica (rota): "+C+f"{str(m.get('equacao_quadratica',''))[:56]}"+RS)
        sub()
        print(DM+"  [1] Definir lucro alvo"+RS)
        print(DM+"  [2] Definir margem alvo (%)"+RS)
        print(DM+"  [3] Definir formula bonus AA (texto livre)"+RS)
        print(DM+"  [4] Definir formula bonus BB (texto livre)"+RS)
        print(DM+"  [5] Definir rota de equacao quadratica (texto livre)"+RS)
        print(DM+"  [6] Limpar valores"+RS)
        print(DM+"  [0] Voltar"+RS)
        op = prompt("Opcao").strip()
        if op == "0":
            salvar_config(cfg)
            return
        if op == "1":
            v = prompt("Lucro alvo (R$) ou vazio para null", "")
            m["lucro_alvo"] = None if not v.strip() else float(str(v).replace(",", "."))
        elif op == "2":
            v = prompt("Margem alvo (%) ou vazio para null", "")
            m["margem_alvo_pct"] = None if not v.strip() else float(str(v).replace(",", "."))
        elif op == "3":
            m["bonus_aa_formula"] = prompt("Formula bonus AA (texto)", "")
        elif op == "4":
            m["bonus_bb_formula"] = prompt("Formula bonus BB (texto)", "")
        elif op == "5":
            m["equacao_quadratica"] = prompt("Rota/observacao equacao quadratica", "")
        elif op == "6":
            m["lucro_alvo"] = None
            m["margem_alvo_pct"] = None
            m["bonus_aa_formula"] = ""
            m["bonus_bb_formula"] = ""
            m["equacao_quadratica"] = ""
            ok("Valores limpos.")
        else:
            aviso("Opcao invalida.")

# ──────────────────────────────────────────────
#  RESOLVERS FINANCEIROS
# ──────────────────────────────────────────────
def _mediana_campo(tarifas, campo):
    vals = []
    for v in (tarifas or {}).values():
        if not isinstance(v, dict):
            continue
        try:
            x = float(v.get(campo, 0))
            if x > 0:
                vals.append(x)
        except (TypeError, ValueError):
            pass
    return median(vals) if vals else None

def resolver_preco_ha(cfg, tarifas, t_nome, strict=False):
    if t_nome in (tarifas or {}):
        p = tarifas[t_nome].get("preco_ha") or tarifas[t_nome].get("preco_unit")
        try:
            pf = float(p)
            if pf > 0:
                return pf
        except (TypeError, ValueError):
            pass
    if strict:
        return None
    fb = cfg.get("preco_ha_fallback")
    if fb:
        try:
            f = float(fb)
            if f > 0:
                return f
        except (TypeError, ValueError):
            pass
    med = _mediana_campo(tarifas, "preco_ha") or _mediana_campo(tarifas, "preco_unit")
    return med if med and med > 0 else 0.0

def resolver_custo_hora(cfg, tarifas, t_nome, strict=False):
    if t_nome in (tarifas or {}):
        c = tarifas[t_nome].get("custo_hora")
        try:
            cf = float(c)
            if cf > 0:
                return cf
        except (TypeError, ValueError):
            pass
    if strict:
        return None
    fb = cfg.get("custo_hora_tf")
    if fb:
        try:
            f = float(fb)
            if f > 0:
                return f
        except (TypeError, ValueError):
            pass
    med = _mediana_campo(tarifas, "custo_hora")
    return med if med and med > 0 else 0.0

def resolver_chave_tarifa(cfg, tarifas, atv):
    """
    Resolve a chave de tarifa para uma atividade do micro.
    Prioridade:
    1) de_para[atividade] quando existir e estiver em tarifas;
    2) nome original da atividade quando existir em tarifas;
    3) fallback para a chave mapeada (mesmo ausente) para manter diagnostico claro.
    """
    de_para = cfg.get("de_para", {}) or {}
    t_map = str(de_para.get(atv, atv))
    if t_map in (tarifas or {}):
        return t_map
    if atv in (tarifas or {}):
        return atv
    return t_map

def modulo_mapeamentos_de_para(cfg, df_micro=None):
    """CRUD de_para: nome no microplanejamento -> nome da tarifa em config.tarifas."""
    tarifas = cfg.get("tarifas", {})
    nomes_tarifa = sorted(tarifas.keys(), key=lambda x: str(x))
    atividades_micro = []
    if df_micro is not None and getattr(df_micro, "columns", None) is not None and "atividade" in df_micro.columns:
        atividades_micro = sorted(df_micro["atividade"].dropna().unique().tolist(), key=str)

    while True:
        cabecalho("MAPEAMENTOS de_para (micro -> tarifa)")
        d = cfg.get("de_para", {})
        pairs = [(k, v) for k, v in d.items() if not str(k).startswith("_")]
        if not pairs:
            print(DM+"  Nenhum par (o sistema usa nome micro = nome na tarifa, ou default 8 h/ha)."+RS)
        else:
            for k, v in sorted(pairs, key=lambda x: str(x[0]))[:35]:
                print(G+f"  {str(k)[:36]:36} -> "+C+f"{str(v)[:36]}"+RS)
            if len(pairs) > 35:
                print(DM+f"  ... +{len(pairs)-35} pares no arquivo"+RS)
        sub()
        print(DM+"  [1] Incluir ou alterar par"+RS)
        print(DM+"  [2] Remover par"+RS)
        print(DM+"  [3] Listar catalogo de TARIFAS (nomes em config)"+RS)
        print(DM+"  [0] Voltar"+RS)
        op = prompt("Opcao").strip()
        if op == "0":
            return
        if op == "1":
            chave_micro = ""
            if atividades_micro and confirmar("Escolher atividade da planilha carregada?", default=True):
                idx = selecionar_paginado("ATIVIDADE no micro", atividades_micro, page_size=8)
                if idx >= 0:
                    chave_micro = atividades_micro[idx]
            if not chave_micro:
                chave_micro = prompt("Nome EXATO da atividade no microplanejamento", "")
            if not chave_micro:
                aviso("Nome vazio."); continue
            val_tarifa = ""
            if nomes_tarifa and confirmar("Escolher tarifa na lista importada?", default=True):
                idx = selecionar_paginado("TARIFA (orcamento)", nomes_tarifa, page_size=8)
                if idx >= 0:
                    val_tarifa = nomes_tarifa[idx]
            if not val_tarifa:
                val_tarifa = prompt("Nome da TARIFA (chave em tarifas)", "")
            if not val_tarifa:
                aviso("Tarifa vazio."); continue
            if val_tarifa not in tarifas:
                if not confirmar(
                    f"  '{str(val_tarifa)[:42]}' nao esta em tarifas. Gravar mesmo assim?", default=False
                ):
                    continue
            cfg.setdefault("de_para", {})
            cfg["de_para"][chave_micro] = val_tarifa
            salvar_config(cfg)
            ok("Mapeamento salvo em config.json.")
        elif op == "2":
            keys = sorted([k for k in d.keys() if not str(k).startswith("_")], key=str)
            if not keys:
                aviso("Nada para remover.")
                continue
            idx = selecionar_paginado("REMOVER mapeamento", [str(k) for k in keys])
            if idx >= 0:
                del cfg["de_para"][keys[idx]]
                salvar_config(cfg)
                ok("Removido.")
        elif op == "3":
            if not nomes_tarifa:
                aviso("Nenhuma tarifa em config. Use menu [2] Importar.")
            else:
                for i, n in enumerate(nomes_tarifa[:60], 1):
                    print(DM+f"  {i:3}. {str(n)[:58]}"+RS)
                if len(nomes_tarifa) > 60:
                    print(DM+f"  ... +{len(nomes_tarifa)-60}"+RS)
            input(DM+"\n  [ENTER] "+RS)
        else:
            aviso("Opcao invalida.")

# ──────────────────────────────────────────────
#  COLUMN MAPPING: KNOWN FIRST, THEN FALLBACK
# ──────────────────────────────────────────────
def encontrar_coluna(cols, campo):
    """Try KNOWN_COLUMNS exact matches first, then fuzzy fallback."""
    known = KNOWN_COLUMNS.get(campo, [])
    for k in known:
        if k in cols:
            return k
    # Fuzzy fallback
    cn_map = {remover_acentos(c): c for c in cols}
    for k in known:
        kn = remover_acentos(k)
        for cn, original in cn_map.items():
            if kn in cn or cn in kn:
                return original
    return None

# ──────────────────────────────────────────────
#  FILE SELECTOR
# ──────────────────────────────────────────────
def buscar_arquivos_excel():
    return [f for f in os.listdir(DIR)
            if not f.startswith('~') and any(f.endswith(e) for e in ['.xlsx','.xls','.csv','.xlsm'])]

def _find_default_micro_path(cfg=None):
    """
    Prioridade: config.arquivo_micro (se existir no disco) > lista fixa > primeiro .xlsx com
    'inovesa'/'consolidado'/'exame'/'micro' no nome.
    """
    if cfg:
        pref = cfg.get("arquivo_micro")
        if pref and isinstance(pref, str):
            p = os.path.join(DIR, pref.strip())
            if os.path.exists(p):
                return p
    candidatos = [
        "MICROPLANEJAMENTO_CONSOLIDADO_INOVESA 1.xlsx",
        "exame.xlsx", "EXAME.xlsx", "Exame.xlsx",
        "microplanejamento.xlsx", "MICROPLANEJAMENTO.xlsx",
    ]
    for c in candidatos:
        p = os.path.join(DIR, c)
        if os.path.exists(p):
            return p
    for f in buscar_arquivos_excel():
        n = remover_acentos(f)
        if "inovesa" in n or "consolidado" in n:
            return os.path.join(DIR, f)
    for f in buscar_arquivos_excel():
        n = remover_acentos(f)
        if "exame" in n or "micro" in n:
            return os.path.join(DIR, f)
    return None

def _prefer_micro_sheet(abas):
    for a in abas:
        if "microplanejamento_abril_junho" in remover_acentos(a).replace(" ", ""):
            return a
    for a in abas:
        if "microplanejamento" in remover_acentos(a):
            return a
    for a in abas:
        if "inovesa" in remover_acentos(a) or "consolidado" in remover_acentos(a):
            return a
    return abas[0] if abas else None

def _find_default_ct_path():
    for f in buscar_arquivos_excel():
        if f == STG_FILENAME:
            continue
        n = remover_acentos(f)
        if "ct_313" in n or ("ct" in n and "313" in n):
            return os.path.join(DIR, f)
    return None

def selecionar_arquivo(titulo="Selecione um Arquivo"):
    """
    Seletor com navegacao de pastas.
    - [DIR] .. : sobe um nivel
    - [DIR] nome: entra na pasta
    - [ARQ] nome: seleciona arquivo
    """
    exts = (".xlsx", ".xls", ".xlsm", ".csv")
    cwd = DIR
    while True:
        try:
            nomes = sorted(os.listdir(cwd), key=lambda s: normalizar_chave(str(s)))
        except Exception as ex:
            erro(f"Nao foi possivel abrir pasta: {ex}")
            return None

        entries = []
        parent = os.path.dirname(cwd.rstrip("\\/"))
        if parent and parent != cwd:
            entries.append(("dir_up", "..", parent))

        for n in nomes:
            p = os.path.join(cwd, n)
            if os.path.isdir(p):
                if n.startswith("."):
                    continue
                entries.append(("dir", n, p))
        for n in nomes:
            p = os.path.join(cwd, n)
            if os.path.isfile(p):
                ln = n.lower()
                if n.startswith("~") or n.startswith("~$"):
                    continue
                if ln.endswith(exts):
                    entries.append(("file", n, p))

        if not entries:
            aviso(f"Nada encontrado em: {cwd}")
            if cwd == DIR:
                return None
            cwd = parent if parent else DIR
            continue

        labels = []
        for kind, name, _ in entries:
            if kind == "dir_up":
                labels.append("[DIR] ..")
            elif kind == "dir":
                labels.append(f"[DIR] {name}")
            else:
                labels.append(f"[ARQ] {name}")

        cabecalho(titulo)
        print(DM+f"  Pasta atual: {cwd}"+RS)
        sub()
        idx = selecionar_paginado(titulo, labels, page_size=12, zero_label="Cancelar")
        if idx < 0:
            return None
        kind, name, path = entries[idx]
        if kind in ("dir_up", "dir"):
            cwd = path
            continue
        return path

# ──────────────────────────────────────────────
#  MICROPLANEJAMENTO
# ──────────────────────────────────────────────
def carregar_planilha_microplanejamento(cfg, caminho=None, modo_auto=False):
    if not caminho:
        caminho = selecionar_arquivo("MICROPLANEJAMENTO (ex. exame.xlsx)")
    if not caminho: return None
    try:
        xls = pd.ExcelFile(caminho)
        abas = xls.sheet_names
        if len(abas) == 1:
            aba = abas[0]
        else:
            if modo_auto:
                aba = _prefer_micro_sheet(abas)
            else:
                aba = selecionar("SELECIONE A ABA", abas)
                if aba is None: return None

        df = pd.read_excel(caminho, sheet_name=aba, header=0)
        cols = df.columns.tolist()

        # Mapear com KNOWN_COLUMNS primeiro
        faz_col  = encontrar_coluna(cols, "fazenda")
        chv_col  = encontrar_coluna(cols, "chave")
        area_col = encontrar_coluna(cols, "area")
        atv_col  = encontrar_coluna(cols, "atividade")

        # BETA: planilhas de testes podem nao ter CHAVE POLIGONO; usar NUCLEO como fallback automatico.
        if modo_auto and not chv_col:
            for c in cols:
                if "nucleo" in normalizar_chave(c):
                    chv_col = c
                    break

        # Se algum nao bateu, pedir manual
        mapeados = {"Fazenda": faz_col, "Talhao": chv_col, "Area(ha)": area_col, "Atividade": atv_col}
        todos_ok = all(v is not None for v in mapeados.values())

        sub()
        print(G+BL+"  MAPEAMENTO AUTOMATICO:"+RS)
        for label, val in mapeados.items():
            cor = G if val else R
            print(cor+f"  {label:12}: "+C+f"{val or '??? NAO ENCONTRADO'}"+RS)
        sub()

        if todos_ok and (modo_auto or confirmar("Usar este mapeamento?", default=True)):
            pass  # tudo certo
        else:
            if modo_auto:
                erro("Nao foi possivel mapear colunas do micro automaticamente.")
                return None
            print(G+"\n  Selecione manualmente cada coluna:\n"+RS)
            if not faz_col:
                idx = selecionar_paginado("COLUNA DA FAZENDA", cols)
                faz_col = cols[idx] if idx >= 0 else cols[0]
            if not chv_col:
                idx = selecionar_paginado("COLUNA DO TALHAO/CHAVE", cols)
                chv_col = cols[idx] if idx >= 0 else cols[1]
            if not area_col:
                idx = selecionar_paginado("COLUNA DA AREA (HECTARES)", cols)
                area_col = cols[idx] if idx >= 0 else cols[2]
            if not atv_col:
                idx = selecionar_paginado("COLUNA DA ATIVIDADE", cols)
                atv_col = cols[idx] if idx >= 0 else cols[3]

        sel_cols = [faz_col, chv_col, area_col, atv_col]
        sel_names = ["fazenda", "chave", "area_ha", "atividade"]

        equipe_col = None
        for c in cols:
            if normalizar_chave(c) in ("equipe", "nome equipe", "empresa"):
                equipe_col = c
                break
        if equipe_col:
            sel_cols.append(equipe_col)
            sel_names.append("equipe")
            print(G+f"  Coluna EQUIPE detectada: "+C+f"{equipe_col}"+RS)

        df_filtro = df[sel_cols].copy()
        df_filtro.columns = sel_names
        for c_txt in ("fazenda", "chave", "atividade", "equipe"):
            if c_txt in df_filtro.columns:
                df_filtro[c_txt] = (
                    df_filtro[c_txt]
                    .astype(str)
                    .str.replace(r"\s+", " ", regex=True)
                    .str.strip()
                )
                df_filtro.loc[df_filtro[c_txt].str.lower().isin(["nan", "none", ""]), c_txt] = None
        df_filtro = df_filtro.dropna(subset=["atividade", "area_ha"])
        df_filtro["area_ha"] = df_filtro["area_ha"].apply(lambda x: _to_float_br(x, default=0.0))

        validos = df_filtro[df_filtro["area_ha"] > 0]
        ok(f"Carregadas {len(validos)} atividades validas.")
        return validos
    except Exception as e:
        erro(f"Erro ao ler microplanejamento: {e}")
        return None

# ──────────────────────────────────────────────
#  IMPORTADOR CT_313
# ──────────────────────────────────────────────
def modulo_importar_tarifas(cfg):
    cabecalho("IMPORTAR TARIFAS ORCADAS (CT_313)")
    caminho = selecionar_arquivo("PLANILHA DE ORCAMENTO (CT_313 ou Tarifas)")
    if not caminho: return

    try:
        print(DM+"  Carregando arquivo..."+RS)
        xls = pd.ExcelFile(caminho)
        aba = selecionar("SELECIONE A ABA (ex: Preco Final)", xls.sheet_names)
        if aba is None: return

        print(DM+f"  Lendo aba '{aba}'..."+RS)
        df = pd.read_excel(caminho, sheet_name=aba, nrows=1000)
        cols_ct = df.columns.tolist()

        # Tentar mapear automaticamente
        col_atv = encontrar_coluna(cols_ct, "atividade")
        sub()
        print(G+BL+"  MAPEAMENTO:"+RS)
        print(G+f"  Atividade: "+C+f"{col_atv or '???'}"+RS)
        sub()

        if not col_atv or not confirmar("Usar este mapeamento?", default=True):
            idx = selecionar_paginado("COLUNA DA ATIVIDADE", cols_ct)
            col_atv = cols_ct[idx] if idx >= 0 else None
            if not col_atv: aviso("Atividade obrigatoria."); return

        # Para HH e Preco, perguntar diretamente
        print(G+"\n  Selecione as colunas adicionais (0 = ignorar):\n"+RS)
        idx = selecionar_paginado("COLUNA DE HH/HA", cols_ct)
        col_hh = cols_ct[idx] if idx >= 0 else None
        idx = selecionar_paginado("COLUNA DE PRECO UNITARIO", cols_ct)
        col_preco = cols_ct[idx] if idx >= 0 else None

        tarifas = cfg.get("tarifas", {})
        importadas = 0
        for _, row in df.iterrows():
            nome = str(row.get(col_atv, "")).strip()
            if not nome or nome.lower() == "nan": continue
            hh = 0 if not col_hh else row.get(col_hh, 0)
            preco = 0 if not col_preco else row.get(col_preco, 0)
            if pd.notna(hh) and str(hh).strip() != "":
                hh_val = float(str(hh).replace(",", "."))
            else:
                hh_val = resolver_rendimento_hh(cfg, tarifas, nome)
            preco_val = float(str(preco).replace(',','.')) if pd.notna(preco) else 0.0
            tarifas[nome] = {
                "rendimento_hh": hh_val, "preco_unit": preco_val,
                "recurso": "homem", "eficiencia": 1.0
            }
            importadas += 1

        cfg["tarifas"] = tarifas
        salvar_config(cfg)
        ok(f"{importadas} tarifas integradas!")
        sem_hh = [k for k, v in tarifas.items() if float(v.get("rendimento_hh", 0) or 0) <= 0]
        sem_preco = [k for k, v in tarifas.items() if float(v.get("preco_unit", 0) or 0) <= 0]
        if sem_hh:
            print(Y+f"  Pos-import: {len(sem_hh)} tarifa(s) com HH zerado."+RS)
            for x in sem_hh[:5]:
                print(DM+f"    - {str(x)[:55]}"+RS)
        if sem_preco:
            print(Y+f"  Pos-import: {len(sem_preco)} tarifa(s) com preco zerado."+RS)
            for x in sem_preco[:5]:
                print(DM+f"    - {str(x)[:55]}"+RS)
    except Exception as e:
        erro(f"Erro ao importar: {e}")

    input(DM+"\n  [ENTER para voltar] "+RS)

# ──────────────────────────────────────────────
#  DECLIVIDADE
# ──────────────────────────────────────────────
def aviso_politica_tarifas_planas():
    """Politica comercial-executiva: base CT sempre 'plana' (Classe I) onde o micro nao discrimina."""
    sub()
    print(Y+BL+"  POLITICA DE DECLIVIDADE E ROÇADA MANUAL (CT)"+RS)
    print(
        DM+"  Na CT, ROÇADA MANUAL CLASSE I = terreno mais plano (menos HH/ha, menor R$/ha); "
        "CLASSE V = declive maximo (mais HH, mais R$/ha — obra mais cara e precos mais altos)."+RS
    )
    print(
        Y+"  Padrao deste app: o exame nao informa a classe por talhao — usamos sempre as linhas "
        "EQUIVALENTES AO CENARIO MAIS PLANO (ex.: ROÇADA MANUAL CLASSE I) no de_para fixo."+RS
    )
    print(
        DM+"  Interpretacao: simulacao conservadora em LUCRO — como se nao houvesse premio de "
        "declividade na mixagem; em campo inclinado real, revise o menu [4] de_para para "
        "Classes II–V conforme a CT."+RS
    )
    sub()

def avaliar_terreno(df_faz):
    print(G+BL+"\n  REFINAMENTO DE DECLIVIDADE\n"+RS)
    print(DM+"  Isto aplica um fator multiplicativo extra sobre HH/ha (1,0 / 1,15 / 1,30), "
              "independente da classe I–V da CT. Classe I vs V ja esta na linha de preco da CT; "
              "este passo e so para penalizar o cronograma se quiser simular declive geral."+RS)
    if not confirmar("Aplicar penalidade por declive?", default=False):
        df_faz["penalidade"] = 1.0
        return df_faz
    terrenos = ["Plano (Base x1.0)", "Misto (x1.15)", "Inclinado (x1.30)"]
    t = selecionar("DECLIVIDADE", terrenos)
    if t and "Inclinado" in t:  df_faz["penalidade"] = 1.3
    elif t and "Misto" in t:    df_faz["penalidade"] = 1.15
    else:                       df_faz["penalidade"] = 1.0
    return df_faz

def filtrar_atividades_por_texto(atividades, texto):
    """Nomes cuja versao sem acento contem o filtro (substring)."""
    t = remover_acentos(texto)
    if not t:
        return []
    out = []
    for a in atividades:
        if t in remover_acentos(str(a)):
            out.append(a)
    return out

def _to_float_br(v, default=0.0):
    """Converte numero em formato BR/EN para float de forma tolerante."""
    try:
        if v is None:
            return float(default)
        if isinstance(v, (int, float)) and not pd.isna(v):
            return float(v)
        s = str(v).strip()
        if not s or s.lower() == "nan":
            return float(default)
        s = s.replace(" ", "")
        if "," in s and "." in s:
            # Assume separador de milhar + decimal (pt-BR): 1.234,56
            s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            s = s.replace(",", ".")
        return float(s)
    except Exception:
        return float(default)

def fazendas_unicas_micro(df):
    """Nomes unicos da coluna fazenda (micro), ordenados."""
    return sorted(
        {str(x).strip() for x in df["fazenda"].dropna().unique() if str(x).strip()},
        key=lambda s: normalizar_chave(s),
    )

def _resolver_fazenda_demo_ulianopolis(df):
    """
    Fazenda alvo no modo DEMO:
    1) Nome que contenha 'ulianopolis' (ex.: fazenda cadastrada como Ulianopolis SWG);
    2) Senao, a fazenda com mais linhas no micro (planilhas municipio Ulianopolis sem 'Ulianopolis' no nome da fazenda).
    """
    fazendas = fazendas_unicas_micro(df)
    for f in fazendas:
        if "ulianopolis" in normalizar_chave(f):
            return f
    best, nmax = None, -1
    for f in fazendas:
        n = len(df[df["fazenda"] == f])
        if n > nmax:
            nmax, best = n, f
    return best

def garantir_fazenda_ulianopolis_no_ct(cfg, df):
    """
    Modo DEMO: acrescenta em fazendas_ct todas as fazendas presentes no micro demo,
    para o aviso micro-vs-CT nao bloquear (CT pode nao listar municipio).
    Retorna quantos nomes novos foram adicionados.
    """
    idx = _indice_fazendas_ct(cfg)
    ad = 0
    for f in fazendas_unicas_micro(df):
        nk = normalizar_chave(f)
        if not nk:
            continue
        if nk not in idx:
            cfg.setdefault("fazendas_ct", []).append(f)
            idx[nk] = f
            ad += 1
    return ad


def reconstruir_demo_ulianopolis_a_partir_da_fonte():
    """
    Le USEESTAPLANILHAULIANOPOLIS.xlsx (Planilha1), filtra municipio Ulianopolis,
    gera ulianopolisswg.xlsx no formato esperado pelo carregador (NOME FAZENDA, CHAVE POLIGONO, ...).
    Retorna (n_linhas, n_atividades_unicas) ou None se a fonte nao existir.
    """
    src = os.path.join(DIR, DEMO_MICRO_SOURCE_FILENAME)
    if not os.path.exists(src):
        return None
    try:
        raw = pd.read_excel(src, sheet_name=0)
    except Exception:
        return None
    if raw.shape[1] < 11:
        return None
    # Colunas fixas pela ordem do export escritorio (USEESTA...)
    muni = raw.iloc[:, 3]
    mask = muni.astype(str).str.contains("ulian", case=False, na=False)
    sub = raw.loc[mask].copy()
    if sub.empty:
        sub = raw.copy()
    cod = sub.iloc[:, 0]
    nome_faz = sub.iloc[:, 1]
    nucleo = sub.iloc[:, 2]
    atividades = sub.iloc[:, 9]
    area = sub.iloc[:, 10]
    chaves = []
    for i in range(len(sub)):
        chaves.append(f"{str(cod.iloc[i]).strip()}_{str(nucleo.iloc[i]).strip()}_{i:04d}")
    out = pd.DataFrame({
        "NOME FAZENDA": nome_faz.astype(str).str.strip(),
        "CHAVE POLÍGONO": chaves,
        "ÁREA TRABALHADA ESTIMADA (HECTARE)": pd.to_numeric(area, errors="coerce"),
        "ATIVIDADES": atividades.astype(str).str.strip(),
    })
    out = out.dropna(subset=["ATIVIDADES"])
    out = out[pd.to_numeric(out["ÁREA TRABALHADA ESTIMADA (HECTARE)"], errors="coerce").fillna(0) > 0]
    dest = os.path.join(DIR, DEMO_MICRO_FILENAME)
    out.to_excel(dest, index=False, sheet_name="MICROPLANEJAMENTO_ULIANOPOLIS")
    return len(out), out["ATIVIDADES"].nunique()

def simular_comparativo_robo_rocador(demandas, dias_baseline, executores, jornada, prod_ha_h=0.18, custo_h=0.0):
    """
    Comparativo operacional pós-cronograma para substituição TOTAL das atividades de ROÇADA por robô.
    Cenário A: remove HH humana de roçada (equivalente teórico).
    Cenário B: humano sem roçada em paralelo com fila do robô.
    """
    total_hh = 0.0
    hh_rocada = 0.0
    area_rocada = 0.0
    for tarefas in demandas.values():
        for t in tarefas:
            hh = float(t.get("hh_total", 0) or 0)
            ar = float(t.get("area", 0) or 0)
            total_hh += hh
            if "rocada" in normalizar_chave(t.get("atividade", "")):
                hh_rocada += hh
                area_rocada += ar

    cap_hum_dia = float(executores) * float(jornada)
    hh_sem_rocada = max(0.0, total_hh - hh_rocada)
    dias_hum_sem_rocada = int(math.ceil(hh_sem_rocada / cap_hum_dia)) if cap_hum_dia > 0 else 0

    cap_robo_dia_ha = float(prod_ha_h) * float(jornada)
    if cap_robo_dia_ha > 0 and area_rocada > 0:
        dias_robo = int(math.ceil(area_rocada / cap_robo_dia_ha))
        horas_robo = area_rocada / float(prod_ha_h)
    else:
        dias_robo = 0
        horas_robo = 0.0

    dias_cenario_a = dias_hum_sem_rocada
    dias_cenario_b = max(dias_hum_sem_rocada, dias_robo)
    ganho_a = int(dias_baseline) - int(dias_cenario_a)
    ganho_b = int(dias_baseline) - int(dias_cenario_b)

    return {
        "dias_baseline": int(dias_baseline),
        "total_hh": float(total_hh),
        "hh_rocada_hum": float(hh_rocada),
        "area_rocada_ha": float(area_rocada),
        "executores": int(executores),
        "jornada_h": float(jornada),
        "prod_robo_ha_h": float(prod_ha_h),
        "custo_robo_h": float(custo_h),
        "cap_hum_dia_hh": float(cap_hum_dia),
        "cap_robo_dia_ha": float(cap_robo_dia_ha),
        "dias_hum_sem_rocada": int(dias_hum_sem_rocada),
        "dias_robo": int(dias_robo),
        "dias_cenario_a": int(dias_cenario_a),
        "dias_cenario_b": int(dias_cenario_b),
        "ganho_dias_a": int(ganho_a),
        "ganho_dias_b": int(ganho_b),
        "horas_robo_total": float(horas_robo),
        "custo_robo_total": float(horas_robo) * float(custo_h),
    }

def _eh_rocada(atv_nome):
    return "rocada" in normalizar_chave(atv_nome)

def construir_cronograma_humano_sem_rocada(cronograma_base, turmas, jornada, executores):
    """
    Remove atividades de roçada do cronograma humano e recompõe dias por turma
    mantendo ordem original de execução dentro de cada turma.
    """
    turmas_ops = {t["nome"]: int(t["operarios"]) for t in turmas}
    # Mantem ordem original de aparicao no cronograma base
    por_turma = defaultdict(list)
    for c in cronograma_base:
        atv = str(c.get("Atividade", ""))
        if _eh_rocada(atv):
            continue
        por_turma[str(c.get("Turma", ""))].append(dict(c))

    novo = []
    for nm_turma, itens in por_turma.items():
        if not itens:
            continue
        if nm_turma == "Pelotao_Unificado":
            n_ops = int(executores)
        else:
            n_ops = int(turmas_ops.get(nm_turma, 1))
        cap_dia = max(0.01, float(n_ops) * float(jornada))
        dia = 1
        saldo = cap_dia
        for it in itens:
            hh_rest = float(it.get("HH", 0) or 0)
            if hh_rest <= 0.01:
                continue
            while hh_rest > 0.01:
                if saldo <= 0.01:
                    dia += 1
                    saldo = cap_dia
                cons = min(hh_rest, saldo)
                hh_rest -= cons
                saldo -= cons
                row = dict(it)
                row["Dia"] = dia
                row["HH"] = round(cons, 2)
                # custo ja vem por HH na base; reescala proporcionalmente
                c_old = float(it.get("Custo_MO", 0) or 0)
                hh_old = float(it.get("HH", 0) or 0)
                row["Custo_MO"] = round((cons / hh_old) * c_old, 2) if hh_old > 0.01 else 0.0
                novo.append(row)
    novo = sorted(novo, key=lambda r: (int(r.get("Dia", 0)), str(r.get("Turma", ""))))
    return novo

def construir_cronograma_robo_rocada(demandas, fazenda, jornada, prod_ha_h=0.18, custo_h=0.0):
    """
    Fila dedicada do robô para TODAS as atividades de roçada (paralelo ao humano).
    Capacidade diária em área: prod_ha_h * jornada.
    """
    tarefas = []
    for talhao, ls in demandas.items():
        for t in ls:
            atv = str(t.get("atividade", ""))
            if not _eh_rocada(atv):
                continue
            area = float(t.get("area", 0) or 0)
            if area <= 0.0001:
                continue
            tarefas.append({"Talhao": talhao, "Atividade": atv, "Area_ha": area})

    if not tarefas:
        return []

    cap_area_dia = max(0.0001, float(prod_ha_h) * float(jornada))
    dia = 1
    area_saldo_dia = cap_area_dia
    out = []
    for t in tarefas:
        area_rest = float(t["Area_ha"])
        while area_rest > 0.0001:
            if area_saldo_dia <= 0.0001:
                dia += 1
                area_saldo_dia = cap_area_dia
            area_exec = min(area_rest, area_saldo_dia)
            area_rest -= area_exec
            area_saldo_dia -= area_exec
            hh = area_exec / float(prod_ha_h) if float(prod_ha_h) > 0 else 0.0
            out.append({
                "Dia": int(dia),
                "Fazenda": fazenda,
                "Talhao": t["Talhao"],
                "Atividade": t["Atividade"],
                "Turma": "ROBO_ROCADOR",
                "Operarios": 1,
                "HH": round(hh, 2),
                "Custo_MO": round(hh * float(custo_h), 2),
                "Modo": "RoboRocador",
                "Area_ha": round(area_exec, 4),
            })
    return out

def construir_cronograma_mecanizado(demandas, fazenda, jornada, recursos_mec):
    """
    Fila dedicada para cada recurso mecanizado.
    recursos_mec: [{"nome", "prod_ha_h", "custo_h", "atividades": set}]
    """
    todas_filas = []
    for rec in recursos_mec:
        nome_rec = rec["nome"]
        prod = float(rec.get("prod_ha_h") or 0.18)
        custo = float(rec.get("custo_h") or 0)
        tarefas = []
        for talhao, ls in demandas.items():
            for t in ls:
                atv = str(t.get("atividade", ""))
                if atv not in rec.get("atividades", set()):
                    continue
                area = float(t.get("area", 0) or 0)
                if area <= 0.0001:
                    continue
                tarefas.append({"Talhao": talhao, "Atividade": atv, "Area_ha": area})
        if not tarefas:
            continue
        cap_area_dia = max(0.0001, prod * float(jornada))
        dia = 1
        saldo = cap_area_dia
        for t in tarefas:
            rest = float(t["Area_ha"])
            while rest > 0.0001:
                if saldo <= 0.0001:
                    dia += 1
                    saldo = cap_area_dia
                exe = min(rest, saldo)
                rest -= exe
                saldo -= exe
                hh = exe / prod if prod > 0 else 0.0
                todas_filas.append({
                    "Dia": int(dia),
                    "Fazenda": fazenda,
                    "Talhao": t["Talhao"],
                    "Atividade": t["Atividade"],
                    "Turma": f"MEC_{nome_rec}",
                    "Operarios": 1,
                    "HH": round(hh, 2),
                    "Custo_MO": round(hh * custo, 2),
                    "Modo": "Mecanizado",
                    "Area_ha": round(exe, 4),
                })
    return sorted(todas_filas, key=lambda r: (int(r.get("Dia", 0)), str(r.get("Turma", ""))))


def construir_cronograma_humano_sem_mecanizadas(cronograma_base, turmas, jornada, executores, atividades_mec):
    """Remove atividades assumidas por mecanizados do cronograma humano e recompoe dias."""
    turmas_ops = {t["nome"]: int(t["operarios"]) for t in turmas}
    por_turma = defaultdict(list)
    for c in cronograma_base:
        atv = str(c.get("Atividade", ""))
        if atv in atividades_mec:
            continue
        por_turma[str(c.get("Turma", ""))].append(dict(c))
    novo = []
    for nm_turma, itens in por_turma.items():
        if not itens:
            continue
        n_ops = int(executores) if nm_turma == "Pelotao_Unificado" else int(turmas_ops.get(nm_turma, 1))
        cap_dia = max(0.01, float(n_ops) * float(jornada))
        dia = 1
        saldo = cap_dia
        for it in itens:
            hh_rest = float(it.get("HH", 0) or 0)
            if hh_rest <= 0.01:
                continue
            while hh_rest > 0.01:
                if saldo <= 0.01:
                    dia += 1
                    saldo = cap_dia
                cons = min(hh_rest, saldo)
                hh_rest -= cons
                saldo -= cons
                row = dict(it)
                row["Dia"] = dia
                row["HH"] = round(cons, 2)
                c_old = float(it.get("Custo_MO", 0) or 0)
                hh_old = float(it.get("HH", 0) or 0)
                row["Custo_MO"] = round((cons / hh_old) * c_old, 2) if hh_old > 0.01 else 0.0
                novo.append(row)
    return sorted(novo, key=lambda r: (int(r.get("Dia", 0)), str(r.get("Turma", ""))))


def _cadastrar_recursos_mecanizados_sn(atividades_reais, cfg=None):
    """Cadastrar N recursos mecanizados com seleção de atividades via S/N."""
    cand_mec = atividades_candidatas_mecanizado(atividades_reais, cfg)
    pool = list(atividades_reais)
    if cand_mec:
        sub()
        print(G+BL+"  LISTA DE ATIVIDADES (modo mecanizado)"+RS)
        print(
            DM
            + f"  Encontradas {len(cand_mec)} candidata(s) (nome: trator, mec., solo mec, etc.; ou tipo HM na tarifa)."
            + RS
        )
        if confirmar("  Mostrar apenas candidatas a mecanizado na pergunta S/N abaixo?", default=True):
            pool = cand_mec
        else:
            pool = list(atividades_reais)
    elif cfg:
        aviso("Nenhuma candidata automatica; listando todas as atividades da fazenda.")
    recursos = []
    while True:
        sub()
        print(G+BL+f"  MODO MECANIZADO — recurso #{len(recursos)+1}"+RS)
        nome = prompt("Nome do recurso (ex: Robo Rocador, Trator X)", f"Mecanizado_{len(recursos)+1}")
        prod = pedir_float("Produtividade (ha/h)", 0.18)
        custo = pedir_float("Custo (R$/h, 0 se placeholder)", 0.0, allow_zero=True)
        print(G+BL+f"\n  Selecionar atividades para '{nome}' (S/N):"+RS)
        print(DM+"  s=sim  n=nao  a=nao e encerrar  ok=sim e encerrar"+RS)
        atvs = set()
        cur_all = sorted(pool, key=str)
        for i, a in enumerate(cur_all, 1):
            v = prompt(f"[{i}/{len(cur_all)}] Vincular '{str(a)[:54]}'? (s/n/a/ok)", "")
            v = str(v).strip().lower()
            if v in ("s", "sim", "y", "yes"):
                atvs.add(a)
            elif v == "a":
                ok("Selecao encerrada (sem vincular esta).")
                break
            elif v == "ok":
                atvs.add(a)
                ok("Selecao encerrada por comando rapido.")
                break
        if not atvs:
            aviso("Nenhuma atividade selecionada para este recurso.")
        else:
            recursos.append({"nome": nome, "prod_ha_h": prod, "custo_h": custo, "atividades": atvs})
            ok(f"Recurso '{nome}': {len(atvs)} atividades, {prod} ha/h, R$ {custo}/h")
        if not confirmar("Adicionar mais um recurso mecanizado?", default=False):
            break
    return recursos


def _parse_lista_numeros(txt, as_int=False):
    out = []
    for p in str(txt).replace(";", ",").split(","):
        s = p.strip()
        if not s:
            continue
        try:
            v = float(s.replace(",", "."))
            if as_int:
                v = int(round(v))
            if v > 0:
                out.append(v)
        except Exception:
            pass
    return sorted(set(out))


def coletar_config_comparativo_multifator(executores_base, jornada_base):
    """Coleta grade de cenarios (jornada/equipe) de forma antecipada e explicita."""
    sub()
    print(C+BL+"  [CENARIOS] CONFIGURAR COMPARATIVO MULTI-FATOR"+RS)
    print(DM+"  O comparativo sera exportado no Excel (COMPARATIVO_CENARIOS)."+RS)
    print(DM+"  Exemplo entradas: jornadas 4.3,5.3,8 | equipes 4,6,8,10"+RS)
    jornadas_txt = prompt("  Jornadas (h/dia) separadas por virgula", f"{jornada_base}")
    equipes_txt = prompt("  Equipes (executores) separadas por virgula", f"{executores_base}")
    jornadas = _parse_lista_numeros(jornadas_txt, as_int=False)
    equipes = _parse_lista_numeros(equipes_txt, as_int=True)
    if not jornadas:
        jornadas = [float(jornada_base)]
    if not equipes:
        equipes = [int(executores_base)]
    ok(f"Comparativo configurado: {len(jornadas)} jornada(s) x {len(equipes)} equipe(s).")
    return {"jornadas": jornadas, "equipes": equipes}


def simular_cenarios_multifator(
    total_hh,
    receita_total,
    custo_hora_tf,
    dias_meta,
    executores_base,
    jornada_base,
    jornadas_in=None,
    equipes_in=None,
    interativo=True,
):
    """
    Simulador de cenarios em lote (aproximacao operacional):
    dias ~ HH total / (executores * jornada)
    """
    jornadas = sorted(set(float(x) for x in (jornadas_in or [] if not interativo else [])))
    equipes = sorted(set(int(x) for x in (equipes_in or [] if not interativo else [])))
    if interativo:
        sub()
        print(C+BL+"  [CENARIOS] COMPARATIVO MULTI-FATOR"+RS)
        print(DM+"  Exemplo entradas: jornadas 4.3,5.3 | equipes 6,8,10"+RS)
        jornadas_txt = prompt("  Jornadas (h/dia) separadas por virgula", f"{jornada_base}")
        equipes_txt = prompt("  Equipes (executores) separadas por virgula", f"{executores_base}")
        jornadas = _parse_lista_numeros(jornadas_txt, as_int=False)
        equipes = _parse_lista_numeros(equipes_txt, as_int=True)
    if not jornadas:
        jornadas = [float(jornada_base)]
    if not equipes:
        equipes = [int(executores_base)]

    rows = []
    for j in jornadas:
        for e in equipes:
            cap = float(e) * float(j)
            dias = int(math.ceil(float(total_hh) / cap)) if cap > 0.01 else 0
            meses = dias / 22.0 if dias > 0 else 0.0
            custo = float(total_hh) * float(custo_hora_tf)
            lucro = float(receita_total) - custo
            margem = (lucro / float(receita_total) * 100.0) if float(receita_total) > 0.01 else 0.0
            ganho = int(dias_meta) - int(dias)
            rows.append({
                "Equipe": int(e),
                "Jornada_h_dia": float(j),
                "Dias_Simulados": int(dias),
                "Meses_Simulados": round(meses, 2),
                "Ganho_vs_Meta_dias": int(ganho),
                "HH_Total": round(float(total_hh), 2),
                "Receita_Total": round(float(receita_total), 2),
                "Custo_MO_Aprox": round(custo, 2),
                "Lucro_Aprox": round(lucro, 2),
                "Margem_Aprox_pct": round(margem, 2),
            })
    rows = sorted(rows, key=lambda r: (r["Dias_Simulados"], -r["Equipe"], -r["Jornada_h_dia"]))
    return rows

def _indice_fazendas_ct(cfg):
    """Chave normalizada -> nome canonico (ultimo da lista)."""
    d = {}
    for x in cfg.get("fazendas_ct") or []:
        s = str(x).strip()
        if not s:
            continue
        k = normalizar_chave(s)
        if k:
            d[k] = s
    return d

def micro_fazendas_ausentes_na_lista_ct(cfg, df):
    """
    Fazendas que aparecem no micro mas nao em config.fazendas_ct (comparacao por normalizar_chave).
    Se fazendas_ct vazia, retorna None (validacao desativada).
    """
    idx = _indice_fazendas_ct(cfg)
    if not idx:
        return None
    out = []
    for m in fazendas_unicas_micro(df):
        if normalizar_chave(m) not in idx:
            out.append(m)
    return out

def aviso_fazendas_micro_sem_cadastro_ct(cfg, df):
    """Aviso no startup / apos recarregar micro."""
    falta = micro_fazendas_ausentes_na_lista_ct(cfg, df)
    if falta is None:
        sub()
        print(
            DM + "  [Fazendas vs CT] Lista `fazendas_ct` vazia no config — "
            "cadastre em menu [6] as fazendas cobertas pelo orcamento CT (ex.: todas menos Ulianopolis)."+RS
        )
        return
    if not falta:
        return
    sub()
    print(
        Y + "  !  Fazendas no MICRO sem correspondencia na lista `fazendas_ct` (orcamento CT): "+RS
    )
    for x in falta:
        print(Y + f"      - {str(x)[:72]}" + RS)
    print(
        DM + "  Corrija o CT no escritorio ou adicione excecao em [6] se a fazenda estiver coberta."+RS
    )

def modulo_validar_fazendas_ct(cfg, df):
    """
    CRUD de `fazendas_ct`: nomes de fazenda que o orcamento CT considera cadastrados.
    O micro pode ter fazendas a mais (esquecimento no CT) — o scan compara por nome normalizado.
    """
    while True:
        cabecalho("FAZENDAS — micro vs lista CT (orcamento)")
        micro_list = fazendas_unicas_micro(df)
        ct_list = list(cfg.get("fazendas_ct") or [])
        idx = _indice_fazendas_ct(cfg)
        print(G + f"  No micro agora: {len(micro_list)} fazenda(s)" + RS)
        print(G + f"  Na lista fazendas_ct: {len(ct_list)}" + RS)
        falta = micro_fazendas_ausentes_na_lista_ct(cfg, df)
        if falta is None:
            print(Y + "  Lista CT vazia — nenhuma comparacao possivel." + RS)
        elif falta:
            print(Y + f"  Ausentes na lista CT ({len(falta)}):" + RS)
            for x in falta[:25]:
                print(Y + f"    - {str(x)[:68]}" + RS)
            if len(falta) > 25:
                print(DM + f"    ... +{len(falta)-25}" + RS)
        else:
            ok("  Todas as fazendas do micro constam em fazendas_ct.")
        sub()
        print(DM + "  [1] Ver / listar fazendas_ct"+RS)
        print(DM + "  [2] Adicionar uma fazenda a fazendas_ct"+RS)
        print(DM + "  [3] Importar TODAS as fazendas do micro para fazendas_ct (substitui lista)"+RS)
        print(DM + "  [4] Remover uma fazenda da lista"+RS)
        print(DM + "  [5] Colar varios nomes (virgula ou ponto-e-virgula)"+RS)
        print(DM + "  [6] Limpar lista (fazendas_ct = [])"+RS)
        print(DM + "  [0] Voltar"+RS)
        op = prompt("Opcao").strip()
        if op == "0":
            return
        if op == "1":
            if not ct_list:
                aviso("Lista vazia.")
            else:
                for i, x in enumerate(ct_list, 1):
                    print(G + f"  {i:2}. " + C + str(x)[:68] + RS)
            input(DM + "\n  [ENTER] "+RS)
        elif op == "2":
            nome = prompt("Nome EXATO como no micro ou na CT", "")
            if nome.strip():
                k = normalizar_chave(nome)
                if k in idx:
                    aviso("Ja consta (mesma chave normalizada).")
                else:
                    cfg.setdefault("fazendas_ct", []).append(nome.strip())
                    salvar_config(cfg)
                    ok("Adicionado.")
        elif op == "3":
            if not confirmar(
                "Substituir fazendas_ct pelas "
                f"{len(micro_list)} fazendas unicas do micro? (nao altera a planilha CT .xlsm)",
                default=False,
            ):
                continue
            cfg["fazendas_ct"] = micro_list[:]
            salvar_config(cfg)
            ok(f"fazendas_ct = {len(micro_list)} nomes (copia do micro).")
        elif op == "4":
            if not ct_list:
                aviso("Lista vazia.")
                continue
            idx_r = selecionar_paginado("REMOVER", ct_list, page_size=10)
            if idx_r >= 0:
                cfg["fazendas_ct"].pop(idx_r)
                salvar_config(cfg)
                ok("Removido.")
        elif op == "5":
            txt = prompt("Nomes separados por virgula ou ;", "")
            partes = []
            for chunk in txt.replace(";", ",").split(","):
                s = str(chunk).strip()
                if s:
                    partes.append(s)
            ad = 0
            seen = _indice_fazendas_ct(cfg)
            for s in partes:
                k = normalizar_chave(s)
                if k and k not in seen:
                    cfg.setdefault("fazendas_ct", []).append(s)
                    seen[k] = s
                    ad += 1
            if ad:
                salvar_config(cfg)
                ok(f"+{ad} nome(s) novos.")
            else:
                aviso("Nenhum nome novo.")
        elif op == "6":
            if confirmar("Zerar fazendas_ct?", default=False):
                cfg["fazendas_ct"] = []
                salvar_config(cfg)
                ok("Lista limpa.")
        else:
            aviso("Opcao invalida.")

def menu_vincular_atividades_turma(turma, atividades_reais):
    """
    Vincula atividades a uma turma.
    Padrao: percurso S/N atividade-por-atividade.
    Fallback: filtro/lista/paginacao acessiveis via menu auxiliar.
    """
    atv_set = set(turma["atividades"])

    def _percurso_sn():
        cur_all = sorted(atividades_reais, key=lambda x: str(x))
        print(G+BL+f"\n  TURMA '{turma['nome']}' — percurso S/N ({len(cur_all)} atividades)"+RS)
        print(DM+"  s=vincular  n=desvincular  a=nao e encerrar  ok=sim e encerrar  ENTER=manter atual"+RS+"\n")
        for i, a in enumerate(cur_all, 1):
            mk = "X" if a in atv_set else " "
            v = prompt(f"[{i}/{len(cur_all)}] [{mk}] '{str(a)[:54]}' (s/n/a/ok)", "")
            v = str(v).strip().lower()
            if v in ("s", "sim", "y", "yes"):
                atv_set.add(a)
            elif v in ("n", "nao", "não", "no"):
                atv_set.discard(a)
            elif v == "a":
                ok("Percurso encerrado (sem alterar esta atividade).")
                break
            elif v == "ok":
                atv_set.add(a)
                ok("Percurso encerrado por comando rapido.")
                break
        ok(f"Percurso concluido. Vinculadas: {len(atv_set)}")

    def _assistente_sn_vinculos():
        """
        Revisao guiada S/N das atividades da turma:
        - ENTER: manter
        - n: remover
        - t: trocar por outra atividade
        - a: adicionar nova atividade agora
        - ok: encerrar assistente
        """
        while True:
            cur_all = sorted(atividades_reais, key=lambda x: str(x))
            cur_v = sorted(atv_set, key=lambda x: str(x))
            print(G+BL+f"\n  ASSISTENTE S/N — TURMA '{turma['nome']}'"+RS)
            print(DM+"  ENTER=manter  n=remover  t=trocar  a=adicionar  ok=encerrar"+RS+"\n")
            for i, a in enumerate(cur_all, 1):
                if a not in atv_set:
                    continue
                v = prompt(f"[{i}/{len(cur_all)}] '{str(a)[:54]}' (ENTER/n/t/a/ok)", "").strip().lower()
                if not v:
                    continue
                if v in ("ok",):
                    ok("Assistente encerrado.")
                    return
                if v in ("n", "nao", "não", "no"):
                    atv_set.discard(a)
                    continue
                if v in ("a",):
                    disp_add = [x for x in cur_all if x not in atv_set]
                    if not disp_add:
                        aviso("Nao ha atividade disponivel para adicionar.")
                        continue
                    idx_add = selecionar_paginado("ADICIONAR ATIVIDADE", disp_add)
                    if idx_add >= 0:
                        atv_set.add(disp_add[idx_add])
                        ok("Adicionada.")
                    continue
                if v in ("t", "trocar"):
                    disp = [x for x in cur_all if x != a]
                    idxd = selecionar_paginado("DESTINO DA TROCA", disp)
                    if idxd >= 0:
                        dest = disp[idxd]
                        atv_set.discard(a)
                        atv_set.add(dest)
                        ok(f"Troca: '{str(a)[:40]}' -> '{str(dest)[:40]}'.")
                    continue
            if not confirmar("Repassar assistente S/N novamente?", default=False):
                return

    _percurso_sn()

    while True:
        cur = sorted(atv_set, key=lambda x: str(x))
        sub()
        print(G+BL+f"  TURMA: {turma['nome']} ({turma['operarios']} ops) — {len(cur)} atividade(s) vinculadas"+RS)
        print(DM+"  [1] Refazer percurso S/N"+RS)
        print(DM+"  [2] Adicionar por filtro de texto"+RS)
        print(DM+"  [3] Adicionar por lista/indices (fallback)"+RS)
        print(DM+"  [4] Remover por filtro"+RS)
        print(DM+"  [5] Remover UMA (lista)"+RS)
        print(DM+"  [6] Ver vinculadas"+RS)
        print(DM+"  [7] Trocar atividade (substituir 1:1)"+RS)
        print(DM+"  [8] Assistente inteligente S/N (revisao guiada)"+RS)
        print(DM+"  [0] Concluir esta turma"+RS)
        sub()
        op = prompt("Opcao", "0").strip()
        if op == "0":
            turma["atividades"] = sorted(atv_set, key=lambda x: str(x))
            return
        if op == "1":
            _percurso_sn()
        elif op == "2":
            filtro = prompt("Texto no nome (ex: roçada)", "")
            if not str(filtro).strip():
                aviso("Filtro vazio.")
                continue
            matches = filtrar_atividades_por_texto(atividades_reais, filtro)
            if not matches:
                aviso("Nenhuma atividade bateu com o filtro.")
                continue
            print(G+f"\n  {len(matches)} encontrada(s):"+RS)
            for m in matches[:12]:
                print(DM+f"    - {str(m)[:62]}"+RS)
            if len(matches) > 12:
                print(DM+f"    ... +{len(matches)-12}"+RS)
            if confirmar("Adicionar TODAS ao vinculo desta turma?", default=True):
                for m in matches:
                    atv_set.add(m)
                ok(f"+{len(matches)} atividades.")
            else:
                for i, m in enumerate(matches, 1):
                    if confirmar(f"  [{i}] {str(m)[:55]}", default=False):
                        atv_set.add(m)
        elif op == "3":
            disp = [a for a in atividades_reais if a not in atv_set]
            if not disp:
                aviso("Ja estao todas vinculadas ou lista vazia.")
                continue
            print(DM+f"\n  Indices de 1 a {len(disp)} (ex.: 1,3,5-8). ENTER = lista paginada"+RS)
            multi = prompt("Indices", "")
            if str(multi).strip():
                idxs = parse_intervalos_escolha(multi, len(disp))
                if not idxs:
                    aviso("Nenhum indice valido.")
                else:
                    for i in idxs:
                        atv_set.add(disp[i])
                    ok(f"+{len(idxs)} atividades.")
                continue
            idx = selecionar_paginado("ADICIONAR ATIVIDADE", disp)
            if idx >= 0:
                atv_set.add(disp[idx])
                ok("Adicionada.")
        elif op == "4":
            filtro = prompt("Remover cujo nome contem", "")
            if not str(filtro).strip():
                aviso("Filtro vazio.")
                continue
            rem = filtrar_atividades_por_texto(list(atv_set), filtro)
            if not rem:
                aviso("Nenhuma vinculada bateu com o filtro.")
                continue
            if confirmar(f"Remover {len(rem)} da turma '{turma['nome']}'?", default=True):
                for r in rem:
                    atv_set.discard(r)
                ok("Removidas.")
        elif op == "5":
            cur2 = sorted(atv_set, key=lambda x: str(x))
            if not cur2:
                aviso("Nada vinculado ainda.")
                continue
            idx = selecionar_paginado("REMOVER ATIVIDADE", cur2)
            if idx >= 0:
                atv_set.discard(cur2[idx])
                ok("Removida.")
        elif op == "6":
            cur2 = sorted(atv_set, key=lambda x: str(x))
            print(G+f"\n  Vinculadas ({len(cur2)}): "+RS)
            for x in cur2[:40]:
                print(DM+f"    - {str(x)[:62]}"+RS)
            if len(cur2) > 40:
                print(DM+f"    ... +{len(cur2)-40}"+RS)
            input(DM+"\n  [ENTER] "+RS)
        elif op == "7":
            cur2 = sorted(atv_set, key=lambda x: str(x))
            if not cur2:
                aviso("Nada vinculado para trocar.")
                continue
            old = selecionar_paginado("ATIVIDADE ORIGEM (será removida)", cur2)
            if old < 0:
                continue
            origem = cur2[old]
            disp = sorted([a for a in atividades_reais if a != origem], key=lambda x: str(x))
            if not disp:
                aviso("Nao ha atividade destino disponivel.")
                continue
            print(DM+"  Dica: ENTER para lista paginada ou use filtro por texto."+RS)
            filtro = prompt("Filtro do destino (opcional)", "")
            if str(filtro).strip():
                candidatos = filtrar_atividades_por_texto(disp, filtro)
                if not candidatos:
                    aviso("Nenhum destino bateu com o filtro.")
                    continue
                destino = selecionar("ATIVIDADE DESTINO", candidatos)
            else:
                idxd = selecionar_paginado("ATIVIDADE DESTINO", disp)
                destino = disp[idxd] if idxd >= 0 else None
            if not destino:
                continue
            if confirmar(f"Trocar '{str(origem)[:48]}' por '{str(destino)[:48]}'?", default=True):
                atv_set.discard(origem)
                atv_set.add(destino)
                ok("Troca aplicada.")
        elif op == "8":
            _assistente_sn_vinculos()
        else:
            aviso("Opcao invalida.")

def resolver_conflitos_e_reatribuir(turmas, atividades_reais):
    """
    Atividades com mais de uma turma: paralelo ou exclusivo.
    Reatribuicao: qualquer atividade -> turma executora (reforco / outra funcao).
    Retorna reatribuicao, paralelo, primaria.
    """
    reatribuicao = {}
    paralelo = {}
    primaria = {}

    def candidatos(atv):
        return [t["nome"] for t in turmas if atv in t["atividades"]]

    for atv in atividades_reais:
        c = candidatos(atv)
        if len(c) <= 1:
            continue
        sub()
        print(Y+f"  Conflito: '{str(atv)[:58]}'"+RS)
        print(DM+f"  Turmas: {', '.join(c)}"+RS)
        if confirmar("  Varias turmas em PARALELO (dividem a mesma demanda no tempo)?", default=True):
            paralelo[atv] = True
        else:
            paralelo[atv] = False
            p = selecionar("  Turma EXCLUSIVA para esta atividade", c)
            if p:
                primaria[atv] = p

    if confirmar("\n  Reatribuir atividades (reforco: outra turma executa, ex. adubacao faz uma roçada)?", default=False):
        nomes_turmas = [t["nome"] for t in turmas]
        while True:
            idx = selecionar_paginado("REATRIBUIR — escolha a ATIVIDADE", atividades_reais, page_size=6)
            if idx < 0:
                break
            atv = atividades_reais[idx]
            print(G+f"\n  Atividade: {str(atv)[:62]}"+RS)
            t_alvo = selecionar("  Turma que EXECUTA (capacidade desta turma)", nomes_turmas)
            if t_alvo:
                reatribuicao[atv] = t_alvo
                ok(f"Executora: '{t_alvo}' (sobrescreve vinculos anteriores para o cronograma).")

    return reatribuicao, paralelo, primaria

def turmas_que_executam(atv, turmas, reatribuicao, paralelo, primaria):
    """Lista de nomes de turma que trabalham nesta atividade no simulador."""
    if atv in reatribuicao:
        return [reatribuicao[atv]]
    c = [t["nome"] for t in turmas if atv in t["atividades"]]
    if not c:
        return []
    if len(c) == 1:
        return c
    if paralelo.get(atv, True):
        return c
    p = primaria.get(atv)
    return [p] if p else c

def atividades_por_filtro(atividades_reais, filtros_texto):
    """Retorna atividades cujo nome contem algum filtro (sem acento)."""
    filtros = [remover_acentos(x).strip() for x in (filtros_texto or []) if str(x).strip()]
    out = set()
    for atv in atividades_reais:
        nome = remover_acentos(str(atv))
        if any(f in nome for f in filtros):
            out.add(atv)
    return sorted(out, key=lambda x: str(x))


_FILTROS_NOME_CANDIDATAS_MECANIZADO = [
    "mecaniz",
    "maquina",
    "máquina",
    "trator",
    "motocovead",
    "motocultor",
    "robo",
    "robô",
    "pulveriz",
    "atomiz",
    "implemento",
    "coveador",
    "solo mec",
    "mec c/",
    "mec s/",
    "esteira",
    "drone",
    "máq",
    "maq.",
]

def atividades_candidatas_mecanizado(atividades_reais, cfg=None):
    """Atividades provavelmente mecanizadas: palavras-chave no nome e/ou tipo HM na tarifa CT."""
    cfg = cfg or {}
    tarifas = cfg.get("tarifas", {}) or {}
    merged = set(atividades_por_filtro(atividades_reais, _FILTROS_NOME_CANDIDATAS_MECANIZADO))
    for atv in atividades_reais:
        t_nome = resolver_chave_tarifa(cfg, tarifas, atv)
        row = tarifas.get(t_nome)
        if not isinstance(row, dict):
            continue
        tipo = str(row.get("tipo", "")).lower()
        try:
            hm = float(row.get("rendimento_hm", 0) or 0)
        except (TypeError, ValueError):
            hm = 0.0
        if "mecaniz" in tipo or hm > 0:
            merged.add(atv)
    return sorted(merged, key=str)


def sequencia_manutencao_seco_placeholder(cfg):
    aviso("Modo manutencao_seco: regras de sequencia ainda nao definidas (stub). Cascata desligada nesta execucao.")


def sequencia_manutencao_umido_placeholder(cfg):
    aviso("Modo manutencao_umido: regras de sequencia ainda nao definidas (stub). Cascata desligada nesta execucao.")


def _match_filtros_fase(nome_atv, filtros, exclusoes=None):
    """True se nome contem algum filtro (normalizar_chave) e nenhuma exclusao."""
    kn = normalizar_chave(nome_atv)
    for ex in exclusoes or []:
        exn = normalizar_chave(ex)
        if exn and exn in kn:
            return False
    for f in filtros or []:
        fn = normalizar_chave(f)
        if fn and fn in kn:
            return True
    return False


def eh_limpeza_quimica_pos_plantio(atv, seq_cfg):
    """Limpeza quimica pos-plantio: todos filtros presentes e nenhuma exclusao."""
    kn = normalizar_chave(atv)
    filtros = seq_cfg.get("limpeza_quimica_filtros") or ["limpeza", "quim"]
    exclusoes = seq_cfg.get("limpeza_quimica_exclusoes") or []
    for f in filtros:
        fn = normalizar_chave(f)
        if not fn or fn not in kn:
            return False
    for ex in exclusoes:
        exn = normalizar_chave(ex)
        if exn and exn in kn:
            return False
    return True


def _fases_ordem_config(seq_cfg, modo):
    if modo == "personalizado" and seq_cfg.get("personalizado_ordem"):
        return seq_cfg["personalizado_ordem"]
    if modo == "manutencao_swg":
        return seq_cfg.get("swg_fases") or []
    return seq_cfg.get("implantacao_fases") or []


def classificar_fase_cascata_valor(atv, seq_cfg, modo, atividades_plantio, atividades_irrig):
    """
    Retorna indice numerico de fase para cascata global.
    manutencao_seco/umido: 0.0 (sem cascata).
    """
    if modo in ("manutencao_seco", "manutencao_umido"):
        return 0.0
    if eh_limpeza_quimica_pos_plantio(atv, seq_cfg):
        return 8.0
    if atv in atividades_plantio or _match_filtros_fase(atv, seq_cfg.get("filtros_plantio") or ["plantio"], None):
        return 6.0
    if atv in atividades_irrig or _match_filtros_fase(atv, seq_cfg.get("filtros_irrigacao") or ["irrig"], None):
        return 7.0
    fases = _fases_ordem_config(seq_cfg, modo)
    for i, fase in enumerate(fases):
        if _match_filtros_fase(atv, fase.get("filtros") or [], fase.get("exclusoes")):
            return float(i)
    try:
        return float(seq_cfg.get("implantacao_outras_fase", 5.5))
    except (TypeError, ValueError):
        return 5.5


def _demanda_plantio_talhao(talhao, demanda_global, atividades_plantio):
    for p in atividades_plantio:
        if demanda_global.get((talhao, p), 0) > 0.01:
            return True
    return False


def limpeza_permitida_por_talhao(
    talhao, dia, seq_cfg, dia_termino_plantio, tem_plantio_previsto_no_talhao,
):
    """Se nao ha plantio no talhao, limpeza pos-plantio nao exige offset."""
    if not tem_plantio_previsto_no_talhao:
        return True
    d = dia_termino_plantio.get(talhao)
    if d is None:
        return False
    off = int(seq_cfg.get("offset_limpeza_quimica_dias", 30) or 30)
    return dia >= d + off


def _min_fase_cascata(
    demanda_global,
    seq_cfg,
    modo,
    usar_cascata,
    usar_bloqueio_global,
    atividades_bloqueadas,
    atividades_plantio,
    atividades_irrig,
    dia,
    dia_termino_plantio,
    tem_plantio_por_talhao,
):
    """Menor fase entre demandas ainda > 0 e elegiveis neste dia."""
    if not usar_cascata or modo in ("manutencao_seco", "manutencao_umido"):
        return None
    bloqueadas = set(atividades_bloqueadas or [])
    vals = []
    for (talhao, atv), hh in demanda_global.items():
        if hh <= 0.01:
            continue
        if usar_bloqueio_global and atv in bloqueadas:
            if _ha_trabalho_nao_bloqueado(demanda_global, bloqueadas):
                continue
        if eh_limpeza_quimica_pos_plantio(atv, seq_cfg):
            if not limpeza_permitida_por_talhao(
                talhao, dia, seq_cfg, dia_termino_plantio,
                tem_plantio_por_talhao.get(talhao, False),
            ):
                continue
        fv = classificar_fase_cascata_valor(atv, seq_cfg, modo, atividades_plantio, atividades_irrig)
        vals.append(fv)
    if not vals:
        return None
    return min(vals)


def pode_agendar_atividade_cascata(
    talhao, atv, demanda_global, seq_cfg, modo, usar_cascata,
    usar_bloqueio_global, atividades_bloqueadas,
    atividades_plantio, atividades_irrig,
    dia, dia_termino_plantio, tem_plantio_por_talhao, min_fase_dia,
):
    """Regras combinadas: cascata, bloqueio global, plantio antes de irrigacao, limpeza pos-plantio."""
    hh = demanda_global.get((talhao, atv), 0)
    if hh <= 0.01:
        return False
    if usar_bloqueio_global and atv in atividades_bloqueadas:
        if _ha_trabalho_nao_bloqueado(demanda_global, atividades_bloqueadas):
            return False
    if eh_limpeza_quimica_pos_plantio(atv, seq_cfg):
        if not limpeza_permitida_por_talhao(
            talhao, dia, seq_cfg, dia_termino_plantio,
            tem_plantio_por_talhao.get(talhao, False),
        ):
            return False
    if atv in atividades_irrig or _match_filtros_fase(atv, seq_cfg.get("filtros_irrigacao") or ["irrig"], None):
        if _demanda_plantio_talhao(talhao, demanda_global, atividades_plantio):
            return False
    if usar_cascata and modo not in ("manutencao_seco", "manutencao_umido"):
        fv = classificar_fase_cascata_valor(atv, seq_cfg, modo, atividades_plantio, atividades_irrig)
        if min_fase_dia is not None and abs(fv - min_fase_dia) > 1e-6 and fv > min_fase_dia + 1e-6:
            return False
    return True


def diagnosticar_sequencia_atividades(atividades_reais, seq_cfg, modo):
    """Avisos: atividades classificadas como 'outras' (fase intermediaria) e lista."""
    if modo in ("manutencao_seco", "manutencao_umido"):
        return
    ap = set(atividades_por_filtro(atividades_reais, seq_cfg.get("filtros_plantio") or ["plantio"]))
    ai = set(atividades_por_filtro(atividades_reais, seq_cfg.get("filtros_irrigacao") or ["irrig"]))
    outras = []
    for atv in atividades_reais:
        if eh_limpeza_quimica_pos_plantio(atv, seq_cfg):
            continue
        if atv in ap or atv in ai:
            continue
        fases = _fases_ordem_config(seq_cfg, modo)
        ok_fase = False
        for fase in fases:
            if _match_filtros_fase(atv, fase.get("filtros") or [], fase.get("exclusoes")):
                ok_fase = True
                break
        if not ok_fase:
            outras.append(atv)
    if outras:
        print(DM+f"\n  Sequencia ({modo}): {len(outras)} atividade(s) em fase generica 'demais (antes plantio)'."+RS)
        print(DM+"  Executam antes de plantio, sem fase fixa na cascata. Para priorizar, adicione filtros em config.sequencia."+RS)
        for a in sorted(outras, key=str)[:15]:
            print(DM+f"    - {str(a)[:70]}"+RS)
        if len(outras) > 15:
            print(DM+f"    ... +{len(outras)-15}"+RS)
    else:
        ok("Todas as atividades possuem fase explicita na sequencia.")


def _ha_trabalho_nao_bloqueado(demanda_global, atividades_bloqueadas):
    """True se ainda existe demanda >0 para atividade fora do grupo bloqueado."""
    bloqueadas = set(atividades_bloqueadas or [])
    for (_, atv), hh in demanda_global.items():
        if hh > 0.01 and atv not in bloqueadas:
            return True
    return False

def auditar_cadeia_dados(cfg, demandas, atividades_reais, session_hh=None):
    """Auditoria unificada: micro → de_para → tarifa CT → HH/Preço/Custo_hora."""
    tarifas = cfg.get("tarifas", {})
    de_para = cfg.get("de_para", {})
    sem_depara = []
    sem_tarifa = []
    sem_hh = []
    sem_preco = []
    total = 0
    for talhao, lista in demandas.items():
        for t in lista:
            total += 1
            atv = t["atividade"]
            chave = resolver_chave_tarifa(cfg, tarifas, atv)
            if atv not in de_para and atv != chave:
                pass
            elif atv not in de_para and chave == atv and chave not in tarifas:
                sem_depara.append(str(atv)[:55])
            if chave not in tarifas:
                sem_tarifa.append(f"{str(atv)[:40]} → {str(chave)[:40]}")
            else:
                entry = tarifas[chave]
                rh_e = float(entry.get("rendimento_hh", 0) or 0)
                if session_hh and (atv in session_hh or chave in session_hh):
                    rh_e = max(rh_e, 1e-6)
                if rh_e <= 0:
                    sem_hh.append(f"{str(chave)[:55]}")
                if float(entry.get("preco_unit", 0) or 0) <= 0:
                    sem_preco.append(f"{str(chave)[:55]}")
    sub()
    print(G+BL+"  AUDITORIA CADEIA DE DADOS"+RS)
    print(G+f"  Demandas: {total} | Atividades unicas: {len(atividades_reais)} | de_para: {len(de_para)} | Tarifas CT: {len(tarifas)}"+RS)
    if session_hh:
        print(DM+f"  Overrides HH/ha nesta execucao: {len(session_hh)} chave(s) (nao gravados no config)."+RS)
    if sem_depara:
        u = sorted(set(sem_depara))
        print(Y+f"\n  Sem de_para ({len(u)}) — atividade micro nao mapeada:"+RS)
        for x in u[:10]:
            print(Y+f"    - {x}"+RS)
        if len(u) > 10:
            print(DM+f"    ... +{len(u)-10}"+RS)
    if sem_tarifa:
        u = sorted(set(sem_tarifa))
        print(Y+f"\n  Sem tarifa CT ({len(u)}) — chave nao encontrada no orcamento importado:"+RS)
        for x in u[:10]:
            print(Y+f"    - {x}"+RS)
        if len(u) > 10:
            print(DM+f"    ... +{len(u)-10}"+RS)
    if sem_hh:
        u = sorted(set(sem_hh))
        print(Y+f"\n  HH zerado ({len(u)}) — rendimento_hh = 0 na tarifa:"+RS)
        for x in u[:10]:
            print(Y+f"    - {x}"+RS)
        if len(u) > 10:
            print(DM+f"    ... +{len(u)-10}"+RS)
    if sem_preco:
        u = sorted(set(sem_preco))
        print(Y+f"\n  Preco zerado ({len(u)}) — preco_unit = 0 na tarifa:"+RS)
        for x in u[:10]:
            print(Y+f"    - {x}"+RS)
        if len(u) > 10:
            print(DM+f"    ... +{len(u)-10}"+RS)
    if not sem_depara and not sem_tarifa and not sem_hh and not sem_preco:
        ok("Cadeia de dados completa — nenhuma lacuna detectada.")
    else:
        total_lacunas = len(set(sem_depara)) + len(set(sem_tarifa)) + len(set(sem_hh)) + len(set(sem_preco))
        aviso(f"Total de lacunas: {total_lacunas}. Corrija via menu [4] de_para ou [2] importar tarifas.")
    sub()


def auto_mapear_de_para(cfg, atividades_reais):
    """
    Mapeia automaticamente atividades do micro para chaves de tarifa por similaridade textual.
    Usa normalizar_chave para comparacao robusta.
    """
    tarifas = cfg.get("tarifas", {})
    if not tarifas:
        return 0
    de_para = cfg.setdefault("de_para", {})
    tarif_norm = {k: normalizar_chave(k) for k in tarifas.keys()}
    novos = 0
    for atv in atividades_reais:
        if atv in de_para:
            continue
        an = normalizar_chave(atv)
        melhor = None
        melhor_score = 0
        for tk, tn in tarif_norm.items():
            score = 0
            if an == tn:
                score = 1000
            elif an in tn or tn in an:
                score = min(len(an), len(tn))
            else:
                toks_a = set(x for x in an.split() if len(x) > 2)
                toks_t = set(x for x in tn.split() if len(x) > 2)
                inter = len(toks_a & toks_t)
                if inter >= 3:
                    score = inter
            if score > melhor_score:
                melhor_score = score
                melhor = tk
        if melhor and melhor_score >= 3:
            de_para[atv] = melhor
            novos += 1
    if novos > 0:
        salvar_config(cfg)
    return novos

def aplicar_depara_padrao_exame(cfg, atividades_reais):
    """
    Aplica mapeamento fixo (hardcoded) do prototipo EXAME->CT_313.
    1) Dicionario exato por normalizar_chave; 2) heuristica por palavras-chave (APPN, parenteses, etc.).
    """
    tarifas = cfg.get("tarifas", {})
    if not tarifas:
        return 0
    de_para = cfg.setdefault("de_para", {})
    novo = 0
    for atv in atividades_reais:
        kn = normalizar_chave(atv)
        alvo = None
        if kn in DEFAULT_DEPARA_EXAME_CT313:
            alvo = DEFAULT_DEPARA_EXAME_CT313[kn]
        else:
            alvo = _depara_heuristico_exame_ct313(kn, tarifas)
        if alvo and alvo in tarifas and de_para.get(atv) != alvo:
            de_para[atv] = alvo
            novo += 1
    if novo:
        salvar_config(cfg)
    return novo

def _somente_bloqueado_restante(demanda_global, atividades_bloqueadas):
    """True se so resta demanda em atividades bloqueadas (ex. plantio/irrigacao)."""
    bloqueadas = set(atividades_bloqueadas or [])
    tem_nao_bloqueado = False
    tem_bloqueado = False
    for (_, atv), hh in demanda_global.items():
        if hh <= 0.01:
            continue
        if atv in bloqueadas:
            tem_bloqueado = True
        else:
            tem_nao_bloqueado = True
    return tem_bloqueado and not tem_nao_bloqueado

def menu_ajustes_hh_apenas_sessao(atividades_reais, cfg, session_hh):
    """Edita HH/ha por atividade apenas na memoria (nao salva config.json)."""
    if session_hh is None:
        return
    tarifas = cfg.get("tarifas", {})
    strict = cfg.get("orcamento_estrito", True)
    sub()
    print(G+BL+"  AJUSTE DE HH/ha — APENAS ESTA EXECUCAO"+RS)
    print(DM+"  Nao grava em config. ENTER = manter valor atual."+RS+"\n")
    n = 0
    for atv in sorted(set(atividades_reais), key=str):
        t_nome = resolver_chave_tarifa(cfg, tarifas, atv)
        cur = resolver_rendimento_hh(cfg, tarifas, t_nome, strict=strict, session_hh=session_hh, atv_micro=atv)
        if cur is None:
            cur = 0.0
        v = prompt(f"  [{str(atv)[:46]}]  CT:{str(t_nome)[:36]}  HH/ha [{cur}]", "")
        vs = str(v).strip()
        if not vs:
            continue
        try:
            nv = float(vs.replace(",", "."))
            session_hh[atv] = nv
            session_hh[t_nome] = nv
            n += 1
        except (TypeError, ValueError):
            aviso("Valor invalido, ignorado.")
    if n:
        ok(f"{n} override(s) HH/ha nesta sessao.")
    else:
        print(DM+"  Nenhum override informado."+RS)


def validar_e_completar_orcamento(cfg, atividades_reais, session_hh=None):
    """
    Modo orcamento_estrito: toda atividade do micro precisa de chave em tarifas com dados CT.
    Lacunas: escolher tarifa na lista ou informar HH/preco/custo manualmente.
    session_hh: dict opcional; HH informado para linha zerada pode ir apenas para a sessao (nao grava tarifas).
    Retorna False se usuario abortar.
    """
    if not cfg.get("orcamento_estrito", True):
        return True
    tarifas = cfg.setdefault("tarifas", {})
    de_para = cfg.setdefault("de_para", {})
    nomes_tarifa = sorted(tarifas.keys(), key=str)
    if not nomes_tarifa:
        erro("Nenhuma tarifa em config. Normalize a CT [3] ou importe [2].")
        return False

    for atv in sorted(set(atividades_reais), key=str):
        t_nome = resolver_chave_tarifa(cfg, tarifas, atv)
        if t_nome not in tarifas:
            sub()
            print(Y+f"  [ESTRITO] Sem tarifa CT para atividade do micro:"+RS)
            print(Y+f"    {str(atv)[:70]}"+RS)
            print(DM+f"    Chave atual: {t_nome}"+RS)
            if confirmar("  Escolher uma linha existente em tarifas (recomendado)?", default=True):
                idx = selecionar_paginado("TARIFA CT (orcamento)", nomes_tarifa, page_size=8)
                if idx < 0:
                    return False
                de_para[atv] = nomes_tarifa[idx]
                salvar_config(cfg)
            else:
                hh_m = pedir_float("  HH/ha (manual)", 8.0)
                pr_m = pedir_float("  Preco R$/ha (manual)", 0.0, allow_zero=True)
                ch_m = pedir_float("  Custo R$/h (manual)", float(cfg.get("custo_hora_tf") or 50), allow_zero=True)
                chave = prompt("  Nome da chave a gravar em tarifas (ex.: alias)", t_nome[:48])
                if not chave:
                    chave = t_nome
                tarifas[chave] = {
                    "rendimento_hh": hh_m, "preco_ha": pr_m, "preco_unit": pr_m,
                    "custo_hora": ch_m, "custo_ha": hh_m * ch_m if ch_m > 0 else 0,
                    "tipo": "Manual", "recurso": "homem", "eficiencia": 1.0,
                }
                de_para[atv] = chave
                salvar_config(cfg)
        t_nome = resolver_chave_tarifa(cfg, tarifas, atv)
        row = tarifas.get(t_nome, {})
        try:
            rh = float(row.get("rendimento_hh", 0))
        except (TypeError, ValueError):
            rh = 0.0
        tipo = str(row.get("tipo", "")).lower()
        hm = float(row.get("rendimento_hm", 0) or 0)
        is_mec = "mecaniz" in tipo or hm > 0
        if rh <= 0.01 and not is_mec:
            sub()
            print(Y+f"  [ESTRITO] rendimento_hh zero ou invalido na tarifa '{str(t_nome)[:50]}'"+RS)
            hh_m = pedir_float("  Informe HH/ha para esta linha (ou 0 se so maquina)", 0.0, allow_zero=True)
            if session_hh is not None and confirmar("  Aplicar SO nesta execucao (nao gravar em config.json)?", default=True):
                session_hh[atv] = float(hh_m)
                session_hh[t_nome] = float(hh_m)
            else:
                tarifas[t_nome]["rendimento_hh"] = float(hh_m)
                salvar_config(cfg)
    return True

# ──────────────────────────────────────────────
#  SMART SCHEDULER v2:
#  Filas por TURMA, sequenciais por talhao.
#  Atividades podem ser paralelas (varias turmas) ou exclusivas (uma turma).
#  Reatribuicao: qualquer atividade pode ser executada por qualquer turma.
# ──────────────────────────────────────────────
def dias_uteis_no_periodo(mes_ini, ano_ini, meses):
    dias = 0
    m, a = mes_ini, ano_ini
    for _ in range(int(math.ceil(meses))):
        for sem in calendar.monthcalendar(a, m):
            for i, d in enumerate(sem):
                if d != 0 and i < 5: dias += 1
        m += 1
        if m > 12: m = 1; a += 1
    return dias

_SEQUENCIAS_DISPONIVEIS = [
    ("implantacao",      "Rocada > Formiga > Coroamento > Coveamento > Adubacao > Plantio > Irrigacao (cascata)"),
    ("manutencao_swg",   "Rocada manual > Limpeza de area > Capina de coroa > Formigas > Coveamento > Adubacao > Plantio > Irrigacao (ordem SWG)"),
    ("manutencao_seco",  "[EM PROGRESSO] Manutencao periodo seco — regras ainda nao definidas"),
    ("manutencao_umido", "[EM PROGRESSO] Manutencao periodo umido — regras ainda nao definidas"),
    ("personalizado",    "Ordem livre (sem bloqueio global plantio/irrigacao)"),
]

def _selecionar_sequencia_padrao_sn(cfg, seq_cfg):
    sub()
    print(G+BL+"  SELECIONAR SEQUENCIA PADRAO:"+RS)
    print(DM+"  Responda S para a sequencia desejada (apenas UMA):"+RS+"\n")
    escolhido = None
    for modo_id, descr in _SEQUENCIAS_DISPONIVEIS:
        resp = confirmar(f"  {modo_id}: {descr}", default=(modo_id == seq_cfg.get("modo", "implantacao")))
        if resp:
            escolhido = modo_id
            break
    if not escolhido:
        aviso("Nenhuma sequencia selecionada. Repetindo...")
        return _selecionar_sequencia_padrao_sn(cfg, seq_cfg)
    seq_cfg["modo"] = escolhido
    ok(f"Sequencia: {escolhido}")
    if confirmar("  Salvar como padrao para proximas execucoes?", default=True):
        cfg["sequencia"] = seq_cfg
        salvar_config(cfg)
    return escolhido


def _norm_atv(x):
    """Normaliza nome de atividade para cruzamento template x micro (NA-safe, str strip)."""
    if x is None:
        return ""
    try:
        if pd.isna(x):
            return ""
    except (TypeError, ValueError):
        pass
    return str(x).strip()


def _slug_ficheiro_seguro(s, max_len=48):
    """Nome seguro para ficheiros Windows (sem acentos problematicos)."""
    t = remover_acentos(str(s).strip()) if s else ""
    t = re.sub(r'[<>:"/\\|?*]+', "_", t)
    t = re.sub(r"\s+", "_", t)
    t = re.sub(r"_+", "_", t).strip("_")
    return (t[:max_len] if t else "escopo")


def _distribuir_atividades_faltantes_turmas(turmas, atividades_reais, fazenda):
    """
    Atribui à turma mais numerosa as atividades da fazenda que não entraram no template
    (ex.: equipe só irrigação em micro sem irrigação — evita cronograma vazio).
    """
    if not turmas or not atividades_reais:
        return
    cobertura = set()
    for t in turmas:
        for a in t.get("atividades") or []:
            na = _norm_atv(a)
            if na:
                cobertura.add(na)
    farm_set = set(atividades_reais)
    orfas = [a for a in atividades_reais if a not in cobertura]
    if not orfas:
        return
    tpl_extra = sorted(cobertura - farm_set, key=str)
    if tpl_extra:
        print(
            DM
            + f"  Modelo de turmas: {len(tpl_extra)} atividade(s) sem demanda nesta fazenda "
            + f"('{str(fazenda)[:42]}') — ignoradas no micro atual." + RS
        )
        for x in tpl_extra[:5]:
            print(DM + f"    ~ {str(x)[:58]}" + RS)
        if len(tpl_extra) > 5:
            print(DM + f"    ... +{len(tpl_extra) - 5}" + RS)
    alvo = max(turmas, key=lambda t: int(t.get("operarios", 0) or 0))
    cur = {_norm_atv(x) for x in alvo.get("atividades", []) if _norm_atv(x)}
    cur |= set(orfas)
    alvo["atividades"] = sorted(cur, key=str)
    ok(
        f"{len(orfas)} atividade(s) desta fazenda sem turma no modelo foram atribuidas a '{alvo['nome']}' "
        f"(evita cronograma vazio quando o template nao cruza o micro)."
    )


def calcular_cronograma_inteligente(cfg, df_faz, fazenda, esperar_enter=True, ctx=None, escopo_meta=None):
    """
    ctx: optional dict with preconfigured session state for batch mode.
    When ctx is provided, interactive setup questions are skipped.
    """
    _batch = ctx is not None
    comparativo_cfg = None
    if not _batch:
        cabecalho(f"SMART SCHEDULER - {fazenda}")
        df_faz = avaliar_terreno(df_faz)
        aviso_politica_tarifas_planas()
    else:
        sub()
        print(G+BL+f"  SMART SCHEDULER - {fazenda}"+RS)
        df_faz["penalidade"] = float(ctx.get("penalidade", 1.0))

    # ── Extrair atividades REAIS da fazenda ──
    df_faz = df_faz.copy()
    df_faz["atividade"] = df_faz["atividade"].map(
        lambda x: _norm_atv(x) if pd.notna(x) else x
    )
    if not _batch and confirmar("Ajustar escopo de atividades (substituir/remover/adicionar) nesta execucao?", default=False):
        df_faz = _menu_ajustar_escopo_atividades(df_faz)
    atividades_reais = sorted(
        {a for a in df_faz["atividade"].dropna().unique().tolist() if _norm_atv(a)},
        key=str,
    )
    talhoes_ordenados = sorted(df_faz["chave"].dropna().unique().tolist())
    escopo_talhoes = []
    if isinstance(escopo_meta, dict):
        escopo_talhoes = list(escopo_meta.get("talhoes") or [])

    novos_fixos = aplicar_depara_padrao_exame(cfg, atividades_reais)
    if novos_fixos > 0:
        ok(f"de_para PADRAO aplicado: {novos_fixos} mapeamento(s) fixos EXAME->CT_313.")
    if not cfg.get("orcamento_estrito", True):
        novos_de_para = auto_mapear_de_para(cfg, atividades_reais)
        if novos_de_para > 0:
            ok(f"de_para complementar: {novos_de_para} mapeamento(s) adicionais.")

    if not _batch:
        sub()
        print(DM+"  Orcamento estrito (sem mediana silenciosa; lacunas pedem input): "+C+str(cfg.get("orcamento_estrito", True))+RS)
        if confirmar("  Alternar orcamento_estrito para esta execucao?", default=False):
            cfg["orcamento_estrito"] = not cfg.get("orcamento_estrito", True)
            salvar_config(cfg)
            ok(f"orcamento_estrito = {cfg['orcamento_estrito']}")

    sub()
    print(G+BL+"  ATIVIDADES ENCONTRADAS NESTA FAZENDA:"+RS)
    for i, a in enumerate(atividades_reais, 1):
        print(G+f"  {i:2}. "+C+a+RS)
    print(G+f"\n  Talhoes: "+C+f"{len(talhoes_ordenados)}"+RS)
    if escopo_talhoes:
        n_show = min(8, len(escopo_talhoes))
        base = ", ".join(str(x)[:24] for x in escopo_talhoes[:n_show])
        if len(escopo_talhoes) > n_show:
            base += f", ... (+{len(escopo_talhoes)-n_show})"
        print(DM + f"  Escopo talhoes selecionados: {base}" + RS)
    sub()

    seq_cfg = cfg.get("sequencia") or {}
    _merge_sequencia_defaults(seq_cfg)
    cfg["sequencia"] = seq_cfg

    if _batch:
        modo_seq = ctx["modo_seq"]
    else:
        modo_seq = _selecionar_sequencia_padrao_sn(cfg, seq_cfg)
    if modo_seq == "manutencao_seco":
        sequencia_manutencao_seco_placeholder(cfg)
    elif modo_seq == "manutencao_umido":
        sequencia_manutencao_umido_placeholder(cfg)
    usar_cascata = modo_seq in ("implantacao", "manutencao_swg", "personalizado")
    diagnosticar_sequencia_atividades(atividades_reais, seq_cfg, modo_seq)

    if _batch:
        usar_bloqueio_global = ctx.get("usar_bloqueio_global", False)
        atividades_bloqueadas = set()
        if usar_bloqueio_global:
            filtros_bloqueio = cfg.get("filtros_bloqueio_global", ["plantio", "irrig"])
            atividades_bloqueadas = set(
                atividades_por_filtro(atividades_reais, filtros_bloqueio)
            )
        usar_reforco_automatico = ctx.get("usar_reforco_automatico", True)
        usar_pool_pos_bloqueio = ctx.get("usar_pool_pos_bloqueio", False)
    else:
        filtros_bloqueio = cfg.get("filtros_bloqueio_global", ["plantio", "irrig"])
        candidatas_bloqueio = atividades_por_filtro(atividades_reais, filtros_bloqueio)
        usar_bloqueio_global = False
        atividades_bloqueadas = set()
        if modo_seq == "personalizado":
            print(DM+"  Modo PERSONALIZADO: bloqueio global plantio/irrigacao DESLIGADO."+RS)
        elif candidatas_bloqueio:
            usar_bloqueio_global = confirmar(
                "Aplicar BLOQUEIO GLOBAL (plantio/irrigacao so iniciam quando TODO o resto zerar na fazenda)?",
                default=True,
            )
            if usar_bloqueio_global:
                atividades_bloqueadas = set(candidatas_bloqueio)
                print(Y+f"\n  BLOQUEADAS ATE LIBERACAO GLOBAL ({len(atividades_bloqueadas)}):"+RS)
                for a in sorted(atividades_bloqueadas, key=lambda x: str(x))[:20]:
                    print(Y+f"    - {str(a)[:58]}"+RS)
                if len(atividades_bloqueadas) > 20:
                    print(DM+f"    ... +{len(atividades_bloqueadas)-20}"+RS)
                if confirmar("Salvar estes filtros de bloqueio no config para proximas execucoes?", default=True):
                    cfg["filtros_bloqueio_global"] = filtros_bloqueio
                    salvar_config(cfg)
        usar_reforco_automatico = confirmar(
            "Ativar REFORCO AUTOMATICO (turma ociosa ajuda outras atividades nao bloqueadas)?",
            default=True,
        )
        usar_pool_pos_bloqueio = False
        if usar_bloqueio_global:
            usar_pool_pos_bloqueio = confirmar(
                "Usar PELOTAO UNIFICADO (todos os executores) so em plantio/irrigacao apos liberacao global?",
                default=True,
            )

    if _batch:
        prazo_meses = ctx["prazo_meses"]
        mes_ref = ctx["mes_ref"]
        ano_ref = ctx["ano_ref"]
        jornada = ctx["jornada"]
        executores = ctx["executores"]
        comparativo_cfg = ctx.get("comparativo_cfg")
        turmas = []
        for t in ctx["turmas"]:
            turmas.append(
                {
                    "nome": t["nome"],
                    "operarios": t["operarios"],
                    "atividades": [
                        _norm_atv(a)
                        for a in (t.get("atividades") or [])
                        if _norm_atv(a)
                    ],
                }
            )
    else:
        # ── Config equipe ──
        print(G+BL+"\n  CONFIGURACAO DO PROJETO"+RS+"\n")

        prazo_meses = pedir_float("Prazo META para conclusao (meses)", 6.0)
        hoje = datetime.datetime.now()
        print(DM+"  Referencia do calendario para DIAS UTEIS da meta (meses corridos a partir de): "+RS)
        mes_ref = pedir_int("Mes inicial (1-12)", hoje.month)
        mes_ref = max(1, min(12, int(mes_ref)))
        ano_ref = pedir_int("Ano inicial", hoje.year)
        j_def = float(cfg.get("jornada_horas") or 4.6)
        if j_def <= 0:
            j_def = 4.6
        colab_total = pedir_int("Tamanho TOTAL da equipe HOJE", 10)
        supervisores = pedir_int("Quantos LIDERES (nao executam)", 1, allow_zero=True)
        jornada = pedir_float("Jornada efetiva diaria (horas no campo)", round(j_def, 2))

        executores = colab_total - supervisores
        if executores <= 0:
            erro("Precisa de pelo menos 1 executor."); return
        print(G+f"\n  Equipe Operacional: {executores} operarios @ {jornada}h/dia"+RS)
        if confirmar("Configurar COMPARATIVO MULTI-FATOR agora (para exportar no Excel)?", default=True):
            comparativo_cfg = coletar_config_comparativo_multifator(executores, jornada)

        # ──────────────────────────────────────────
        #  ETAPA 1: CRIAR TURMAS
        # ──────────────────────────────────────────
        sub()
        print(G+BL+"  ETAPA 1: CRIAR TURMAS / FUNCOES"+RS)
        print(DM+"  Defina grupos de trabalho (ex: Rocadores, Adubadores, Geral)."+RS)
        print(DM+"  Depois voce vinculara quais atividades cada turma executa.\n"+RS)

        turmas = []
        restantes = executores

        while restantes > 0:
            print(G+f"  Operarios disponiveis: {restantes}"+RS)
            nome_turma = prompt("Nome da turma (ex: Rocadores)", f"Turma {len(turmas)+1}")
            def_pad = min(restantes, max(1, restantes // 2 or restantes))
            qtd = pedir_int(f"  Quantos operarios na turma '{nome_turma}'", def_pad)
            if qtd > restantes:
                aviso(f"Maximo disponivel: {restantes}. Ajustando.")
                qtd = restantes
            turmas.append({"nome": nome_turma, "operarios": qtd, "atividades": []})
            restantes -= qtd
            if restantes > 0:
                if not confirmar(f"Criar outra turma? ({restantes} restantes)", default=True):
                    turmas.append({"nome": "Geral", "operarios": restantes, "atividades": []})
                    restantes = 0

        sub()
        print(G+BL+"  TURMAS CRIADAS:"+RS)
        for t in turmas:
            print(G+f"  - {t['nome']}: "+C+f"{t['operarios']} operarios"+RS)
        sub()

        # ──────────────────────────────────────────
        #  ETAPA 2: VINCULAR ATIVIDADES AS TURMAS
        # ──────────────────────────────────────────
        print(G+BL+"\n  ETAPA 2: VINCULAR ATIVIDADES AS TURMAS"+RS)
        print(DM+"  Use FILTRO por texto para ligar varias de uma vez (ex: todas com 'roçada')."+RS)
        print(DM+"  Depois: conflitos (paralelo vs uma turma) e opcao de REATRIBUIR a outra turma.\n"+RS)

    for turma in turmas:
        if not _batch:
            menu_vincular_atividades_turma(turma, atividades_reais)
        else:
            existing = {_norm_atv(a) for a in (turma.get("atividades") or []) if _norm_atv(a)}
            matched = existing & set(atividades_reais)
            turma["atividades"] = sorted(matched, key=str)

    def _cobertura_atual_turmas():
        s = set()
        for t in turmas:
            for a in t.get("atividades") or []:
                na = _norm_atv(a)
                if na:
                    s.add(na)
        return s

    cob_pre = _cobertura_atual_turmas()
    orfas_pre = [a for a in atividades_reais if a not in cob_pre]
    preencher_orfas = False
    if orfas_pre:
        if _batch:
            preencher_orfas = bool(ctx.get("preencher_orfas_template", False))
        elif modo_seq == "personalizado":
            preencher_orfas = confirmar(
                "Esta fazenda tem demandas sem turma no modelo. Preencher na turma com mais operarios? "
                "(N = equipe especializada; HH dessas atividades nao entram no cronograma)",
                default=False,
            )
    if preencher_orfas:
        _distribuir_atividades_faltantes_turmas(turmas, atividades_reais, fazenda)

    for turma in turmas:
        if not turma["atividades"]:
            aviso(f"Turma '{turma['nome']}' ficou sem atividades!")

    def coletar_vinculadas():
        s = set()
        for t in turmas:
            s.update(t["atividades"])
        return s

    atividades_vinculadas = coletar_vinculadas()
    orfas = [a for a in atividades_reais if a not in atividades_vinculadas]
    if orfas:
        print(Y+f"\n  ATENCAO: {len(orfas)} atividades sem turma vinculada:"+RS)
        for o in orfas:
            print(Y+f"    - {str(o)[:55]}"+RS)
        if confirmar("Vincular todas as orfas a uma turma existente?", default=True):
            nomes = [t["nome"] for t in turmas]
            turma_alvo = selecionar("TURMA PARA ORFAS", nomes)
            if turma_alvo:
                for t in turmas:
                    if t["nome"] == turma_alvo:
                        t["atividades"] = sorted(set(t["atividades"]) | set(orfas), key=lambda x: str(x))
                        ok(f"{len(orfas)} atividades vinculadas a '{turma_alvo}'.")

    # ──────────────────────────────────────────
    #  ETAPA 3: Conflitos (paralelo / exclusivo) + reatribuicao opcional
    # ──────────────────────────────────────────
    print(G+BL+"\n  ETAPA 3: CONFLITOS E REATRIBUICAO"+RS)
    reatribuicao, paralelo, primaria = resolver_conflitos_e_reatribuir(turmas, atividades_reais)

    session_hh = {}
    if ctx and isinstance(ctx.get("session_hh"), dict):
        session_hh.update(ctx["session_hh"])

    # ── Validacao orcamento estrito (antes das demandas) ──
    if not validar_e_completar_orcamento(cfg, atividades_reais, session_hh=session_hh):
        input(DM+"\n  [ENTER para voltar] "+RS)
        return
    if confirmar("Ajustar HH/ha por atividade APENAS nesta execucao (nao grava config)?", default=False):
        menu_ajustes_hh_apenas_sessao(atividades_reais, cfg, session_hh)

    tarifas = cfg.get("tarifas", {})
    de_para = cfg.get("de_para", {})
    strict = cfg.get("orcamento_estrito", True)

    # ── Construir demandas por talhao ──
    demandas = OrderedDict()  # {talhao: [{atividade, area, hh_total}, ...]}
    total_hh = 0.0

    for talhao in talhoes_ordenados:
        df_t = df_faz[df_faz["chave"] == talhao]
        tarefas = []
        for _, row in df_t.iterrows():
            atv = row["atividade"]
            area = float(row["area_ha"])
            pen = float(row["penalidade"])

            t_nome = resolver_chave_tarifa(cfg, tarifas, atv)
            rend_base = resolver_rendimento_hh(cfg, tarifas, t_nome, strict=strict, session_hh=session_hh, atv_micro=atv)
            if rend_base is None:
                erro(f"Rendimento invalido para '{t_nome}'.")
                return
            rend_hh_ha = float(rend_base) * pen
            preco = resolver_preco_ha(cfg, tarifas, t_nome, strict=strict)
            custo_h = resolver_custo_hora(cfg, tarifas, t_nome, strict=strict)
            if preco is None:
                preco = 0.0
            if custo_h is None:
                custo_h = 0.0
            in_tarifa = t_nome in tarifas

            horas = area * rend_hh_ha
            total_hh += horas
            receita_item = area * float(preco or 0)
            custo_item = horas * float(custo_h or 0)

            tarifa_row = tarifas.get(t_nome, {})
            tipo_tarifa = str(tarifa_row.get("tipo", "")).lower()
            hm_tarifa = float(tarifa_row.get("rendimento_hm", 0) or 0)
            is_mec = "mecaniz" in tipo_tarifa or (hm_tarifa > 0 and rend_hh_ha <= 0)

            if strict:
                origem_linha = "CT"
                rfonte = "CT"
            else:
                origem_linha = "tarifa" if in_tarifa else "fallback"
                rfonte = "CT" if in_tarifa else "estimado"
            tarefas.append({
                "atividade": atv, "area": area, "hh_total": horas,
                "preco_ha": preco, "custo_hora": custo_h,
                "receita": receita_item, "custo_mo": custo_item,
                "chave_tarifa": t_nome, "origem": origem_linha,
                "rendimento_fonte": rfonte,
                "tipo": "Mecanizada" if is_mec else "Manual",
            })
        demandas[talhao] = tarefas

    print(DM+f"\n  Total HH da fazenda (bruto): {total_hh:.1f} horas-homem"+RS)

    auditar_cadeia_dados(cfg, demandas, atividades_reais, session_hh=session_hh)

    sem_tarifa = []
    for talhao, tarefas in demandas.items():
        for t in tarefas:
            atv = t["atividade"]
            t_nome = resolver_chave_tarifa(cfg, tarifas, atv)
            if t_nome not in tarifas:
                sem_tarifa.append((str(atv)[:50], str(t_nome)[:50]))
    if not strict and sem_tarifa:
        est_fb = resolver_rendimento_hh(cfg, tarifas, "!__chave_inexistente__!", strict=False)
        print(Y+"\n  !  Chave de tarifa NAO encontrada no orcamento importado (desencontro de nome)."+RS)
        print(Y+f"     Rendimento estimado aplicado: ~{est_fb:.2f} h/ha (mediana/config; ver doc)."+RS)
        visto = set()
        for a, tn in sem_tarifa:
            key = (a, tn)
            if key in visto:
                continue
            visto.add(key)
            print(Y+f"    micro: {a}  ->  chave buscada: {tn}"+RS)
        print(DM+"    Correcao: menu [4] de_para ou importe tarifas [2] — no orcamento o homem/ha existe."+RS)

    sem_executor = []
    for talhao, tarefas in demandas.items():
        for t in tarefas:
            if t["hh_total"] < 0.01:
                continue
            atv = t["atividade"]
            if not turmas_que_executam(atv, turmas, reatribuicao, paralelo, primaria):
                sem_executor.append(atv)
    if sem_executor:
        print(R+"\n  X  Atividades com demanda mas SEM turma executora:"+RS)
        for a in sorted(set(str(x) for x in sem_executor))[:15]:
            print(R+f"    - {a[:58]}"+RS)
        if len(set(sem_executor)) > 15:
            print(DM+f"    ... +{len(set(sem_executor))-15}"+RS)
        if not confirmar("  Continuar mesmo assim (essas HH nao serao agendadas)?", default=False):
            input(DM+"\n  [ENTER para voltar] "+RS)
            return
        for talhao, tarefas in demandas.items():
            for t in tarefas:
                atv = t["atividade"]
                if t["hh_total"] > 0.01 and not turmas_que_executam(atv, turmas, reatribuicao, paralelo, primaria):
                    t["hh_total"] = 0.0
        total_hh = sum(t["hh_total"] for tarefas in demandas.values() for t in tarefas)
        aviso("HH sem executora foram zeradas no cronograma.")
        print(DM+f"  Total HH agendavel: {total_hh:.1f} horas-homem"+RS)

    sub()
    print(G+BL+"  GERANDO CRONOGRAMA (talhao a talhao)..."+RS+"\n")

    # ──────────────────────────────────────────
    #  SCHEDULER: Filas por TURMA, sequenciais por talhao.
    #  Cada turma tem uma fila de trabalho construida a partir
    #  das atividades vinculadas, na ordem dos talhoes.
    #  Quando a turma termina no talhao 1, avanca pro 2.
    # ──────────────────────────────────────────

    # Build per-turma work queue: list of {talhao, atividade, hh_rest}
    turma_filas = {}
    for turma in turmas:
        fila = []
        for talhao in talhoes_ordenados:
            for tarefa in demandas.get(talhao, []):
                atv = tarefa["atividade"]
                if tarefa["hh_total"] > 0.01 and turma["nome"] in turmas_que_executam(
                        atv, turmas, reatribuicao, paralelo, primaria):
                    fila.append({
                        "talhao": talhao,
                        "atividade": atv,
                        "hh_rest": tarefa["hh_total"]
                    })
        turma_filas[turma["nome"]] = fila

    # Uma entrada por (talhao, atividade): todas as turmas autorizadas
    # consomem o mesmo saldo (paralelo) ou so uma turma (exclusivo/reatribuido).
    demanda_global = {}  # key=(talhao,atividade) -> remaining hh
    for talhao, tarefas in demandas.items():
        for t in tarefas:
            demanda_global[(talhao, t["atividade"])] = t["hh_total"]

    atividades_plantio = set(atividades_por_filtro(atividades_reais, seq_cfg.get("filtros_plantio") or ["plantio"]))
    atividades_irrig = set(atividades_por_filtro(atividades_reais, seq_cfg.get("filtros_irrigacao") or ["irrig"]))
    tem_plantio_por_talhao = {}
    for th in talhoes_ordenados:
        tem_plantio_por_talhao[th] = any(
            t["atividade"] in atividades_plantio and t["hh_total"] > 0.01
            for t in demandas.get(th, [])
        )
    dia_termino_plantio = {}

    if usar_cascata:
        for _tn, fila in turma_filas.items():
            fila.sort(key=lambda x: (
                classificar_fase_cascata_valor(x["atividade"], seq_cfg, modo_seq, atividades_plantio, atividades_irrig),
                str(x["talhao"]),
                str(x["atividade"]),
            ))

    cronograma = []
    dia = 0
    MAX_DIAS = 10000

    def _registrar_fim_plantio_talhao(th, dia_atual):
        if dia_termino_plantio.get(th) is not None:
            return
        if not _demanda_plantio_talhao(th, demanda_global, atividades_plantio):
            dia_termino_plantio[th] = dia_atual

    while dia < MAX_DIAS:
        tem_trabalho = any(v > 0.01 for v in demanda_global.values())
        if not tem_trabalho:
            break

        dia += 1
        pool_only = (
            usar_bloqueio_global and usar_pool_pos_bloqueio
            and _somente_bloqueado_restante(demanda_global, atividades_bloqueadas)
        )
        if pool_only:
            cap_pool = float(executores) * float(jornada)
            while cap_pool > 0.01:
                fez = False
                min_fase_dia = _min_fase_cascata(
                    demanda_global, seq_cfg, modo_seq, usar_cascata,
                    usar_bloqueio_global, atividades_bloqueadas,
                    atividades_plantio, atividades_irrig,
                    dia, dia_termino_plantio, tem_plantio_por_talhao,
                )
                for talhao in talhoes_ordenados:
                    tlist = list(demandas.get(talhao, []))
                    tlist.sort(key=lambda t: (
                        0 if t["atividade"] in atividades_plantio else (
                            1 if t["atividade"] in atividades_irrig else 2),
                        str(t["atividade"]),
                    ))
                    for t in tlist:
                        atv = t["atividade"]
                        if atv not in atividades_bloqueadas:
                            continue
                        key = (talhao, atv)
                        rest = demanda_global.get(key, 0.0)
                        if rest <= 0.01:
                            continue
                        if not pode_agendar_atividade_cascata(
                            talhao, atv, demanda_global, seq_cfg, modo_seq, usar_cascata,
                            usar_bloqueio_global, atividades_bloqueadas,
                            atividades_plantio, atividades_irrig,
                            dia, dia_termino_plantio, tem_plantio_por_talhao, min_fase_dia,
                        ):
                            continue
                        consumo = min(rest, cap_pool)
                        demanda_global[key] -= consumo
                        cap_pool -= consumo
                        fez = True
                        _registrar_fim_plantio_talhao(talhao, dia)
                        t_nome_cr = resolver_chave_tarifa(cfg, tarifas, atv)
                        ch = resolver_custo_hora(cfg, tarifas, t_nome_cr, strict=False) or 0.0
                        cronograma.append({
                            "Dia": dia,
                            "Fazenda": fazenda,
                            "Talhao": talhao,
                            "Atividade": atv,
                            "Turma": "Pelotao_Unificado",
                            "Operarios": executores,
                            "HH": round(consumo, 2),
                            "Custo_MO": round(consumo * ch, 2),
                            "Modo": "PoolPosBloqueio",
                        })
                        if cap_pool <= 0.01:
                            break
                    if cap_pool <= 0.01:
                        break
                if not fez:
                    break
            for turma in turmas:
                fila = turma_filas[turma["nome"]]
                while fila and demanda_global.get((fila[0]["talhao"], fila[0]["atividade"]), 0) < 0.01:
                    fila.pop(0)
            continue

        for turma in turmas:
            fila = turma_filas[turma["nome"]]
            n_ops = turma["operarios"]
            cap_dia = n_ops * jornada

            # Process items in queue order
            idx = 0
            while cap_dia > 0.01 and idx < len(fila):
                min_fase_dia = _min_fase_cascata(
                    demanda_global, seq_cfg, modo_seq, usar_cascata,
                    usar_bloqueio_global, atividades_bloqueadas,
                    atividades_plantio, atividades_irrig,
                    dia, dia_termino_plantio, tem_plantio_por_talhao,
                )
                item = fila[idx]
                key = (item["talhao"], item["atividade"])
                rest = demanda_global.get(key, 0)

                if rest < 0.01:
                    idx += 1  # Already done (by another turma perhaps)
                    continue

                if not pode_agendar_atividade_cascata(
                    item["talhao"], item["atividade"], demanda_global, seq_cfg, modo_seq, usar_cascata,
                    usar_bloqueio_global, atividades_bloqueadas,
                    atividades_plantio, atividades_irrig,
                    dia, dia_termino_plantio, tem_plantio_por_talhao, min_fase_dia,
                ):
                    idx += 1
                    continue

                consumo = min(rest, cap_dia)
                demanda_global[key] -= consumo
                cap_dia -= consumo
                _registrar_fim_plantio_talhao(item["talhao"], dia)

                atv_key = item["atividade"]
                t_nome_cr = de_para.get(atv_key, atv_key)
                ch = resolver_custo_hora(cfg, tarifas, t_nome_cr, strict=False) or 0.0

                cronograma.append({
                    "Dia": dia,
                    "Fazenda": fazenda,
                    "Talhao": item["talhao"],
                    "Atividade": item["atividade"],
                    "Turma": turma["nome"],
                    "Operarios": n_ops,
                    "HH": round(consumo, 2),
                    "Custo_MO": round(consumo * ch, 2),
                })

                if demanda_global[key] < 0.01:
                    idx += 1  # Move to next item in queue
                # else stay on same item (partially done today)

            # Clean up completed items from front of queue
            while fila and demanda_global.get((fila[0]["talhao"], fila[0]["atividade"]), 0) < 0.01:
                fila.pop(0)

            # Mutirao/realloc automatico:
            # se ainda sobrou capacidade no dia, ajuda demanda de outras atividades nao bloqueadas.
            if usar_reforco_automatico and cap_dia > 0.01:
                for talhao in talhoes_ordenados:
                    if cap_dia <= 0.01:
                        break
                    tarefas_t = list(demandas.get(talhao, []))
                    if usar_cascata:
                        tarefas_t.sort(key=lambda t: (
                            classificar_fase_cascata_valor(t["atividade"], seq_cfg, modo_seq, atividades_plantio, atividades_irrig),
                            str(t["atividade"]),
                        ))
                    for t in tarefas_t:
                        min_fase_dia = _min_fase_cascata(
                            demanda_global, seq_cfg, modo_seq, usar_cascata,
                            usar_bloqueio_global, atividades_bloqueadas,
                            atividades_plantio, atividades_irrig,
                            dia, dia_termino_plantio, tem_plantio_por_talhao,
                        )
                        atv = t["atividade"]
                        key_ref = (talhao, atv)
                        rest_ref = demanda_global.get(key_ref, 0.0)
                        if rest_ref <= 0.01:
                            continue
                        if not pode_agendar_atividade_cascata(
                            talhao, atv, demanda_global, seq_cfg, modo_seq, usar_cascata,
                            usar_bloqueio_global, atividades_bloqueadas,
                            atividades_plantio, atividades_irrig,
                            dia, dia_termino_plantio, tem_plantio_por_talhao, min_fase_dia,
                        ):
                            continue
                        consumo_ref = min(rest_ref, cap_dia)
                        if consumo_ref <= 0.01:
                            continue
                        demanda_global[key_ref] -= consumo_ref
                        cap_dia -= consumo_ref
                        _registrar_fim_plantio_talhao(talhao, dia)
                        t_nome_ref = resolver_chave_tarifa(cfg, tarifas, atv)
                        ch_ref = resolver_custo_hora(cfg, tarifas, t_nome_ref, strict=False) or 0.0
                        cronograma.append({
                            "Dia": dia,
                            "Fazenda": fazenda,
                            "Talhao": talhao,
                            "Atividade": atv,
                            "Turma": turma["nome"],
                            "Operarios": n_ops,
                            "HH": round(consumo_ref, 2),
                            "Custo_MO": round(consumo_ref * ch_ref, 2),
                            "Modo": "Reforco",
                        })

    dias_simulado = dia

    # ── Diagnostico ──
    dias_meta = dias_uteis_no_periodo(mes_ref, ano_ref, prazo_meses)
    exec_teoricos = math.ceil(total_hh / (dias_meta * jornada)) if (dias_meta * jornada) > 0 else 1
    meses_simulado = dias_simulado / 22.0 if dias_simulado > 0 else 0

    # ── Tabela semanal ──
    table = Table(title=f"Cronograma - {fazenda} ({executores} Exec.)")
    table.add_column("Semana", justify="center", style="cyan")
    table.add_column("Dias", justify="center")
    table.add_column("Talhoes / Atividades", style="green")

    semanas = defaultdict(lambda: {"dias": set(), "acoes": set()})
    for c in cronograma:
        sem = math.ceil(c["Dia"] / 5)
        semanas[sem]["dias"].add(c["Dia"])
        semanas[sem]["acoes"].add(f"[{c['Talhao']}] {c['Atividade'][:18]}")

    for sem in sorted(semanas.keys())[:8]:
        d = semanas[sem]
        dias_str = f"Dia {min(d['dias'])} a {max(d['dias'])}"
        acoes = ", ".join(list(d['acoes'])[:3])
        if len(d['acoes']) > 3: acoes += " (+)"
        table.add_row(f"Sem {sem}", dias_str, acoes)

    console.print(table)
    if len(semanas) > 8:
        print(DM+f"  ... e mais {len(semanas)-8} semanas no Excel."+RS)

    # ── Metricas financeiras ──
    hh_por_turma = defaultdict(float)
    custo_por_turma = defaultdict(float)
    for c in cronograma:
        hh_por_turma[c["Turma"]] += float(c["HH"])
        custo_por_turma[c["Turma"]] += float(c.get("Custo_MO", 0))

    hh_agendada_total = sum(hh_por_turma.values())
    custo_mo_efetivo = sum(custo_por_turma.values())
    cap_total = sum(float(dias_simulado) * float(t["operarios"]) * float(jornada) for t in turmas)
    hh_ociosa = max(0.0, cap_total - hh_agendada_total)
    custo_medio_hora = (custo_mo_efetivo / hh_agendada_total) if hh_agendada_total > 0.01 else 0.0
    custo_mo_ociosidade = hh_ociosa * custo_medio_hora
    custo_mo_total = custo_mo_efetivo + custo_mo_ociosidade

    receita_total = sum(
        t.get("receita", 0) for tarefas in demandas.values() for t in tarefas
    )
    receita_mo = sum(
        t.get("receita", 0) for tarefas in demandas.values() for t in tarefas
        if t.get("tipo") != "Mecanizada"
    )
    receita_mec = sum(
        t.get("receita", 0) for tarefas in demandas.values() for t in tarefas
        if t.get("tipo") == "Mecanizada"
    )
    lucro_direto = receita_total - custo_mo_total
    custos_globais = cfg.get("custos_globais", {}) or {}
    custo_direto_global = float(custos_globais.get("valor_direto_total", 0) or 0.0)
    custo_indireto_global = float(custos_globais.get("valor_indireto_total", 0) or 0.0)
    custo_global_total = custo_direto_global + custo_indireto_global
    lucro_operacional = lucro_direto - custo_global_total
    margem = (lucro_operacional / receita_total * 100) if receita_total > 0.01 else 0.0
    n_demandas = sum(1 for tarefas in demandas.values() for t in tarefas)
    n_fb = sum(1 for tarefas in demandas.values() for t in tarefas if t.get("origem") == "fallback")
    pct_fallback = (100.0 * n_fb / n_demandas) if n_demandas > 0 else 0.0

    recursos_mec = []
    cronograma_mec = []
    cronograma_com_mec = []
    atividades_mec_set = set()
    sub()
    print(C+BL+"  ATIVAR MODO MECANIZADO"+RS)
    print(DM+"  Cadastre recursos mecanizados (robo, trator, etc.) que assumem atividades do humano."+RS)
    if confirmar("  Ativar modo mecanizado?", default=False):
        recursos_mec = _cadastrar_recursos_mecanizados_sn(atividades_reais, cfg)
        for rec in recursos_mec:
            atividades_mec_set.update(rec.get("atividades", set()))
        if recursos_mec and atividades_mec_set:
            cronograma_mec = construir_cronograma_mecanizado(demandas, fazenda, jornada, recursos_mec)
            crono_hum_sem_mec = construir_cronograma_humano_sem_mecanizadas(
                cronograma, turmas, jornada, executores, atividades_mec_set
            )
            cronograma_com_mec = sorted(
                crono_hum_sem_mec + cronograma_mec,
                key=lambda r: (int(r.get("Dia", 0)), str(r.get("Turma", ""))),
            )
            d_hum = max([int(x.get("Dia", 0)) for x in crono_hum_sem_mec], default=0)
            d_mec = max([int(x.get("Dia", 0)) for x in cronograma_mec], default=0)
            d_comb = max(d_hum, d_mec)
            t_mec = Table(title="Comparativo Operacional - Modo Mecanizado")
            t_mec.add_column("Metrica", style="cyan")
            t_mec.add_column("Valor", justify="right")
            t_mec.add_row("Dias baseline (humano)", str(dias_simulado))
            t_mec.add_row("Dias humano sem atividades mecanizadas", str(d_hum))
            t_mec.add_row("Dias recursos mecanizados (filas dedicadas)", str(d_mec))
            t_mec.add_row("Dias cenario combinado (humano || mecanizado)", str(d_comb))
            t_mec.add_row("Ganho de prazo (dias)", f"{int(dias_simulado) - int(d_comb):+d}")
            for rec in recursos_mec:
                t_mec.add_row(f"  Recurso: {rec['nome']}", f"{rec['prod_ha_h']} ha/h | R$ {rec['custo_h']}/h")
                t_mec.add_row(f"  Atividades ({rec['nome']})", str(len(rec.get('atividades', set()))))
            hh_mec_total = sum(float(x.get("HH", 0)) for x in cronograma_mec)
            custo_mec_total = sum(float(x.get("Custo_MO", 0)) for x in cronograma_mec)
            t_mec.add_row("HH total mecanizada", f"{hh_mec_total:.1f}")
            t_mec.add_row("Custo total mecanizado", f"R$ {custo_mec_total:,.2f}")
            console.print(t_mec)

            t_alt = Table(title="Cronograma Alternativo (Humano + Mecanizado)")
            t_alt.add_column("Semana", justify="center", style="cyan")
            t_alt.add_column("Dias", justify="center")
            t_alt.add_column("Acoes", style="green")
            sem_alt = defaultdict(lambda: {"dias": set(), "acoes": set()})
            for c in cronograma_com_mec:
                s = int(math.ceil(float(c.get("Dia", 0)) / 5.0)) if c.get("Dia") else 0
                if s <= 0:
                    continue
                sem_alt[s]["dias"].add(int(c["Dia"]))
                txt = f"[{str(c.get('Talhao',''))[:18]}] {str(c.get('Atividade',''))[:18]} ({c.get('Turma','')})"
                sem_alt[s]["acoes"].add(txt)
            for s in sorted(sem_alt.keys())[:8]:
                d = sem_alt[s]
                dias_str = f"Dia {min(d['dias'])} a {max(d['dias'])}"
                acoes = ", ".join(list(d["acoes"])[:3])
                if len(d["acoes"]) > 3:
                    acoes += " (+)"
                t_alt.add_row(f"Sem {s}", dias_str, acoes)
            console.print(t_alt)

    # ── Tabela ocupacao ──
    sub()
    print(G+BL+"  OCUPACAO POR TURMA"+RS)
    t_occ = Table()
    t_occ.add_column("Turma", style="cyan")
    t_occ.add_column("HH", justify="right")
    t_occ.add_column("Cap. max", justify="right")
    t_occ.add_column("Uso %", justify="right")
    t_occ.add_column("Custo MO", justify="right")
    crit_nm, crit_pct = "", 0.0
    for turma in turmas:
        nm = turma["nome"]
        cap = float(dias_simulado) * float(turma["operarios"]) * float(jornada)
        us = hh_por_turma.get(nm, 0.0)
        pct = (100.0 * us / cap) if cap > 0.01 else 0.0
        if pct > crit_pct:
            crit_pct, crit_nm = pct, nm
        t_occ.add_row(nm, f"{us:.1f}", f"{cap:.1f}", f"{pct:.0f}%", f"R$ {custo_por_turma.get(nm, 0):,.2f}")
    if hh_por_turma.get("Pelotao_Unificado", 0) > 0.01:
        d_pool = len(set(c["Dia"] for c in cronograma if c.get("Turma") == "Pelotao_Unificado"))
        pu = hh_por_turma["Pelotao_Unificado"]
        cu = float(custo_por_turma.get("Pelotao_Unificado", 0))
        cap_p = float(d_pool) * float(executores) * float(jornada)
        pct_p = (100.0 * pu / cap_p) if cap_p > 0.01 else 0.0
        t_occ.add_row(
            "Pelotao_Unificado", f"{pu:.1f}", f"{cap_p:.1f}", f"{pct_p:.0f}%", f"R$ {cu:,.2f}"
        )
    console.print(t_occ)
    print(DM+"  Uso % = HH no cronograma com o nome da turma / (dias simulados x operarios x jornada)."+RS)
    print(DM+"  Reforco nao aumenta n_ops; bloqueio global impede reforco em plantio/irrigacao ate liberar tudo."+RS)
    if usar_pool_pos_bloqueio and usar_bloqueio_global:
        print(DM+"  Pelotao_Unificado: plantio/irrigacao apos liberacao usam todos os executores num so pelotao."+RS)
    if crit_nm:
        print(DM+f"  Heuristica caminho critico (maior Uso %): turma '{crit_nm}' (~{crit_pct:.0f}%)."+RS)

    # ── Painel financeiro ──
    sub()
    print(G+BL+"  DOSSIER FINANCEIRO"+RS)
    t_fin = Table()
    t_fin.add_column("Metrica", style="cyan")
    t_fin.add_column("Valor", justify="right")
    t_fin.add_row("Receita Bruta Total", f"R$ {receita_total:,.2f}")
    if receita_mec > 0.01:
        t_fin.add_row("  Receita Mao de Obra", f"R$ {receita_mo:,.2f}")
        t_fin.add_row("  Receita Mecanizada", f"R$ {receita_mec:,.2f}")
    t_fin.add_row("Custo MO (Trabalho Ativo)", f"R$ {custo_mo_efetivo:,.2f}")
    t_fin.add_row("Custo MO (Ociosidade)", f"R$ {custo_mo_ociosidade:,.2f}")
    t_fin.add_row("Custo MO Total", f"R$ {custo_mo_total:,.2f}")
    t_fin.add_row("Lucro Direto (Receita - Custo MO)", f"R$ {lucro_direto:,.2f}")
    if custo_global_total > 0.01:
        t_fin.add_row("Custo Direto Global (bruto)", f"R$ {custo_direto_global:,.2f}")
        t_fin.add_row("Custo Indireto Global (bruto)", f"R$ {custo_indireto_global:,.2f}")
        t_fin.add_row("Custo Global Total", f"R$ {custo_global_total:,.2f}")
        t_fin.add_row("Lucro Operacional (apos custos globais)", f"R$ {lucro_operacional:,.2f}")
        t_fin.add_row("Margem Operacional", f"{margem:.1f} %")
    else:
        t_fin.add_row("Margem Bruta", f"{margem:.1f} %")
    if n_fb > 0:
        t_fin.add_row("Itens com fallback (sem tarifa CT)", f"{pct_fallback:.0f}% ({n_fb}/{n_demandas})")
    t_fin.add_row("Fonte dos dados", "100% CT" if pct_fallback < 0.01 else f"{100-pct_fallback:.0f}% CT")
    metas = cfg.get("metas", {})
    if metas:
        if metas.get("lucro_alvo") is not None:
            t_fin.add_row("Meta de Lucro (R$)", f"R$ {float(metas.get('lucro_alvo') or 0):,.2f}")
        if metas.get("margem_alvo_pct") is not None:
            t_fin.add_row("Meta de Margem (%)", f"{float(metas.get('margem_alvo_pct') or 0):.1f}%")
        if metas.get("bonus_aa_formula"):
            t_fin.add_row("Bonus AA (rota)", str(metas.get("bonus_aa_formula"))[:38])
        if metas.get("bonus_bb_formula"):
            t_fin.add_row("Bonus BB (rota)", str(metas.get("bonus_bb_formula"))[:38])
        if metas.get("equacao_quadratica"):
            t_fin.add_row("Equacao quadratica (rota)", str(metas.get("equacao_quadratica"))[:38])
    console.print(t_fin)

    def _render_tabela_cenarios(rows, label):
        if not rows:
            return
        t_sc = Table(title=f"Comparativo de Cenários (Equipe x Jornada) - {label}")
        t_sc.add_column("Equipe", justify="right")
        t_sc.add_column("Jornada", justify="right")
        t_sc.add_column("Dias", justify="right")
        t_sc.add_column("Meses", justify="right")
        t_sc.add_column("Ganho vs Meta", justify="right")
        for r in rows[:40]:
            t_sc.add_row(
                str(r["Equipe"]),
                f"{r['Jornada_h_dia']:.2f}",
                str(r["Dias_Simulados"]),
                f"{r['Meses_Simulados']:.2f}",
                f"{r['Ganho_vs_Meta_dias']:+d}",
            )
        console.print(t_sc)

    cenarios_rows = []
    if comparativo_cfg is not None and isinstance(comparativo_cfg, dict):
        hh_base_multi = float(total_hh)
        lbl_base_multi = "Sem mecanizado"
        if recursos_mec and cronograma_com_mec:
            hh_hum_pos_mec = sum(
                float(x.get("HH", 0) or 0)
                for x in cronograma_com_mec
                if not str(x.get("Turma", "")).startswith("MEC_")
            )
            base_opt = selecionar(
                "BASE DO COMPARATIVO MULTI-FATOR",
                ["Sem mecanizado (HH total atual)", "Com mecanizado (HH humano remanescente)"],
            )
            if base_opt and base_opt.startswith("Com mecanizado"):
                hh_base_multi = float(hh_hum_pos_mec)
                lbl_base_multi = "Com mecanizado"
        print(DM+f"  Base selecionada: {lbl_base_multi} | HH={hh_base_multi:.1f}"+RS)
        cenarios_rows = simular_cenarios_multifator(
            total_hh=hh_base_multi,
            receita_total=receita_total,
            custo_hora_tf=float(cfg.get("custo_hora_tf") or 0.0),
            dias_meta=dias_meta,
            executores_base=executores,
            jornada_base=jornada,
            jornadas_in=comparativo_cfg.get("jornadas"),
            equipes_in=comparativo_cfg.get("equipes"),
            interativo=False,
        )
        _render_tabela_cenarios(cenarios_rows, lbl_base_multi)

    while confirmar("Recalcular comparativo multi-fator com novos valores agora?", default=False):
        hh_base_multi = float(total_hh)
        lbl_base_multi = "Sem mecanizado"
        if recursos_mec and cronograma_com_mec:
            hh_hum_pos_mec = sum(
                float(x.get("HH", 0) or 0)
                for x in cronograma_com_mec
                if not str(x.get("Turma", "")).startswith("MEC_")
            )
            base_opt = selecionar(
                "BASE DO COMPARATIVO MULTI-FATOR",
                ["Sem mecanizado (HH total atual)", "Com mecanizado (HH humano remanescente)"],
            )
            if base_opt and base_opt.startswith("Com mecanizado"):
                hh_base_multi = float(hh_hum_pos_mec)
                lbl_base_multi = "Com mecanizado"
        print(DM+f"  Base selecionada: {lbl_base_multi} | HH={hh_base_multi:.1f}"+RS)
        cenarios_rows = simular_cenarios_multifator(
            total_hh=hh_base_multi,
            receita_total=receita_total,
            custo_hora_tf=float(cfg.get("custo_hora_tf") or 0.0),
            dias_meta=dias_meta,
            executores_base=executores,
            jornada_base=jornada,
            interativo=True,
        )
        _render_tabela_cenarios(cenarios_rows, lbl_base_multi)

    atividades_escopo = sorted({str(a).strip() for a in df_faz["atividade"].dropna().tolist() if str(a).strip()}, key=str)
    escopo_set = set(atividades_escopo)
    ag_hum_set = {str(x.get("Atividade", "")).strip() for x in cronograma if str(x.get("Atividade", "")).strip()}
    ag_mec_set = {str(x.get("Atividade", "")).strip() for x in cronograma_mec if str(x.get("Atividade", "")).strip()}
    ag_any_set = ag_hum_set | ag_mec_set
    faltantes_set = sorted(list(escopo_set - ag_any_set), key=str)
    sem_executor_set = {str(x).strip() for x in sem_executor if str(x).strip()}
    hh_por_atividade = defaultdict(float)
    for tarefas in demandas.values():
        for t in tarefas:
            hh_por_atividade[str(t.get("atividade", "")).strip()] += float(t.get("hh_total", 0) or 0)

    def _motivo_faltante(atv_nome):
        hhv = float(hh_por_atividade.get(atv_nome, 0) or 0)
        if atv_nome in sem_executor_set:
            return "sem turma executora"
        if hhv <= 0.01:
            return "hh zerado/no escopo sem demanda util"
        return "nao agendada no cronograma"

    rows_audit = []
    for a in atividades_escopo:
        if a in ag_hum_set:
            status = "agendada_humana"
        elif a in ag_mec_set:
            status = "agendada_mecanizada"
        else:
            status = "nao_agendada"
        rows_audit.append(
            {
                "Atividade": a,
                "HH_Escopo": round(float(hh_por_atividade.get(a, 0) or 0), 2),
                "Status": status,
                "Motivo": "" if status != "nao_agendada" else _motivo_faltante(a),
            }
        )
    df_audit = pd.DataFrame(rows_audit)
    sub()
    print(G + BL + "  AUDITORIA DO ESCOPO (ANTES DA EXPORTACAO)" + RS)
    print(DM + f"  Atividades no escopo: {len(atividades_escopo)}" + RS)
    print(DM + f"  Agendadas no humano: {len(ag_hum_set & escopo_set)}" + RS)
    print(DM + f"  Agendadas no mecanizado: {len(ag_mec_set & escopo_set)}" + RS)
    print(DM + f"  Nao agendadas: {len(faltantes_set)}" + RS)
    rocadas_escopo = [a for a in atividades_escopo if "rocada" in normalizar_chave(a)]
    if rocadas_escopo:
        for rcv in rocadas_escopo:
            if rcv in ag_hum_set:
                st = "agendada_humana"
            elif rcv in ag_mec_set:
                st = "agendada_mecanizada"
            else:
                st = "nao_agendada"
            print(DM + f"    rocada: {rcv[:56]} -> {st}" + RS)

    # ── Export Dossier Excel (financeiro + operacional) ──
    if cronograma:
        try:
            def _slug_nome(v):
                return str(v).replace("/", "_").replace(" ", "_")

            scope_tag = "__FAZENDA_TODOS"
            if isinstance(escopo_meta, dict):
                modo_th = str(escopo_meta.get("modo_talhao") or "")
                ths = [str(x) for x in (escopo_meta.get("talhoes") or []) if str(x).strip()]
                if modo_th in ("unico", "parcial") and len(ths) == 1:
                    scope_tag = f"__TH_{_slug_nome(ths[0])}"
                elif modo_th == "parcial" and len(ths) > 1:
                    scope_tag = f"__TH_MULTI_{len(ths)}"
                elif modo_th in ("todos", "fallback_todos"):
                    scope_tag = "__FAZENDA_TODOS"
            nome_base = f"Dossier_{_slug_nome(fazenda)}{scope_tag}"
            nome_fin = f"{nome_base}_FINANCEIRO.xlsx"
            nome_op = f"{nome_base}_OPERACIONAL.xlsx"
            pasta_dossier = os.path.join(DIR, DOSSIER_DIRNAME)
            os.makedirs(pasta_dossier, exist_ok=True)
            nome_fin, caminho_fin = _proximo_caminho_livre(pasta_dossier, nome_fin)
            nome_op, caminho_op = _proximo_caminho_livre(pasta_dossier, nome_op)

            df_crono = pd.DataFrame(cronograma)
            if "Dia" in df_crono.columns:
                df_crono["Semana"] = df_crono["Dia"].apply(
                    lambda d: int(math.ceil(float(d) / 5.0)) if pd.notna(d) else ""
                )

            # Aba 2: custo por atividade (pivot)
            atv_fin = []
            for tarefas in demandas.values():
                for t in tarefas:
                    atv_fin.append({
                        "Atividade": t["atividade"],
                        "Tipo": t.get("tipo", ""),
                        "Area_ha": t["area"],
                        "HH": t["hh_total"],
                        "Receita_Orcada": t.get("receita", 0),
                        "Custo_MO": t.get("custo_mo", 0),
                        "Lucro": t.get("receita", 0) - t.get("custo_mo", 0),
                        "Chave_Tarifa": t.get("chave_tarifa", ""),
                        "Origem": t.get("origem", ""),
                    })
            df_atv = pd.DataFrame(atv_fin)
            if not df_atv.empty:
                df_pivot = df_atv.groupby(["Atividade", "Tipo"]).agg({
                    "Area_ha": "sum", "HH": "sum", "Receita_Orcada": "sum",
                    "Custo_MO": "sum", "Lucro": "sum",
                }).reset_index()
                df_pivot["Margem_%"] = (df_pivot["Lucro"] / df_pivot["Receita_Orcada"] * 100).fillna(0).round(1)
                df_pivot = df_pivot.sort_values("Lucro", ascending=False)
            else:
                df_pivot = pd.DataFrame()

            # Aba 1: resumo financeiro
            rows_resumo = [
                {"Metrica": "Fazenda", "Valor": fazenda},
                {"Metrica": "Dossier operacional (cronograma+cascata, Gantt — sem R$)", "Valor": nome_op},
                {"Metrica": "Executores", "Valor": executores},
                {"Metrica": "Jornada (h/dia)", "Valor": jornada},
                {"Metrica": "Prazo Meta (meses)", "Valor": prazo_meses},
                {"Metrica": "Dias Uteis Meta", "Valor": dias_meta},
                {"Metrica": "Duracao Simulada (dias uteis)", "Valor": dias_simulado},
                {"Metrica": "Duracao Simulada (meses)", "Valor": f"{meses_simulado:.1f}"},
                {"Metrica": "", "Valor": ""},
                {"Metrica": "Receita Bruta Total", "Valor": f"R$ {receita_total:,.2f}"},
            ]
            if receita_mec > 0.01:
                rows_resumo.append({"Metrica": "  Receita Mao de Obra", "Valor": f"R$ {receita_mo:,.2f}"})
                rows_resumo.append({"Metrica": "  Receita Mecanizada", "Valor": f"R$ {receita_mec:,.2f}"})
            rows_resumo += [
                {"Metrica": "Custo MO (Trabalho Ativo)", "Valor": f"R$ {custo_mo_efetivo:,.2f}"},
                {"Metrica": "Custo MO (Ociosidade Logistica)", "Valor": f"R$ {custo_mo_ociosidade:,.2f}"},
                {"Metrica": "Custo MO Total Equipe", "Valor": f"R$ {custo_mo_total:,.2f}"},
                {"Metrica": "Lucro Direto (Receita - Custo MO)", "Valor": f"R$ {lucro_direto:,.2f}"},
                {"Metrica": "", "Valor": ""},
                {"Metrica": "HH Total Simulado", "Valor": f"{total_hh:,.1f}"},
                {"Metrica": "Custo/hora TF (CT_313)", "Valor": f"R$ {float(cfg.get('custo_hora_tf', 0)):,.2f}"},
                {"Metrica": "Fonte dos dados", "Valor": "100% CT" if pct_fallback < 0.01 else f"{100-pct_fallback:.0f}% CT ({n_fb} fallbacks)"},
            ]
            if custo_global_total > 0.01:
                rows_resumo += [
                    {"Metrica": "Custo Direto Global (bruto)", "Valor": f"R$ {custo_direto_global:,.2f}"},
                    {"Metrica": "Custo Indireto Global (bruto)", "Valor": f"R$ {custo_indireto_global:,.2f}"},
                    {"Metrica": "Custo Global Total", "Valor": f"R$ {custo_global_total:,.2f}"},
                    {"Metrica": "Lucro Operacional (apos custos globais)", "Valor": f"R$ {lucro_operacional:,.2f}"},
                    {"Metrica": "Margem Operacional", "Valor": f"{margem:.1f}%"},
                ]
            else:
                rows_resumo += [
                    {"Metrica": "Margem Bruta", "Valor": f"{margem:.1f}%"},
                ]
            metas = cfg.get("metas", {})
            if metas:
                rows_resumo.append({"Metrica": "", "Valor": ""})
                rows_resumo.append({"Metrica": "Rotas de Metas/Bonificacao", "Valor": "Preparatorio"})
                if metas.get("lucro_alvo") is not None:
                    rows_resumo.append({"Metrica": "Meta Lucro (R$)", "Valor": f"R$ {float(metas.get('lucro_alvo') or 0):,.2f}"})
                if metas.get("margem_alvo_pct") is not None:
                    rows_resumo.append({"Metrica": "Meta Margem (%)", "Valor": f"{float(metas.get('margem_alvo_pct') or 0):.1f}%"})
                if metas.get("bonus_aa_formula"):
                    rows_resumo.append({"Metrica": "Bonus AA (rota)", "Valor": str(metas.get("bonus_aa_formula"))})
                if metas.get("bonus_bb_formula"):
                    rows_resumo.append({"Metrica": "Bonus BB (rota)", "Valor": str(metas.get("bonus_bb_formula"))})
                if metas.get("equacao_quadratica"):
                    rows_resumo.append({"Metrica": "Equacao quadratica (rota)", "Valor": str(metas.get("equacao_quadratica"))})
            if recursos_mec and cronograma_mec:
                rows_resumo.append({"Metrica": "", "Valor": ""})
                for rec in recursos_mec:
                    rows_resumo.append({"Metrica": f"Mecanizado: {rec['nome']}", "Valor": f"{rec['prod_ha_h']} ha/h | R$ {rec['custo_h']}/h"})
                    rows_resumo.append({"Metrica": f"  Atividades ({rec['nome']})", "Valor": str(len(rec.get('atividades', set())))})
            resumo = pd.DataFrame(rows_resumo)

            rows_op = [
                {"Metrica": "Fazenda", "Valor": fazenda},
                {"Metrica": "Dossier financeiro (valores R$)", "Valor": nome_fin},
                {"Metrica": "Executores", "Valor": executores},
                {"Metrica": "Jornada (h/dia)", "Valor": jornada},
                {"Metrica": "Prazo Meta (meses)", "Valor": prazo_meses},
                {"Metrica": "Dias Uteis Meta", "Valor": dias_meta},
                {"Metrica": "Duracao Simulada (dias uteis)", "Valor": dias_simulado},
                {"Metrica": "Duracao Simulada (meses)", "Valor": f"{meses_simulado:.1f}"},
                {"Metrica": "HH Total Simulado", "Valor": f"{total_hh:,.1f}"},
                {"Metrica": "Fonte dos dados", "Valor": "100% CT" if pct_fallback < 0.01 else f"{100-pct_fallback:.0f}% CT ({n_fb} fallbacks)"},
                {"Metrica": "", "Valor": ""},
                {"Metrica": "Atividades no escopo", "Valor": len(atividades_escopo)},
                {"Metrica": "Agendadas (humano)", "Valor": len(ag_hum_set & escopo_set)},
                {"Metrica": "Agendadas (mecanizado)", "Valor": len(ag_mec_set & escopo_set)},
                {"Metrica": "Nao agendadas", "Valor": len(faltantes_set)},
            ]
            if isinstance(escopo_meta, dict):
                rows_op.append(
                    {
                        "Metrica": "Escopo talhoes",
                        "Valor": ", ".join(str(x) for x in (escopo_meta.get("talhoes") or [])) or "todos",
                    }
                )
            metas_op = cfg.get("metas", {})
            if metas_op:
                rows_op += [
                    {"Metrica": "", "Valor": ""},
                    {"Metrica": "Metas / rotas (sem valores monetarios)", "Valor": ""},
                ]
                if metas_op.get("margem_alvo_pct") is not None:
                    rows_op.append({"Metrica": "Meta Margem (%)", "Valor": f"{float(metas_op.get('margem_alvo_pct') or 0):.1f}%"})
                if metas_op.get("bonus_aa_formula"):
                    rows_op.append({"Metrica": "Bonus AA (rota)", "Valor": str(metas_op.get("bonus_aa_formula"))})
                if metas_op.get("bonus_bb_formula"):
                    rows_op.append({"Metrica": "Bonus BB (rota)", "Valor": str(metas_op.get("bonus_bb_formula"))})
                if metas_op.get("equacao_quadratica"):
                    rows_op.append({"Metrica": "Equacao quadratica (rota)", "Valor": str(metas_op.get("equacao_quadratica"))})
            if recursos_mec:
                rows_op += [{"Metrica": "", "Valor": ""}]
                for rec in recursos_mec:
                    rows_op.append({"Metrica": f"Mecanizado: {rec['nome']}", "Valor": f"{rec['prod_ha_h']} ha/h"})
                    rows_op.append({"Metrica": f"  Atividades ({rec['nome']})", "Valor": str(len(rec.get("atividades", set())))})
            resumo_op = pd.DataFrame(rows_op)

            df_cascata = _gerar_aba_cascata_explicada(cronograma, jornada)
            df_ocupacao = _gerar_aba_ocupacao_turmas(cronograma, turmas, jornada, dias_simulado)
            df_crono_op = _df_crono_operacional(df_crono)

            with pd.ExcelWriter(caminho_fin, engine="openpyxl") as writer_fin:
                resumo.to_excel(writer_fin, sheet_name="RESUMO_FINANCEIRO", index=False)
                if not df_pivot.empty:
                    df_pivot.to_excel(writer_fin, sheet_name="CUSTO_POR_ATIVIDADE", index=False)
                if cenarios_rows:
                    pd.DataFrame(cenarios_rows).to_excel(writer_fin, sheet_name="COMPARATIVO_CENARIOS", index=False)
                wb_fin = writer_fin.book
                try:
                    from srf_excel_format import aplicar_formatacao_financeiro
                    aplicar_formatacao_financeiro(
                        wb_fin, lucro_operacional, margem, dias_simulado, fazenda, receita_total, custo_mo_total
                    )
                except Exception:
                    pass

            with pd.ExcelWriter(caminho_op, engine="openpyxl") as writer_op:
                resumo_op.to_excel(writer_op, sheet_name="RESUMO_OPERACIONAL", index=False)
                df_crono_op.to_excel(writer_op, sheet_name="CRONOGRAMA_DETALHADO", index=False)
                if not df_cascata.empty:
                    df_cascata.to_excel(writer_op, sheet_name="CASCATA_EXPLICADA", index=False)
                if not df_ocupacao.empty:
                    df_ocupacao.to_excel(writer_op, sheet_name="OCUPACAO_TURMAS_DIA", index=False)
                if recursos_mec and cronograma_mec:
                    df_mec_crono = _df_crono_operacional(pd.DataFrame(cronograma_mec))
                    df_mec_crono.to_excel(writer_op, sheet_name="CRONOGRAMA_MECANIZADO", index=False)
                    df_combinado = _df_crono_operacional(pd.DataFrame(cronograma_com_mec))
                    df_combinado.to_excel(writer_op, sheet_name="CRONOGRAMA_COMBINADO", index=False)
                if not df_audit.empty:
                    df_audit.to_excel(writer_op, sheet_name="AUDITORIA_ESCOPO", index=False)
                wb_op = writer_op.book
                _aplicar_cores_ocupacao_excel(wb_op, "OCUPACAO_TURMAS_DIA")
                try:
                    from srf_excel_format import aplicar_formatacao_operacional
                    aplicar_formatacao_operacional(wb_op, dias_simulado, cronograma)
                except Exception:
                    pass

            ok(f"Dossier financeiro exportado: {nome_fin}")
            ok(f"Dossier operacional exportado: {nome_op}")

            if cenarios_rows:
                nome_xlsx_cmp = f"Dossier_{fazenda.replace('/','_').replace(' ','_')}_COMPARATIVO_CENARIOS.xlsx"
                nome_xlsx_cmp, caminho_xlsx_cmp = _proximo_caminho_livre(pasta_dossier, nome_xlsx_cmp)
                with pd.ExcelWriter(caminho_xlsx_cmp, engine="openpyxl") as writer3:
                    pd.DataFrame(cenarios_rows).to_excel(writer3, sheet_name="COMPARATIVO_CENARIOS", index=False)
                ok(f"Dossier comparativo de cenarios exportado: {nome_xlsx_cmp}")

            if recursos_mec and cronograma_com_mec:
                nome_mec_fin = f"{nome_base}_COM_MECANIZADO_FINANCEIRO.xlsx"
                nome_mec_op = f"{nome_base}_COM_MECANIZADO_OPERACIONAL.xlsx"
                nome_mec_fin, caminho_mec_fin = _proximo_caminho_livre(pasta_dossier, nome_mec_fin)
                nome_mec_op, caminho_mec_op = _proximo_caminho_livre(pasta_dossier, nome_mec_op)
                df_mec_full = pd.DataFrame(cronograma_com_mec)
                if "Dia" in df_mec_full.columns:
                    df_mec_full["Semana"] = df_mec_full["Dia"].apply(
                        lambda d: int(math.ceil(float(d) / 5.0)) if pd.notna(d) else ""
                    )
                d_comb = max([int(x.get("Dia", 0)) for x in cronograma_com_mec], default=0)
                resumo_mec_rows = [
                    {"Metrica": "Fazenda", "Valor": fazenda},
                    {"Metrica": "Dossier operacional (cenario mecanizado)", "Valor": nome_mec_op},
                    {"Metrica": "Cenario", "Valor": "Humano + Mecanizado"},
                    {"Metrica": "Dias baseline (humano)", "Valor": dias_simulado},
                    {"Metrica": "Dias cenario combinado", "Valor": d_comb},
                    {"Metrica": "Ganho de prazo (dias)", "Valor": int(dias_simulado) - int(d_comb)},
                ]
                for rec in recursos_mec:
                    resumo_mec_rows.append({"Metrica": f"Recurso: {rec['nome']}", "Valor": f"{rec['prod_ha_h']} ha/h | R$ {rec['custo_h']}/h"})
                rows_mec_op = [
                    {"Metrica": "Fazenda", "Valor": fazenda},
                    {"Metrica": "Dossier financeiro (cenario mecanizado, R$)", "Valor": nome_mec_fin},
                    {"Metrica": "Cenario", "Valor": "Humano + Mecanizado"},
                    {"Metrica": "Dias baseline (humano)", "Valor": dias_simulado},
                    {"Metrica": "Dias cenario combinado", "Valor": d_comb},
                    {"Metrica": "Ganho de prazo (dias)", "Valor": int(dias_simulado) - int(d_comb)},
                ]
                for rec in recursos_mec:
                    rows_mec_op.append({"Metrica": f"Recurso: {rec['nome']}", "Valor": f"{rec['prod_ha_h']} ha/h"})
                df_cascata_mec = _gerar_aba_cascata_explicada(cronograma_com_mec, jornada)
                df_mec_op = _df_crono_operacional(df_mec_full)
                with pd.ExcelWriter(caminho_mec_fin, engine="openpyxl") as writer_mf:
                    pd.DataFrame(resumo_mec_rows).to_excel(writer_mf, sheet_name="RESUMO_FINANCEIRO", index=False)
                    wb_mf = writer_mf.book
                    try:
                        from srf_excel_format import aplicar_formatacao_financeiro
                        aplicar_formatacao_financeiro(
                            wb_mf, lucro_operacional, margem, d_comb, fazenda, receita_total, custo_mo_total
                        )
                    except Exception:
                        pass
                with pd.ExcelWriter(caminho_mec_op, engine="openpyxl") as writer_mo:
                    pd.DataFrame(rows_mec_op).to_excel(writer_mo, sheet_name="RESUMO_OPERACIONAL", index=False)
                    df_mec_op.to_excel(writer_mo, sheet_name="CRONOGRAMA_DETALHADO", index=False)
                    if not df_cascata_mec.empty:
                        df_cascata_mec.to_excel(writer_mo, sheet_name="CASCATA_EXPLICADA", index=False)
                    wb_mo = writer_mo.book
                    try:
                        from srf_excel_format import aplicar_formatacao_operacional
                        aplicar_formatacao_operacional(wb_mo, d_comb, cronograma_com_mec)
                    except Exception:
                        pass
                ok(f"Dossier cenario mecanizado (financeiro): {nome_mec_fin}")
                ok(f"Dossier cenario mecanizado (operacional): {nome_mec_op}")
        except Exception as ex:
            aviso(f"Nao foi possivel salvar Dossier: {ex}")

    # ── Diagnostico final ──
    linha()
    print(G+BL+"  DIAGNOSTICO DE PRAZO"+RS)
    sub()
    print(G+f"  Meta informada             : {prazo_meses} meses ({dias_meta} dias uteis a partir de {mes_ref:02d}/{ano_ref})"+RS)
    print(G+f"  Duracao simulada           : {dias_simulado} dias ({meses_simulado:.1f} meses)"+RS)
    if recursos_mec and cronograma_com_mec:
        d_mc = max([int(x.get("Dia", 0)) for x in cronograma_com_mec], default=0)
        m_mc = d_mc / 22.0 if d_mc > 0 else 0.0
        print(C+f"  Duracao cenario mecanizado : {d_mc} dias ({m_mc:.1f} meses)"+RS)
        print(C+f"  Ganho operacional estimado : {int(dias_simulado) - int(d_mc):+d} dias"+RS)
    sub()

    if meses_simulado <= prazo_meses:
        print(G+BL+"  STATUS: DENTRO DO PRAZO"+RS)
        print(G+f"  Equipe de {executores} executores conclui antes da meta."+RS)
    else:
        print(Y+BL+"  STATUS: PRAZO EXCEDIDO"+RS)
        print(Y+f"  Equipe atual levara {meses_simulado:.1f} meses (meta: {prazo_meses})."+RS)
        print(C+f"  [SUGESTAO] ~{exec_teoricos} executores @ {jornada}h/dia cumpririam a meta."+RS)
        if dias_meta > 0 and total_hh > 0.01:
            ex5 = math.ceil(total_hh / (dias_meta * 5.0))
            ex6 = math.ceil(total_hh / (dias_meta * 6.0))
            print(
                DM+f"  [DICA] Com a mesma jornada na meta, ~{ex5} executores @ 5h/dia ou ~{ex6} @ 6h/dia "
                f"(aprox.: HH total / {dias_meta} dias uteis / jornada)."+RS
            )

    linha()
    if esperar_enter:
        input(DM+"\n  [ENTER para voltar ao menu] "+RS)
    d_mc = max([int(x.get("Dia", 0)) for x in cronograma_com_mec], default=0) if (recursos_mec and cronograma_com_mec) else None
    ganho_mc = (int(dias_simulado) - int(d_mc)) if d_mc is not None else 0
    return {
        "fazenda": fazenda,
        "dias_simulado": int(dias_simulado),
        "meses_simulado": float(meses_simulado),
        "dias_mecanizado": d_mc,
        "ganho_mecanizado_dias": int(ganho_mc),
        "receita_total": float(receita_total),
        "custo_mo_total": float(custo_mo_total),
        "lucro_operacional": float(lucro_operacional),
        "total_hh": float(total_hh),
        "cronograma": cronograma,
        "turmas_snapshot": [{"nome": t["nome"], "operarios": t["operarios"]} for t in turmas],
    }

# ──────────────────────────────────────────────
#  V6: ABAS EXCEL TIMELINE + OCUPACAO + PERFIS
# ──────────────────────────────────────────────

_FASE_CORES_HEX = {
    "rocada":   "4472C4",
    "formiga":  "ED7D31",
    "coroamento": "70AD47",
    "coveamento": "FFC000",
    "adubacao_quimica": "9B59B6",
    "plantio":  "2ECC71",
    "irrigacao": "3498DB",
    "limpeza_quimica": "95A5A6",
    "demais":   "BDC3C7",
    "reforco":  "D5DBDB",
    "pool":     "1ABC9C",
}


def _fase_nome_pt(fase_id):
    m = {
        "rocada": "Rocada",
        "formiga": "Formiga",
        "coroamento": "Coroamento",
        "coveamento": "Coveamento",
        "adubacao_quimica": "Adubacao quim.",
        "plantio": "Plantio",
        "irrigacao": "Irrigacao",
        "limpeza_quimica": "Limpeza quim.",
        "demais": "Demais",
        "reforco": "Reforco",
        "pool": "Pelotao unif.",
    }
    return m.get(fase_id, str(fase_id).capitalize())


def _classificar_fase_nome(atv, seq_cfg, modo, atvs_plantio, atvs_irrig):
    """Retorna (fase_id, fase_valor) para rotulagem no Excel."""
    from collections import OrderedDict as _OD
    if eh_limpeza_quimica_pos_plantio(atv, seq_cfg):
        return "limpeza_quimica", 8.0
    if atv in atvs_plantio or _match_filtros_fase(atv, seq_cfg.get("filtros_plantio") or ["plantio"], None):
        return "plantio", 6.0
    if atv in atvs_irrig or _match_filtros_fase(atv, seq_cfg.get("filtros_irrigacao") or ["irrig"], None):
        return "irrigacao", 7.0
    fases = _fases_ordem_config(seq_cfg, modo)
    for i, fase in enumerate(fases):
        if _match_filtros_fase(atv, fase.get("filtros") or [], fase.get("exclusoes")):
            return fase.get("id", f"fase_{i}"), float(i)
    return "demais", 5.5


def _gerar_aba_timeline(cronograma, seq_cfg, modo_seq, atividades_reais, fazenda):
    """Retorna DataFrame para aba TIMELINE_CASCATA com colunas de visualização."""
    atvs_plantio = set(atividades_por_filtro(atividades_reais, seq_cfg.get("filtros_plantio") or ["plantio"]))
    atvs_irrig = set(atividades_por_filtro(atividades_reais, seq_cfg.get("filtros_irrigacao") or ["irrig"]))
    rows = []
    for c in cronograma:
        atv = c.get("Atividade", "")
        fase_id, fase_val = _classificar_fase_nome(atv, seq_cfg, modo_seq, atvs_plantio, atvs_irrig)
        modo_exec = c.get("Modo", "Normal")
        if modo_exec == "Reforco":
            fase_id_display = "reforco"
        elif modo_exec == "PoolPosBloqueio":
            fase_id_display = "pool"
        else:
            fase_id_display = fase_id
        rows.append({
            "Dia": c.get("Dia"),
            "Semana": int(math.ceil(float(c.get("Dia", 1)) / 5.0)),
            "Fazenda": fazenda,
            "Talhao": c.get("Talhao", ""),
            "Atividade": atv,
            "Fase": _fase_nome_pt(fase_id),
            "Fase_ID": fase_id,
            "Fase_Ordem": fase_val,
            "Turma": c.get("Turma", ""),
            "Operarios": c.get("Operarios", 0),
            "HH": c.get("HH", 0),
            "Custo_MO": c.get("Custo_MO", 0),
            "Modo": modo_exec,
            "Cor_Hex": _FASE_CORES_HEX.get(fase_id_display, "BDC3C7"),
        })
    return pd.DataFrame(rows) if rows else pd.DataFrame()


def _gerar_aba_cascata_explicada(cronograma, jornada):
    """
    Trilha explicativa da cascata por dia/turma/atividade.
    Mostra capacidade, consumo, saldo e pendencia (carry-over) de forma didatica.
    """
    if not cronograma:
        return pd.DataFrame()

    rows_src = []
    for i, c in enumerate(cronograma):
        try:
            dia = int(c.get("Dia", 0) or 0)
            hh = float(c.get("HH", 0) or 0.0)
            ops = float(c.get("Operarios", 0) or 0.0)
        except (TypeError, ValueError):
            continue
        if dia <= 0 or hh <= 0.0:
            continue
        rows_src.append(
            {
                "_ord": i,
                "Dia": dia,
                "Semana": int(math.ceil(float(dia) / 5.0)),
                "Fazenda": c.get("Fazenda", ""),
                "Talhao": c.get("Talhao", ""),
                "Atividade": c.get("Atividade", ""),
                "Turma": c.get("Turma", ""),
                "Operarios": ops,
                "HH": hh,
            }
        )
    if not rows_src:
        return pd.DataFrame()

    # demanda total por atividade/talhao/turma para calcular pendencia durante o consumo
    demanda_total = defaultdict(float)
    for r in rows_src:
        k = (str(r["Talhao"]), str(r["Atividade"]), str(r["Turma"]))
        demanda_total[k] += float(r["HH"])

    df_rows = pd.DataFrame(rows_src).sort_values(["Dia", "Turma", "_ord"]).reset_index(drop=True)
    out = []
    consumido_atividade = defaultdict(float)

    for (dia, turma), grp in df_rows.groupby(["Dia", "Turma"], sort=True):
        ops_dia = max(float(x) for x in grp["Operarios"].tolist()) if len(grp) else 0.0
        cap_dia = max(0.0, float(ops_dia) * float(jornada))
        usado_dia = 0.0
        for _, r in grp.iterrows():
            hh_inicio = max(0.0, cap_dia - usado_dia)
            hh_cons = float(r["HH"])
            usado_dia += hh_cons
            hh_saldo = max(0.0, cap_dia - usado_dia)
            k = (str(r["Talhao"]), str(r["Atividade"]), str(turma))
            consumido_atividade[k] += hh_cons
            pend = max(0.0, float(demanda_total[k]) - float(consumido_atividade[k]))
            op = float(r["Operarios"] or 0.0)
            hh_equiv_op = (hh_cons / op) if op > 0.01 else 0.0
            out.append(
                {
                    "Tipo_Linha": "ATIVIDADE",
                    "Dia": int(dia),
                    "Semana": int(r["Semana"]),
                    "Fazenda": r["Fazenda"],
                    "Talhao": r["Talhao"],
                    "Atividade": r["Atividade"],
                    "Turma": turma,
                    "Operarios": int(round(op)) if op > 0 else 0,
                    "Jornada_h_dia": round(float(jornada), 2),
                    "Capacidade_Dia_HH": round(cap_dia, 2),
                    "HH_Disponivel_Inicio_Dia": round(hh_inicio, 2),
                    "HH_Atividade_Demandado": round(float(demanda_total[k]), 2),
                    "HH_Atividade_Consumido": round(hh_cons, 2),
                    "HH_Consumido_Por_Operador_Equiv": round(hh_equiv_op, 3),
                    "HH_Saldo_Apos_Atividade": round(hh_saldo, 2),
                    "HH_Pendente_Atividade": round(pend, 2),
                    "Fechou_Dia": "S" if hh_saldo <= 0.01 else "N",
                    "Calculo_Dia": f"{cap_dia:.2f} - {usado_dia - hh_cons:.2f} - {hh_cons:.2f} = {hh_saldo:.2f}",
                }
            )
        out.append(
            {
                "Tipo_Linha": "RESUMO_DIA",
                "Dia": int(dia),
                "Semana": int(math.ceil(float(dia) / 5.0)),
                "Fazenda": "",
                "Talhao": "",
                "Atividade": "__RESUMO_DIA__",
                "Turma": turma,
                "Operarios": int(round(ops_dia)) if ops_dia > 0 else 0,
                "Jornada_h_dia": round(float(jornada), 2),
                "Capacidade_Dia_HH": round(cap_dia, 2),
                "HH_Disponivel_Inicio_Dia": round(cap_dia, 2),
                "HH_Atividade_Demandado": "",
                "HH_Atividade_Consumido": round(usado_dia, 2),
                "HH_Consumido_Por_Operador_Equiv": round((usado_dia / ops_dia), 3) if ops_dia > 0.01 else 0.0,
                "HH_Saldo_Apos_Atividade": round(max(0.0, cap_dia - usado_dia), 2),
                "HH_Pendente_Atividade": "",
                "Fechou_Dia": "S" if max(0.0, cap_dia - usado_dia) <= 0.01 else "N",
                "Calculo_Dia": f"{cap_dia:.2f} - {usado_dia:.2f} = {max(0.0, cap_dia - usado_dia):.2f}",
            }
        )
    return pd.DataFrame(out)


def _gerar_aba_ocupacao_turmas(cronograma, turmas, jornada, dias_simulado):
    """Retorna DataFrame pivot: dia x turma com HH, Cap, Uso%, Status."""
    if not cronograma or dias_simulado < 1:
        return pd.DataFrame()
    turma_nomes = sorted(set(t["nome"] for t in turmas))
    turma_ops = {t["nome"]: t["operarios"] for t in turmas}
    hh_dia_turma = defaultdict(lambda: defaultdict(float))
    for c in cronograma:
        hh_dia_turma[c["Dia"]][c.get("Turma", "")] += float(c.get("HH", 0))
    rows = []
    for dia in range(1, dias_simulado + 1):
        row = {"Dia": dia, "Semana": int(math.ceil(dia / 5.0))}
        hh_total_dia = 0.0
        cap_total_dia = 0.0
        for tn in turma_nomes:
            hh = hh_dia_turma[dia].get(tn, 0.0)
            cap = turma_ops.get(tn, 0) * jornada
            pct = (hh / cap * 100) if cap > 0.01 else 0.0
            row[f"{tn}_HH"] = round(hh, 2)
            row[f"{tn}_Cap"] = round(cap, 2)
            row[f"{tn}_Uso%"] = round(pct, 1)
            if pct >= 90:
                row[f"{tn}_Status"] = "ALTO"
            elif pct >= 50:
                row[f"{tn}_Status"] = "MEDIO"
            elif pct > 0.01:
                row[f"{tn}_Status"] = "BAIXO"
            else:
                row[f"{tn}_Status"] = "OCIOSO"
            hh_total_dia += hh
            cap_total_dia += cap
        row["Total_HH"] = round(hh_total_dia, 2)
        row["Total_Cap"] = round(cap_total_dia, 2)
        row["Total_Uso%"] = round((hh_total_dia / cap_total_dia * 100) if cap_total_dia > 0.01 else 0.0, 1)
        rows.append(row)
    return pd.DataFrame(rows)


def _df_crono_operacional(df_crono):
    """Remove colunas monetarias do cronograma para export operacional."""
    drop = [c for c in ("Custo_MO",) if c in df_crono.columns]
    return df_crono.drop(columns=drop, errors="ignore")


def _escrever_cronograma_e_cascata(writer, df_crono_op, df_timeline, sheet_name="CRONOGRAMA_E_CASCATA"):
    """
    Cronograma e timeline na mesma folha (linha em branco entre blocos).
    Retorna a linha 1-based do cabecalho da timeline (para colorir Fase), ou None.
    """
    df_crono_op.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
    if df_timeline is None or getattr(df_timeline, "empty", True):
        return None
    start = len(df_crono_op) + 2
    df_timeline.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start)
    return start + 1


def _aplicar_cores_timeline_excel(wb, sheet_name="TIMELINE_CASCATA", header_row=1):
    """Colorir coluna Cor_Hex como fill real na coluna da Fase (header_row = linha do cabecalho da timeline, 1-based)."""
    try:
        from openpyxl.styles import PatternFill, Font
    except ImportError:
        return
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    header = [cell.value for cell in ws[header_row]]
    if "Cor_Hex" not in header or "Fase" not in header:
        return
    idx_cor = header.index("Cor_Hex") + 1
    idx_fase = header.index("Fase") + 1
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        hex_val = str(row[idx_cor - 1].value or "BDC3C7")
        if len(hex_val) == 6:
            fill = PatternFill(start_color=hex_val, end_color=hex_val, fill_type="solid")
            row[idx_fase - 1].fill = fill
            row[idx_fase - 1].font = Font(color="FFFFFF", bold=True)


def _aplicar_cores_ocupacao_excel(wb, sheet_name="OCUPACAO_TURMAS_DIA"):
    """Colorir Status (ALTO/MEDIO/BAIXO/OCIOSO) na aba de ocupação."""
    try:
        from openpyxl.styles import PatternFill, Font
    except ImportError:
        return
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    header = [cell.value for cell in ws[1]]
    status_cols = [i for i, h in enumerate(header) if h and str(h).endswith("_Status")]
    fills = {
        "ALTO":   PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid"),
        "MEDIO":  PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid"),
        "BAIXO":  PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid"),
        "OCIOSO": PatternFill(start_color="95A5A6", end_color="95A5A6", fill_type="solid"),
    }
    font_w = Font(color="FFFFFF", bold=True)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for ci in status_cols:
            val = str(row[ci].value or "")
            if val in fills:
                row[ci].fill = fills[val]
                row[ci].font = font_w


PERFIS_DIR = os.path.join(DIR, "perfis_equipe")


def _salvar_perfil_equipe(turmas, executores, jornada, nome_perfil):
    os.makedirs(PERFIS_DIR, exist_ok=True)
    dados = {
        "nome": nome_perfil,
        "executores": executores,
        "jornada": jornada,
        "turmas": [
            {"nome": t["nome"], "operarios": t["operarios"],
             "atividades": list(t.get("atividades") or [])}
            for t in turmas
        ],
    }
    caminho = os.path.join(PERFIS_DIR, f"{_slug_ficheiro_seguro(nome_perfil)}.json")
    with open(caminho, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)
    return caminho


def _listar_perfis_equipe():
    if not os.path.isdir(PERFIS_DIR):
        return []
    out = []
    for fn in sorted(os.listdir(PERFIS_DIR)):
        if fn.endswith(".json"):
            try:
                with open(os.path.join(PERFIS_DIR, fn), "r", encoding="utf-8") as f:
                    d = json.load(f)
                out.append(d)
            except Exception:
                pass
    return out


def _carregar_perfil_equipe_menu():
    perfis = _listar_perfis_equipe()
    if not perfis:
        aviso("Nenhum perfil de equipe salvo ainda.")
        return None
    nomes = [p.get("nome", "?") for p in perfis]
    sel = selecionar("PERFIL DE EQUIPE", nomes)
    if not sel:
        return None
    for p in perfis:
        if p.get("nome") == sel:
            return p
    return None


# ──────────────────────────────────────────────
#  EQUIPE PADRAO + BATCH ORCHESTRATOR
# ──────────────────────────────────────────────

def _checkpoint_editar_template(turmas, atividades_reais_global):
    """Checkpoint before each farm: allow quick edits to team template."""
    sub()
    print(G+BL+"  CHECKPOINT — Equipe Padrao"+RS)
    for t in turmas:
        print(G+f"  - {t['nome']}: "+C+f"{t['operarios']} ops, {len(t.get('atividades',[]))} atividades"+RS)
    print(DM+"  [0] Continuar sem alterar"+RS)
    print(DM+"  [1] Editar operarios de uma turma"+RS)
    print(DM+"  [2] Adicionar nova turma"+RS)
    print(DM+"  [3] Redistribuir atividades (S/N) de uma turma"+RS)
    sub()
    op = prompt("Opcao", "0").strip()
    if op == "1":
        nomes = [t["nome"] for t in turmas]
        nm = selecionar("TURMA PARA EDITAR", nomes)
        if nm:
            for t in turmas:
                if t["nome"] == nm:
                    t["operarios"] = pedir_int(f"  Novos operarios para '{nm}'", t["operarios"])
    elif op == "2":
        nome = prompt("Nome da nova turma", f"Turma {len(turmas)+1}")
        qtd = pedir_int("Quantos operarios", 1)
        nova = {"nome": nome, "operarios": qtd, "atividades": []}
        if atividades_reais_global:
            menu_vincular_atividades_turma(nova, atividades_reais_global)
        turmas.append(nova)
        ok(f"Turma '{nome}' adicionada.")
    elif op == "3":
        nomes = [t["nome"] for t in turmas]
        nm = selecionar("TURMA PARA REDISTRIBUIR", nomes)
        if nm and atividades_reais_global:
            for t in turmas:
                if t["nome"] == nm:
                    menu_vincular_atividades_turma(t, atividades_reais_global)
    return turmas


def _recomendar_equipes_padrao(total_hh, dias_meta, cap_ep_dia, jornada, prazo_absoluto):
    """Compute how many standard-team sets are needed and return recommendation dict."""
    if cap_ep_dia <= 0.01 or dias_meta <= 0:
        return None
    hh_capacidade_ep = cap_ep_dia * dias_meta
    if hh_capacidade_ep >= total_hh:
        return {"status": "suficiente", "ep_necessarias": 1, "ep_extras": 0, "trabalhadores_extras": 0}
    ep_necessarias = math.ceil(total_hh / hh_capacidade_ep)
    ep_extras = ep_necessarias - 1
    trab_necessarios = math.ceil(total_hh / (dias_meta * jornada))
    trab_extras = max(0, trab_necessarios - int(cap_ep_dia / jornada))
    return {
        "status": "insuficiente",
        "ep_necessarias": ep_necessarias,
        "ep_extras": ep_extras,
        "trabalhadores_extras": trab_extras,
        "trab_total_necessario": trab_necessarios,
    }


def _imprimir_recomendacao_ep(rec, fazenda, prazo_absoluto):
    """Print equipe padrao recommendation for one farm."""
    if not rec:
        return
    if rec["status"] == "suficiente":
        print(G+f"  Equipe padrao SUFICIENTE para '{fazenda}'."+RS)
    else:
        print(Y+f"  Equipe padrao INSUFICIENTE para '{fazenda}'."+RS)
        if prazo_absoluto:
            print(C+f"  [SUGESTAO] +{rec['ep_extras']} equipe(s) padrao (total {rec['ep_necessarias']}) cumpririam a meta."+RS)
            print(C+f"  [ALTERNATIVA] +{rec['trabalhadores_extras']} trabalhador(es) extra(s) (total {rec['trab_total_necessario']})."+RS)


def _exportar_excel_consolidado_lote(resultados, empresa_filtro=None, nome_arquivo_micro="", extras=None):
    """Exporta workbook consolidado do lote com cascata inter-fazendas e timeline unificada."""
    if not resultados:
        return
    extras = extras or {}
    try:
        pasta = os.path.join(DIR, DOSSIER_DIRNAME)
        os.makedirs(pasta, exist_ok=True)
        emp_slug = _slug_ficheiro_seguro(empresa_filtro) if empresa_filtro else "Todas_empresas"
        nome_xlsx = f"Consolidado_SmartScheduler_{emp_slug}.xlsx"
        caminho = os.path.join(pasta, nome_xlsx)
        meta_rows = [
            {"Campo": "Empresa_filtro_EQUIPE", "Valor": empresa_filtro or "(todas)"},
            {
                "Campo": "Microplanejamento",
                "Valor": os.path.basename(nome_arquivo_micro) if nome_arquivo_micro else "",
            },
        ]
        for k, v in extras.items():
            meta_rows.append({"Campo": str(k), "Valor": str(v)})

        dias_acum_total = max(
            (int(x.get("dia_fim_acumulado", 0)) for x in resultados), default=0
        )
        dias_meta_val = int(extras.get("Dias_meta", 0) or 0)
        resumo_rows = [
            {"Metrica": "Fazendas processadas", "Valor": len(resultados)},
            {
                "Metrica": "HH total (soma)",
                "Valor": round(sum(float(x.get("total_hh", 0)) for x in resultados), 1),
            },
            {
                "Metrica": "Dias acumulados lote continuo",
                "Valor": dias_acum_total,
            },
            {
                "Metrica": "Dias meta",
                "Valor": dias_meta_val,
            },
            {
                "Metrica": "Saldo meta (dias)",
                "Valor": max(0, dias_meta_val - dias_acum_total),
            },
            {
                "Metrica": "Status meta global",
                "Valor": "DENTRO" if dias_acum_total <= dias_meta_val else "EXCEDIDO",
            },
            {
                "Metrica": "Receita total",
                "Valor": round(sum(float(x.get("receita_total", 0)) for x in resultados), 2),
            },
            {
                "Metrica": "Custo MO total",
                "Valor": round(sum(float(x.get("custo_mo_total", 0)) for x in resultados), 2),
            },
            {
                "Metrica": "Lucro operacional (soma)",
                "Valor": round(sum(float(x.get("lucro_operacional", 0)) for x in resultados), 2),
            },
        ]
        d_mec_vals = [int(x.get("dias_mecanizado") or 0) for x in resultados if x.get("dias_mecanizado")]
        if d_mec_vals:
            resumo_rows.append({"Metrica": "Dias cenario mecanizado (max)", "Valor": max(d_mec_vals)})

        rows_faz = []
        for x in resultados:
            rec = x.get("rec_ep") or {}
            rows_faz.append({
                "Fazenda": x.get("fazenda"),
                "Dias_simulado": x.get("dias_simulado"),
                "Dia_inicio_acum": x.get("dia_inicio_acumulado"),
                "Dia_fim_acum": x.get("dia_fim_acumulado"),
                "Meta_consumida_%": x.get("pct_meta_consumida"),
                "Saldo_meta_dias": x.get("saldo_meta_apos"),
                "Status_meta": x.get("status_meta_continuo"),
                "Total_HH": x.get("total_hh"),
                "Receita_total": x.get("receita_total"),
                "Custo_MO_total": x.get("custo_mo_total"),
                "Lucro_operacional": x.get("lucro_operacional"),
            })

        curva_rows = []
        for x in resultados:
            curva_rows.append({
                "Fazenda": x.get("fazenda"),
                "Dia_fim_acumulado": x.get("dia_fim_acumulado", 0),
                "Meta_dias": dias_meta_val,
                "Consumido_%": x.get("pct_meta_consumida", 0),
                "HH_acumulado": round(
                    sum(float(r.get("total_hh", 0)) for r in resultados[:resultados.index(x) + 1]), 1
                ),
            })

        crono_all_rows = []
        for x in resultados:
            offset = int(x.get("dia_inicio_acumulado", 1)) - 1
            for c in (x.get("cronograma") or []):
                row = dict(c)
                row["Dia_Lote"] = int(c.get("Dia", 0)) + offset
                row["Semana_Lote"] = int(math.ceil(row["Dia_Lote"] / 5.0))
                crono_all_rows.append(row)

        with pd.ExcelWriter(caminho, engine="openpyxl") as w:
            pd.DataFrame(meta_rows).to_excel(w, sheet_name="METADADOS", index=False)
            pd.DataFrame(resumo_rows).to_excel(w, sheet_name="RESUMO", index=False)
            pd.DataFrame(rows_faz).to_excel(w, sheet_name="CASCATA_FAZENDAS", index=False)
            pd.DataFrame(curva_rows).to_excel(w, sheet_name="CURVA_CONSUMO_META", index=False)
            if crono_all_rows:
                pd.DataFrame(crono_all_rows).to_excel(
                    w, sheet_name="CRONOGRAMA_LOTE", index=False
                )
            try:
                wb = w.book
                from openpyxl.styles import PatternFill, Font
                if "CASCATA_FAZENDAS" in wb.sheetnames:
                    ws = wb["CASCATA_FAZENDAS"]
                    header = [cell.value for cell in ws[1]]
                    if "Status_meta" in header:
                        idx_st = header.index("Status_meta") + 1
                        fills_st = {
                            "OK": PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid"),
                            "RISCO": PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid"),
                            "EXCEDIDO": PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid"),
                        }
                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                            val = str(row[idx_st - 1].value or "")
                            if val in fills_st:
                                row[idx_st - 1].fill = fills_st[val]
                                row[idx_st - 1].font = Font(color="FFFFFF", bold=True)
            except Exception:
                pass
        ok(f"Consolidado Excel exportado: {nome_xlsx}")
    except Exception as ex:
        aviso(f"Nao foi possivel exportar consolidado Excel: {ex}")


def _executar_lote_fazendas(cfg, df_scope, fazendas, empresa_filtro=None, nome_arquivo_micro=""):
    """Orchestrate all-farms batch: one-time setup, per-farm checkpoint, consolidated report."""

    # ── One-time global setup ──
    cabecalho("CONFIGURACAO GLOBAL — TODAS AS FAZENDAS")

    todas_atvs = sorted(
        {_norm_atv(x) for x in df_scope["atividade"].dropna().unique() if _norm_atv(x)},
        key=str,
    )

    # Sequence selection
    seq_cfg = cfg.get("sequencia") or {}
    _merge_sequencia_defaults(seq_cfg)
    cfg["sequencia"] = seq_cfg
    modo_seq = _selecionar_sequencia_padrao_sn(cfg, seq_cfg)

    # Bloqueio / reforco / pool
    usar_bloqueio_global = False
    if modo_seq != "personalizado":
        usar_bloqueio_global = confirmar(
            "Aplicar BLOQUEIO GLOBAL (plantio/irrigacao so iniciam quando TODO o resto zerar)?",
            default=True,
        )
    usar_reforco_automatico = confirmar("Ativar REFORCO AUTOMATICO?", default=True)
    usar_pool_pos_bloqueio = False
    if usar_bloqueio_global:
        usar_pool_pos_bloqueio = confirmar("Usar PELOTAO UNIFICADO apos liberacao global?", default=True)

    # Deadline
    prazo_meses = pedir_float("Prazo META para conclusao (meses)", 6.0)
    prazo_absoluto = confirmar(
        f"  {prazo_meses} meses e o periodo ABSOLUTO? Se sim, havera sugestoes se necessario",
        default=True,
    )
    hoje = datetime.datetime.now()
    mes_ref = pedir_int("Mes inicial (1-12)", hoje.month)
    mes_ref = max(1, min(12, int(mes_ref)))
    ano_ref = pedir_int("Ano inicial", hoje.year)
    j_def = float(cfg.get("jornada_horas") or 4.6)
    if j_def <= 0:
        j_def = 4.6
    jornada = pedir_float("Jornada efetiva diaria (horas no campo)", round(j_def, 2))

    # Team template — carregar perfil ou criar
    sub()
    print(G+BL+"  CONFIGURAR EQUIPE PADRAO"+RS)
    print(DM+"  Defina as turmas que serao reutilizadas em todas as fazendas."+RS)
    print(DM+"  Voce podera ajustar antes de cada fazenda no checkpoint.\n"+RS)

    perfil_carregado = None
    perfis_existentes = _listar_perfis_equipe()
    if perfis_existentes:
        if confirmar("Carregar perfil de equipe salvo anteriormente?", default=False):
            perfil_carregado = _carregar_perfil_equipe_menu()

    if perfil_carregado:
        turmas = [
            {"nome": t["nome"], "operarios": t["operarios"],
             "atividades": list(t.get("atividades") or [])}
            for t in perfil_carregado.get("turmas", [])
        ]
        executores = perfil_carregado.get("executores", sum(t["operarios"] for t in turmas))
        ok(f"Perfil '{perfil_carregado['nome']}' carregado: {executores} executores, {len(turmas)} turma(s).")
        for t in turmas:
            print(G+f"  - {t['nome']}: "+C+f"{t['operarios']} ops, {len(t.get('atividades', []))} atividades"+RS)
        if confirmar("Editar este perfil antes de usar?", default=False):
            for turma in turmas:
                menu_vincular_atividades_turma(turma, todas_atvs)
    else:
        colab_total = pedir_int("Tamanho TOTAL da equipe padrao HOJE", 10)
        supervisores = pedir_int("Quantos LIDERES (nao executam)", 1, allow_zero=True)
        executores = colab_total - supervisores
        if executores <= 0:
            erro("Precisa de pelo menos 1 executor.")
            return

        turmas = []
        restantes = executores
        while restantes > 0:
            print(G+f"  Operarios disponiveis: {restantes}"+RS)
            nome_turma = prompt("Nome da turma", f"Turma {len(turmas)+1}")
            def_pad = min(restantes, max(1, restantes // 2 or restantes))
            qtd = pedir_int(f"  Quantos operarios na turma '{nome_turma}'", def_pad)
            if qtd > restantes:
                qtd = restantes
            turmas.append({"nome": nome_turma, "operarios": qtd, "atividades": []})
            restantes -= qtd
            if restantes > 0:
                if not confirmar(f"Criar outra turma? ({restantes} restantes)", default=True):
                    turmas.append({"nome": "Geral", "operarios": restantes, "atividades": []})
                    restantes = 0

        sub()
        print(G+BL+"  VINCULAR ATIVIDADES (usa todas as atividades do escopo)"+RS+"\n")
        for turma in turmas:
            menu_vincular_atividades_turma(turma, todas_atvs)

    if confirmar("Salvar este perfil de equipe para reusar depois?", default=False):
        nome_p = prompt("Nome do perfil", "padrao")
        cam_p = _salvar_perfil_equipe(turmas, executores, jornada, nome_p)
        ok(f"Perfil salvo: {cam_p}")

    sub()
    print(G+BL+"  LOTE: TEMPLATE vs MICRO (lacunas)"+RS)
    print(DM+"  Template estreito (ex. so irrigacao): outras demandas podem ficar sem turma."+RS)
    print(DM+"  Com N, a turma especializada nao recebe tarefas que voce nao vinculou no modelo."+RS)
    preencher_orfas_template = confirmar(
        "  Por fazenda: distribuir automaticamente demandas sem turma para a turma com mais operarios?",
        default=False,
    )

    cap_ep_dia = float(executores) * float(jornada)
    dias_meta = dias_uteis_no_periodo(mes_ref, ano_ref, prazo_meses)

    ctx_base = {
        "modo_seq": modo_seq,
        "usar_bloqueio_global": usar_bloqueio_global,
        "usar_reforco_automatico": usar_reforco_automatico,
        "usar_pool_pos_bloqueio": usar_pool_pos_bloqueio,
        "prazo_meses": prazo_meses,
        "mes_ref": mes_ref,
        "ano_ref": ano_ref,
        "jornada": jornada,
        "executores": executores,
        "turmas": turmas,
        "penalidade": 1.0,
        "preencher_orfas_template": preencher_orfas_template,
    }

    # ── Per-farm loop (lote continuo) ──
    resultados = []
    dias_acumulados = 0
    for i_f, fz in enumerate(fazendas, 1):
        linha()
        print(C+BL+f"  [{i_f}/{len(fazendas)}] FAZENDA: {fz}"+RS)
        if prazo_absoluto:
            saldo_pre = dias_meta - dias_acumulados
            pct_consumido = (dias_acumulados / dias_meta * 100) if dias_meta > 0 else 0.0
            print(
                DM + f"  Meta: {dias_meta} dias | Consumido: {dias_acumulados} dias "
                f"({pct_consumido:.0f}%) | Saldo: {saldo_pre} dias" + RS
            )
            if pct_consumido >= 100:
                print(Y + BL + "  !! META GLOBAL JA EXCEDIDA antes desta fazenda !!" + RS)
            elif pct_consumido >= 80:
                print(Y + f"  ! Atencao: {pct_consumido:.0f}% da meta ja consumida." + RS)
        linha()

        if i_f > 1:
            turmas = _checkpoint_editar_template(turmas, todas_atvs)
            ctx_base["turmas"] = turmas
            ctx_base["executores"] = sum(t["operarios"] for t in turmas)

        r = calcular_cronograma_inteligente(
            cfg, df_scope[df_scope["fazenda"] == fz].copy(), fz,
            esperar_enter=False, ctx=dict(ctx_base),
        )
        if r:
            dias_faz = int(r.get("dias_simulado", 0))
            dia_inicio_acum = dias_acumulados + 1
            dias_acumulados += dias_faz
            r["dia_inicio_acumulado"] = dia_inicio_acum
            r["dia_fim_acumulado"] = dias_acumulados
            r["saldo_meta_apos"] = max(0, dias_meta - dias_acumulados)
            r["pct_meta_consumida"] = round(
                (dias_acumulados / dias_meta * 100) if dias_meta > 0 else 0.0, 1
            )
            if dias_acumulados > dias_meta:
                r["status_meta_continuo"] = "EXCEDIDO"
            elif dias_acumulados >= dias_meta * 0.8:
                r["status_meta_continuo"] = "RISCO"
            else:
                r["status_meta_continuo"] = "OK"

            hh_faz = float(r.get("total_hh", 0))
            rec = _recomendar_equipes_padrao(hh_faz, dias_meta, cap_ep_dia, jornada, prazo_absoluto)
            r["rec_ep"] = rec
            if rec and prazo_absoluto:
                _imprimir_recomendacao_ep(rec, fz, prazo_absoluto)

            if prazo_absoluto:
                st_lbl = r["status_meta_continuo"]
                cor_st = G if st_lbl == "OK" else (Y if st_lbl == "RISCO" else R)
                sub()
                print(cor_st + BL + f"  LOTE CONTINUO — apos '{fz}':" + RS)
                print(
                    cor_st
                    + f"  Dia {dia_inicio_acum} a {dias_acumulados} | "
                    f"Saldo: {r['saldo_meta_apos']} dias | "
                    f"Consumo: {r['pct_meta_consumida']:.0f}% | "
                    f"Status: {st_lbl}"
                    + RS
                )
            resultados.append(r)

    # ── Consolidated final report ──
    if not resultados:
        return
    linha()
    print(G+BL+"  CONSOLIDADO FINAL (TODAS AS FAZENDAS)"+RS)
    tit_cons = (
        f"Consolidado — {empresa_filtro}"
        if empresa_filtro
        else "Consolidado — todas as empresas (sem filtro EQUIPE)"
    )
    t_all = Table(title=tit_cons)
    t_all.add_column("Metrica", style="cyan")
    t_all.add_column("Valor", justify="right")
    t_all.add_row("Fazendas processadas", str(len(resultados)))
    t_all.add_row("HH total (soma)", f"{sum(float(x.get('total_hh', 0)) for x in resultados):,.1f}")
    dias_max_isolado = max(int(x.get("dias_simulado", 0)) for x in resultados)
    t_all.add_row("Dias simulados (maior fazenda isolada)", str(dias_max_isolado))
    t_all.add_row("Dias acumulados lote continuo", str(dias_acumulados))
    t_all.add_row("Meta (dias uteis)", str(dias_meta))
    if dias_meta > 0:
        saldo_final = max(0, dias_meta - dias_acumulados)
        st_final = "DENTRO" if dias_acumulados <= dias_meta else "EXCEDIDO"
        cor_final = "[green]" if st_final == "DENTRO" else "[red]"
        t_all.add_row("Saldo apos todas as fazendas", f"{cor_final}{saldo_final} dias[/]")
        t_all.add_row("Status meta global", f"{cor_final}{st_final}[/]")
    d_mec_vals = [int(x.get("dias_mecanizado") or 0) for x in resultados if x.get("dias_mecanizado")]
    if d_mec_vals:
        t_all.add_row("Dias cenario mecanizado (max)", str(max(d_mec_vals)))
        t_all.add_row(
            "Ganho mecanizado total (dias)",
            f"{sum(int(x.get('ganho_mecanizado_dias', 0)) for x in resultados):+d}",
        )
    t_all.add_row("Receita total", f"R$ {sum(float(x.get('receita_total', 0)) for x in resultados):,.2f}")
    t_all.add_row("Custo MO total", f"R$ {sum(float(x.get('custo_mo_total', 0)) for x in resultados):,.2f}")
    t_all.add_row("Lucro operacional", f"R$ {sum(float(x.get('lucro_operacional', 0)) for x in resultados):,.2f}")
    console.print(t_all)

    if prazo_absoluto:
        sub()
        print(G+BL+"  ANALISE EQUIPE PADRAO — CONSOLIDADO"+RS)
        ep_cap = sum(t["operarios"] for t in turmas)
        print(G+f"  Equipe padrao: {ep_cap} executores @ {jornada}h/dia = {cap_ep_dia:.1f} HH/dia"+RS)
        print(G+f"  Meta: {prazo_meses} meses = {dias_meta} dias uteis (ABSOLUTO)"+RS)

        t_ep = Table(title=f"Cascata de execucao — {tit_cons}")
        t_ep.add_column("Fazenda", style="cyan")
        t_ep.add_column("HH", justify="right")
        t_ep.add_column("Dias", justify="right")
        t_ep.add_column("Inicio", justify="right")
        t_ep.add_column("Fim", justify="right")
        t_ep.add_column("Meta consumida", justify="right")
        t_ep.add_column("Saldo", justify="right")
        t_ep.add_column("Status", justify="center")
        for r in resultados:
            pct = r.get("pct_meta_consumida", 0)
            st = r.get("status_meta_continuo", "?")
            if st == "OK":
                cor_st = "[green]"
            elif st == "RISCO":
                cor_st = "[yellow]"
            else:
                cor_st = "[red]"
            t_ep.add_row(
                str(r["fazenda"])[:28],
                f"{float(r.get('total_hh', 0)):,.1f}",
                str(r.get("dias_simulado", 0)),
                f"Dia {r.get('dia_inicio_acumulado', '?')}",
                f"Dia {r.get('dia_fim_acumulado', '?')}",
                f"{pct:.0f}%",
                f"{r.get('saldo_meta_apos', '?')} dias",
                f"{cor_st}{st}[/]",
            )
        hh_total_all = sum(float(x.get("total_hh", 0)) for x in resultados)
        st_global = "OK" if dias_acumulados <= dias_meta else "EXCEDIDO"
        cor_g = "[green]" if st_global == "OK" else "[red]"
        t_ep.add_row(
            "TOTAL", f"{hh_total_all:,.1f}",
            str(dias_acumulados),
            "Dia 1",
            f"Dia {dias_acumulados}",
            f"{(dias_acumulados/dias_meta*100) if dias_meta > 0 else 0:.0f}%",
            f"{max(0, dias_meta - dias_acumulados)} dias",
            f"{cor_g}{st_global}[/]",
        )
        console.print(t_ep)

    _exportar_excel_consolidado_lote(
        resultados,
        empresa_filtro=empresa_filtro,
        nome_arquivo_micro=nome_arquivo_micro,
        extras={
            "Prazo_meses": prazo_meses,
            "Meta_absoluta": prazo_absoluto,
            "Modo_sequencia": modo_seq,
            "Jornada_h": jornada,
            "Executores_equipe_padrao": sum(t["operarios"] for t in turmas),
            "Preencher_orfas_auto": preencher_orfas_template,
            "Dias_meta": dias_meta,
            "Dias_acumulados_lote": dias_acumulados,
        },
    )

    linha()
    input(DM+"\n  [ENTER para voltar ao menu] "+RS)


# ──────────────────────────────────────────────
#  V6: MODO MULTI-EQUIPES
# ──────────────────────────────────────────────

def _executar_multi_equipes(cfg, df_scope, fazendas, empresa_filtro=None, nome_arquivo_micro=""):
    """Modo avançado: N equipes independentes, cada uma com carteira de fazendas e meta própria."""
    cabecalho("MODO MULTI-EQUIPES")
    print(DM+"  Cada equipe tera sua propria configuracao, meta e carteira de fazendas."+RS)
    print(DM+"  Ao final, um consolidado comparativo mostra a situacao de cada equipe.\n"+RS)

    n_equipes = pedir_int("Quantas equipes independentes?", 2)
    if n_equipes < 1:
        aviso("Precisa de pelo menos 1 equipe.")
        return

    todas_atvs = sorted(
        {_norm_atv(x) for x in df_scope["atividade"].dropna().unique() if _norm_atv(x)},
        key=str,
    )

    seq_cfg = cfg.get("sequencia") or {}
    _merge_sequencia_defaults(seq_cfg)
    cfg["sequencia"] = seq_cfg
    modo_seq = _selecionar_sequencia_padrao_sn(cfg, seq_cfg)

    hoje = datetime.datetime.now()
    mes_ref = pedir_int("Mes inicial (1-12)", hoje.month)
    mes_ref = max(1, min(12, int(mes_ref)))
    ano_ref = pedir_int("Ano inicial", hoje.year)

    equipes_config = []
    fazendas_restantes = list(fazendas)

    for ie in range(1, n_equipes + 1):
        sub()
        print(G+BL+f"  EQUIPE {ie}/{n_equipes}"+RS)
        nome_eq = prompt(f"Nome da equipe {ie}", f"Equipe {ie}")

        prazo_eq = pedir_float(f"Prazo meta para '{nome_eq}' (meses)", 3.0)
        j_eq = pedir_float(f"Jornada diaria '{nome_eq}' (horas)", 4.3)
        exec_eq = pedir_int(f"Executores '{nome_eq}'", 10)

        perfil_carregado = None
        perfis = _listar_perfis_equipe()
        if perfis and confirmar(f"Carregar perfil de equipe para '{nome_eq}'?", default=False):
            perfil_carregado = _carregar_perfil_equipe_menu()

        if perfil_carregado:
            turmas_eq = [
                {"nome": t["nome"], "operarios": t["operarios"],
                 "atividades": list(t.get("atividades") or [])}
                for t in perfil_carregado.get("turmas", [])
            ]
            exec_eq = sum(t["operarios"] for t in turmas_eq)
            ok(f"Perfil carregado: {len(turmas_eq)} turma(s), {exec_eq} executores.")
        else:
            turmas_eq = [{"nome": nome_eq, "operarios": exec_eq, "atividades": []}]
            menu_vincular_atividades_turma(turmas_eq[0], todas_atvs)

        if not fazendas_restantes:
            aviso("Todas as fazendas ja foram atribuidas. Esta equipe ficara vazia.")
            faz_eq = []
        elif ie == n_equipes:
            faz_eq = list(fazendas_restantes)
            ok(f"Restantes ({len(faz_eq)}) atribuidas a '{nome_eq}'.")
        else:
            print(G+f"\n  Fazendas disponiveis ({len(fazendas_restantes)}):"+RS)
            for idx_f, f in enumerate(fazendas_restantes, 1):
                print(G+f"  {idx_f:3}. "+C+f+RS)
            sel_txt = prompt(
                f"Indices das fazendas para '{nome_eq}' (ex: 1,3,5-7) ou ENTER=todas restantes",
                "",
            )
            if not sel_txt.strip():
                faz_eq = list(fazendas_restantes)
            else:
                idxs = parse_intervalos_escolha(sel_txt, len(fazendas_restantes))
                faz_eq = [fazendas_restantes[i] for i in idxs]
            for f in faz_eq:
                if f in fazendas_restantes:
                    fazendas_restantes.remove(f)
            ok(f"{len(faz_eq)} fazenda(s) para '{nome_eq}'.")

        equipes_config.append({
            "nome": nome_eq,
            "prazo_meses": prazo_eq,
            "jornada": j_eq,
            "executores": exec_eq,
            "turmas": turmas_eq,
            "fazendas": faz_eq,
            "modo_seq": modo_seq,
            "mes_ref": mes_ref,
            "ano_ref": ano_ref,
        })

    all_eq_results = []
    for ec in equipes_config:
        linha()
        print(G+BL+f"  PROCESSANDO EQUIPE: {ec['nome']} ({len(ec['fazendas'])} fazendas)"+RS)
        linha()

        dias_meta_eq = dias_uteis_no_periodo(ec["mes_ref"], ec["ano_ref"], ec["prazo_meses"])
        cap_eq_dia = float(ec["executores"]) * float(ec["jornada"])
        eq_resultados = []
        dias_acum_eq = 0

        ctx_eq = {
            "modo_seq": ec["modo_seq"],
            "usar_bloqueio_global": False,
            "usar_reforco_automatico": True,
            "usar_pool_pos_bloqueio": False,
            "prazo_meses": ec["prazo_meses"],
            "mes_ref": ec["mes_ref"],
            "ano_ref": ec["ano_ref"],
            "jornada": ec["jornada"],
            "executores": ec["executores"],
            "turmas": ec["turmas"],
            "penalidade": 1.0,
            "preencher_orfas_template": True,
        }

        for fz in ec["fazendas"]:
            r = calcular_cronograma_inteligente(
                cfg, df_scope[df_scope["fazenda"] == fz].copy(), fz,
                esperar_enter=False, ctx=dict(ctx_eq),
            )
            if r:
                dias_faz = int(r.get("dias_simulado", 0))
                r["dia_inicio_acumulado"] = dias_acum_eq + 1
                dias_acum_eq += dias_faz
                r["dia_fim_acumulado"] = dias_acum_eq
                r["saldo_meta_apos"] = max(0, dias_meta_eq - dias_acum_eq)
                r["pct_meta_consumida"] = round(
                    (dias_acum_eq / dias_meta_eq * 100) if dias_meta_eq > 0 else 0.0, 1
                )
                r["status_meta_continuo"] = (
                    "EXCEDIDO" if dias_acum_eq > dias_meta_eq
                    else ("RISCO" if dias_acum_eq >= dias_meta_eq * 0.8 else "OK")
                )
                eq_resultados.append(r)

        all_eq_results.append({
            "equipe": ec["nome"],
            "executores": ec["executores"],
            "jornada": ec["jornada"],
            "prazo_meses": ec["prazo_meses"],
            "dias_meta": dias_meta_eq,
            "dias_acumulados": dias_acum_eq,
            "hh_total": sum(float(x.get("total_hh", 0)) for x in eq_resultados),
            "receita": sum(float(x.get("receita_total", 0)) for x in eq_resultados),
            "custo_mo": sum(float(x.get("custo_mo_total", 0)) for x in eq_resultados),
            "n_fazendas": len(ec["fazendas"]),
            "status": "DENTRO" if dias_acum_eq <= dias_meta_eq else "EXCEDIDO",
            "resultados_fazendas": eq_resultados,
        })

    linha()
    print(G+BL+"  CONSOLIDADO MULTI-EQUIPES"+RS)
    t_meq = Table(title="Comparativo entre equipes")
    t_meq.add_column("Equipe", style="cyan")
    t_meq.add_column("Exec.", justify="right")
    t_meq.add_column("Fazendas", justify="right")
    t_meq.add_column("HH", justify="right")
    t_meq.add_column("Dias acum.", justify="right")
    t_meq.add_column("Meta (dias)", justify="right")
    t_meq.add_column("Saldo", justify="right")
    t_meq.add_column("Status", justify="center")
    for eq in all_eq_results:
        saldo = max(0, eq["dias_meta"] - eq["dias_acumulados"])
        st = eq["status"]
        cor = "[green]" if st == "DENTRO" else "[red]"
        t_meq.add_row(
            eq["equipe"],
            str(eq["executores"]),
            str(eq["n_fazendas"]),
            f"{eq['hh_total']:,.1f}",
            str(eq["dias_acumulados"]),
            str(eq["dias_meta"]),
            f"{saldo} dias",
            f"{cor}{st}[/]",
        )
    console.print(t_meq)

    try:
        pasta = os.path.join(DIR, DOSSIER_DIRNAME)
        os.makedirs(pasta, exist_ok=True)
        emp_slug = _slug_ficheiro_seguro(empresa_filtro) if empresa_filtro else "Todas"
        nome_xlsx = f"MultiEquipes_{emp_slug}.xlsx"
        caminho = os.path.join(pasta, nome_xlsx)
        rows_eq = []
        for eq in all_eq_results:
            for r in eq["resultados_fazendas"]:
                rows_eq.append({
                    "Equipe": eq["equipe"],
                    "Fazenda": r.get("fazenda"),
                    "Dias": r.get("dias_simulado"),
                    "Dia_inicio_acum": r.get("dia_inicio_acumulado"),
                    "Dia_fim_acum": r.get("dia_fim_acumulado"),
                    "Meta_consumida_%": r.get("pct_meta_consumida"),
                    "Saldo": r.get("saldo_meta_apos"),
                    "Status": r.get("status_meta_continuo"),
                    "HH": r.get("total_hh"),
                    "Receita": r.get("receita_total"),
                    "Custo_MO": r.get("custo_mo_total"),
                })
        rows_sumario = []
        for eq in all_eq_results:
            rows_sumario.append({
                "Equipe": eq["equipe"],
                "Executores": eq["executores"],
                "Jornada": eq["jornada"],
                "Fazendas": eq["n_fazendas"],
                "HH_total": eq["hh_total"],
                "Dias_acumulados": eq["dias_acumulados"],
                "Meta_dias": eq["dias_meta"],
                "Status": eq["status"],
            })
        with pd.ExcelWriter(caminho, engine="openpyxl") as w:
            pd.DataFrame(rows_sumario).to_excel(w, sheet_name="SUMARIO_EQUIPES", index=False)
            pd.DataFrame(rows_eq).to_excel(w, sheet_name="DETALHE_POR_FAZENDA", index=False)
        ok(f"Multi-equipes exportado: {nome_xlsx}")
    except Exception as ex:
        aviso(f"Erro ao exportar multi-equipes: {ex}")

    linha()
    input(DM+"\n  [ENTER para voltar ao menu] "+RS)


# ──────────────────────────────────────────────
#  MENU PRINCIPAL
# ──────────────────────────────────────────────
def _aplicar_filtro_empresa_e_escopo(df):
    """Filtro por EQUIPE (empresa) e escopo (uma fazenda / todas). Retorna (df_filtrado, empresa ou None)."""
    tem_equipe = "equipe" in df.columns
    df_filt = df.copy()
    empresa_filtro = None
    if tem_equipe:
        raw_eq = [str(x).strip() for x in df["equipe"].dropna().tolist() if str(x).strip()]
        norm_to_raw = {}
        for e in raw_eq:
            nk = normalizar_chave(e)
            if nk and nk not in norm_to_raw:
                norm_to_raw[nk] = e
        equipes = sorted(norm_to_raw.values(), key=str)
        if equipes:
            print(G+BL+"\n  FILTRO POR EMPRESA (EQUIPE)"+RS)
            print(DM+f"  {len(equipes)} empresa(s) encontrada(s) no micro."+RS)
            equipes_disp = ["TODAS"] + equipes
            eq = selecionar("EMPRESA / EQUIPE", equipes_disp)
            if eq and eq != "TODAS":
                empresa_filtro = eq
                nk_sel = normalizar_chave(eq)
                sem_eq = df_filt["equipe"].isna() | (df_filt["equipe"].astype(str).str.strip() == "")
                n_sem = int(sem_eq.sum())
                if n_sem:
                    print(DM+f"  Excluindo {n_sem} linha(s) sem EQUIPE preenchida (nao entram no filtro por empresa)."+RS)
                df_filt = df_filt[~sem_eq]
                df_filt = df_filt[
                    df_filt["equipe"].astype(str).apply(lambda x: normalizar_chave(x.strip()) == nk_sel)
                ]
                ok(
                    f"Filtrado por equipe: {eq} ({len(df_filt)} registros, "
                    f"{df_filt['atividade'].nunique()} atividade(s), {df_filt['fazenda'].nunique()} fazenda(s))"
                )
    return df_filt, empresa_filtro


def _selecionar_talhoes_fazenda(df_faz, fazenda):
    """Permite recorte por talhao dentro da fazenda selecionada."""
    if df_faz is None or df_faz.empty:
        return df_faz, {"fazenda": fazenda, "modo_talhao": "vazio", "talhoes": []}
    if "chave" not in df_faz.columns:
        return df_faz, {"fazenda": fazenda, "modo_talhao": "sem_coluna", "talhoes": []}

    talhoes = sorted({str(x).strip() for x in df_faz["chave"].dropna().tolist() if str(x).strip()}, key=str)
    if not talhoes:
        return df_faz, {"fazenda": fazenda, "modo_talhao": "sem_talhoes", "talhoes": []}
    if len(talhoes) == 1:
        ok(f"Talhao unico na fazenda: {talhoes[0]}")
        return df_faz, {"fazenda": fazenda, "modo_talhao": "unico", "talhoes": talhoes[:]}

    print(G + BL + "\n  ESCOPO POR TALHAO" + RS)
    print(DM + f"  Fazenda: {fazenda}" + RS)
    print(DM + f"  {len(talhoes)} talhao(oes) disponivel(is)." + RS)
    op = selecionar(
        "ESCOPO DOS TALHOES",
        ["TODOS OS TALHOES", "SELECIONAR TALHOES POR LISTA", "FILTRAR TALHOES POR TEXTO"],
    )
    if not op or op == "TODOS OS TALHOES":
        return df_faz, {"fazenda": fazenda, "modo_talhao": "todos", "talhoes": talhoes[:]}

    selecionados = []
    if op == "SELECIONAR TALHOES POR LISTA":
        print(DM + "  Digite numeros separados por virgula (ex.: 1,3,7)." + RS)
        for i, t in enumerate(talhoes, 1):
            print(G + f"  [{i:2}] " + C + str(t) + RS)
        raw = prompt("Talhoes", "")
        idxs = []
        for p in str(raw).replace(";", ",").split(","):
            p = p.strip()
            if p.isdigit():
                iv = int(p)
                if 1 <= iv <= len(talhoes):
                    idxs.append(iv - 1)
        idxs = sorted(set(idxs))
        selecionados = [talhoes[i] for i in idxs]
    else:
        filtro = normalizar_chave(prompt("Texto para filtrar talhoes", ""))
        if filtro:
            selecionados = [t for t in talhoes if filtro in normalizar_chave(t)]

    if not selecionados:
        aviso("Nenhum talhao selecionado; mantendo TODOS os talhoes da fazenda.")
        return df_faz, {"fazenda": fazenda, "modo_talhao": "fallback_todos", "talhoes": talhoes[:]}

    df_sel = df_faz[df_faz["chave"].astype(str).isin(set(selecionados))].copy()
    ok(f"Escopo por talhao aplicado: {len(selecionados)} selecionado(s), {len(df_sel)} linha(s) no micro.")
    return df_sel, {"fazenda": fazenda, "modo_talhao": "parcial", "talhoes": selecionados}


def _menu_ajustar_escopo_atividades(df_faz):
    """
    Ajustes por execucao no escopo atual:
      - substituir atividade
      - remover atividade
      - adicionar atividade em talhao(es)
    """
    if df_faz is None or df_faz.empty:
        return df_faz

    if "atividade" not in df_faz.columns or "chave" not in df_faz.columns:
        aviso("Escopo sem colunas necessarias para ajuste de atividade.")
        return df_faz

    out = df_faz.copy()
    if "area_ha" not in out.columns:
        out["area_ha"] = 0.0
    if "penalidade" not in out.columns:
        out["penalidade"] = 1.0

    def _atividades():
        return sorted({str(x).strip() for x in out["atividade"].dropna().tolist() if str(x).strip()}, key=str)

    def _talhoes():
        return sorted({str(x).strip() for x in out["chave"].dropna().tolist() if str(x).strip()}, key=str)

    while True:
        atvs = _atividades()
        tls = _talhoes()
        sub()
        print(G + BL + "  AJUSTE DE ATIVIDADES (APENAS NESTA EXECUCAO)" + RS)
        print(DM + f"  Atividades no escopo: {len(atvs)} | Talhoes no escopo: {len(tls)}" + RS)
        op = selecionar(
            "OPERACAO DE AJUSTE",
            ["Substituir atividade", "Remover atividade", "Adicionar atividade", "Concluir ajustes"],
        )
        if not op or op == "Concluir ajustes":
            break

        if op == "Substituir atividade":
            if not atvs:
                aviso("Sem atividades para substituir.")
                continue
            src = selecionar("ATIVIDADE ORIGEM", atvs)
            if not src:
                continue
            dst_opt = selecionar("DESTINO", atvs + ["[DIGITAR NOVA ATIVIDADE]"])
            if not dst_opt:
                continue
            dst = prompt("Nova atividade", "") if dst_opt == "[DIGITAR NOVA ATIVIDADE]" else dst_opt
            dst = _norm_atv(dst)
            if not dst:
                aviso("Destino invalido.")
                continue
            out.loc[out["atividade"].astype(str) == str(src), "atividade"] = dst
            ok(f"Substituida atividade: '{src}' -> '{dst}'.")
            continue

        if op == "Remover atividade":
            if not atvs:
                aviso("Sem atividades para remover.")
                continue
            rm = selecionar("ATIVIDADE PARA REMOVER", atvs)
            if not rm:
                continue
            n0 = len(out)
            out = out[out["atividade"].astype(str) != str(rm)].copy()
            ok(f"Atividade removida do escopo: '{rm}' ({n0-len(out)} linha(s)).")
            continue

        if op == "Adicionar atividade":
            base_opt = selecionar("NOVA ATIVIDADE", atvs + ["[DIGITAR NOVA ATIVIDADE]"])
            if not base_opt:
                continue
            nova = prompt("Nome da atividade", "") if base_opt == "[DIGITAR NOVA ATIVIDADE]" else base_opt
            nova = _norm_atv(nova)
            if not nova:
                aviso("Atividade invalida.")
                continue
            op_t = selecionar("APLICAR EM", ["Todos os talhoes do escopo", "Talhoes por lista", "Talhoes por texto"])
            if not op_t:
                continue
            sel_talhoes = tls[:]
            if op_t == "Talhoes por lista":
                print(DM + "  Digite numeros separados por virgula (ex.: 1,3,7)." + RS)
                for i, t in enumerate(tls, 1):
                    print(G + f"  [{i:2}] " + C + str(t) + RS)
                raw = prompt("Talhoes", "")
                idxs = []
                for p in str(raw).replace(";", ",").split(","):
                    p = p.strip()
                    if p.isdigit():
                        iv = int(p)
                        if 1 <= iv <= len(tls):
                            idxs.append(iv - 1)
                idxs = sorted(set(idxs))
                sel_talhoes = [tls[i] for i in idxs]
            elif op_t == "Talhoes por texto":
                fx = normalizar_chave(prompt("Texto no talhao", ""))
                sel_talhoes = [t for t in tls if fx and fx in normalizar_chave(t)]

            if not sel_talhoes:
                aviso("Nenhum talhao selecionado para adicionar atividade.")
                continue

            area_def = float(out["area_ha"].median() or 1.0)
            area_nova = pedir_float("Area/ha para nova atividade (por talhao)", round(area_def, 2), allow_zero=True)
            pen_def = float(out["penalidade"].median() or 1.0)
            pen_nova = pedir_float("Penalidade de terreno da nova atividade", round(pen_def, 2), allow_zero=False)

            add_rows = []
            for th in sel_talhoes:
                ja = out[(out["chave"].astype(str) == str(th)) & (out["atividade"].astype(str) == str(nova))]
                if not ja.empty:
                    continue
                ref = out[out["chave"].astype(str) == str(th)].head(1)
                row = ref.iloc[0].to_dict() if not ref.empty else {}
                row["chave"] = th
                row["atividade"] = nova
                row["area_ha"] = float(area_nova)
                row["penalidade"] = float(pen_nova)
                add_rows.append(row)
            if add_rows:
                out = pd.concat([out, pd.DataFrame(add_rows)], ignore_index=True)
            ok(f"Atividade '{nova}' adicionada em {len(add_rows)} talhao(es).")

    return out


def _proximo_caminho_livre(pasta, nome_arquivo):
    """
    Retorna nome/caminho sem colisao:
    - se nao existe, usa o nome original
    - se existe, gera sufixo _v2, _v3...
    """
    base, ext = os.path.splitext(str(nome_arquivo))
    candidato_nome = str(nome_arquivo)
    candidato_path = os.path.join(pasta, candidato_nome)
    if not os.path.exists(candidato_path):
        return candidato_nome, candidato_path
    i = 2
    while True:
        candidato_nome = f"{base}_v{i}{ext}"
        candidato_path = os.path.join(pasta, candidato_nome)
        if not os.path.exists(candidato_path):
            return candidato_nome, candidato_path
        i += 1


def menu_principal(cfg, df, nome_arquivo_micro="", demo_mode=False):
    opcoes = [
        ("1", "Smart Scheduler + Dossier Financeiro"),
        ("2", "Importar Tarifas (CT_313 manual)"),
        ("3", "Normalizar CT_313 -> STG (auto)"),
        ("4", "Mapeamentos de_para (micro -> tarifa)"),
        ("5", "Trocar planilha de microplanejamento (.xlsx)"),
        ("6", "Fazendas micro vs CT (lista fazendas_ct)"),
        ("7", "Importar Planilha de Preco (PRECO_FINAL + custos)"),
        ("8", "Importar Custos Globais Brutos (sem rateio por atividade)"),
        ("9", "Rotas de Metas/Bonificacao/Equacoes (preparatorio)"),
        ("0", "Sair"),
    ]
    while True:
        cabecalho()
        nf = df["fazenda"].nunique(); nu = df["chave"].nunique(); na = df["atividade"].nunique()
        stg_existe = os.path.exists(os.path.join(DIR, STG_FILENAME))
        nt = len(cfg.get("tarifas", {}))
        print(G+f"  Base: "+C+f"{nf} fazendas  |  {nu} talhoes  |  {na} atividades"+RS)
        print(G+f"  Tarifas: "+C+f"{nt} carregadas"+G+f"  |  STG: "+C+f"{'Sim' if stg_existe else 'Nao'}"+RS)
        print(G+f"  Orcamento estrito: "+C+("Sim" if cfg.get("orcamento_estrito", True) else "Nao")+RS)
        if "equipe" in df.columns:
            eq_list = sorted(df["equipe"].dropna().unique().tolist(), key=str)
            print(G+f"  Empresas (EQUIPE): "+C+f"{len(eq_list)} ({', '.join(str(e)[:20] for e in eq_list[:5])}{'...' if len(eq_list)>5 else ''})"+RS)
        if nome_arquivo_micro:
            print(G+f"  Microplanejamento: "+C+os.path.basename(nome_arquivo_micro)+RS)
        cg = cfg.get("custos_globais", {}) or {}
        if float(cg.get("valor_direto_total", 0) or 0) > 0 or float(cg.get("valor_indireto_total", 0) or 0) > 0:
            print(
                G
                + "  Custos globais ativos: "
                + C
                + f"Direto R$ {float(cg.get('valor_direto_total', 0) or 0):,.2f} | "
                + f"Indireto R$ {float(cg.get('valor_indireto_total', 0) or 0):,.2f}"
                + RS
            )
        if demo_mode and _is_demo_micro_path(nome_arquivo_micro):
            print(Y+f"  DEMO: opcao [1] = maior fazenda do micro (municipio Ulianopolis), tarifas = CT 313."+RS)
        sub()
        for cod, desc in opcoes:
            print(G+f"  [{cod}] "+C+desc+RS)
        sub()
        v = prompt("Opcao").strip()
        if v == "1":
            if demo_mode and _is_demo_micro_path(nome_arquivo_micro):
                faz = _resolver_fazenda_demo_ulianopolis(df)
                if not faz:
                    aviso("DEMO: nenhuma fazenda com 'Ulianópolis' na coluna fazenda do micro.")
                else:
                    ok(f"DEMO: fazenda {faz}")
                    calcular_cronograma_inteligente(cfg, df[df["fazenda"] == faz].copy(), faz)
            else:
                df_scope, empresa_filtro = _aplicar_filtro_empresa_e_escopo(df)
                if df_scope is None or df_scope.empty:
                    aviso("Nenhum dado apos filtros.")
                    continue
                fazendas = sorted(df_scope["fazenda"].unique().tolist())
                if len(fazendas) == 1:
                    faz = fazendas[0]
                    ok(f"Fazenda unica no escopo: {faz}")
                    df_faz = df_scope[df_scope["fazenda"] == faz].copy()
                    df_faz, meta_escopo = _selecionar_talhoes_fazenda(df_faz, faz)
                    calcular_cronograma_inteligente(cfg, df_faz, faz, escopo_meta=meta_escopo)
                else:
                    op_faz = ["TODAS AS FAZENDAS (equipe unica)", "MULTI-EQUIPES (carteiras separadas)"] + fazendas
                    faz = selecionar("SELECIONE A FAZENDA OU MODO", op_faz)
                    if faz == "TODAS AS FAZENDAS (equipe unica)":
                        _executar_lote_fazendas(
                            cfg,
                            df_scope,
                            fazendas,
                            empresa_filtro=empresa_filtro,
                            nome_arquivo_micro=nome_arquivo_micro,
                        )
                    elif faz == "MULTI-EQUIPES (carteiras separadas)":
                        _executar_multi_equipes(
                            cfg,
                            df_scope,
                            fazendas,
                            empresa_filtro=empresa_filtro,
                            nome_arquivo_micro=nome_arquivo_micro,
                        )
                    elif faz:
                        df_faz = df_scope[df_scope["fazenda"] == faz].copy()
                        df_faz, meta_escopo = _selecionar_talhoes_fazenda(df_faz, faz)
                        calcular_cronograma_inteligente(cfg, df_faz, faz, escopo_meta=meta_escopo)
        elif v == "2":
            modulo_importar_tarifas(cfg)
        elif v == "3":
            modulo_normalizar_ct(cfg)
        elif v == "4":
            modulo_mapeamentos_de_para(cfg, df)
        elif v == "5":
            p = selecionar_arquivo("NOVO MICROPLANEJAMENTO (.xlsx)")
            if p:
                ndf = carregar_planilha_microplanejamento(cfg, caminho=p, modo_auto=True)
                if ndf is None:
                    aviso("Nao foi possivel carregar automaticamente. Tente de novo sem modo_auto (o app pedira colunas).")
                    ndf = carregar_planilha_microplanejamento(cfg, caminho=p, modo_auto=False)
                if ndf is not None:
                    df = ndf
                    nome_arquivo_micro = p
                    cfg["arquivo_micro"] = os.path.basename(p)
                    salvar_config(cfg)
                    atividades_reais = sorted(
                        str(x).strip() for x in df["atividade"].dropna().unique() if str(x).strip()
                    )
                    novos = aplicar_depara_padrao_exame(cfg, atividades_reais)
                    ok(
                        f"Micro atualizado: {os.path.basename(p)} | {len(df)} registros | "
                        f"{df['fazenda'].nunique()} fazendas | de_para +{novos} novos mapeamentos."
                    )
                    if demo_mode and _is_demo_micro_path(p):
                        n = garantir_fazenda_ulianopolis_no_ct(cfg, df)
                        if n:
                            salvar_config(cfg)
                            ok(f"DEMO: +{n} fazenda(s) em fazendas_ct.")
                    aviso_fazendas_micro_sem_cadastro_ct(cfg, df)
                    input(DM+"  [ENTER] "+RS)
        elif v == "6":
            modulo_validar_fazendas_ct(cfg, df)
        elif v == "7":
            modulo_importar_precos_contrato(cfg)
        elif v == "8":
            modulo_importar_custos_globais_brutos(cfg)
        elif v == "9":
            modulo_rotas_metas_bonus(cfg)
        elif v == "0":
            print(G+"\n  Sistema encerrado.\n"+RS); break
        else:
            aviso("Opcao invalida.")

def main():
    demo = _is_demo_mode()
    beta = _is_beta_mode()
    legacy = _is_legacy_mode()
    sub_titulo = (
        f"DEMO Ulianópolis ({DEMO_MICRO_SOURCE_FILENAME} -> {DEMO_MICRO_FILENAME} + CT 313)"
        if demo else ""
    )
    if beta:
        if sub_titulo:
            sub_titulo += " | PADRAO"
        else:
            sub_titulo = "PADRAO - carga robusta micro + comparativos operacionais"
    if legacy:
        if sub_titulo:
            sub_titulo += " | LEGACY"
        else:
            sub_titulo = "LEGACY - comportamento anterior sem comparativos padrao"
    cabecalho(sub_titulo)
    print(DM+"  Inicializando sistema...\n"+RS)
    cfg = carregar_config()
    salvar_config(cfg)

    if demo:
        rebuilt = reconstruir_demo_ulianopolis_a_partir_da_fonte()
        if rebuilt:
            ok(
                f"DEMO: {DEMO_MICRO_FILENAME} atualizado a partir de {DEMO_MICRO_SOURCE_FILENAME} "
                f"({rebuilt[0]} linhas, {rebuilt[1]} atividades unicas)."
            )
        micro_padrao = os.path.join(DIR, DEMO_MICRO_FILENAME)
        if not os.path.exists(micro_padrao):
            erro(
                f"Modo DEMO: coloque {DEMO_MICRO_SOURCE_FILENAME} (gera {DEMO_MICRO_FILENAME}) "
                f"ou o proprio {DEMO_MICRO_FILENAME} em:\n  {DIR}"
            )
            sys.exit(1)
    else:
        micro_padrao = _find_default_micro_path(cfg)
    ct_padrao = _find_default_ct_path()
    contrato_cfg = cfg.get("precos_contrato", {})
    usar_contrato = bool(contrato_cfg and contrato_cfg.get("arquivo") and len(cfg.get("tarifas", {})) > 0)

    if usar_contrato:
        ok(f"Preco de contrato ativo: {contrato_cfg.get('arquivo')} | {len(cfg.get('tarifas', {}))} tarifas")
    elif ct_padrao:
        try:
            stg_path, n, custo_h = normalizar_ct313(ct_padrao)
            if stg_path and n > 0:
                cfg["tarifas"] = carregar_stg_tarifas(stg_path)
                cfg["custo_hora_tf"] = round(custo_h, 4)
                salvar_config(cfg)
                ok(f"CT auto: {os.path.basename(ct_padrao)} -> {n} tarifas | custo/h TF = R${custo_h:.2f}")
        except Exception as ex:
            aviso(f"Falha no auto-carregamento CT: {ex}")

    if micro_padrao:
        df = carregar_planilha_microplanejamento(cfg, caminho=micro_padrao, modo_auto=True)
        if df is None:
            aviso("Falha no auto-carregamento do micro padrao; abrindo modo manual.")
            df = carregar_planilha_microplanejamento(cfg)
    else:
        df = carregar_planilha_microplanejamento(cfg)

    if df is not None:
        if micro_padrao:
            cfg["arquivo_micro"] = os.path.basename(micro_padrao)
            salvar_config(cfg)
        atividades_reais = sorted(str(x).strip() for x in df["atividade"].dropna().unique() if str(x).strip())
        novos = aplicar_depara_padrao_exame(cfg, atividades_reais)
        if demo and micro_padrao and _is_demo_micro_path(micro_padrao):
            n = garantir_fazenda_ulianopolis_no_ct(cfg, df)
            if n:
                salvar_config(cfg)
                ok(f"DEMO: +{n} fazenda(s) em fazendas_ct.")
        aviso_fazendas_micro_sem_cadastro_ct(cfg, df)
        dp = {k: v for k, v in cfg.get("de_para", {}).items() if not str(k).startswith("_")}
        ok(f"{len(df)} registros | "
           f"{df['fazenda'].nunique()} fazendas | {df['chave'].nunique()} talhoes | "
           f"{len(dp)} de_para mapeados ({novos} novos)")
        input(DM+"  [ENTER para continuar] "+RS)
        menu_principal(cfg, df, micro_padrao or "", demo_mode=demo)
    else:
        aviso("Nenhuma planilha selecionada.")

if __name__ == "__main__":
    main()
